"""
Unit tests for the MF module pure logic (Phase 5) — no DB/Redis.

Covers: CAS parse normalization (ISIN filter, txn parse), SHA-256 dedup hash,
report assembly (disclosure injected, NO unified_score numeric), scoring-bridge
FactorInputs mapping (no risk_profile), holdings→snapshot mapping, and that the
upload route is wired to the mf_analytics consent gate (B20).
"""

from __future__ import annotations

import enum
import io
import re
from dataclasses import fields
from datetime import date

import pytest
from openpyxl import Workbook

from dhanradar.mf.cas import CasParseError, classify_cas_failure, parse_cas
from dhanradar.mf.schemas import FundReportItem, PortfolioReport
from dhanradar.mf.scoring_bridge import FundSignals, to_factor_inputs
from dhanradar.mf.service import assemble_report, cas_sha256, dedup_key
from dhanradar.tasks.mf import (
    _drop_over_covered_funds,
    _extract_sebi_row,
    _fetch_parse_upsert_files,
    _normalize_fraction_weight_groups,
    _parse_sebi_xlsx,
    _pick_canonical_plan_isin,
    _process_amc_paginated_query,
    parsed_to_snapshot_holdings,
)


# --- CAS parse ---------------------------------------------------------------
def _fake_cas(_path, _password):
    # Transactions are in casparser STATEMENT convention: purchase amounts are
    # positive (amount sign follows units sign). parse_cas must normalize them to
    # INVESTOR convention (purchases → negative outflow). See B65.
    return {
        "folios": [
            {"folio": "F1", "schemes": [
                {"isin": "INF001", "amfi": "100", "scheme": "Big Cap", "close": 100.0,
                 "valuation": {"nav": 50.0, "value": 5000.0, "cost": 4000.0, "date": "2026-06-01"},
                 "transactions": [{"date": "2024-01-01", "amount": 4000.0, "type": "PURCHASE"}]},
                {"isin": "", "scheme": "Not an MF row"},  # no ISIN → skipped
            ]},
        ]
    }


def _fake_cdsl_cas(_path, _password):
    """Simulates casparser 1.1.0 CDSL CAS output (accounts structure)."""
    return {
        "file_type": "CDSL",
        "accounts": [
            {
                "name": "Test Account", "type": "CDSL Demat Account",
                "mutual_funds": [
                    {
                        "isin": "INF846K01K35",
                        "amfi": "120505",
                        "name": "AXIS AMC LTD#AXIS MF-AXIS SMALL CAP FUND-DIRECT GROWTH",
                        "balance": 214.1,
                        "nav": 123.36,
                        "value": 26411.38,
                        "total_cost": 20000.0,
                        "folio": "12345678",
                    },
                    {
                        "isin": "",  # no ISIN — skipped
                        "name": "EQUITY SHARES",
                        "balance": 100,
                        "amfi": None,
                    },
                ],
            }
        ],
    }


def test_parse_cas_walks_folios_and_skips_non_isin():
    holdings, _ = parse_cas("x.pdf", "pw", reader=_fake_cas)
    assert len(holdings) == 1  # the no-ISIN row was skipped
    h = holdings[0]
    assert h.isin == "INF001" and h.folio_number == "F1"
    assert h.units == 100.0 and h.value == 5000.0 and h.cost == 4000.0
    assert h.amfi_code == "100" and h.as_of_date == date(2026, 6, 1)
    assert len(h.txns) == 1 and h.txns[0].amount == -4000.0


def test_parse_cas_failure_raises():
    import pytest

    from dhanradar.mf.cas import CasParseError

    def _boom(_p, _pw):
        raise ValueError("bad password")

    with pytest.raises(CasParseError):
        parse_cas("x.pdf", "wrong", reader=_boom)


def test_parse_cas_wrong_password_classifies_as_incorrect_password():
    """A casparser IncorrectPasswordError (the founder-reported prod case, 2026-07-03)
    must classify to 'incorrect_password' — never the opaque 'parse_failed' catch-all —
    so the FE can show a specific, actionable message instead of a raw code."""
    import pytest
    from casparser.exceptions import IncorrectPasswordError

    def _wrong_password(_p, _pw):
        raise IncorrectPasswordError("Incorrect PDF password!")

    with pytest.raises(CasParseError) as exc_info:
        parse_cas("x.pdf", "wrong", reader=_wrong_password)
    assert classify_cas_failure(exc_info.value) == "incorrect_password"


def test_classify_cas_failure_unreadable_file():
    """Corrupt/wrong-format failures (unsupported extension, empty statement, a raw
    casparser CASParseError) classify to 'unreadable_file', distinct from a wrong password."""
    assert classify_cas_failure(
        CasParseError("Unsupported file type '.docx'. Upload a CAS PDF.")
    ) == "unreadable_file"
    assert classify_cas_failure(CasParseError("No data rows found in x.txt")) == "unreadable_file"
    assert classify_cas_failure(
        CasParseError("CASParseError: Unhandled error while opening PDF: bad")
    ) == "unreadable_file"


def test_classify_cas_failure_falls_back_to_parse_failed():
    """An unrecognised failure mode keeps the generic 'parse_failed' fallback — the FE's
    catch-all copy — rather than guessing."""
    assert classify_cas_failure(CasParseError("TypeError: something odd")) == "parse_failed"


def test_parse_cas_normalises_pydantic_model_output():
    """casparser >= 1.0 returns a CASData pydantic MODEL for output="dict"
    (0.7.x returned a dict). parse_cas must model_dump() it before the dict-walk,
    else every successful parse 500s on `model.get(...)`. Regression guard."""

    class _FakeCasData:
        def model_dump(self, mode="python"):
            return _fake_cas("x.pdf", "pw")

    holdings, _ = parse_cas("x.pdf", "pw", reader=lambda _p, _pw: _FakeCasData())
    assert len(holdings) == 1  # walked the model_dump() output, not crashed
    assert holdings[0].isin == "INF001" and holdings[0].value == 5000.0


def test_parse_cas_handles_cdsl_accounts_structure():
    """casparser 1.1.0 CDSL CAS uses accounts[].mutual_funds[] not folios[].
    parse_cas must walk the accounts path when folios is absent/empty."""
    holdings, _ = parse_cas("x.pdf", "pw", reader=_fake_cdsl_cas)
    assert len(holdings) == 1  # the no-ISIN entry was skipped
    h = holdings[0]
    assert h.isin == "INF846K01K35"
    assert h.amfi_code == "120505"
    assert h.units == 214.1
    assert h.nav == 123.36
    assert h.value == 26411.38
    assert h.cost == 20000.0
    assert h.folio_number == "12345678"
    assert h.txns == []  # CDSL has no transaction history
    # Scheme name should not contain the "AXIS AMC LTD#" prefix
    assert "AXIS AMC LTD" not in h.scheme_name
    assert "AXIS" in h.scheme_name  # still mentions AXIS fund


# --- §39.4 statement-period extraction ----------------------------------------


def test_parse_cas_extracts_statement_period():
    """casparser's PDF statement_period (dd-Mmm-yyyy strings, model_dump field name `from_`)
    parses into ParsedCasIdentity.stmt_from/stmt_to."""

    def _reader(_path, _password):
        d = _fake_cas(_path, _password)
        d["statement_period"] = {"from_": "01-Apr-2023", "to": "30-Jun-2023"}
        return d

    _, identity = parse_cas("x.pdf", "pw", reader=_reader)
    assert identity.stmt_from == date(2023, 4, 1)
    assert identity.stmt_to == date(2023, 6, 30)


def test_parse_cas_statement_period_absent_is_none():
    """No statement_period in the raw output (or an empty one) → stmt_from/stmt_to stay None —
    §39.4 says never guess."""
    _, identity = parse_cas("x.pdf", "pw", reader=_fake_cas)
    assert identity.stmt_from is None
    assert identity.stmt_to is None


# --- dedup hash --------------------------------------------------------------
def test_cas_sha256_is_deterministic():
    assert cas_sha256(b"abc") == cas_sha256(b"abc")
    assert cas_sha256(b"abc") != cas_sha256(b"abd")
    assert len(cas_sha256(b"abc")) == 64


def test_dedup_key_is_namespaced_per_user_and_portfolio():
    # Same CAS bytes → DIFFERENT dedup keys across users AND across a user's
    # portfolios (no cross-user leak; the same statement can go to two portfolios).
    h = cas_sha256(b"same-cas-bytes")
    assert dedup_key("userA", "pf1", h) != dedup_key("userB", "pf1", h)  # cross-user
    assert dedup_key("userA", "pf1", h) != dedup_key("userA", "pf2", h)  # cross-portfolio
    assert dedup_key("userA", "pf1", h) == dedup_key("userA", "pf1", h)  # stable
    assert "userA" in dedup_key("userA", "pf1", h) and h in dedup_key("userA", "pf1", h)


class _FakeRedis:
    def __init__(self) -> None:
        self._d: dict[str, str] = {}

    async def get(self, k):
        return self._d.get(k)

    async def set(self, k, v, ex=None):
        self._d[k] = v

    async def delete(self, k):
        self._d.pop(k, None)

    async def exists(self, k):
        return 1 if k in self._d else 0


async def test_dedup_clear_removes_record_so_reupload_reprocesses():
    """A re-upload after a failed/stuck job must NOT be deduped to the dead job:
    the upload route drops the stale dedup record (only a `done` job short-circuits).
    dedup_clear is the primitive that enables that reprocess path."""
    from dhanradar.mf import service

    r = _FakeRedis()
    await service.dedup_record(r, "u1", "pf1", "hash1", "job-old")
    assert await service.dedup_lookup(r, "u1", "pf1", "hash1") == "job-old"
    await service.dedup_clear(r, "u1", "pf1", "hash1")
    assert await service.dedup_lookup(r, "u1", "pf1", "hash1") is None


async def test_can_return_existing_requires_done_and_cached_report():
    """Regression (TTL gap): the dedup key lives 24h (`_DEDUP_TTL`) but the report
    cache only 2h (`_REPORT_TTL`). A re-upload in that 22h gap must NOT short-circuit
    to the done job whose report has EXPIRED — doing so bounced the user to a
    GET /report 404 ('report_expired'), the very 'dead job' the dedup self-heal was
    meant to prevent. Short-circuit ONLY when the job is done AND its report is still
    cached; otherwise fall through to reprocess the freshly-uploaded bytes."""
    from dhanradar.mf import service

    r = _FakeRedis()
    job = "job-done-1"
    # done, but report cache has expired (absent) → must NOT short-circuit.
    assert await service.can_return_existing(r, "done", job) is False
    # report still cached → short-circuit is allowed.
    await r.set(f"{service._REPORT_PREFIX}{job}", "{}")
    assert await service.can_return_existing(r, "done", job) is True
    # non-done statuses never short-circuit, even if some report key exists.
    await r.set(f"{service._REPORT_PREFIX}job-x", "{}")
    assert await service.can_return_existing(r, "failed", "job-x") is False
    assert await service.can_return_existing(r, None, "job-x") is False


# --- report assembly: disclosure injected, NO numeric score ------------------
def test_report_has_no_unified_score_field():
    item_fields = {f for f in FundReportItem.model_fields}
    report_fields = {f for f in PortfolioReport.model_fields}
    assert "unified_score" not in item_fields  # no numeric score to the client
    assert "unified_score" not in report_fields
    assert "verb_label" in item_fields and "confidence_band" in item_fields


def test_assemble_report_injects_disclosure_and_strips_score():
    report = assemble_report(
        job_id="J1", status="done",
        snapshot={"total_invested": 4000.0, "current_value": 5000.0, "xirr_pct": 12.0,
                  "category_allocation": {"Equity": 100.0}, "overlap_matrix": {}},
        funds=[{"isin": "INF001", "scheme_name": "Big Cap", "folio_number": "F1", "units": 100.0,
                "invested_amount": 4000.0, "current_value": 5000.0,
                "verb_label": "on_track", "confidence_band": "medium",
                "contributing_signals": ["category match"], "contradicting_signals": []}],
        model_version="v1", generated_at="2026-06-06T00:00:00Z",
    )
    assert report.disclosure and report.not_advice == "NOT_ADVICE"
    assert report.current_value == 5000.0 and report.xirr_pct == 12.0  # user's own facts OK
    assert report.funds[0].verb_label == "on_track"
    # The serialized report must not contain a numeric unified score anywhere.
    assert "unified_score" not in report.model_dump_json()


# --- scoring bridge ----------------------------------------------------------
def test_fund_signals_exclude_risk_profile():
    names = {f.name for f in fields(FundSignals)}
    assert "risk_profile" not in names and "user_id" not in names


def test_to_factor_inputs_maps_axes_and_labels():
    fi = to_factor_inputs(FundSignals(isin="INF001", quality=70.0, outperform_1y=True,
                                      contributing=["a", "b"]))
    assert fi.instrument_type == "mf" and fi.identifier == "INF001"
    assert len(fi.axes) == 5  # all five axes present (missing values → None subfactor)
    assert fi.label_signals.outperform_1y is True
    # FactorInputs itself never carries a risk_profile (engine non-neg #3).
    assert not hasattr(fi, "risk_profile")


# --- holdings → snapshot mapping --------------------------------------------
def test_parsed_to_snapshot_applies_nav_and_builds_cashflows():
    holdings, _ = parse_cas("x.pdf", None, reader=_fake_cas)
    snap_holdings = parsed_to_snapshot_holdings(holdings, nav_map={"INF001": 55.0})
    h = snap_holdings[0]
    assert h.current_value == 100.0 * 55.0  # NAV applied (units × latest NAV)
    assert h.invested_amount == 4000.0
    # cashflows = the purchase txn + a terminal positive current-value flow
    assert any(cf.amount < 0 for cf in h.cashflows) and any(cf.amount > 0 for cf in h.cashflows)


def test_parsed_to_snapshot_falls_back_to_cas_value_without_nav():
    holdings, _ = parse_cas("x.pdf", None, reader=_fake_cas)
    snap_holdings = parsed_to_snapshot_holdings(holdings, nav_map={})  # no NAV feed
    assert snap_holdings[0].current_value == 5000.0  # CAS-reported valuation used


def test_parsed_to_snapshot_fills_category_from_map():
    # Holding category is filled from the mf_funds master so the portfolio
    # category-allocation + per-fund Category column are real (else every holding
    # buckets as "uncategorized" → a meaningless 100% donut).
    holdings, _ = parse_cas("x.pdf", None, reader=_fake_cas)
    filled = parsed_to_snapshot_holdings(
        holdings, category_map={"INF001": "Equity Scheme - Small Cap Fund"}
    )
    assert filled[0].category == "Equity Scheme - Small Cap Fund"
    # An ISIN absent from the master stays honestly uncategorized (not guessed).
    bare = parsed_to_snapshot_holdings(holdings, category_map={})
    assert bare[0].category == "uncategorized"


# --- consent gate wiring (B20) ----------------------------------------------
def test_upload_route_is_consent_gated_for_mf_analytics():
    from dhanradar.mf import router as mf_router

    assert mf_router._require_mf_consent.purpose == "mf_analytics"


# --- reap_stuck_cas_jobs unit tests -----------------------------------------
# Helpers shared across the three reaper tests.

class _ReaperRow:
    """Mimics a SQLAlchemy Row returned by the reaper SELECT."""

    def __init__(self, job_id, user_id, portfolio_id, source_hash):
        self.job_id = job_id
        self.user_id = user_id
        self.portfolio_id = portfolio_id
        self.source_hash = source_hash


class _ReaperResult:
    def __init__(self, rows):
        self._rows = rows

    def all(self):
        return self._rows


def _make_reaper_session(rows):
    """Return an async context-manager session that yields `rows` on every execute."""
    execute_calls: list = []

    class _Sess:
        async def execute(self, stmt):
            execute_calls.append(stmt)
            return _ReaperResult(rows)

        async def commit(self):
            pass

        async def __aenter__(self):
            return self

        async def __aexit__(self, *_):
            pass

    return _Sess(), execute_calls


async def test_reap_stuck_cas_jobs_marks_old_queued_job_failed():
    """An old queued job (> 10 min, completed_at IS NULL) must be reaped and the
    summary must report count=1.  The session must receive both a SELECT and an
    UPDATE (two execute calls)."""
    import uuid
    from unittest.mock import AsyncMock, patch

    from dhanradar.tasks.mf import _reap_stuck_cas_jobs

    old_row = _ReaperRow(
        job_id=uuid.uuid4(),
        user_id=uuid.uuid4(),
        portfolio_id=uuid.uuid4(),
        source_hash="deadbeef",
    )
    sess, calls = _make_reaper_session([old_row])
    dedup_clear_mock = AsyncMock()

    with (
        patch("dhanradar.db.admin_task_session", return_value=sess),
        patch("dhanradar.redis_client.get_redis", return_value=AsyncMock()),
        patch("dhanradar.mf.service.dedup_clear", dedup_clear_mock),
    ):
        summary = await _reap_stuck_cas_jobs()

    assert "1" in summary and "stuck" in summary
    # SELECT + UPDATE = at least 2 execute calls.
    assert len(calls) >= 2


async def test_reap_stuck_cas_jobs_does_not_touch_recent_job():
    """When the DB returns no rows (recent job does not cross the cutoff), the reaper
    must return a zero-count summary and must NOT issue an UPDATE."""
    from unittest.mock import patch

    from dhanradar.tasks.mf import _reap_stuck_cas_jobs

    sess, calls = _make_reaper_session([])  # empty → nothing to reap

    with patch("dhanradar.db.admin_task_session", return_value=sess):
        summary = await _reap_stuck_cas_jobs()

    assert "0" in summary
    # Only the SELECT should have been issued (early return after empty rows check).
    assert len(calls) == 1


async def test_reap_stuck_cas_jobs_does_not_touch_done_job():
    """A 'done' job is excluded by the WHERE status IN (...) predicate; the DB
    returns zero rows, so the reaper produces a zero-count summary."""
    from unittest.mock import patch

    from dhanradar.tasks.mf import _reap_stuck_cas_jobs

    # Simulate DB correctly excluding 'done' jobs → empty result set.
    sess, calls = _make_reaper_session([])

    with patch("dhanradar.db.admin_task_session", return_value=sess):
        summary = await _reap_stuck_cas_jobs()

    assert "0" in summary
    assert len(calls) == 1


# --- B61: AMFI batch deduplication ------------------------------------------

class _FakeNavRow:
    """Minimal NavRow stand-in — carries only the attributes the helpers read."""

    def __init__(
        self,
        isin_growth: str | None,
        isin_reinvest: str | None,
        nav_date: object,
        nav: float,
        amfi_code: str,
        scheme_name: str,
        category: str,
    ) -> None:
        self.isin_growth = isin_growth
        self.isin_reinvest = isin_reinvest
        self.nav_date = nav_date
        self.nav = nav
        self.amfi_code = amfi_code
        self.scheme_name = scheme_name
        self.category = category


def test_navrows_to_nav_upserts_deduplicates_same_isin_nav_date():
    """Duplicate (isin, nav_date) in the batch → exactly one output row, last-seen wins."""
    from datetime import date as _date

    from dhanradar.tasks.mf import _navrows_to_nav_upserts

    d = _date(2026, 6, 10)
    rows = [
        _FakeNavRow("INF001", None, d, 100.0, "A001", "Fund A", "Equity"),
        _FakeNavRow("INF001", None, d, 105.0, "A001", "Fund A", "Equity"),  # duplicate
    ]
    result = _navrows_to_nav_upserts(rows)

    assert len(result) == 1, "duplicate (isin, nav_date) must be collapsed to one row"
    assert result[0]["isin"] == "INF001"
    assert result[0]["nav"] == 105.0, "last-seen value must win"
    assert result[0]["nav_date"] == d
    assert result[0]["source"] == "amfi"


def test_navrows_to_nav_upserts_skips_none_isin():
    """Rows where both ISINs are None must be silently dropped."""
    from datetime import date as _date

    from dhanradar.tasks.mf import _navrows_to_nav_upserts

    rows = [
        _FakeNavRow(None, None, _date(2026, 6, 10), 99.0, "X", "Y", "Z"),
        _FakeNavRow("INF002", None, _date(2026, 6, 10), 50.0, "B001", "Fund B", "Debt"),
    ]
    result = _navrows_to_nav_upserts(rows)
    assert len(result) == 1
    assert result[0]["isin"] == "INF002"


def test_navrows_to_fund_upserts_deduplicates_same_isin():
    """Duplicate isin in the batch → exactly one output row, last-seen wins."""
    from datetime import date as _date

    from dhanradar.tasks.mf import _navrows_to_fund_upserts

    d = _date(2026, 6, 10)
    rows = [
        _FakeNavRow("INF003", None, d, 200.0, "C001", "Fund C v1", "Hybrid"),
        _FakeNavRow("INF003", None, d, 210.0, "C001", "Fund C v2", "Hybrid"),  # duplicate
    ]
    result = _navrows_to_fund_upserts(rows)

    assert len(result) == 1, "duplicate isin must be collapsed to one row"
    assert result[0]["isin"] == "INF003"
    assert result[0]["scheme_name"] == "Fund C v2", "last-seen scheme_name must win"


# --- FIX 2: control-char sanitization ----------------------------------------

def test_clean_text_strips_control_chars_from_scheme_name():
    """_clean_text must remove ASCII control characters and preserve real punctuation."""
    from dhanradar.mf.cas import _clean_text

    raw = "AXIS\x02FUND"          # U+0002 STX embedded in the name
    cleaned = _clean_text(raw)

    assert "\x02" not in cleaned, "control char must be removed"
    assert "AXIS" in cleaned and "FUND" in cleaned, "words must be preserved"
    assert "  " not in cleaned, "double-space must be collapsed"


def test_clean_text_preserves_legitimate_punctuation():
    """_clean_text must not alter hyphens, slashes, parentheses, or periods."""
    from dhanradar.mf.cas import _clean_text

    name = "HDFC Top-100 Fund (Direct) / Growth 3.5%"
    assert _clean_text(name) == name


def test_clean_text_handles_falsy_input():
    """_clean_text must return the original value unchanged when s is falsy."""
    from dhanradar.mf.cas import _clean_text

    assert _clean_text("") == ""


def test_parse_cas_strips_control_chars_from_scheme_name():
    """End-to-end: parse_cas must deliver a scheme_name free of control chars."""

    def _reader_with_ctrl(_path, _password):
        return {
            "folios": [
                {"folio": "F99", "schemes": [
                    {"isin": "INF999", "amfi": "999", "scheme": "AXIS\x02FUND",
                     "close": 10.0,
                     "valuation": {"nav": 50.0, "value": 500.0, "cost": 400.0,
                                   "date": "2026-06-10"},
                     "transactions": []},
                ]},
            ]
        }

    holdings, _ = parse_cas("x.pdf", "pw", reader=_reader_with_ctrl)
    assert len(holdings) == 1
    h = holdings[0]
    assert "\x02" not in h.scheme_name
    assert "AXIS" in h.scheme_name and "FUND" in h.scheme_name


# --- B65: statement→investor sign normalization ------------------------------

def test_b65_statement_purchases_normalize_to_outflows():
    """PURCHASE txns with positive statement-convention amounts must become negative
    investor-convention outflows after parse_cas normalization (B65)."""

    def _reader(_p, _pw):
        return {"folios": [{"folio": "F1", "schemes": [
            {"isin": "INF100", "amfi": "100", "scheme": "Fund A", "close": 50.0,
             "valuation": {"nav": 80.0, "value": 4000.0, "cost": 8000.0, "date": "2026-06-01"},
             "transactions": [
                 {"date": "2024-01-01", "amount": 4000.0, "type": "PURCHASE"},
                 {"date": "2024-07-01", "amount": 4000.0, "type": "PURCHASE"},
             ]},
        ]}]}

    holdings, _ = parse_cas("x.pdf", None, reader=_reader)
    txns = holdings[0].txns
    assert len(txns) == 2
    assert txns[0].amount == -4000.0
    assert txns[1].amount == -4000.0


def test_b65_all_purchase_portfolio_xirr_computable():
    """THE B65 bug repro: an all-purchase portfolio must have a computable XIRR after
    normalization (before B65, all-positive flows triggered the all-same-sign guard
    in xirr() and returned None)."""
    from dhanradar.mf.snapshot import build_snapshot

    def _reader(_p, _pw):
        return {"folios": [{"folio": "F1", "schemes": [
            {"isin": "INF200", "amfi": "200", "scheme": "Fund B", "close": 100.0,
             "valuation": {"nav": 90.0, "value": 9000.0, "cost": 8000.0, "date": "2026-06-01"},
             "transactions": [
                 {"date": "2024-01-01", "amount": 4000.0, "type": "PURCHASE"},
                 {"date": "2024-07-01", "amount": 4000.0, "type": "PURCHASE"},
             ]},
        ]}]}

    holdings, _ = parse_cas("x.pdf", None, reader=_reader)
    # current_value > invested (9000 > 8000) → XIRR should be computable and positive.
    snap_holdings = parsed_to_snapshot_holdings(holdings, nav_map={"INF200": 90.0})
    snap = build_snapshot(snap_holdings)
    # Before B65 this returned None (all purchase flows were positive → all-same-sign guard).
    assert snap.xirr_pct is not None
    assert snap.xirr_pct > 0


def test_b65_redemption_becomes_inflow():
    """REDEMPTION txns have negative amounts in statement convention (units leave).
    After normalization (negate) they must be positive investor-convention inflows."""

    def _reader(_p, _pw):
        return {"folios": [{"folio": "F1", "schemes": [
            {"isin": "INF300", "amfi": "300", "scheme": "Fund C", "close": 0.0,
             "valuation": {"nav": 90.0, "value": 0.0, "cost": 0.0, "date": "2026-06-01"},
             "transactions": [
                 {"date": "2025-01-01", "amount": -2500.0, "type": "REDEMPTION"},
             ]},
        ]}]}

    holdings, _ = parse_cas("x.pdf", None, reader=_reader)
    assert holdings[0].txns[0].amount == 2500.0


def test_b65_dividend_payout_kept_as_inflow():
    """DIVIDEND_PAYOUT is cash credited to the investor — kept as printed (positive),
    not negated."""

    def _reader(_p, _pw):
        return {"folios": [{"folio": "F1", "schemes": [
            {"isin": "INF400", "amfi": "400", "scheme": "Fund D", "close": 50.0,
             "valuation": {"nav": 10.0, "value": 500.0, "cost": 500.0, "date": "2026-06-01"},
             "transactions": [
                 {"date": "2025-03-01", "amount": 120.0, "type": "DIVIDEND_PAYOUT"},
             ]},
        ]}]}

    holdings, _ = parse_cas("x.pdf", None, reader=_reader)
    assert holdings[0].txns[0].amount == 120.0


def test_b65_internal_and_tax_rows_excluded():
    """B65 cashflow exclusion preserved + B3 ledger capture:
    - STAMP_DUTY_TAX / STT_TAX / MISC carry no external cash AND no units → fully excluded.
    - DIVIDEND_REINVEST carries reinvested UNITS but no external cash → captured for the ledger (B3)
      with amount=0, so it stays XIRR-neutral (the cashflow path filters amount==0). The PURCHASE is
      the only cashflow-bearing txn."""

    def _reader(_p, _pw):
        return {"folios": [{"folio": "F1", "schemes": [
            {"isin": "INF500", "amfi": "500", "scheme": "Fund E", "close": 100.0,
             "valuation": {"nav": 50.0, "value": 5000.0, "cost": 4000.0, "date": "2026-06-01"},
             "transactions": [
                 {"date": "2024-01-01", "amount": 4000.0, "type": "PURCHASE"},
                 {"date": "2024-01-01", "amount": 8.0,    "type": "STAMP_DUTY_TAX"},
                 {"date": "2024-01-01", "amount": 0.5,    "type": "STT_TAX"},
                 {"date": "2024-06-01", "amount": 200.0, "units": 4.0, "type": "DIVIDEND_REINVEST"},
                 {"date": "2025-01-01", "amount": 5.0,    "type": "MISC"},
             ]},
        ]}]}

    holdings, _ = parse_cas("x.pdf", None, reader=_reader)
    txns = holdings[0].txns
    by_type = {t.txn_type: t for t in txns}
    # tax/misc fully excluded; purchase (cashflow) + dividend_reinvest (ledger) survive.
    assert set(by_type) == {"purchase", "dividend_reinvest"}, by_type
    assert by_type["purchase"].amount == -4000.0
    # reinvest captured for the ledger: units kept, amount 0 (XIRR-neutral).
    assert by_type["dividend_reinvest"].amount == 0.0
    assert by_type["dividend_reinvest"].units == 4.0
    # the only cashflow-bearing (non-zero amount) txn is the purchase — B65 preserved.
    assert [t for t in txns if t.amount] == [by_type["purchase"]]


def test_b65_missing_type_defaults_to_negate():
    """A txn dict with no 'type' key (casparser 0.7.x plain-dict compatibility)
    defaults to the negate path — outflow convention (amount negated)."""

    def _reader(_p, _pw):
        return {"folios": [{"folio": "F1", "schemes": [
            {"isin": "INF600", "amfi": "600", "scheme": "Fund F", "close": 10.0,
             "valuation": {"nav": 10.0, "value": 100.0, "cost": 100.0, "date": "2026-06-01"},
             "transactions": [
                 {"date": "2024-01-01", "amount": 1000.0},  # no "type" key
             ]},
        ]}]}

    holdings, _ = parse_cas("x.pdf", None, reader=_reader)
    assert holdings[0].txns[0].amount == -1000.0


def test_b65_enum_like_type_handled():
    """casparser 1.0+ model_dump(mode="python") leaves 'type' as a str-subclass enum
    member (e.g. TransactionType.PURCHASE where .value == 'PURCHASE').
    parse_cas must handle this via getattr(.value) — not break on the enum object."""

    # Deliberately (str, Enum) — NOT StrEnum: str(member) yields
    # "_FakeType.PURCHASE" (exactly as casparser's TransactionType does), so
    # only the getattr(.value) path resolves the bare "PURCHASE". A str()-based
    # implementation would fail this test.
    class _FakeType(str, enum.Enum):  # noqa: UP042 — deliberately (str, Enum), not StrEnum: mirrors casparser's TransactionType where str(member) != .value
        PURCHASE = "PURCHASE"

    def _reader(_p, _pw):
        return {"folios": [{"folio": "F1", "schemes": [
            {"isin": "INF700", "amfi": "700", "scheme": "Fund G", "close": 10.0,
             "valuation": {"nav": 10.0, "value": 100.0, "cost": 500.0, "date": "2026-06-01"},
             "transactions": [
                 {"date": "2024-01-01", "amount": 500.0, "type": _FakeType.PURCHASE},
             ]},
        ]}]}

    holdings, _ = parse_cas("x.pdf", None, reader=_reader)
    # The getattr(.value) path must resolve PURCHASE → negate → -500.0.
    assert holdings[0].txns[0].amount == -500.0


def _b65_single_txn_reader(ttype: str | None, amount: float):
    """Build a fake reader with one txn of the given type/amount (B65 helpers)."""
    txn: dict = {"date": "2024-05-01", "amount": amount}
    if ttype is not None:
        txn["type"] = ttype

    def _reader(_p, _pw):
        return {"folios": [{"folio": "F1", "schemes": [
            {"isin": "INF800", "amfi": "800", "scheme": "Fund H", "close": 10.0,
             "valuation": {"nav": 10.0, "value": 100.0, "cost": 100.0, "date": "2026-06-01"},
             "transactions": [txn]},
        ]}]}

    return _reader


def test_b65_switch_and_reversal_types_negated():
    """Each units-bearing type is negated: switches become out/inflows that cancel
    pairwise at portfolio level; reversal rows invert their statement sign."""
    cases = [
        ("SWITCH_IN", 3000.0, -3000.0),        # outflow into the target fund
        ("SWITCH_OUT", -3000.0, 3000.0),       # inflow back to the investor
        ("SWITCH_IN_MERGER", 1500.0, -1500.0),
        ("SWITCH_OUT_MERGER", -1500.0, 1500.0),
        ("REVERSAL", -700.0, 700.0),           # cancels its original when in-window
    ]
    for ttype, statement_amount, expected in cases:
        holdings, _ = parse_cas("x.pdf", None, reader=_b65_single_txn_reader(ttype, statement_amount))
        assert holdings[0].txns[0].amount == expected, ttype


def test_b65_remaining_excluded_types():
    """TDS_TAX / SEGREGATION / UNKNOWN rows carry no usable external cash flow and
    are excluded (completes the _TXN_FLOW_EXCLUDED coverage)."""
    for ttype in ("TDS_TAX", "SEGREGATION", "UNKNOWN"):
        holdings, _ = parse_cas("x.pdf", None, reader=_b65_single_txn_reader(ttype, 50.0))
        assert holdings[0].txns == [], ttype


# --- Constituents parser: section-header / subtotal row filter --------------
# Bug repro: mf.mf_fund_constituents for INF789F01WY2 (UTI), as_of_month
# 2026-05-01, had 107 rows whose weight_pct summed to ~199.66% because
# section-header/subtotal disclosure-sheet rows were ingested as holdings.
# See docs/rca/README.md.


def _col_map(*headers: str) -> dict[str, int]:
    return {h.lower(): i for i, h in enumerate(headers)}


def test_extract_sebi_row_skips_lettered_section_header():
    """A section header like "(a) Listed/awaiting listing on Stock Exchanges" carries
    the section's own subtotal weight — must not be ingested alongside the holdings
    it summarizes."""
    col_map = _col_map("Name of Instrument", "ISIN", "% to NAV", "Market Value")
    row = ["(a)  Listed/awaiting listing on Stock Exchanges", "", "99.83", "23456.78"]
    assert _extract_sebi_row(row, col_map, "UTI Equity Fund", "UTI", date(2026, 5, 1)) is None


def test_extract_sebi_row_skips_unlisted_header():
    col_map = _col_map("Name of Instrument", "ISIN", "% to NAV", "Market Value")
    row = ["Unlisted", "", "0.17", "40.12"]
    assert _extract_sebi_row(row, col_map, "UTI Equity Fund", "UTI", date(2026, 5, 1)) is None


def test_extract_sebi_row_skips_subtotal_by_keyword():
    col_map = _col_map("Name of Instrument", "ISIN", "% to NAV", "Market Value")
    row = ["Sub Total", "", "99.83", "23456.78"]
    assert _extract_sebi_row(row, col_map, "UTI Equity Fund", "UTI", date(2026, 5, 1)) is None


def test_extract_sebi_row_skips_label_only_row_with_no_data():
    """A bare category label with no ISIN and no numbers at all is structurally
    not a holding, even when its name matches no "total"/header keyword."""
    col_map = _col_map("Name of Instrument", "ISIN", "% to NAV", "Market Value")
    row = ["Money Market Instruments", "", "", ""]
    assert _extract_sebi_row(row, col_map, "UTI Equity Fund", "UTI", date(2026, 5, 1)) is None


def test_extract_sebi_row_keeps_no_isin_row_with_real_data():
    """Cash/receivable lines legitimately carry no ISIN but do carry a weight — the
    structural no-ISIN guard must not drop these (no over-strip)."""
    col_map = _col_map("Name of Instrument", "ISIN", "% to NAV", "Market Value")
    row = ["Net Receivables/(Payables)", "", "0.42", "98.5"]
    result = _extract_sebi_row(row, col_map, "UTI Equity Fund", "UTI", date(2026, 5, 1))
    assert result is not None
    assert result["constituent_name"] == "Net Receivables/(Payables)"


def test_extract_sebi_row_strips_eq_prefix():
    """UTI writes "EQ - ABB INDIA LTD." in the Name-of-Instrument cell — strip the
    display-only instrument-type prefix so the UI shows a clean stock name."""
    col_map = _col_map("Name of Instrument", "ISIN", "% to NAV", "Market Value", "Sector")
    row = ["EQ - ABB INDIA LTD.", "INE117A01022", "5.234", "1234.56", "Capital Goods"]
    result = _extract_sebi_row(row, col_map, "UTI Equity Fund", "UTI", date(2026, 5, 1))
    assert result is not None
    assert result["constituent_name"] == "ABB INDIA LTD."
    assert result["constituent_isin"] == "INE117A01022"
    assert result["weight_pct"] == 5.234
    assert result["market_value_cr"] == 12.3456


# --- AUM extraction (root-cause fix 2026-07-05: aum_crore was 100% NULL) -----
def test_extract_sebi_row_keeps_grand_total_as_total_row():
    """NIPPON's literal "GRAND TOTAL" row must survive both the no-ISIN guard and
    the section-header drop, flagged is_total_row=True, so _upsert_constituents can
    use it as the scheme's AUM — the root cause of the 100%-NULL aum_crore bug was
    this row being unconditionally dropped before ever reaching that logic."""
    col_map = _col_map("Name of Instrument", "ISIN", "% to NAV", "Market Value")
    row = ["GRAND TOTAL", "", "100.00", "3587039.00"]
    result = _extract_sebi_row(row, col_map, "Nippon India Liquid Fund", "NIPPON", date(2026, 6, 1))
    assert result is not None
    assert result["is_total_row"] is True
    assert result["market_value_cr"] == 35870.39


def test_extract_sebi_row_keeps_net_assets_as_total_row():
    col_map = _col_map("Name of Instrument", "ISIN", "% to NAV", "Market Value")
    row = ["Net Assets", "", "100.00", "96750.00"]
    result = _extract_sebi_row(
        row, col_map, "Nippon India Conservative Hybrid Fund", "NIPPON", date(2026, 6, 1)
    )
    assert result is not None
    assert result["is_total_row"] is True
    assert result["market_value_cr"] == 967.5


def test_extract_sebi_row_keeps_tata_portfolio_total_as_total_row():
    """TATA's literal "PORTFOLIO TOTAL" row is its true scheme-level grand
    total (equivalent to NIPPON's "GRAND TOTAL") — must be kept, flagged
    is_total_row=True, so it can feed the AUM heuristic."""
    col_map = _col_map("Name of Instrument", "ISIN", "% to NAV", "Market Value")
    row = ["PORTFOLIO TOTAL", "", "99.93", "1114790.50"]
    result = _extract_sebi_row(
        row, col_map, "Tata Silver ETF Fund of Fund", "TATA", date(2026, 5, 1)
    )
    assert result is not None
    assert result["is_total_row"] is True


def test_extract_sebi_row_drops_tata_asset_class_subtotal_entirely():
    """Real production bug found live 2026-07-08 (first prod run of the TATA
    scraper, B87/B88): TATA labels EVERY asset-class subtotal as a TRAILING
    "... TOTAL" suffix ("EQUITY & EQUITY RELATED TOTAL") — the opposite shape
    of UTI's "TOTAL: ..." prefix, which the section-header regex already
    caught. Without a suffix match, this row was silently counted as a REAL
    holding for every TATA scheme, inflating weight_pct_sum to 280-400% and
    tripping `_drop_over_covered_funds`'s >105% fail-closed guard for 66/67
    real schemes in the real May-2026 file — every constituent row for the
    fund was discarded, not just the subtotal. Confirmed via a direct
    weight_pct_sum recompute against the real file: 0/67 schemes exceed 105%
    after this fix (was 66/67 before)."""
    col_map = _col_map("Name of Instrument", "ISIN", "% to NAV", "Market Value")
    row = ["EQUITY & EQUITY RELATED TOTAL", "", "99.93", "1114790.50"]
    result = _extract_sebi_row(
        row, col_map, "Tata Silver ETF Fund of Fund", "TATA", date(2026, 5, 1)
    )
    assert result is None


def test_extract_sebi_row_does_not_flag_bare_subtotal_as_total_row():
    """A bare asset-class subtotal (UTI: "TOTAL: EQUITY AND EQUITY RELATED") must NOT
    be flagged is_total_row — only "grand total"/"net assets" identify the scheme's
    true AUM. Broadening to bare "total" was tried and reverted: it caused a
    sub-category subtotal to be written as the scheme's AUM for UTI (2026-07-05)."""
    col_map = _col_map("Name of Instrument", "ISIN", "% to NAV", "Market Value")
    row = ["TOTAL:  EQUITY AND EQUITY RELATED", "", "50.00", "194304.71"]
    assert (
        _extract_sebi_row(row, col_map, "UTI - Unit Linked Insurance Plan", "UTI", date(2026, 6, 1))
        is None
    )


def test_extract_sebi_row_market_fair_value_header_variant():
    """NIPPON debt schemes header the value column "Market/Fair Value\\n( Rs. in
    Lacs)" — the embedded "/fair" breaks a plain "market value" substring match."""
    col_map = _col_map(
        "Name of Instrument", "ISIN", "% to NAV", "Market/Fair Value\n( Rs. in Lacs)"
    )
    row = ["7.18% GOI 2033", "IN0020230019", "10.00", "9433.32"]
    result = _extract_sebi_row(
        row, col_map, "Nippon India Corporate Bond Fund", "NIPPON", date(2026, 6, 1)
    )
    assert result is not None
    assert result["market_value_cr"] == pytest.approx(94.3332)


def test_extract_sebi_row_hyphenated_market_value_header_variant():
    """UTI hyphenates the value column header: "MARKET-VALUE"."""
    col_map = _col_map("Name of Instrument", "ISIN", "% to NAV", "MARKET-VALUE")
    row = ["HDFC BANK LIMITED", "INE040A01034", "5.00", "11454.10"]
    result = _extract_sebi_row(row, col_map, "UTI - Large Cap Fund", "UTI", date(2026, 6, 1))
    assert result is not None
    assert result["market_value_cr"] == 114.541


def test_extract_sebi_row_hdfc_market_fair_value_space_variant():
    """HDFC's per-scheme manual-ingest files (2026-07-06 triage) header the value
    column "Market/ Fair Value (Rs. in Lacs.)" -- WITH a space after the slash,
    which the existing no-space "market/fair value" (NIPPON) variant doesn't match."""
    col_map = _col_map(
        "ISIN",
        "Coupon (%)",
        "Name Of the Instrument",
        "Market/ Fair Value (Rs. in Lacs.)",
    )
    row = ["INE134E08MC7", "7.77", "Power Finance Corporation Ltd.^", "28016.41"]
    result = _extract_sebi_row(
        row, col_map, "HDFC Liquid Fund", "HDFC", date(2026, 6, 1)
    )
    assert result is not None
    assert result["market_value_cr"] == pytest.approx(280.1641)


def test_parse_sebi_xlsx_hdfc_per_scheme_merged_title_banner():
    """HDFC's per-scheme manual-ingest files (~88 files, 2026-07-06 triage) title
    each sheet with a scheme-name banner that openpyxl's read_only mode repeats
    across MOST (not all) of the row's columns (a merged cell), with a couple of
    unrelated trailing metadata values (e.g. "Income", "Hybrid") -- not the single
    non-empty cell the original single-scheme-row detection expected. The banner
    also carries a boilerplate SEBI scheme-TYPE disclaimer in parens that must be
    stripped so the cleaned name resolves via pg_trgm fuzzy matching."""
    banner = "HDFC Liquid Fund (An Open ended Liquid scheme)"
    wb = Workbook()
    ws = wb.active
    ws.append([banner] * 10 + ["Income", "Hybrid"])
    ws.append(["Portfolio as on 15-Jun-2026"] * 10 + [None, None])
    ws.append([None] * 12)
    ws.append(
        [
            None,
            "ISIN",
            "Coupon (%)",
            "Name Of the Instrument",
            "Industry+ /Rating",
            "Quantity",
            "Market/ Fair Value (Rs. in Lacs.)",
            "% to NAV",
        ]
    )
    ws.append(
        [
            "",
            "INE134E08MC7",
            7.77,
            "Power Finance Corporation Ltd.^",
            "CRISIL - AAA",
            28000,
            28016.41,
            0.4,
        ]
    )
    ws.append(
        ["", "INE134E08ML8", 7.55, "REC LTD", "CRISIL - AAA", 25000, 25011.43, 0.36]
    )
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "HDFC")

    assert len(rows) == 2
    assert all(r["scheme_name"] == "HDFC Liquid Fund" for r in rows)
    assert rows[0]["constituent_isin"] == "INE134E08MC7"
    assert rows[0]["as_of_month"] == date(2026, 6, 1)
    assert rows[0]["market_value_cr"] == pytest.approx(280.1641)


def test_parse_sebi_xlsx_hdfc_close_ended_a_not_an_grammar():
    """HDFC's close-ended scheme banners are grammatically correct English --
    "(A Close Ended ...)" not "(An Close Ended ...)" -- e.g. real files "HDFC
    Charity Fund for Cancer Cure" and every HDFC FMP (confirmed 2026-07-09, 13
    residual manual-ingest failures). The strip regex only matched literal "an",
    so the whole tenure/risk-rating disclaimer stayed attached to the scheme
    name and pg_trgm resolution against the bare fund name in mf_funds failed."""
    banner = (
        "HDFC Charity Fund for Cancer Cure (A Close Ended Income Scheme With "
        "Tenure 1196 Days. A Relatively High Interest Rate Risk and Relatively "
        "Low Credit Risk.)"
    )
    wb = Workbook()
    ws = wb.active
    ws.append([banner] * 10 + ["Income", "Hybrid"])
    ws.append(["Portfolio as on 15-Jun-2026"])
    ws.append([None])
    ws.append(["ISIN", "Name Of the Instrument", "Quantity", "Market/Fair Value(Rs. In Lacs)"])
    ws.append(["IN1520220097", "7.49% Gujarat SDL Mat 280926", 10000000, 10052.77])
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "HDFC")

    assert len(rows) == 1
    assert rows[0]["scheme_name"] == "HDFC Charity Fund for Cancer Cure"


def test_parse_sebi_xlsx_icici_fund_size_banner_not_scheme_name():
    """ICICI's per-scheme files (confirmed 2026-07-09, 134 residual manual-
    ingest failures) pair a "Figures as on <date>" cell with a "Fund Size Rs.
    <n> in Lakhs" cell in the same 2-value row, immediately after the real
    scheme-name banner row. "Fund Size" contains the same "fund" keyword a
    real scheme name would, so before the fix it won the single-keyword-hit
    disambiguation used for ABSL-style fund-code/name rows and silently
    overwrote the already-correctly-detected current_scheme."""
    banner = "ICICI PRUDENTIAL HOUSING OPPORTUNITIES FUND (An open ended equity scheme)"
    wb = Workbook()
    ws = wb.active
    ws.append(["ICICI Prudential Mutual Fund"])
    ws.append([banner])
    ws.append([None])
    ws.append(["Figures as on Mar 31,2026", "Fund Size Rs. 242353.39 in Lakhs"])
    ws.append([None])
    ws.append(["Company/Issuer/Instrument Name", "ISIN", "Quantity", "Exposure/Market Value (Rs. In Lakhs)"])
    ws.append(["NTPC Ltd.", "INE733E01010", 5467947, 20266.95])
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "ICICI_PRU")

    assert len(rows) == 1
    assert rows[0]["scheme_name"] == "ICICI PRUDENTIAL HOUSING OPPORTUNITIES FUND"


def test_parse_sebi_xlsx_icici_fof_disclaimer_with_embedded_newline():
    """ICICI's Fund-of-Funds scheme banners can wrap the boilerplate disclaimer
    onto a second line WITHIN the same cell (confirmed 2026-07-09, e.g. real
    file "ICICI Prudential Multi Sector Passive FOF.xlsx"). `.` never matches
    `\\n` without re.DOTALL, so the strip regex silently failed to match past
    the embedded newline, leaving the entire multi-line disclaimer attached.
    Also covers the "fof" scheme-type keyword: once the disclaimer (which
    contained "Fund of Funds", the only occurrence of the "fund" keyword) is
    correctly stripped, a bare "... FOF" name no longer contains "fund" at
    all, so "fof" must be its own accepted keyword or the correctly-stripped
    name is silently discarded."""
    banner = (
        "ICICI Prudential Multi Sector Passive FOF  (An open ended Fund of "
        "Funds scheme investing predominantly in Units of passive domestic "
        "sector/multi sector based \nEquity Oriented Exchange Traded Funds (ETFs))"
    )
    wb = Workbook()
    ws = wb.active
    ws.append(["ICICI Prudential Mutual Fund"])
    ws.append([banner])
    ws.append([None])
    ws.append(["Figures as on Mar 31,2026", "Fund Size Rs. 1000.00 in Lakhs"])
    ws.append([None])
    ws.append(["Company/Issuer/Instrument Name", "ISIN", "Quantity", "Exposure/Market Value (Rs. In Lakhs)"])
    ws.append(["NTPC Ltd.", "INE733E01010", 5467947, 20266.95])
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "ICICI_PRU")

    assert len(rows) == 1
    assert rows[0]["scheme_name"] == "ICICI Prudential Multi Sector Passive FOF"


def test_parse_sebi_xlsx_icici_fof_disclaimer_without_wrapping_parens():
    """Some ICICI FOF banners state the scheme-type disclaimer with NO
    wrapping parens of its own (confirmed 2026-07-09, real file "ICICI
    Prudential Thematic Advantage Fund (FOF).xlsx") -- the "(FOF)" paren
    belongs to the real scheme name, and "An open ended ..." simply follows
    with no parens around it. The strip regex previously required a literal
    opening paren immediately before "a|an", so this shape was never
    stripped at all."""
    banner = (
        "ICICI Prudential Thematic Advantage Fund (FOF)  An open ended fund "
        "of funds scheme investing predominantly in Sectoral / Thematic "
        "Schemes)\nschemes)"
    )
    wb = Workbook()
    ws = wb.active
    ws.append(["ICICI Prudential Mutual Fund"])
    ws.append([banner])
    ws.append([None])
    ws.append(["Figures as on Mar 31,2026", "Fund Size Rs. 1000.00 in Lakhs"])
    ws.append([None])
    ws.append(["Company/Issuer/Instrument Name", "ISIN", "Quantity", "Exposure/Market Value (Rs. In Lakhs)"])
    ws.append(["NTPC Ltd.", "INE733E01010", 5467947, 20266.95])
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "ICICI_PRU")

    assert len(rows) == 1
    assert rows[0]["scheme_name"] == "ICICI Prudential Thematic Advantage Fund (FOF)"


def test_parse_sebi_xlsx_hdfc_fmp_keyword_accepted():
    """HDFC's Fixed Maturity Plan (FMP) close-ended scheme banners correctly
    strip the "(A Close Ended ...)" disclaimer (B83, PR #508) down to a bare
    name like "HDFC FMP 1269D March 2023" -- but "fmp" alone matched none of
    the scheme-type keywords ("fund"/"scheme"/"plan"/"etf"/"index"/"growth"/
    "idcw"/"direct"/"regular"/"fof"), so the correctly-stripped candidate
    failed the final acceptance gate and `current_scheme` was never set at
    all -- a stricter failure (status='unsupported', ZERO rows extracted)
    than the zero_rows_upserted_scheme_unresolved bug B83 fixed. Confirmed
    2026-07-09 against 10 real HDFC FMP files sitting in the unsupported
    bucket despite the B83 fix already being deployed."""
    banner = (
        "HDFC FMP 1269D March 2023 (A Close Ended Income Scheme With Tenure "
        "1269 Days. A Relatively High Interest Rate Risk And Relatively Low "
        "Credit Risk)"
    )
    wb = Workbook()
    ws = wb.active
    ws.append([banner] * 10 + ["Income", "Hybrid"])
    ws.append(["Portfolio as on 15-Jun-2026"])
    ws.append([None])
    ws.append(["ISIN", "Name Of the Instrument", "Quantity", "Market/Fair Value(Rs. In Lacs)"])
    ws.append(["IN1520220097", "7.49% Gujarat SDL Mat 280926", 10000000, 10052.77])
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "HDFC")

    assert len(rows) == 1
    assert rows[0]["scheme_name"] == "HDFC FMP 1269D March 2023"


def test_parse_sebi_xlsx_tata_two_digit_year_as_of_date():
    """TATA's own per-scheme sheet banner ("Portfolio as on 31/05/26" — real
    May-2026 disclosure file, confirmed 2026-07-08, B87) uses a 2-DIGIT year in
    the DD/MM/YY slash format. The DD/MM/YYYY fallback regex previously
    required a strict 4-digit year, so this never matched -- as_of_month
    stayed None for every row with zero error, and _upsert_constituents drops
    any row with no as_of_month (a silent zero-rows-written failure even
    though the scheme name resolves fine, the same failure shape as the ICICI
    comma-format bug). `(\\d{2,4})` accepts both; a 2-digit year widens to the
    2000s."""
    wb = Workbook()
    ws = wb.active
    ws.append(["TATA SILVER ETF FUND OF FUND"])
    ws.append(["Portfolio as on 31/05/26"])
    ws.append([None])
    ws.append(["ISIN", "Name Of the Instrument", "Quantity", "Market/Fair Value(Rs. In Lacs)"])
    ws.append(["INF277KA1984", "TATA SILVER EXCHANGE TRADED FUND", 100000, 5000.0])
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "TATA")

    assert len(rows) == 1
    assert rows[0]["as_of_month"] == date(2026, 5, 1)


def test_parse_sebi_xlsx_zerodha_for_month_year_banner_no_as_on_phrase():
    """B103/Fix-4: ZERODHA's real SEBI banner (confirmed 2026-07-12 against
    all 19 live scheme files from assets.zerodhafundhouse.com) reads "MONTHLY
    PORTFOLIO STATEMENT OF ZERODHA NIFTY 100 ETF FOR JUNE 2026" -- no "as
    on"/"as of" phrase and no day-of-month anywhere in the file, so the
    phrase-gate above never even fired and as_of_month stayed None for every
    row. _upsert_constituents then silently dropped every row (`if
    row.get("as_of_month") is None: continue`) -- 0 rows written to the DB
    despite _parse_sebi_xlsx extracting every ISIN/name/weight correctly, the
    same zero-rows-written failure shape as the TATA/ICICI cases above but on
    a NEW trigger-phrase gap (the gate itself never opened, not just the
    date-format regex inside it)."""
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            None,
            None,
            "PURSUANT TO REGULATION 59A OF SECURITIES & EXCHANGE BOARD OF "
            "INDIA (MUTUAL FUNDS) REGULATIONS, 1996",
        ]
    )
    ws.append([None, None, "MONTHLY PORTFOLIO STATEMENT OF ZERODHA NIFTY 100 ETF FOR JUNE 2026"])
    ws.append([None])
    ws.append(
        [
            None,
            None,
            "Name of the Instrument",
            "ISIN",
            "Rating / Industry^",
            "Quantity",
            "Market value\n(Rs. in Lakhs)",
            "% to NAV",
        ]
    )
    ws.append([None])
    ws.append([None, None, "EQUITY & EQUITY RELATED"])
    ws.append([None, None, "a) Listed/awaiting listing on Stock Exchanges"])
    ws.append([None, None, "HDFC Bank Limited", "INE040A01034", "Banks", 179572, 1432.894774, 0.0911])
    ws.append([None, None, "ICICI Bank Limited", "INE090A01021", "Banks", 84005, 1155.23676, 0.0734])
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "ZERODHA")

    assert len(rows) == 2
    assert rows[0]["as_of_month"] == date(2026, 6, 1)
    assert rows[0]["constituent_isin"] == "INE040A01034"
    assert rows[0]["weight_pct"] == pytest.approx(0.0911)
    assert rows[1]["constituent_isin"] == "INE090A01021"


def test_parse_sebi_xlsx_zerodha_bare_month_year_banner_no_for_keyword():
    """B103/Fix-4: ZEN50's own file (confirmed 2026-07-12, the 19th real
    file) drops the word "for" entirely -- "MONTHLY PORTFOLIO STATEMENT OF
    ZERODHA NIFTY 50 ETF JUNE 2026" -- the one inconsistency among the 19 real
    files. The fallback anchors "(for) <Month> <YYYY>" to the END of the row
    (not a bare search anywhere in it), so it still resolves here without
    ever floating a mid-string date match -- see _max_disclosure_month's
    target-maturity-scheme-name incident docstring for the real bug class
    that end-anchoring guards against."""
    wb = Workbook()
    ws = wb.active
    ws.append([None, None, "MONTHLY PORTFOLIO STATEMENT OF ZERODHA NIFTY 50 ETF JUNE 2026"])
    ws.append([None])
    ws.append(
        [
            None,
            None,
            "Name of the Instrument",
            "ISIN",
            "Rating / Industry^",
            "Quantity",
            "Market value\n(Rs. in Lakhs)",
            "% to NAV",
        ]
    )
    ws.append([None, None, "HDFC Bank Limited", "INE040A01034", "Banks", 85862, 685.135829, 0.1113])
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "ZERODHA")

    assert len(rows) == 1
    assert rows[0]["as_of_month"] == date(2026, 6, 1)


def test_parse_sebi_xlsx_zerodha_banner_only_no_holdings_yields_zero_rows():
    """Negative case: a Zerodha-style banner + header with NO data rows
    beneath it (a genuinely empty/malformed disclosure) must still yield 0
    rows -- the as_of_month gate fix must not manufacture holdings out of a
    banner alone."""
    wb = Workbook()
    ws = wb.active
    ws.append([None, None, "MONTHLY PORTFOLIO STATEMENT OF ZERODHA NIFTY 100 ETF FOR JUNE 2026"])
    ws.append([None])
    ws.append(
        [
            None,
            None,
            "Name of the Instrument",
            "ISIN",
            "Rating / Industry^",
            "Quantity",
            "Market value\n(Rs. in Lakhs)",
            "% to NAV",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "ZERODHA")

    assert rows == []


class _FakeConstituentFetchResponse:
    """Minimal stand-in for an httpx.Response — _fetch_parse_upsert_files only
    calls `.raise_for_status()`, reads `.headers.get("content-type", ...)` and
    `.content`."""

    def __init__(self, content: bytes, content_type: str = "") -> None:
        self.content = content
        self.headers = {"content-type": content_type}

    def raise_for_status(self) -> None:
        pass


class _FakeConstituentFetchClient:
    """Minimal stand-in for httpx.AsyncClient — _fetch_parse_upsert_files only
    calls `await client.get(url, headers=...)`."""

    def __init__(self, content: bytes) -> None:
        self._content = content

    async def get(self, url, headers=None):  # noqa: ANN001 — test double
        return _FakeConstituentFetchResponse(self._content)


async def test_fetch_parse_upsert_files_logs_parsed_zero_written_outcome(monkeypatch, caplog):
    """B103/Fix-4 robustness: a file can genuinely PARSE (xlsx opens, rows
    extracted by _parse_sebi_xlsx) but still have every row dropped
    downstream in _upsert_constituents (bad as_of_month / unresolved scheme
    name) -- before this fix that outcome was silently indistinguishable in
    the logs from a truly-empty file (both just produced total_rows=0, and
    the file was still counted in `parsed_files`). Assert the dedicated
    WARNING + outcome=parsed_zero_written line fires whenever
    _parse_sebi_xlsx returns rows but the upsert writes none, so this failure
    class shows up in log-based monitoring instead of hiding as a silent
    0-yield AMC."""
    wb = Workbook()
    ws = wb.active
    ws.append([None, None, "MONTHLY PORTFOLIO STATEMENT OF ZERODHA NIFTY 100 ETF FOR JUNE 2026"])
    ws.append(
        [
            None,
            None,
            "Name of the Instrument",
            "ISIN",
            "Rating / Industry^",
            "Quantity",
            "Market value\n(Rs. in Lakhs)",
            "% to NAV",
        ]
    )
    ws.append([None, None, "HDFC Bank Limited", "INE040A01034", "Banks", 179572, 1432.894774, 0.0911])
    buf = io.BytesIO()
    wb.save(buf)

    async def _fake_upsert_writes_zero(parsed_rows, amc_name, run_id=None):
        return 0, 0  # simulate every extracted row getting dropped downstream

    monkeypatch.setattr("dhanradar.tasks.mf._upsert_constituents", _fake_upsert_writes_zero)

    client = _FakeConstituentFetchClient(buf.getvalue())
    with caplog.at_level("WARNING"):
        total_rows, total_aum = await _fetch_parse_upsert_files(
            client, "ZERODHA", ["https://assets.zerodhafundhouse.com/x.xlsx"]
        )

    assert total_rows == 0
    assert total_aum == 0
    assert any("outcome=parsed_zero_written" in r.message for r in caplog.records)


class _FakeStaticResponse:
    def __init__(self, text: str, status_code: int = 200):
        self.text = text
        self.status_code = status_code

    def raise_for_status(self) -> None:
        if self.status_code != 200:
            raise RuntimeError(f"status={self.status_code}")


class _FakeStaticClient:
    """Minimal stand-in for httpx.AsyncClient — _discover_all_urls_static only
    calls `await client.get(url, headers=...)`."""

    def __init__(self, html: str):
        self._html = html

    async def get(self, url, headers=None):  # noqa: ANN001 — test double
        return _FakeStaticResponse(self._html)


async def test_discover_all_urls_static_matches_both_xlsx_and_xls():
    """PPFAS (B87) mixes .xlsx and .xls extensions for different scheme/month
    files on the SAME archive page. The discovery regex previously matched
    only .xlsx (`href=["\\']([^"\\']+\\.xlsx...`), silently dropping every
    .xls link with zero error — confirmed 2026-07-08 against the real May-2026
    PPFAS_Monthly_Portfolio_Report file, which is .xls while its sibling
    scheme files that same month are .xlsx."""
    from dhanradar.tasks.mf import _discover_all_urls_static

    html = (
        '<a href="/downloads/PPFAS_Monthly_Portfolio_Report_May_31_2026.xls?09062026">May .xls</a>'
        '<a href="/downloads/PPFCF_PPFAS_Monthly_Portfolio_Report_May_31_2026.xlsx?09062026">May .xlsx</a>'
    )
    client = _FakeStaticClient(html)

    links = await _discover_all_urls_static(client, "https://amc.ppfas.com/downloads/", "PPFAS")

    assert any(link.endswith(".xls?09062026") for link in links)
    assert any(link.endswith(".xlsx?09062026") for link in links)


async def test_discover_all_urls_static_matches_embedded_json_not_just_href():
    """TATA (B87) is a Next.js SSR page whose disclosure links are NOT real
    <a href> anchors at all — they only appear inside an embedded, HTML-escaped
    JSON payload (`\\"field_media_document\\":\\"https://...xlsx\\"`). An
    href=-anchored regex finds zero links here even though a plain GET (no JS)
    returns the URLs as literal text (confirmed live 2026-07-08). The escaped
    JSON also risks a trailing backslash being captured from the closing
    `\\"` — must be stripped."""
    from dhanradar.tasks.mf import _discover_all_urls_static

    html = (
        '{"field_description":"Portfolio as on 31st May, 2026",'
        '"field_media_document":"https://betacms.tatamutualfund.com/system/files/'
        '2026-06/Monthly%20Portfolio%20as%20on%2031st%20May%202026.xlsx",'
        '"field_section_flag":"On"}'
    ).replace('"', '\\"')
    client = _FakeStaticClient(html)

    links = await _discover_all_urls_static(
        client, "https://www.tatamutualfund.com/schemes-related/portfolio", "TATA"
    )

    assert len(links) == 1
    assert links[0].endswith(".xlsx")
    assert "\\" not in links[0]


async def test_discover_all_urls_static_excludes_non_portfolio_same_month_document():
    """Real production bug found live 2026-07-08 (first deploy of the TATA
    scraper): TATA's `/schemes-related/portfolio` page's embedded JSON carries
    OTHER document categories in the SAME payload as the real portfolio files
    (an "Debt Index Replication Factor" note, AAUM annexures, ...). Once the
    href=-anchor requirement was dropped (previous test), a same-month/year
    document that ISN'T a portfolio disclosure at all can false-match the
    month/year filter and get selected INSTEAD of the real file — confirmed
    live: a "Debt Index Replication Factor as on 30th Jun 2026.xlsx" beat the
    real "Monthly Portfolio as on 31st May 2026.xlsx" for a June-2026 target
    (both "jun" and "2026" appear in its URL), and the real portfolio scraper
    run silently wrote ZERO rows because that file isn't a SEBI holdings
    disclosure. Every AMC's real portfolio file has "portfolio" in its own
    filename — restricting to links containing "portfolio" BEFORE the
    month/year filter fixes this without dropping any real file."""
    from dhanradar.tasks.mf import _discover_all_urls_static

    html = (
        '{"field_media_document":"https://betacms.tatamutualfund.com/system/files/'
        '2026-07/Debt%20Index%20Replication%20Factor%20as%20on%2030th%20Jun%202026.xlsx"},'
        '{"field_media_document":"https://betacms.tatamutualfund.com/system/files/'
        '2026-06/Monthly%20Portfolio%20as%20on%2031st%20May%202026.xlsx"}'
    ).replace('"', '\\"')
    client = _FakeStaticClient(html)

    links = await _discover_all_urls_static(
        client,
        "https://www.tatamutualfund.com/schemes-related/portfolio",
        "TATA",
        target_month=date(2026, 6, 1),
    )

    assert not any("Debt" in link or "Replication" in link for link in links)
    assert any("Portfolio" in link and "May" in link for link in links)


def test_pick_canonical_plan_isin_prefers_direct_growth_among_tied_candidates():
    """A single-scheme portfolio disclosure never states which plan/option ISIN
    its holdings belong to (holdings are identical across all Direct/Regular x
    Growth/IDCW/Bonus variants of the same base scheme). Among near-tied pg_trgm
    candidates, prefer Direct Plan + Growth over a same-score Regular/Bonus/IDCW
    variant -- e.g. confirmed 2026-07-09 real case where a "Bonus" plan (sim
    0.925) outscored "Growth" (sim 0.841) purely due to shorter-suffix string
    similarity, both matches for the SAME underlying scheme."""
    rows = [
        ("ISIN_BONUS", "ICICI Prudential Medium Term Bond Fund - Bonus", 0.925),
        ("ISIN_GROWTH", "ICICI Prudential Medium Term Bond Fund - Growth", 0.841),
        ("ISIN_DIRECT_GROWTH", "ICICI Prudential Medium Term Bond Fund - Direct Plan - Growth", 0.673),
        ("ISIN_IDCW", "ICICI Prudential Medium Term Bond Fund - Annual IDCW", 0.787),
    ]
    assert _pick_canonical_plan_isin(rows, tie_margin=0.30) == "ISIN_DIRECT_GROWTH"


def test_pick_canonical_plan_isin_falls_back_to_growth_only_when_no_direct_tied():
    """When no Direct-Plan-Growth variant is among the tied candidates, fall
    back to any plain Growth variant rather than the raw top (possibly
    Bonus/IDCW) match."""
    rows = [
        ("ISIN_BONUS", "ICICI Prudential Multi Sector Passive FOF - Bonus", 0.90),
        ("ISIN_GROWTH", "ICICI Prudential Multi Sector Passive FOF - Growth", 0.85),
    ]
    assert _pick_canonical_plan_isin(rows, tie_margin=0.10) == "ISIN_GROWTH"


def test_pick_canonical_plan_isin_returns_raw_top_when_no_tie():
    """When the top candidate is not tied with anything else (a clear, unique
    best match), return it unchanged -- never second-guess a decisive match."""
    rows = [
        ("ISIN_EXACT", "ICICI Prudential Bluechip Fund - Growth", 0.99),
        ("ISIN_UNRELATED", "ICICI Prudential Value Discovery Fund - Growth", 0.20),
    ]
    assert _pick_canonical_plan_isin(rows, tie_margin=0.10) == "ISIN_EXACT"


def test_parse_sebi_xlsx_sbi_scheme_name_label_value_row():
    """SBI's per-scheme manual-ingest files (B80, 2026-07-07 triage — 461 files)
    put the real scheme name in a SEPARATE cell from its "SCHEME NAME :" label,
    inside a "Portfolio Details" sheet:
        ['', '', 'SCHEME NAME :', 'SBI Test Equity Fund', ...]
    Before the fix this row's 2 distinct non-empty values failed the
    merged-banner heuristic (candidate stayed ""), so `current_scheme` was never
    set from the real name — and because the sheet is literally named
    "Portfolio Details" (which contains the "portfolio" keyword), the per-scheme
    sheet-name FALLBACK kicked in and used the sheet name itself as the scheme,
    producing scheme_name="Portfolio Details" for every row (never resolves in
    mf_funds, however good the pg_trgm threshold). This test proves the label
    row is recovered and the sheet-name fallback never fires.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Details"
    ws.append(["", "", "SBI MUTUAL FUND", "101", "", "", "", ""])
    ws.append(["", "", "SCHEME NAME :", "SBI Test Equity Fund", "", "", "", ""])
    ws.append(["", "", "Portfolio as on Mar 31, 2020", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(
        [
            "",
            "",
            "Name of the Instrument / Issuer",
            "ISIN",
            "Rating / Industry ^",
            "Quantity",
            "Market value (Rs. in Lakhs)",
            "% to AUM",
        ]
    )
    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(
        ["", "100006", "HDFC Bank Ltd.", "INE040A01034", "Banks", "64293567", "554146.25", "10.42"]
    )
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "SBI")

    assert len(rows) == 1
    assert rows[0]["scheme_name"] == "SBI Test Equity Fund"
    assert rows[0]["as_of_month"] == date(2020, 3, 1)


def test_parse_sebi_xlsx_rejects_section_header_as_scheme_name_midfile():
    """A SEBI section-header row ("a) Mutual Fund Units / Exchange Traded
    Funds") appearing between two asset-class blocks of the SAME scheme must
    never overwrite `current_scheme` — confirmed 2026-07-07 in a real SBI
    multi-asset-allocation file, where this exact row satisfied the
    single-value scheme-name-candidate heuristic (and the "fund" keyword gate)
    and silently became the new scheme_name for every subsequent holding row.
    Reuses the existing `_SECTION_HEADER_RE` (already used to keep these rows
    out of constituent_name) as the rejection gate for scheme-name candidates.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Details"
    ws.append(["", "", "SCHEME NAME :", "SBI Test Multi Asset Fund", "", "", "", ""])
    ws.append(["", "", "Portfolio as on Mar 31, 2020", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(
        [
            "",
            "",
            "Name of the Instrument / Issuer",
            "ISIN",
            "Rating / Industry ^",
            "Quantity",
            "Market value (Rs. in Lakhs)",
            "% to AUM",
        ]
    )
    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(
        ["", "100006", "HDFC Bank Ltd.", "INE040A01034", "Banks", "64293567", "554146.25", "10.42"]
    )
    ws.append(["", "", "", "", "", "", "", ""])  # blank row — resets col_map
    ws.append(["", "", "a) Mutual Fund Units / Exchange Traded Funds", "", "", "", "", ""])
    ws.append(
        [
            "",
            "",
            "Name of the Instrument / Issuer",
            "ISIN",
            "Rating / Industry ^",
            "Quantity",
            "Market value (Rs. in Lakhs)",
            "% to AUM",
        ]
    )
    ws.append(
        ["", "100010", "SBI Liquid Fund - Direct - Growth", "INE123A01011", "MF", "1000", "100.0", "1.0"]
    )
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "SBI")

    assert len(rows) == 2
    assert all(r["scheme_name"] == "SBI Test Multi Asset Fund" for r in rows)


# --- SBI monthly-portfolio parse-depth fixes (2026-07-12) --------------------
# Real June-2026 file (docs/Sample/amc-data/SBI/All-Schemes-Monthly-Portfolio---
# as-on-30th-June-2026.xlsx, 4,437 rows / 120 schemes): per-scheme weight_pct
# sum median 96.1%, min 0.81% before these two fixes.


def test_parse_sebi_xlsx_sbi_alternative_investment_funds_bare_label_not_new_scheme():
    """MUST-NOT: "Alternative Investment Funds" is SBI's own SEBI-standard
    "OTHERS" sub-category label (every debt-ish scheme's mandatory CDMDF
    line) — the SAME EDELWEISS-class bug (see
    test_parse_sebi_xlsx_edelweiss_investment_in_fund_label_not_new_scheme
    below), SBI's own wording. Confirmed real June-2026 file, sheet SFRDF
    (SBI Floating Rate Debt Fund): whenever a scheme genuinely HOLDS
    something in this sub-category, the label row prints with NO trailing
    "NIL NIL" padding, so it's a bare single-cell row — `len(non_empty)==1`
    won the candidate check unconditionally and "Alternative Investment
    Funds" contains "fund", clobbering `current_scheme` and silently
    dropping every real holding after it (confirmed real: 26 rows of
    interest-rate-swap holdings from MULTIPLE real schemes ended up
    attributed to a fake scheme literally named "Alternative Investment
    Funds" instead). Real quoted row text below (byte-for-byte, only the
    scheme/AMFI codes changed for test isolation)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Details"
    ws.append(["", "", "SCHEME NAME :", "SBI Test Floating Rate Fund", "", "", "", "", ""])
    ws.append(["", "", "PORTFOLIO STATEMENT AS ON :", "2026-06-30", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", "", ""])
    ws.append(
        [
            "",
            "",
            "Name of the Instrument / Issuer",
            "ISIN",
            "Rating / Industry^",
            "Quantity",
            "Market value\n(Rs. in Lakhs)",
            "% to AUM",
            "YTM %",
        ]
    )
    ws.append(["", "", "", "", "", "", "", "", ""])
    ws.append(
        [
            "",
            "705489",
            "Aditya Birla Housing Finance Ltd.",
            "INE831R07607",
            "CRISIL AAA",
            "3000",
            "2985.99",
            "4.43",
            "7.565",
        ]
    )
    ws.append(["", "", "Total", "", "", "", "2985.99", "4.43", ""])
    ws.append(["", "", "", "", "", "", "", "", ""])  # blank row — arms pending_scheme_check
    ws.append(["", "", "Alternative Investment Funds"])  # bare label — NO NIL padding
    ws.append(
        [
            "",
            "6400002",
            "Corporate Debt Market Development Fund-A2",
            "INF0RQ622028",
            "CDMDF",
            "4940.353",
            "586.49",
            "0.87",
            "5.51",
        ]
    )
    ws.append(["", "", "Total", "", "", "", "586.49", "0.87", ""])
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "SBI")
    holdings = [r for r in rows if not r["is_total_row"]]

    assert len(holdings) == 2
    assert all(r["scheme_name"] == "SBI Test Floating Rate Fund" for r in holdings)
    assert any(
        r["constituent_name"] == "Corporate Debt Market Development Fund-A2" for r in holdings
    )


def test_parse_sebi_xlsx_sbi_foreign_securities_bare_label_not_new_scheme():
    """MUST-NOT: same bug class as the Alternative Investment Funds test
    above, SBI's "Foreign Securities and /or overseas ETF" EQUITY-section
    label. Confirmed real June-2026 file, sheet SLMF (SBI Large & Midcap
    Fund): the label prints bare (no NIL padding) whenever the scheme
    actually holds a foreign security, and "etf" satisfies the scheme-
    keyword gate exactly like "fund" does for the AIF case."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Details"
    ws.append(["", "", "SCHEME NAME :", "SBI Test Large Midcap Fund", "", "", "", "", ""])
    ws.append(["", "", "PORTFOLIO STATEMENT AS ON :", "2026-06-30", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", "", ""])
    ws.append(
        [
            "",
            "",
            "Name of the Instrument / Issuer",
            "ISIN",
            "Rating / Industry^",
            "Quantity",
            "Market value\n(Rs. in Lakhs)",
            "% to AUM",
            "YTM %",
        ]
    )
    ws.append(["", "", "", "", "", "", "", "", ""])
    ws.append(
        [
            "",
            "100775",
            "Lemon Tree Hotels Ltd.",
            "INE970X01018",
            "Leisure Services",
            "4210285",
            "4911.3",
            "0.12",
            "",
        ]
    )
    ws.append(["", "", "Total", "", "", "", "4911.3", "0.12", ""])
    ws.append(["", "", "", "", "", "", "", "", ""])  # blank row — arms pending_scheme_check
    ws.append(["", "", "Foreign Securities and /or overseas ETF"])  # bare label — NO NIL padding
    ws.append(
        [
            "",
            "3000019",
            "Epam Systems Inc",
            "US29414B1044",
            "IT - Services",
            "299000",
            "22443.87",
            "0.55",
            "",
        ]
    )
    ws.append(["", "", "Total", "", "", "", "22443.87", "0.55", ""])
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "SBI")
    holdings = [r for r in rows if not r["is_total_row"]]

    assert len(holdings) == 2
    assert all(r["scheme_name"] == "SBI Test Large Midcap Fund" for r in holdings)
    assert any(r["constituent_name"] == "Epam Systems Inc" for r in holdings)


def test_parse_sebi_xlsx_sbi_long_bse_url_not_mistaken_for_header_row():
    """MUST-NOT: SBI's ESG-scored schemes add BRSR-disclosure hyperlink
    columns per holding. Confirmed real June-2026 file, sheet SMEEF (SBI
    ESG Exclusionary Strategy Fund), Ultratech Cement Ltd. row — one older
    BSE URL shape,
    "https://www.bseindia.com/stockinfo/AnnPdfOpen.aspx?Pname=..."
    (byte-for-byte real text below), coincidentally contains BOTH "name"
    (inside "Pname=") and "stock" (inside "stockinfo") as bare substrings —
    an ordinary holding row was misdetected as a NEW header row, wiping
    col_map and rebuilding it from that row's own garbage cell values,
    silently dropping every real holding for the rest of the sheet (only
    11 of 42 equity rows survived, weight_pct sum 51.64% instead of
    ~100%)."""
    long_bse_url = (
        "https://www.bseindia.com/stockinfo/AnnPdfOpen.aspx?Pname="
        "\\a3fa3c06-68a4-408b-9d3e-089ebd48d102.pdf"
    )
    assert len(long_bse_url) > 60  # the condition this test exercises

    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Details"
    ws.append(["", "", "SCHEME NAME :", "SBI Test ESG Fund", "", "", ""])
    ws.append(["", "", "PORTFOLIO STATEMENT AS ON :", "2026-06-30", "", "", ""])
    ws.append(["", "", "", "", "", "", ""])
    ws.append(
        [
            "",
            "",
            "Name of the Instrument / Issuer",
            "ISIN",
            "Rating / Industry^",
            "Market value\n(Rs. in Lakhs)",
            "% to AUM",
        ]
    )
    ws.append(["", "", "", "", "", "", ""])
    ws.append(
        [
            "",
            "100082",
            "Ultratech Cement Ltd.",
            "INE481G01011",
            "Cement & Cement Products",
            "14291.31",
            "2.67",
            long_bse_url,
        ]
    )
    ws.append(
        [
            "",
            "100280",
            "TVS Motor Company Ltd.",
            "INE494B01023",
            "Automobiles",
            "13461.73",
            "2.52",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "SBI")

    assert len(rows) == 2
    names = {r["constituent_name"] for r in rows}
    assert names == {"Ultratech Cement Ltd.", "TVS Motor Company Ltd."}
    assert all(r["scheme_name"] == "SBI Test ESG Fund" for r in rows)


# --- Item 5 (2026-07-12): EDELWEISS constituent parse depth -------------------
# Real May-2026 EDELWEISS monthly portfolio file (manual-ingest inbox,
# 2,840 raw rows) pads a blank row between EVERY asset-class sub-table within
# ONE scheme's per-scheme sheet (Equity, Debt/Government Securities,
# Derivatives, Money Market Instruments, TREPS/Reverse Repo, ...) — the SAME
# blank-row shape EDELWEISS was already known for between header and first
# holding (see the SBI/repeated-header test above), just recurring. The old
# unconditional `col_map = {}` on every blank row silently dropped every
# sub-table after the first (confirmed: sheet "EDBE30" — BHARAT Bond ETF -
# April 2030 — captured only 88.86% of its own 100% GRAND TOTAL; sheet
# "EEDGEF" — Edelweiss Large Cap Fund — captured 96.75%, losing Derivatives +
# Money Market + TREPS + Accrued Interest + Net Receivables entirely).
def test_parse_sebi_xlsx_edelweiss_multi_section_same_scheme_survives_blank_gap():
    """Positive: two asset-class sub-tables of the SAME scheme, separated by
    a blank row and a bare (non-scheme-keyword) section label — no repeated
    header — must both be captured under the one real scheme_name, weight_pct
    summing close to the file's own 100%. Matches the real BHARAT Bond ETF /
    Large Cap Fund shape (Equity/Debt block, blank, Government Securities/
    Derivatives block, no second header row)."""
    wb = Workbook()
    ws = wb.active
    ws.append(["PORTFOLIO STATEMENT OF EDELWEISS TEST FUND AS ON MAY 31, 2026"])
    ws.append(["(An open ended scheme investing in test securities)"])
    ws.append([])
    ws.append(
        [
            "Name of the Instrument",
            "ISIN",
            "Rating/Industry",
            "Quantity",
            "Market/Fair Value(Rs. In Lacs)",
            "% to Net Assets",
            "YIELD",
        ]
    )
    ws.append([])
    ws.append(["Equity & Equity related"])
    ws.append(["HDFC Bank Ltd.", "INE040A01034", "Banks", 1000, 900.0, 0.90, None])
    ws.append(["Sub Total", None, None, None, 900.0, 0.90, None])
    ws.append([])  # blank — no header repeats after this in real EDELWEISS files
    ws.append(["Government Securities"])
    ws.append(
        ["7.10% GOVT OF INDIA RED 18-04-2029", "IN0020220011", "SOVEREIGN", 500, 100.0, 0.10, None]
    )
    ws.append(["Sub Total", None, None, None, 100.0, 0.10, None])
    ws.append([])
    ws.append(["TOTAL", None, None, None, 1000.0, 1.0, None])
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "EDELWEISS")
    holdings = [r for r in rows if not r["is_total_row"]]

    assert len(holdings) == 2
    assert all(r["scheme_name"] == "EDELWEISS TEST FUND" for r in holdings)
    wsum = sum(r["weight_pct"] for r in holdings)
    assert wsum == pytest.approx(1.0)  # 0.90 + 0.10 — Government Securities not dropped


def test_parse_sebi_xlsx_edelweiss_investment_in_fund_label_not_new_scheme():
    """MUST-NOT: "Investment in Mutual fund" / "Investment in Exchange Traded
    Fund" are EDELWEISS's own asset-class sub-section labels (confirmed real
    May-2026 file, sheets EEARBF/EEESSF/EEMAAF/EEMOFF) — both genuinely
    contain "fund"/"etf", so once a blank-row gap is checked for a new
    scheme, this label used to satisfy the keyword gate and wrongly hijack
    `current_scheme`, silently dropping every real holding that followed
    under the WRONG (garbage) scheme name. Confirmed real: Edelweiss
    Arbitrage Fund's weight_pct sum was 28.62% without this exclusion vs
    100.00% with it, purely from 3 real "invested into Edelweiss's own
    Liquid/Money-Market/Low-Duration funds" rows landing on the wrong
    scheme."""
    wb = Workbook()
    ws = wb.active
    ws.append(["PORTFOLIO STATEMENT OF EDELWEISS TEST ARBITRAGE FUND AS ON MAY 31, 2026"])
    ws.append(["(An open ended scheme investing in arbitrage opportunities)"])
    ws.append([])
    ws.append(
        [
            "Name of the Instrument",
            "ISIN",
            "Rating/Industry",
            "Quantity",
            "Market/Fair Value(Rs. In Lacs)",
            "% to Net Assets",
            "YIELD",
        ]
    )
    ws.append([])
    ws.append(["HDFC Bank Ltd.", "INE040A01034", "Banks", 1000, 900.0, 0.28, None])
    ws.append(["Sub Total", None, None, None, 900.0, 0.28, None])
    ws.append([])
    ws.append(["Investment in Mutual fund"])
    ws.append(
        ["EDELWEISS LIQUID FUND - DIRECT PL -GR", "INF754K01GM4", None, 500, 100.0, 0.07, None]
    )
    ws.append(["Sub Total", None, None, None, 100.0, 0.07, None])
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "EDELWEISS")
    holdings = [r for r in rows if not r["is_total_row"]]

    assert len(holdings) == 2
    # Both rows must stay on the REAL scheme, not "Investment in Mutual fund".
    assert all(r["scheme_name"] == "EDELWEISS TEST ARBITRAGE FUND" for r in holdings)
    assert any(r["constituent_name"] == "EDELWEISS LIQUID FUND - DIRECT PL -GR" for r in holdings)


def test_parse_sebi_xlsx_edelweiss_grand_total_stops_footer_leaking_as_holdings():
    """MUST-NOT: the real May-2026 file's per-scheme sheets continue past
    GRAND TOTAL into a NAV-per-plan-option table and numbered SEBI disclosure
    notes — a completely different table shape that happens to share the
    stale holdings col_map once blank rows stop resetting it. Confirmed
    real: 1,298 such rows (e.g. "Direct Plan Growth Option" paired with a NAV
    number landing in the ISIN column) got written as fake constituent rows
    across the May-2026 file before this guard. A GRAND TOTAL row must be the
    definitive end of extraction for that scheme's table."""
    wb = Workbook()
    ws = wb.active
    ws.append(["PORTFOLIO STATEMENT OF EDELWEISS TEST FUND AS ON MAY 31, 2026"])
    ws.append([])
    ws.append(
        [
            "Name of the Instrument",
            "ISIN",
            "Rating/Industry",
            "Quantity",
            "Market/Fair Value(Rs. In Lacs)",
            "% to Net Assets",
            "YIELD",
        ]
    )
    ws.append([])
    ws.append(["HDFC Bank Ltd.", "INE040A01034", "Banks", 1000, 900.0, 1.0, None])
    ws.append(["GRAND TOTAL", None, None, None, 900.0, 1.0, None])
    ws.append([])
    ws.append(["Plan /option (Face Value 10)", "As on", "As on"])
    ws.append(["Direct Plan Growth Option", 21.9151, 21.9397])
    ws.append(
        [
            "5. Investment in Repo of Corporate Debt Securities during the month ended May 31, 2026",
            "NIL",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "EDELWEISS")
    holdings = [r for r in rows if not r["is_total_row"]]

    assert len(holdings) == 1
    assert holdings[0]["constituent_name"] == "HDFC Bank Ltd."


def test_parse_sebi_xlsx_blank_row_gap_still_splits_genuine_new_scheme():
    """Regression guard for the reset's ORIGINAL purpose (multi-scheme-per-
    sheet consolidated files, e.g. UTI): a blank row followed by a REAL new
    scheme banner (its own "fund"/"scheme" keyword, a DIFFERENT name) must
    still start a new scheme and a new header — item 5's deferred-reset
    (`pending_scheme_check`) must not accidentally merge two different
    schemes' holdings together."""
    wb = Workbook()
    ws = wb.active
    ws.append(["SCHEME:UTI Test Equity Fund"])
    ws.append(["Name of the Instrument", "ISIN", "Industry", "Quantity", "Market Value", "% to NAV"])
    ws.append(["HDFC Bank Ltd.", "INE040A01034", "Banks", 1000, 900.0, 90.0])
    ws.append([])
    ws.append(["SCHEME:UTI Test Debt Fund"])
    ws.append(["Name of the Instrument", "ISIN", "Industry", "Quantity", "Market Value", "% to NAV"])
    ws.append(
        ["7.10% GOVT OF INDIA RED 18-04-2029", "IN0020220011", "SOVEREIGN", 500, 100.0, 95.0]
    )
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "UTI")
    holdings = [r for r in rows if not r["is_total_row"]]

    assert len(holdings) == 2
    scheme_names = {r["scheme_name"] for r in holdings}
    assert scheme_names == {"UTI Test Equity Fund", "UTI Test Debt Fund"}


def test_parse_sebi_xlsx_month_dd_yyyy_no_space_after_comma():
    """ICICI_PRU's per-scheme manual-ingest files (B80, 2026-07-07 triage — 269
    files) banner the as-of date WITHOUT a space after the comma:
    "Portfolio as on May 31,2026" / "Figures as on Mar 31,2026" — the prior
    `,?\\s+` regex required at least one space there and never matched, so
    every row kept as_of_month=None and `_upsert_constituents` silently drops
    any row with no as_of_month (zero_rows_upserted_scheme_unresolved), even
    though the scheme name itself resolves fine via pg_trgm."""
    wb = Workbook()
    ws = wb.active
    ws.append(["", "ICICI Prudential Mutual Fund", "", "", "", "", "", "", "", ""])
    ws.append(["", "ICICI Prudential Test ETF", "", "", "", "", "", "", "", ""])
    ws.append(["", "Portfolio as on May 31,2026", "", "", "", "", "", "", "", ""])
    ws.append(
        [
            "",
            "Company/Issuer/Instrument Name",
            "ISIN",
            "Coupon",
            "Industry/Rating",
            "Quantity",
            "Exposure/Market Value(Rs.Lakh)",
            "% to Nav",
            "",
            "",
        ]
    )
    ws.append(
        ["", "ITC Ltd.", "INE154A01025", "", "Diversified Fmcg", "7208379", "20680.84", "26.98", "", ""]
    )
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "ICICI_PRU")

    assert len(rows) == 1
    assert rows[0]["as_of_month"] == date(2026, 5, 1)


def test_parse_sebi_xlsx_closed_ended_series_name_left_unmangled():
    """A genuinely closed-ended/matured scheme name (no live ISIN in mf_funds —
    verified 2026-07-07 against live prod: 'SBI Debt Fund Series C-16' /
    'SBI Debt Fund Series B-35' score <0.27 similarity against every SBI fund,
    confirming these are not in the catalog at all, not a formatting mismatch)
    must be extracted EXACTLY as written — the "SCHEME NAME :" label-row
    recovery must not mangle or truncate a real name it doesn't specifically
    target, so a downstream fuzzy match (or its correct absence) never gets a
    false positive from over-eager cleanup."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Details"
    ws.append(["", "", "SCHEME NAME :", "SBI Debt Fund Series C-16", "", "", "", ""])
    ws.append(["", "", "Portfolio as on Aug 16, 2018", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(
        [
            "",
            "",
            "Name of the Instrument / Issuer",
            "ISIN",
            "Rating / Industry ^",
            "Quantity",
            "Market value (Rs. in Lakhs)",
            "% to AUM",
        ]
    )
    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(
        ["", "100006", "HDFC Bank Ltd.", "INE040A01034", "Banks", "64293567", "554146.25", "10.42"]
    )
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "SBI")

    assert len(rows) == 1
    assert rows[0]["scheme_name"] == "SBI Debt Fund Series C-16"


def test_parse_sebi_xlsx_xlrd_fallback_for_legacy_binary_xls(monkeypatch):
    """A genuine legacy binary .xls (openpyxl raises BadZipFile trying to read
    it as a zip container — confirmed 2026-07-08 against ABSL's real
    ~107-sheet monthly portfolio) falls back to xlrd via the module's shim
    classes, so the SAME parsing loop runs unchanged. Mirrors ABSL's real
    per-scheme banner: a 2-distinct-value row pairing a short fund CODE with
    the real scheme name, followed by a bare scheme-TYPE description row that
    must NOT overwrite the correctly-detected name."""
    import xlrd

    from dhanradar.tasks.mf import _parse_sebi_xlsx as target_fn

    class _FakeSheet:
        def __init__(self, rows: list[list]) -> None:
            self._rows = rows
            self.nrows = len(rows)

        def row_values(self, r: int) -> list:
            return list(self._rows[r])

    class _FakeBook:
        def __init__(self, sheets: dict[str, list[list]]) -> None:
            self._sheets = sheets

        def sheet_names(self) -> list[str]:
            return list(self._sheets)

        def sheet_by_name(self, name: str) -> _FakeSheet:
            return _FakeSheet(self._sheets[name])

    scheme_rows = [
        [
            "ABBSEIIF",
            "ADITYA BIRLA SUN LIFE BSE INDIA INFRASTRUCTURE INDEX FUND",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "",
            "An open ended Index Fund replicating the BSE India Infrastructure Total Return Index",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        ["", "Portfolio Statement as on May 31,2026", "", "", "", "", "", "", ""],
        [
            "",
            "Name of the Instrument",
            "ISIN",
            "Industry^ / Rating",
            "Quantity",
            "Market/Fair Value\n(Rs.in Lacs)",
            "% to Net Assets",
            "Yield",
            "Yield to Call",
        ],
        ["", "Larsen & Toubro Limited", "INE018A01030", "Construction", 8218.0, 335.02, 0.0989, "", ""],
    ]
    fake_book = _FakeBook({"Index": [["Sr", "Code", "Name"]], "ABBSEIIF": scheme_rows})
    monkeypatch.setattr(xlrd, "open_workbook", lambda file_contents: fake_book)

    # Not a real xlsx — openpyxl raises BadZipFile, triggering the xlrd fallback.
    rows = target_fn(b"not a real xlsx file", "ABSL")

    assert len(rows) == 1
    assert rows[0]["scheme_name"] == "ADITYA BIRLA SUN LIFE BSE INDIA INFRASTRUCTURE INDEX FUND"
    assert rows[0]["constituent_isin"] == "INE018A01030"


def test_parse_sebi_xlsx_2value_banner_disambiguates_via_whitespace(monkeypatch):
    """A fund CODE can coincidentally contain a scheme-name keyword as a bare
    substring (e.g. ABSL's 'C10YGETF' ends in 'ETF' — confirmed 2026-07-08),
    making the keyword-match rule ambiguous (both values match). The
    whitespace tie-breaker must still pick the real (spaced) name, since a
    fund code is always a single space-free token."""
    import xlrd

    from dhanradar.tasks.mf import _parse_sebi_xlsx as target_fn

    class _FakeSheet:
        def __init__(self, rows: list[list]) -> None:
            self._rows = rows
            self.nrows = len(rows)

        def row_values(self, r: int) -> list:
            return list(self._rows[r])

    class _FakeBook:
        def __init__(self, sheets: dict[str, list[list]]) -> None:
            self._sheets = sheets

        def sheet_names(self) -> list[str]:
            return list(self._sheets)

        def sheet_by_name(self, name: str) -> _FakeSheet:
            return _FakeSheet(self._sheets[name])

    scheme_rows = [
        ["C10YGETF", "ADITYA BIRLA SUN LIFE CRISIL 10 YEAR GILT ETF", "", "", "", "", "", "", ""],
        [
            "",
            "An open ended Debt Exchange Traded Fund tracking the CRISIL 10 Year Gilt Index.",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        ["", "Portfolio Statement as on May 31,2026", "", "", "", "", "", "", ""],
        [
            "",
            "Name of the Instrument",
            "ISIN",
            "Rating",
            "Quantity",
            "Market/Fair Value\n(Rs.in Lacs)",
            "% to Net Assets",
            "Yield",
            "Yield to Call",
        ],
        ["", "Government of India (06/10/2035)", "IN0020250091", "Sovereign", 3145000.0, 3032.87, 0.9577, "", ""],
    ]
    fake_book = _FakeBook({"Index": [["Sr", "Code", "Name"]], "C10YGETF": scheme_rows})
    monkeypatch.setattr(xlrd, "open_workbook", lambda file_contents: fake_book)

    rows = target_fn(b"not a real xlsx file", "ABSL")

    assert len(rows) == 1
    assert rows[0]["scheme_name"] == "ADITYA BIRLA SUN LIFE CRISIL 10 YEAR GILT ETF"


def test_parse_sebi_xlsx_reraises_when_neither_format_readable():
    """A file that is NEITHER a real .xlsx NOR a real legacy binary .xls —
    e.g. an AMC website 'Fund Details' page saved with a misleading .xls
    extension (confirmed 2026-07-08: ~52 SBI files are plain HTML, not a
    spreadsheet at all — a different data source entirely, tracked
    separately) — must raise, not silently return zero rows. The caller
    (tasks/manual_ingest.py) then marks the file failed with an honest parse
    error instead of the misleading zero_rows_upserted_scheme_unresolved."""
    with pytest.raises(Exception):  # noqa: B017 — exact type varies (xlrd's own error)
        _parse_sebi_xlsx(b"<html><body>not a spreadsheet</body></html>", "SBI")


def test_drop_over_covered_funds_skips_fund_over_105_pct():
    """ADR-0039 fail-closed guard: a fund whose weight_pct rows already sum past
    105% is dropped entirely rather than written half-garbage."""
    batch = [
        {"isin": "INF789F01WY2", "weight_pct": 99.83},
        {"isin": "INF789F01WY2", "weight_pct": 99.83},  # header-leak duplicates the weight
        {"isin": "INF000OTHER", "weight_pct": 40.0},
        {"isin": "INF000OTHER", "weight_pct": 50.0},
    ]
    result = _drop_over_covered_funds(batch, "UTI")
    isins = {r["isin"] for r in result}
    assert "INF789F01WY2" not in isins
    assert "INF000OTHER" in isins
    assert len(result) == 2


# --- B90 (7-AMC enrichment): MOTILAL_OSWAL fixes ------------------------------
def test_extract_sebi_row_motilal_oswal_singular_percent_header_recognized():
    """MOTILAL_OSWAL's header reads singular "% to Net Asset" (every other AMC
    uses a plural/NAV variant already recognized) — the column-name match
    alone must still resolve weight_pct."""
    col_map = _col_map("Name of the Instrument", "ISIN", "% to Net Asset", "Market Value")
    row = ["HDFC Bank Ltd.", "INE040A01034", "11.14", "1234.56"]
    result = _extract_sebi_row(
        row, col_map, "Motilal Oswal Nifty 50 ETF", "MOTILAL_OSWAL", date(2026, 6, 1)
    )
    assert result is not None
    assert result["weight_pct"] == pytest.approx(11.14)


def test_extract_sebi_row_motilal_oswal_not_double_scaled_2026_07_12():
    """MUST-NOT (item 4c, 2026-07-12, B103-class): a row-level `weight_pct *=
    100` used to fire unconditionally for MOTILAL_OSWAL on the theory its "%
    to Net Asset" cell always stores an Excel percentage-format FRACTION.
    Real files fetched live from motilaloswalmf.com's own AEM API (both
    May-2026 and June-2026, confirmed 2026-07-12) show the column already
    holding a plain percentage number (11.14 for an 11.14% HDFC Bank
    holding), not a <1 fraction — the blind multiply turned every row 100x
    too large (11.14 -> 1114.0), so every one of the June file's 86 schemes
    summed to exactly 10000.00% and `_drop_over_covered_funds`'s >105% guard
    dropped all 6,369 rows (parsed_zero_written). weight_pct must now pass
    through unmultiplied; `_normalize_fraction_weight_groups` (group-sum
    based) is the sole place fraction-vs-percent semantics get corrected."""
    col_map = _col_map("Name of the Instrument", "ISIN", "% to Net Asset", "Market Value")
    row = ["HDFC Bank Ltd.", "INE040A01034", "11.14", "1234.56"]
    result = _extract_sebi_row(
        row, col_map, "Motilal Oswal Nifty 50 ETF", "MOTILAL_OSWAL", date(2026, 6, 1)
    )
    assert result is not None
    assert result["weight_pct"] == pytest.approx(11.14)  # NOT 1114.0


def test_motilal_oswal_full_scheme_survives_upsert_pipeline_2026_07_12():
    """MUST-NOT, end-to-end (item 4c): a realistic MOTILAL_OSWAL scheme whose
    per-row weight_pct is already percent-scale (summing to ~100 across its
    holdings, matching the real June-2026 file) must survive both
    `_normalize_fraction_weight_groups` (must NOT treat a ~100 sum as a
    fraction-scale group needing ×100) and `_drop_over_covered_funds` (must
    NOT trip the >105% guard) — i.e. must actually get written, unlike the
    real 6,369-rows-extracted-0-written failure this regresses against."""
    col_map = _col_map("Name of the Instrument", "ISIN", "% to Net Asset", "Market Value")
    # 10 holdings, 10.0% each == 100% total, matching real-file percent scale.
    batch = []
    for i in range(10):
        row = [f"Holding {i}", f"INE{i:06d}0{i % 10}0{i}", "10.0", "50.0"]
        r = _extract_sebi_row(
            row, col_map, "Motilal Oswal Nifty 50 ETF", "MOTILAL_OSWAL", date(2026, 6, 1)
        )
        assert r is not None
        batch.append(
            {
                "isin": "INF247L01AB0",  # stand-in resolved scheme ISIN
                "constituent_name": r["constituent_name"],
                "as_of_month": r["as_of_month"],
                "constituent_isin": r["constituent_isin"],
                "sector": r["sector"],
                "rating": r["rating"],
                "weight_pct": r["weight_pct"],
                "market_value_cr": r["market_value_cr"],
                "source_amc": "MOTILAL_OSWAL",
            }
        )
    batch = _normalize_fraction_weight_groups(batch, "MOTILAL_OSWAL")
    result = _drop_over_covered_funds(batch, "MOTILAL_OSWAL")
    assert len(result) == 10  # nothing dropped
    assert sum(r["weight_pct"] for r in result) == pytest.approx(100.0)


def test_extract_sebi_row_other_amc_plain_percent_not_scaled():
    """A different AMC's plainly-typed, genuinely small weight_pct must NOT be
    multiplied — no AMC-specific row-level scaling exists any more; the
    shared group-sum normalizer is the only place semantics are corrected."""
    col_map = _col_map("Name of the Instrument", "ISIN", "% to NAV", "Market Value")
    row = ["Net Receivables/(Payables)", "", "0.42", "1234.56"]
    result = _extract_sebi_row(row, col_map, "UTI Equity Fund", "UTI", date(2026, 6, 1))
    assert result is not None
    assert result["weight_pct"] == 0.42


def test_parse_sebi_xlsx_motilal_oswal_back_to_index_banner_not_scheme_name():
    """MOTILAL_OSWAL's per-scheme sheet banner row literally reads "Back To
    Index" paired with the real scheme name in the same 2-value row — "index"
    is itself an accepted scheme-name keyword, so BOTH values satisfy the
    keyword-hit AND whitespace-fallback disambiguation before the fix,
    leaving `current_scheme` unset and the whole sheet silently skipped."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Motilal Oswal Mutual Fund"])
    ws.append(["Back To Index", "Motilal Oswal Nifty 50 ETF"])
    ws.append([None])
    ws.append(["Portfolio Statement as on Jun 30,2026"])
    ws.append([None])
    ws.append(["Name of the Instrument", "ISIN", "% to Net Asset", "Market Value"])
    ws.append(["HDFC Bank Ltd.", "INE040A01034", "0.05", "1234.56"])
    buf = io.BytesIO()
    wb.save(buf)

    rows = _parse_sebi_xlsx(buf.getvalue(), "MOTILAL_OSWAL")

    assert len(rows) == 1
    assert rows[0]["scheme_name"] == "Motilal Oswal Nifty 50 ETF"


# --- B103 (2026-07-12): CANARA_ROBECO paginated-query discovery re-fix ------
# Real page fixtures below are TRIMMED excerpts of the live discovery page's
# rendered HTML (captured 2026-07-12) — genuine filenames/paths, just with
# most of the surrounding markup stripped for test-file size.
_CANARA_PAGE1_HTML = """
<div class="documents-listing">
  <a href="https://www.canararobeco.com/wp-content/uploads/2026/07/MI-CR-June-26.xlsx">MI</a>
  <a href="https://www.canararobeco.com/wp-content/uploads/2026/07/GB-CR-June-26.xlsx">GB</a>
</div>
<div class="custom-pagination">
  <a class="active" href="...&amp;pagination=1">1</a>
  <a href="...&amp;pagination=2">2</a>
</div>
"""
_CANARA_PAGE2_HTML = """
<div class="documents-listing">
  <a href="https://www.canararobeco.com/wp-content/uploads/2026/07/LI-CR-June-26.xlsx">LI</a>
</div>
"""
_CANARA_PAGE3_EMPTY_HTML = (
    '<div class="form-container-right-card-pdf"><p>No documents found.</p></div>'
)


class _FakeCanaraResponse:
    def __init__(self, text: str) -> None:
        self.text = text

    def raise_for_status(self) -> None:
        pass


class _FakeCanaraClient:
    """Records every requested URL + header so the test can assert the FIXED
    param shape (filteryear/filtermonth/pagination — NOT the old, wrong
    `searchyear` fiscal-year param) and serves canned pages by pagination #."""

    def __init__(self, pages: dict[int, str]) -> None:
        self.pages = pages
        self.requested_urls: list[str] = []
        self.requested_headers: list[dict] = []

    async def get(self, url, headers=None):  # noqa: ANN001 — test double
        self.requested_urls.append(url)
        self.requested_headers.append(headers or {})
        m = re.search(r"pagination=(\d+)", url)
        page_num = int(m.group(1)) if m else 1
        return _FakeCanaraResponse(self.pages.get(page_num, _CANARA_PAGE3_EMPTY_HTML))


async def test_process_amc_paginated_query_uses_filteryear_filtermonth_not_searchyear(monkeypatch):
    """B103 root cause: the old template also sent a fiscal-year `searchyear`
    param the site's current filter widget never sets, which made the
    listing return zero documents. The fixed template/function must request
    only `filteryear`/`filtermonth`/`pagination`."""
    client = _FakeCanaraClient({1: _CANARA_PAGE1_HTML, 2: _CANARA_PAGE2_HTML})
    url_template = (
        "https://www.canararobeco.com/documents/statutory-disclosures/scheme-dashboard/"
        "scheme-monthly-portfolio/?filteryear={year}&filtermonth={month:02d}&pagination={page}"
    )

    captured: dict = {}

    async def _fake_fetch_parse_upsert(client_, amc_name, file_urls, **kwargs):
        captured["file_urls"] = file_urls
        return len(file_urls) * 10, 0

    monkeypatch.setattr("dhanradar.tasks.mf._fetch_parse_upsert_files", _fake_fetch_parse_upsert)

    total_rows, _ = await _process_amc_paginated_query(client, "CANARA_ROBECO", url_template)

    assert total_rows == 30  # 3 files (2 page1 + 1 page2) * 10
    assert all("searchyear=" not in u for u in client.requested_urls)
    assert all("filteryear=" in u and "filtermonth=" in u for u in client.requested_urls)
    # Discovery GET uses a browser-shaped UA (WAF blocks the honest UA) —
    # every request the fake client saw must carry it.
    assert all("Mozilla" in h.get("User-Agent", "") for h in client.requested_headers)
    assert len(captured["file_urls"]) == 3


async def test_process_amc_paginated_query_stops_pagination_when_no_new_links(monkeypatch):
    """Page 3 (and beyond) returns no new .xlsx links -> the page loop must
    stop instead of walking all 20 hard-cap pages."""
    client = _FakeCanaraClient({1: _CANARA_PAGE1_HTML, 2: _CANARA_PAGE2_HTML})
    url_template = (
        "https://www.canararobeco.com/documents/statutory-disclosures/scheme-dashboard/"
        "scheme-monthly-portfolio/?filteryear={year}&filtermonth={month:02d}&pagination={page}"
    )

    async def _fake_fetch_parse_upsert(client_, amc_name, file_urls, **kwargs):
        return 0, 0

    monkeypatch.setattr("dhanradar.tasks.mf._fetch_parse_upsert_files", _fake_fetch_parse_upsert)

    await _process_amc_paginated_query(client, "CANARA_ROBECO", url_template)

    # 3 requests for the matching month (page1, page2, page3-empty-breaks) —
    # never the full 20-page hard cap.
    assert len(client.requested_urls) == 3


# --- B103 (2026-07-12): NAVI download Referer hardening ----------------------
async def test_fetch_parse_upsert_files_sends_referer_when_provided():
    """NAVI's production CDN 403 could not be reproduced locally (see
    _process_amc_nonce_api docstring), so the defensive `referer` param must
    at minimum actually reach the request when a caller supplies one."""
    buf = io.BytesIO()
    Workbook().save(buf)  # empty-but-real xlsx — _parse_sebi_xlsx must not choke on it
    client = _FakeConstituentFetchClient(buf.getvalue())
    captured_headers: list[dict] = []
    orig_get = client.get

    async def _get_capture(url, headers=None):
        captured_headers.append(headers or {})
        return await orig_get(url, headers=headers)

    client.get = _get_capture  # type: ignore[method-assign]

    navi_referer = "https://navi.com/mutual-fund/downloads/portfolio"
    await _fetch_parse_upsert_files(
        client, "NAVI", ["https://public-assets.prod.navi-tech.in/x.xlsx"], referer=navi_referer
    )

    assert captured_headers[0]["Referer"] == navi_referer


async def test_fetch_parse_upsert_files_omits_referer_by_default():
    """Every other AMC (MOTILAL_OSWAL, CANARA_ROBECO, ...) must see UNCHANGED
    behaviour — no Referer header when the caller doesn't pass one."""
    buf = io.BytesIO()
    Workbook().save(buf)
    client = _FakeConstituentFetchClient(buf.getvalue())
    captured_headers: list[dict] = []
    orig_get = client.get

    async def _get_capture(url, headers=None):
        captured_headers.append(headers or {})
        return await orig_get(url, headers=headers)

    client.get = _get_capture  # type: ignore[method-assign]

    await _fetch_parse_upsert_files(client, "MOTILAL_OSWAL", ["https://example.com/x.xlsx"])

    assert "Referer" not in captured_headers[0]
