"""
Unit tests for mf/disclosure_parsers.py — pure parsers for the non-portfolio
manual-ingest file classes. Synthetic workbooks reproduce the REAL layouts
observed in the founder's first batch (2026-07-07): the SEBI AAUM annexure
(AXIS/SBI shape), ICICI's annual riskometer sheet, and BOTH scheme-performance
layouts (ICICI repeating blocks; SBI per-scheme sheets with explicit labels).

Every negative case asserts the fail-closed outcome (empty result), never a
guess — a wrong AAUM/riskometer/benchmark on a fund page is worse than none.
"""

from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook

from dhanradar.mf.disclosure_parsers import (
    RISKOMETER_BANDS,
    classify_file_class,
    parse_aaum_annexure,
    parse_amfi_aaum,
    parse_riskometer_annual,
    parse_scheme_master_details,
    parse_scheme_performance,
    parse_ter_disclosure,
)


def _xlsx(builder) -> bytes:
    wb = Workbook()
    builder(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# classify_file_class — real filenames from the founder's batch
# ---------------------------------------------------------------------------


def test_classify_real_filenames():
    assert classify_file_class("AXISAverage_Assets_Under_Management_May_2026.xlsx") == "aaum"
    assert classify_file_class("Monthly AAUM_May 2026_09062026_062604_PM.xlsx") == "aaum"
    assert classify_file_class("SBI-monthly-average-asset-under-management-may-2026.xls") == "aaum"
    assert (
        classify_file_class("HDFCMF AMC- Monthly AAUM Disclosure Report - May 2026.xls") == "aaum"
    )
    assert (
        classify_file_class("ICICIAnnual Disclosure of Scheme Riskometer_FY 2025-26.xlsx")
        == "riskometer"
    )
    assert classify_file_class("Risk o meter Summary_June 09, 2026.pdf") == "riskometer"
    assert classify_file_class("ICICIScheme Performance Disclosure.xlsx") == "performance"
    assert classify_file_class("sbimf-scheme-performance---may-2026.xlsx") == "performance"
    # Portfolio disclosures must stay on the constituents path.
    assert classify_file_class("KOTAKConsolidatedSEBIPortfolioMay2026.xlsx") == "portfolio"
    assert classify_file_class("HDFC Flexi Cap Fund - 30-June-2026.xlsx") == "portfolio"


# ---------------------------------------------------------------------------
# parse_aaum_annexure — SEBI Annexure I shape (verified vs AXIS/SBI/EDEL/HDFC)
# ---------------------------------------------------------------------------


def _build_aaum(units_note: str = "") -> bytes:
    def build(wb: Workbook) -> None:
        ws = wb.active
        ws.title = "AAUM disclosure"
        ws.append([])
        ws.append(
            [
                "Sl. No.",
                "Scheme Category/ Scheme Name",
                f"Axis Mutual Fund : Net AAUM as on  2026-05-31 {units_note}",
                None,
                "GRAND TOTAL",
            ]
        )
        ws.append([None, None, "Through Direct Plan", None, None])
        ws.append(["A)", "INCOME / DEBT ORIENTED SCHEMES", None, None, None])
        ws.append([None, "AXIS LIQUID FUND", 100.0, 200.0, 59127.66])
        ws.append([None, "SUB-TOTAL(a)", 100.0, 200.0, 59127.66])
        ws.append([None, "AXIS GILT FUND", 1.0, 2.0, 104.08])
        ws.append([None, "GRAND TOTAL", None, None, 59231.74])

    return _xlsx(build)


def test_aaum_extracts_grand_total_per_scheme_and_period():
    period, pairs = parse_aaum_annexure(_build_aaum(), ".xlsx")
    assert period == date(2026, 5, 1)  # from the header "as on 2026-05-31"
    assert ("AXIS LIQUID FUND", 59127.66) in pairs
    assert ("AXIS GILT FUND", 104.08) in pairs
    # SUB-TOTAL / GRAND TOTAL / category-header rows are never schemes.
    names = {n for n, _v in pairs}
    assert not any("total" in n.lower() for n in names)
    assert "INCOME / DEBT ORIENTED SCHEMES" not in names


def test_aaum_lakh_units_scaled_to_crore():
    _period, pairs = parse_aaum_annexure(_build_aaum(units_note="(Rs. in Lakhs)"), ".xlsx")
    liquid = dict(pairs)["AXIS LIQUID FUND"]
    assert liquid == round(59127.66 * 0.01, 2)


def test_aaum_unrecognized_layout_fails_closed():
    def build(wb: Workbook) -> None:
        ws = wb.active
        ws.append(["Some", "Other", "Sheet"])
        ws.append(["AXIS LIQUID FUND", 59127.66])

    period, pairs = parse_aaum_annexure(_xlsx(build), ".xlsx")
    assert period is None
    assert pairs == []


# ---------------------------------------------------------------------------
# parse_amfi_aaum — AMFI consolidated scheme-wise quarterly AAUM
# (structure mirrors the real portal download, verified 2026-07-10)
# ---------------------------------------------------------------------------


def _build_amfi_aaum(
    title: str = (
        "Average Assets under Management (AAUM) for the quarter of April - June 2026 (Rs in Lakhs)"
    ),
) -> bytes:
    def build(wb: Workbook) -> None:
        ws = wb.active
        ws.append([title])
        ws.append(
            [
                "AMFI Code",
                "Scheme NAV Name",
                "Excluding Fund of Funds - Domestic but including Fund of Funds - Overseas",
                "Fund Of Funds - Domestic",
            ]
        )
        ws.append(["360 ONE Mutual Fund"])  # AMC section row — never a scheme
        ws.append(["Other Scheme - Other  ETFs"])  # category section row — never a scheme
        ws.append(["154366", "360 ONE MSCI India ETF", "277.38", "0"])
        ws.append([153415, "360 ONE Silver ETF", 4784.48, 0])  # numeric cells too
        ws.append(["Hybrid Scheme - Multi Asset Allocation"])
        # FoF scheme: value sits in the LAST column only
        ws.append([132175, "ABSL Aggressive Hybrid Omni FOF - Regular - IDCW", 0, 449.2])
        ws.append(["999999", "Zero AAUM Scheme", 0, 0])  # zero → skipped, never written

    return _xlsx(build)


def test_amfi_aaum_classified_by_content_not_filename():
    data = _build_amfi_aaum()
    # The portal filename is arbitrary and matches no keyword — content wins.
    assert classify_file_class("average-aum2.xlsx", data) == "amfi_aaum"
    # Without bytes there is nothing to sniff — falls through to 'portfolio'.
    assert classify_file_class("average-aum2.xlsx") == "portfolio"


def test_amfi_aaum_parses_codes_lakhs_and_quarter_period():
    period, triples = parse_amfi_aaum(_build_amfi_aaum(), ".xlsx")
    assert period == date(2026, 6, 1)  # quarter Apr–Jun → END month, day 1
    by_code = {c: (n, v) for c, n, v in triples}
    # 277.38 lakhs → 2.7738 crore (string cell)
    assert by_code["154366"] == ("360 ONE MSCI India ETF", 2.7738)
    # numeric cells parse identically
    assert by_code["153415"][1] == round(4784.48 * 0.01, 4)
    # FoF value comes from the domestic-FoF column
    assert by_code["132175"][1] == round(449.2 * 0.01, 4)
    # AMC / category section rows never parse as schemes; zero-AAUM skipped
    names = {n for _c, n, _v in triples}
    assert "360 ONE Mutual Fund" not in names
    assert "Other Scheme - Other  ETFs" not in names
    assert "999999" not in by_code


def test_amfi_aaum_unstated_units_fail_closed():
    data = _build_amfi_aaum(title="Average Assets under Management (AAUM)")
    period, triples = parse_amfi_aaum(data, ".xlsx")
    assert triples == []  # no 'lakh'/'crore' in the title → never guess 100×
    assert period is None


def test_amfi_aaum_periodless_title_returns_none_period():
    data = _build_amfi_aaum(title="Average Assets under Management (AAUM) (Rs in Lakhs)")
    period, triples = parse_amfi_aaum(data, ".xlsx")
    assert period is None  # caller must NOT fabricate aum_as_of (§8.4)
    assert len(triples) == 3


def test_ter_amfi_consolidated_layout_latest_date_wins():
    """AMFI's all-AMC TER workbook ('TER_Revised': one row per scheme per DAY,
    15 'Regular Plan - …'/'Direct Plan - …' columns) parses via the SAME
    generic detection as the per-AMC layouts — per-plan Total TER only, and
    only the LATEST date's row per scheme survives."""

    def build(wb: Workbook) -> None:
        ws = wb.active
        ws.title = "TER_Revised"
        ws.append(
            [
                "NSDL Scheme Code",
                "Scheme Name",
                "Scheme Type",
                "Scheme Category",
                "TER Date",
                "Regular Plan - Base Expense Ratio (BER) (%)",
                "Regular Plan - Brokerage cost (%)",
                "Regular Plan - Transaction Cost incurred for the purpose of execution of trade (%)",
                "Regular Plan - Statutory Levies (including GST) (%)",
                "Regular Plan - Total TER (%)",
                "Direct Plan - Base Expense Ratio (BER) (%)",
                "Direct Plan - Brokerage cost (%)",
                "Direct Plan - Transaction Cost incurred for the purpose of execution of trade (%)",
                "Direct Plan - Statutory Levies (including GST) (%)",
                "Direct Plan - Total TER (%)",
            ]
        )
        code = "360O/O/H/BHF/23/07/0007"
        # TER Date cells are genuine Excel dates (openpyxl datetime), as in
        # the real portal file — not strings.
        from datetime import datetime as _dt

        for day, reg, direct in (
            (_dt(2026, 7, 1), "1.78", "0.63"),
            (_dt(2026, 7, 8), "1.80", "0.62"),
        ):
            ws.append(
                [
                    code,
                    "360 ONE Balanced Hybrid Fund",
                    "Open Ended",
                    "Hybrid Scheme - Balanced Hybrid ",
                    day,
                    reg,
                    "0.03",
                    "0",
                    "0.38",
                    str(float(reg) + 0.41),
                    "0.46",
                    "0.03",
                    "0",
                    "0.14",
                    direct,
                ]
            )

    rows = parse_ter_disclosure(_xlsx(build), ".xlsx")
    assert len(rows) == 1  # latest date only — never one row per day
    name, ter_date, reg, direct = rows[0]
    assert name == "360 ONE Balanced Hybrid Fund"
    assert ter_date == date(2026, 7, 8)
    assert reg == round(1.80 + 0.41, 4)  # the per-plan TOTAL TER, never the BER
    assert direct == 0.62


def test_amfi_aaum_wrong_layout_fails_closed():
    def build(wb: Workbook) -> None:
        ws = wb.active
        ws.append(["Some", "Other", "Sheet"])
        ws.append(["154366", "A Fund", 100.0])

    period, triples = parse_amfi_aaum(_xlsx(build), ".xlsx")
    assert period is None
    assert triples == []


# ---------------------------------------------------------------------------
# parse_riskometer_annual — ICICI annual disclosure shape
# ---------------------------------------------------------------------------


def _build_riskometer() -> bytes:
    def build(wb: Workbook) -> None:
        ws = wb.active
        ws.append(
            [
                "SR No.",
                "Scheme name",
                "Risk-o-meter level as on March 31, 2025",
                "Risk-o-meter level as on March 31, 2026",
                "Number of changes",
            ]
        )
        ws.append(["1", "BHARAT 22 ETF", "Very High", "Very High", "0"])
        ws.append(["2", "ICICI Prudential Savings Fund", "Moderate", "Moderately High", "2"])
        ws.append(["3#", "ICICI Prudential New Fund", "NA", "NA", "0"])  # skipped
        ws.append(["4", "ICICI Prudential Odd Fund", "Low", "Somewhat Risky", "1"])  # invalid band

    return _xlsx(build)


def test_riskometer_takes_latest_column_and_validates_bands():
    pairs = parse_riskometer_annual(_build_riskometer(), ".xlsx")
    d = dict(pairs)
    assert d["BHARAT 22 ETF"] == "Very High"
    # LATEST column (FY-end 2026) wins — not the 2025 value.
    assert d["ICICI Prudential Savings Fund"] == "Moderately High"
    # 'NA' and non-regulatory band words are skipped verbatim — never coerced.
    assert "ICICI Prudential New Fund" not in d
    assert "ICICI Prudential Odd Fund" not in d
    assert set(d.values()) <= set(RISKOMETER_BANDS)


# ---------------------------------------------------------------------------
# parse_scheme_performance — both real layouts
# ---------------------------------------------------------------------------


def _build_icici_performance() -> bytes:
    def build(wb: Workbook) -> None:
        ws = wb.active
        ws.title = "Scheme Performance"
        ws.append([None, "ICICI Prudential Multicap Fund", "ICICI Prudential Multicap Fund"])
        ws.append(["Particulars", "1 Year", "3 Year"])
        ws.append([None, "CAGR (%)", "CAGR (%)"])  # repeated header — must NOT become a scheme
        ws.append(["Scheme", "6.90", "19.67"])
        ws.append(["NIFTY 500 Multicap 50:25:25 TRI (Benchmark)", "1.43", "16.43"])
        ws.append([])
        ws.append([None, "ICICI Prudential Liquid Fund", "ICICI Prudential Liquid Fund"])
        ws.append(["Particulars", "1 Year", "3 Year"])
        ws.append(["Scheme", "7.10", "6.50"])
        ws.append(["CRISIL Liquid Debt A-I Index (Benchmark)", "7.00", "6.40"])

    return _xlsx(build)


def test_performance_icici_block_layout():
    pairs = parse_scheme_performance(_build_icici_performance(), ".xlsx")
    d = dict(pairs)
    assert d["ICICI Prudential Multicap Fund"] == "NIFTY 500 Multicap 50:25:25 TRI"
    assert d["ICICI Prudential Liquid Fund"] == "CRISIL Liquid Debt A-I Index"
    # The repeated 'CAGR (%)' header row must never be captured as a scheme.
    assert not any("cagr" in name.lower() for name in d)
    assert d["ICICI Prudential Multicap Fund"] == "NIFTY 500 Multicap 50:25:25 TRI"
    assert d["ICICI Prudential Liquid Fund"] == "CRISIL Liquid Debt A-I Index"
    # The repeated 'CAGR (%)' header row must never be captured as a scheme.


def _build_kotak_benchmark_table() -> bytes:
    """Trimmed reproduction of Kotak's real annual riskometer file (2026-07-08):
    a plain table with a literal 'Scheme name' + 'Benchmark Name (Tier 1)'
    header — the risk-o-meter LEVEL column is genuinely blank in every real
    row (a change-COUNT disclosure, not a current-level one), but the
    benchmark column is fully populated."""

    def build(wb: Workbook) -> None:
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["", "Disclosure of Scheme wise change in Risk-o-meter", "", "", "", ""])
        ws.append(
            [
                "Sl No.",
                "Scheme name",
                "Scheme Risk-o-meter level",
                "Changes",
                "Benchmark Name (Tier 1)",
                "Benchmark Risk-o-meter",
            ]
        )
        ws.append(["1", "Kotak Contra Fund", None, "0", "Nifty 500 TRI", None])
        ws.append(["2", "Kotak Active Momentum Fund^", None, "2", "Nifty 500 TRI", None])

    return _xlsx(build)


def test_performance_kotak_table_layout_from_a_riskometer_named_file():
    """The Kotak table shape lives inside a file classified 'riskometer' by
    filename, but its own risk-level column is blank — this parser must
    still extract the benchmark data from the SAME bytes."""
    pairs = parse_scheme_performance(_build_kotak_benchmark_table(), ".xlsx")
    d = dict(pairs)
    assert d["Kotak Contra Fund"] == "Nifty 500 TRI"
    # A trailing footnote marker ("^") must be stripped — never part of the real name.
    assert d["Kotak Active Momentum Fund"] == "Nifty 500 TRI"
    assert "Kotak Active Momentum Fund^" not in d


def test_riskometer_annual_returns_empty_when_level_column_blank():
    """Confirms the riskometer parser itself correctly returns no bands for
    this file (never fabricates one) — the benchmark data above is recovered
    by a SEPARATE parser on the same bytes, not by relaxing this one."""
    assert parse_riskometer_annual(_build_kotak_benchmark_table(), ".xlsx") == []


def _build_sbi_performance() -> bytes:
    def build(wb: Workbook) -> None:
        ws = wb.active
        ws.title = "SMEEF"
        ws.append([None, "SBI Mutual Fund", "007"])
        ws.append([None, "SCHEME NAME :", "SBI ESG Exclusionary Strategy Fund"])
        ws.append([None, "Scheme Performace", "2026-05-31"])
        ws.append([None, "Scheme Name", "CAGR\n%"])  # table header — must not steal the name
        ws.append([None, "Regular Plan", None])
        ws.append([None, "SBI ESG Exclusionary Strategy Fund", "-2.58"])
        ws.append([None, "Scheme Benchmark: NIFTY100 ESG TRI", "0.47"])
        ws.append([None, "Additional Benchmark: BSE Sensex TRI", "-7.23"])  # ignored

    return _xlsx(build)


def test_performance_sbi_sheet_layout_primary_benchmark_only():
    pairs = parse_scheme_performance(_build_sbi_performance(), ".xlsx")
    assert pairs == [("SBI ESG Exclusionary Strategy Fund", "NIFTY100 ESG TRI")]


def test_performance_unrecognized_layout_fails_closed():
    def build(wb: Workbook) -> None:
        ws = wb.active
        ws.append(["Random", "Data"])

    assert parse_scheme_performance(_xlsx(build), ".xlsx") == []


# ---------------------------------------------------------------------------
# parse_scheme_master_details — SBI's "Fund Details" HTML page (2026-07-08)
# ---------------------------------------------------------------------------


def _build_scheme_master_html(
    *, co_manager: bool = False, salutation_no_period: bool = False
) -> bytes:
    """Trimmed synthetic reproduction of the REAL label/value HTML shape
    (single scheme, no real customer data) — mirrors the exact tag nesting
    (a nested <table> inside the 'Options Names' value) that a naive
    <td>...</td> regex would mis-pair, which is why the real parser flattens
    tags to text instead of trying to walk the DOM."""
    if co_manager:
        manager_block = (
            "<tr><td>Fund Manager</td><td><table border=1>"
            "<tr><td>Mr. Ardhendu Bhattacharya (Co-Fund Manager)</td></tr>"
            "<tr><td>Ms. Nidhi Chawla</td></tr></table></td></tr>"
            "<tr><td>Fund Manager Type</td><td><table border=1>"
            "<tr><td>Mr. Ardhendu Bhattacharya (Co-Fund Manager):Primary Debt</td></tr>"
            "<tr><td>Ms. Nidhi Chawla:Primary Equity</td></tr></table></td></tr>"
            "<tr><td>Fund Manager From Date</td><td><table border=1>"
            "<tr><td>Mr. Ardhendu Bhattacharya (Co-Fund Manager):15-Sep-2025</td></tr>"
            "<tr><td>Ms. Nidhi Chawla:15-Sep-2025</td></tr></table></td></tr>"
        )
    else:
        salutation = "Ms" if salutation_no_period else "Ms."
        manager_block = (
            f"<tr><td>Fund Manager</td><td><table border=1><tr><td>{salutation} Ranjana Gupta</td></tr>"
            "</table></td></tr>"
            f"<tr><td>Fund Manager Type</td><td><table border=1><tr><td>{salutation} Ranjana Gupta:</td></tr>"
            "</table></td></tr>"
            "<tr><td>Fund Manager From Date</td><td><table border=1>"
            f"<tr><td>{salutation} Ranjana Gupta:27-Dec-2024</td></tr></table></td></tr>"
        )

    return (
        "\r\n<html><head></head><body>"
        "<div><h2>Fund Details For SBI Test Overnight Fund </h2></div>"
        "<table border=1 align='center'>"
        "<tr><td>Fund Name</td><td>SBI Test Overnight Fund</td></tr>"
        "<tr><td>Options Names</td><td><table border=1>"
        "<tr><td>Regular Plan - Growth</td></tr><tr><td>Direct Plan - Growth</td></tr>"
        "</table></td></tr>"
        "<tr><td>Fund Type</td><td>An open ended debt scheme investing in overnight securities</td></tr>"
        "<tr><td>Riskometer At Launch</td><td>Low</td></tr>"
        "<tr><td>Riskometer As on Date</td><td>LOW</td></tr>"
        "<tr><td>Category as per SEBI categorization Circular</td><td>DEBT</td></tr>"
        "<tr><td>Benchmark(Tier 1)</td><td>NIFTY 1D Rate Index</td></tr>"
        "<tr><td>Benchmark(Tier 2)</td><td>NA</td></tr>"
        f"{manager_block}"
        "<tr><td>Annual Expense (Stated Maximum)</td><td>Regular Plan: 0.13 Direct Plan : 0.07</td></tr>"
        "<tr><td>Exit Load ( If Applicable)</td><td>NIL</td></tr>"
        "<tr><td>Custodian</td><td>Test Custodian Ltd</td></tr>"
        "</table></body></html>"
    ).encode()


def test_classify_scheme_master_by_content_not_filename():
    """A per-scheme filename indistinguishable from a real portfolio
    disclosure (e.g. 'SBI Test Overnight Fund.xls') must classify as
    scheme_master when the BYTES are the Fund Details HTML page — filename
    keywords alone can't tell these apart."""
    data = _build_scheme_master_html()
    assert classify_file_class("SBI Test Overnight Fund.xls", data) == "scheme_master"
    # Without content, filename-only classification is unaffected (existing behavior).
    assert classify_file_class("SBI Test Overnight Fund.xls") == "portfolio"


def test_scheme_master_extracts_all_fields():
    parsed = parse_scheme_master_details(_build_scheme_master_html())
    assert parsed["scheme_name"] == "SBI Test Overnight Fund"
    assert parsed["risk_band"] == "Low"
    assert parsed["benchmark_tier1"] == "NIFTY 1D Rate Index"
    assert parsed["ter_pct"] == 0.13
    assert parsed["manager_pairs"] == [("Ms. Ranjana Gupta", date(2024, 12, 27))]


def test_scheme_master_salutation_without_trailing_period():
    """A real sample used 'Ms Ranjana Gupta' (no period after 'Ms') — the
    manager/date pairing regex must still match."""
    parsed = parse_scheme_master_details(_build_scheme_master_html(salutation_no_period=True))
    assert parsed["manager_pairs"] == [("Ms Ranjana Gupta", date(2024, 12, 27))]


def test_scheme_master_co_manager_designation_not_mistaken_for_next_label():
    """A co-manager designation like '(Co-Fund Manager)' literally contains
    the substring 'Fund Manager' — a naive whole-list boundary re-scan wrongly
    matches it as the next section start, truncating the real value. Both
    managers' own distinct start dates must survive intact."""
    parsed = parse_scheme_master_details(_build_scheme_master_html(co_manager=True))
    assert parsed["manager_pairs"] == [
        ("Mr. Ardhendu Bhattacharya (Co-Fund Manager)", date(2025, 9, 15)),
        ("Ms. Nidhi Chawla", date(2025, 9, 15)),
    ]


def test_scheme_master_missing_fund_name_fails_closed():
    data = b"<html><body><h2>Fund Details For nothing useful</h2><p>no table here</p></body></html>"
    assert parse_scheme_master_details(data) == {}


def test_scheme_master_unavailable_benchmark_is_none_not_literal_na():
    data = _build_scheme_master_html().replace(b"NIFTY 1D Rate Index", b"NA")
    parsed = parse_scheme_master_details(data)
    assert parsed["benchmark_tier1"] is None


# ---------------------------------------------------------------------------
# parse_ter_disclosure — TER (Total Expense Ratio) disclosure, 2026-07-09
# ---------------------------------------------------------------------------


def test_classify_ter_real_filenames():
    assert classify_file_class("TotalExpenseRatio-2026-2027.xlsx") == "ter"
    assert classify_file_class("SBIcurrent-year-ter.xlsx") == "ter"
    assert classify_file_class("SBIhistorical-ter-march2026.xlsx") == "ter"
    assert classify_file_class("Expense Ratio.xlsx") == "ter"
    assert classify_file_class("HDFCMF_SCHEMES_TER_02-06-2026_1.xls") == "ter"
    # HSBC's real filename uses a plain space to delimit "Ter" (confirmed
    # 2026-07-09) — no underscore/hyphen at all, unlike every other AMC's
    # convention, so it needs the `\bter\b` word-boundary fallback.
    assert (
        classify_file_class(
            "D__Camsonline_www_HSBCFiles_Final_HSBC MF Total Exp TER Report_06072026.xlsx"
        )
        == "ter"
    )
    # "ter" must never match as a bare substring inside an unrelated word.
    assert classify_file_class("water_treatment.xlsx") == "portfolio"
    assert classify_file_class("after_hours_trading.xlsx") == "portfolio"


def test_parse_ter_disclosure_edelweiss_2row_group_title_layout():
    """Edelweiss's real layout: a free-text title banner row, a group-title
    row ("Regular Plan"/"Direct Plan") whose cells sit at arbitrary columns
    (NOT spanning the actual sub-column block), then the real 13-column
    sub-header row ending each plan's block with "Total TER (%)"."""

    def build(wb):
        ws = wb.active
        ws.append(["Total Expense Ratio Of Mutual Fund Scheme:-2026-2027"])
        ws.append([])
        ws.append(["Regular Plan", "Direct Plan"])
        ws.append(
            [
                "Scheme Name",
                "Date",
                "Base Expense Ratio (BER) (%)1",
                "Brokerage cost (%)2",
                "Transaction Cost incurred for the purpose of execution of trade (%)3",
                "Statutory Levies (including GST) (%)4",
                "Total TER (%)",
                "Base Expense Ratio (BER) (%)1",
                "Brokerage cost (%)2",
                "Transaction Cost incurred for the purpose of execution of trade (%)3",
                "Statutory Levies (including GST) (%)4",
                "Total TER (%)",
                "NSDL Scheme Code",
            ]
        )
        ws.append(
            [
                "Edelweiss Aggressive Hybrid Fund",
                "01/04/2026",
                1.63,
                0,
                0,
                0.25,
                1.88,
                0.39,
                0,
                0,
                0.07,
                0.46,
                "EDEL/O/H/AHF/09/04/0007",
            ]
        )
        ws.append(
            [
                "Edelweiss Aggressive Hybrid Fund",
                "01/05/2026",
                1.60,
                0,
                0,
                0.24,
                1.84,
                0.38,
                0,
                0,
                0.07,
                0.45,
                "EDEL/O/H/AHF/09/04/0007",
            ]
        )

    data = _xlsx(build)
    result = parse_ter_disclosure(data, ".xlsx")

    assert result == [("Edelweiss Aggressive Hybrid Fund", date(2026, 5, 1), 1.84, 0.45)]


def test_parse_ter_disclosure_sbi_single_row_prefixed_layout():
    """SBI's layout: ONE header row, every column literally prefixed
    "Regular Plan - " / "Direct Plan - "."""

    def build(wb):
        ws = wb.active
        ws.append(
            [
                "NSDL Scheme Code",
                "Scheme Name",
                "TER Date\n(DD/MM/\nYYYY)",
                "Regular Plan - Base Expense Ratio (BER) (%)",
                "Regular Plan - Brokerage cost (%)",
                "Regular Plan - Transaction Cost incurred for the purpose of execution of trade (%)",
                "Regular Plan - Statutory Levies (including GST) (%)",
                "Regular Plan - Total TER (%)",
                "Direct Plan - Base Expense Ratio (BER) (%)",
                "Direct Plan - Brokerage cost (%)",
                "Direct Plan - Transaction Cost incurred for the purpose of execution of trade (%)",
                "Direct Plan - Statutory Levies (including GST) (%)",
                "Direct Plan - Total TER (%)",
            ]
        )
        ws.append(
            [
                "SBIM/O/E/THE/97/01/0001",
                "SBI ESG Exclusionary Strategy Fund",
                "01/04/2026",
                1.61,
                0,
                0,
                0.28,
                1.89,
                1.1,
                0,
                0,
                0.19,
                1.29,
            ]
        )

    data = _xlsx(build)
    result = parse_ter_disclosure(data, ".xlsx")

    assert result == [("SBI ESG Exclusionary Strategy Fund", date(2026, 4, 1), 1.89, 1.29)]


def test_parse_ter_disclosure_absl_2row_bare_group_title_layout():
    """ABSL's layout: a 2-row header with BARE group titles ("Regular"/
    "Direct", no "Plan" suffix) sitting at their REAL (non-contiguous)
    column offsets — confirmed 2026-07-09 against the real file: "Regular"
    at column 3, "Direct" at column 8, with the sub-label row's cells
    correctly aligned underneath (not at columns 0-9) — and value cells
    stored as numeric-looking STRINGS (not genuine floats)."""

    def build(wb):
        ws = wb.active
        ws.append(
            [
                "NSDL Scheme Code",
                "Name of the scheme",
                "Date",
                "Regular",
                None,
                None,
                None,
                None,
                "Direct",
            ]
        )
        ws.append(
            [
                None,
                None,
                None,
                "Base TER(%)1",
                "Additional Expense as per Regulation 52(6A)(b)( %)2",
                "Additional Expense as per Regulation 52(6A)(c)( %)3",
                "GST (%)4",
                "Total TER (%)",
                "Base TER(%)1",
                "Additional Expense as per Regulation 52(6A)(b)( %)2",
                "Additional Expense as per Regulation 52(6A)(c)( %)3",
                "GST (%)4",
                "Total TER (%)",
            ]
        )
        ws.append(
            [
                "ABSL/O/O/DIN/23/11/0157",
                "ABSL CRISIL IBX GILT JUNE 2027 INDEX FUND",
                "01-JAN-2026",
                "0.56",
                "0",
                "0",
                "0",
                "0.56",
                "0.26",
                "0",
                "0",
                "0",
                "0.26",
            ]
        )

    data = _xlsx(build)
    result = parse_ter_disclosure(data, ".xlsx")

    assert result == [("ABSL CRISIL IBX GILT JUNE 2027 INDEX FUND", date(2026, 1, 1), 0.56, 0.26)]


def test_parse_ter_disclosure_ignores_freetext_title_banner_total_ter_match():
    """HDFC's real file titles its single sheet "Total Expense Ratio (TER)
    for Mutual Fund Schemes" — this ONE-cell banner row contains both
    "total" and "ter" as prose, which must NOT be mistaken for a genuine
    "Total TER" column header (confirmed 2026-07-09: without the >2-non-
    empty-cells guard, this produced a spurious 3rd "total ter" match and
    the whole file failed to parse)."""

    def build(wb):
        ws = wb.active
        ws.append(["Total Expense Ratio (TER) for Mutual Fund Schemes"])
        ws.append(["Regular Plan", "Direct Plan"])
        ws.append(
            [
                "Scheme Name",
                "NSDL Scheme Code",
                "Date (DD/MM/YYYY)",
                "Base Expense Ratio (BER) (%)",
                "Brokerage Cost (%)",
                "Transaction Cost incurred for the purpose of execution of trade (%)",
                "Statutory Levies (Including GST) (%)",
                "Total TER (%)",
                "Base Expense Ratio (BER) (%)",
                "Brokerage Cost (%)",
                "Transaction Cost incurred for the purpose of execution of trade (%)",
                "Statutory Levies (Including GST) (%)",
                "Total TER (%)",
            ]
        )
        ws.append(
            [
                "HDFC Arbitrage Fund",
                "HDFC/O/H/ARB/07/08/0017",
                "01-Jun-2026",
                0.79,
                0.06,
                0.09,
                1.09,
                2.03,
                0.34,
                0.06,
                0.09,
                1.02,
                1.51,
            ]
        )

    data = _xlsx(build)
    result = parse_ter_disclosure(data, ".xlsx")

    assert result == [("HDFC Arbitrage Fund", date(2026, 6, 1), 2.03, 1.51)]


def test_parse_ter_disclosure_unrecognized_layout_returns_empty():
    """A file with no Regular/Direct TER structure at all fails closed —
    never a guess."""

    def build(wb):
        ws = wb.active
        ws.append(["Just some notes about the fund house"])
        ws.append(["Nothing resembling a TER table here"])

    data = _xlsx(build)
    assert parse_ter_disclosure(data, ".xlsx") == []


# ---------------------------------------------------------------------------
# fund_performance (2026-07-10) — AMFI's Crisil-powered per-category export:
# ONLY the two factual columns (benchmark + validated riskometer band) are
# extracted; NAV/return columns are never read (house rule — returns are
# computed in-house from NAV, never ingested).
# ---------------------------------------------------------------------------


def _fund_performance_xlsx(rows_after_header=None) -> bytes:
    import io as _io

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Crisil Intelligence"])
    ws.append(["Generated on: 10-Jul-2026 10:47"])
    ws.append([])
    ws.append(["Fund Performance"])
    ws.append(
        [
            "Scheme Name",
            "Benchmark",
            "Riskometer Scheme",
            "Riskometer Benchmark",
            "NAV Date",
            "NAV Regular",
            "NAV Direct",
            "Return 1 Year (%) Regular",
            "Return 1 Year (%) Direct",
            "Return 1 Year (%) Benchmark",
        ]
    )
    for row in rows_after_header or [
        [
            "Aditya Birla Sun Life Large Cap Fund",
            "Nifty 100 TRI",
            "Very High",
            "Very High",
            "2026-07-08",
            510.18,
            567.41,
            -4.95,
            -4.32,
            -3.70,
        ],
        [
            "Axis Large Cap Fund",
            "BSE 100 TRI",
            "VERY HIGH",  # case-insensitive band
            "Very High",
            "2026-07-08",
            58.91,
            68.33,
            -4.0,
            -3.17,
            -3.99,
        ],
        [
            "Broken Row Fund",
            "NA",
            "Not A Real Band",  # nothing factual -> dropped
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
    ]:
        ws.append(row)
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestFundPerformance:
    def test_sniff_and_classify(self):
        from dhanradar.mf.disclosure_parsers import classify_file_class

        data = _fund_performance_xlsx()
        # Filename is arbitrary — only the bytes decide.
        assert classify_file_class("Fund-Performance-10-Jul-2026--1047.xlsx", data) == (
            "fund_performance"
        )
        assert classify_file_class("whatever (3).xlsx", data) == "fund_performance"

    def test_parses_factual_columns_only(self):
        from dhanradar.mf.disclosure_parsers import parse_fund_performance

        triples = parse_fund_performance(_fund_performance_xlsx(), ".xlsx")
        assert triples == [
            ("Aditya Birla Sun Life Large Cap Fund", "Nifty 100 TRI", "Very High"),
            ("Axis Large Cap Fund", "BSE 100 TRI", "Very High"),  # band canonicalized
        ]

    def test_invalid_band_and_na_benchmark_fail_closed(self):
        from dhanradar.mf.disclosure_parsers import parse_fund_performance

        triples = parse_fund_performance(
            _fund_performance_xlsx(
                [
                    ["X Fund", "NA", "Extremely High", "", "", "", "", "", "", ""],
                    ["Y Fund", "Nifty 50 TRI", "garbage", "", "", "", "", "", "", ""],
                ]
            ),
            ".xlsx",
        )
        # X: no valid benchmark AND no valid band -> dropped entirely.
        # Y: benchmark kept, band fail-closed None (never guessed).
        assert triples == [("Y Fund", "Nifty 50 TRI", None)]

    def test_non_performance_workbook_is_not_sniffed(self):
        from dhanradar.mf.disclosure_parsers import _looks_like_fund_performance

        assert not _looks_like_fund_performance(b"%PDF junk")
        import io as _io

        from openpyxl import Workbook

        wb = Workbook()
        wb.active.append(["Name of the Instrument", "ISIN", "Quantity"])
        buf = _io.BytesIO()
        wb.save(buf)
        assert not _looks_like_fund_performance(buf.getvalue())


# ---------------------------------------------------------------------------
# factsheet_pdf (2026-07-11) — per-scheme AMC factsheets: the fund MANAGER
# (+ tenure), closing AUM, exit load and min amount. Fixture lines are
# verbatim pypdf-extracted text from the founder's real ICICI files.
# ---------------------------------------------------------------------------

_FACTSHEET_LINES = [
    "Portfolio as on May 31, 2026",
    "ICICI Prudential Flexicap Fund",
    "(An open ended dynamic equity scheme investing across large cap, mid cap & small cap stocks)",
    "Scheme Details",
    "Monthly AAUM as on 31-May-26 : Rs. 21,132.13 crores",
    "Closing AUM as on 31-May-26 : Rs. 21,188.99 crores `",
    "Fund Managers** :",
    "Rajat Chandak",
    "(Managing this fund since July, 2021",
    "& Overall 18 years of experience)",
    "Inception/Allotment date: 17-Jul-21",
    "Application Amount for fresh Subscription :",
    "Rs. 5,000/- (plus in multiple of Re. 1)",
    "Exit load for Redemption / Switch out :-",
    "Lumpsum & SIP / STP / SWP Option",
    "1% of the applicable NAV - If the amount sought to be",
    "redeemed or switched out is invested for a period of upto 12 months",
    "1. Different plans shall have different expense structure. The performance "
    "details provided herein are of ICICI Prudential Flexicap Fund.",
]


class TestFactsheetPdf:
    def test_parses_all_fields_from_real_lines(self, monkeypatch):
        import datetime as _dt

        from dhanradar.mf import disclosure_parsers as dp

        monkeypatch.setattr(dp, "_flatten_pdf_lines", lambda data: _FACTSHEET_LINES)
        parsed = dp.parse_factsheet_pdf(b"%PDF-fake")
        assert parsed["scheme_name"] == "ICICI Prudential Flexicap Fund"
        assert parsed["manager_pairs"] == [("Rajat Chandak", _dt.date(2021, 7, 1))]
        assert parsed["aum_crore"] == 21188.99
        assert parsed["aum_as_of"] == _dt.date(2026, 5, 31)
        assert parsed["min_lumpsum_amount"] == 5000.0
        assert parsed["exit_load_pct"] == 1.0
        assert parsed["exit_load_days"] == 360  # "upto 12 months"

    def test_disclaimer_name_wins_when_page_order_flips(self, monkeypatch):
        # Real evidence 2026-07-11: pypdf emits the disclaimer FIRST for some
        # files of the same layout — the disclaimer's own fund name is the
        # reliable strategy, never the first title-looking line.
        from dhanradar.mf import disclosure_parsers as dp

        flipped = _FACTSHEET_LINES[-1:] + _FACTSHEET_LINES[:-1]
        monkeypatch.setattr(dp, "_flatten_pdf_lines", lambda data: flipped)
        parsed = dp.parse_factsheet_pdf(b"%PDF-fake")
        assert parsed["scheme_name"] == "ICICI Prudential Flexicap Fund"

    def test_fails_closed_without_a_name(self, monkeypatch):
        from dhanradar.mf import disclosure_parsers as dp

        monkeypatch.setattr(dp, "_flatten_pdf_lines", lambda data: ["random", "text"])
        assert dp.parse_factsheet_pdf(b"%PDF-fake") == {}

    def test_sniff_rejects_non_factsheets(self):
        from dhanradar.mf.disclosure_parsers import looks_like_factsheet_pdf

        assert not looks_like_factsheet_pdf(b"PK\x03\x04 xlsx")
        assert not looks_like_factsheet_pdf(b"%PDF-1.7 garbage without markers")


# ---------------------------------------------------------------------------
# factsheet_compilation (2026-07-11) — whole-AMC factsheet PDFs (HDFC/UTI/
# AXIS), one file covering every scheme. Each AMC's per-scheme extractor
# takes a `reader`-shaped object (`.pages`, each with `.extract_text()`) —
# tested directly against a fake reader built from real-file page-text
# shapes (verbatim pypdf output from the founder's real May/June 2026 HDFC/
# UTI/AXIS files) rather than constructing real PDF bytes, mirroring the
# `_flatten_pdf_lines` monkeypatch seam TestFactsheetPdf uses above.
# ---------------------------------------------------------------------------


class _FakePage:
    def __init__(self, text: str) -> None:
        self._text = text

    def extract_text(self) -> str:
        return self._text


class _FakeReader:
    def __init__(self, page_texts: list[str]) -> None:
        self.pages = [_FakePage(t) for t in page_texts]


_HDFC_SCHEME_A = (
    "For Product label and Riskometers, refer page no: 125-140 \n"
    "8  |  May 2026 \n"
    "HDFC Test Alpha Fund \n"
    "An open ended equity scheme. \n"
    "INVESTMENT OBJECTIVE: To generate capital appreciation. \n"
    "FUND MANAGER ¥ \n"
    "Name Since Total Exp \n"
    "Amit Ganatra February \n"
    "01, 2026 Over 19 years \n"
    " \n"
    "DATE OF ALLOTMENT/INCEPTION DATE \n"
    "January 01, 1995"
)
# No "¥" footnote marker — the real bug (2026-07-11): without slicing the
# block strictly AFTER "Name Since Total Exp", the literal "FUND MANAGER"
# text leaked into the first captured manager name.
_HDFC_SCHEME_B_MULTI = (
    "9  |  May 2026 \n"
    "HDFC Test Beta Fund \n"
    "An open ended hybrid scheme. \n"
    "FUND MANAGER \n"
    "Name Since Total Exp \n"
    "Rakesh Sethia May 16, \n"
    "2024 Over 18 years \n"
    "Anand Laddha July 1, 2021 Over 22 years \n"
    " \n"
    "DATE OF ALLOTMENT/INCEPTION DATE \n"
    "March 01, 2005"
)
_HDFC_NON_SCHEME_PAGE = (
    "6  | May 2026 \n"
    "Glossary \n"
    "Sharpe Ratio is a risk to reward ratio, it measures portfolio returns."
)


class TestFactsheetCompilationHdfc:
    def test_splits_two_schemes_and_skips_non_scheme_page(self) -> None:
        from dhanradar.mf import disclosure_parsers as dp

        reader = _FakeReader([_HDFC_NON_SCHEME_PAGE, _HDFC_SCHEME_A, _HDFC_SCHEME_B_MULTI])
        result = dp._parse_hdfc_compilation(reader)
        assert len(result) == 2
        alpha, beta = result
        assert alpha["scheme_name"] == "HDFC Test Alpha Fund"
        assert alpha["manager_pairs"] == [("Amit Ganatra", date(2026, 2, 1))]
        assert beta["scheme_name"] == "HDFC Test Beta Fund"
        assert beta["manager_pairs"] == [
            ("Rakesh Sethia", date(2024, 5, 1)),
            ("Anand Laddha", date(2021, 7, 1)),
        ]

    def test_never_leaks_fund_manager_label_into_name_without_yen_marker(self) -> None:
        # Regression for the real 2026-07-11 bug: pages that omit the "¥"
        # footnote must still extract a clean manager name, never
        # "FUND MANAGER Rakesh Sethia".
        from dhanradar.mf import disclosure_parsers as dp

        result = dp._parse_hdfc_compilation(_FakeReader([_HDFC_SCHEME_B_MULTI]))
        names = [name for name, _ in result[0]["manager_pairs"]]
        assert names == ["Rakesh Sethia", "Anand Laddha"]
        assert not any("FUND MANAGER" in n for n in names)


_UTI_SUMMARY_PAGE = (
    "79\n"
    "FUND MANAGER SUMMARY\n"
    "Sr. No Name of the Fund Manager Funds Managed Performance data refer page  no.\n"
    "1 Mr. Test Manager UTI Test Alpha Fund (Earlier Known as UTI Old Alpha Fund)@ 08\n"
    "UTI Test Beta Fund 09\n"
)
_UTI_SCHEME_A = (
    "01\n"
    "The objective of the scheme is to generate long term capital appreciation.\n"
    "15th October, 1986\n"
    "BSE 100 TRI\n"
    "Mr. Test Manager, B.Com, CA\n"
    "Managing the scheme since Sep 2022\n"
    "Total Exp: 19 Yrs\n"
    "Investment Objective\n"
    "Date of inception/allotment\n"
    "Benchmark Index\n"
    "Fund Manager\n"
    "Plans/Option (Regular/Direct)\n"
    "UTI TEST ALPHA FUND (Erstwhile UTI Old Alpha Fund)\n"
    "An open-ended equity scheme.\n"
    "Past performance may or may not be sustained in future. The Scheme is "
    "currently managed by Mr. Test Manager since Sep-2022."
)
# Two managers, joined with the SAME "Mr. X since <date>, Mr. Y Managing the
# scheme since <date2>" phrasing the trailing disclaimer paragraph restates —
# the real 2026-07-11 bug produced a bogus extra "manager" out of that prose.
_UTI_SCHEME_B_MULTI = (
    "02\n"
    "Some other objective.\n"
    "Mr. Alpha One - B.Com\n"
    "Managing the scheme since Jan 2020\n"
    "Total Exp: 10 Yrs\n"
    "Mr. Beta Two, CFA\n"
    "Managing the scheme since Feb 2021\n"
    "Total Exp: 8 Yrs\n"
    "Fund Manager\n"
    "UTI TEST BETA FUND\n"
    "An open ended debt scheme.\n"
    "Past performance may or may not be sustained in future. The scheme is "
    "managed by Mr. Alpha One since Jan-2020, Mr. Beta Two Managing the "
    "scheme since Feb-2021."
)
_UTI_NON_SCHEME_PAGE = (
    "UTI MUTUAL FUND IN MEDIA\nAmit Premchandani Senior Vice President and Fund Manager"
)
# Real shape (2026-07-12, UTI Balanced Advantage Fund): the "(Equity
# Portion)"/"(Debt Portion)" role qualifier sits directly after the name with
# NO comma/hyphen, and the tenure phrase is "Managing THIS scheme since"
# (not "the scheme").
_UTI_SCHEME_C_PAREN_QUALIFIER = (
    "20\n"
    "Some hybrid objective text.\n"
    "10th August, 2023\n"
    "Nifty 50 Hybrid Composite Debt 50:50 Index\n"
    "Mr. Test Gamma (Equity Portion) B.Com, MMS, CFA. Managing this scheme since Aug 2023\n"
    "Total Exp: 24 Yrs\n"
    "Mr. Test Delta (Debt Portion) Bcom, MSc, CA\n"
    "Managing this scheme since Aug 2023\n"
    "Total Exp: 15 Yrs\n"
    "Fund Manager\n"
    "UTI TEST GAMMA FUND\n"
    "An open-ended dynamic asset allocation fund\n"
)
# Real shape (2026-07-12, UTI Unit Linked Insurance Plan co-managers): bare
# "Managing since <date>" with neither "this" nor "the scheme" at all, AND a
# footnote-marker-suffixed banner ("...PLAN*").
_UTI_SCHEME_G_BARE_MANAGING_SINCE = (
    "37\n"
    "Hybrid objective text.\n"
    "1st October, 1971\n"
    "NIFTY 50 Hybrid Composite Debt 50:50 Index\n"
    "Mr. Test Eta (Debt Portion) - Bcom, MSc, CA. Managing since Nov 2025.\n"
    "Total Exp: 15 Yrs\n"
    "Fund Manager\n"
    "UTI TEST ETA PLAN*\n"
    "An open ended tax saving cum insurance scheme.\n"
)
_UTI_SUMMARY_PAGE_MARKERS = (
    "79\n"
    "FUND MANAGER SUMMARY\n"
    "Sr. No Name of the Fund Manager Funds Managed Performance data refer page  no.\n"
    "1 Mr. Test Epsilon UTI Test Epsilon Fund@ 50\n"
    "2 Mr. Test Zeta UTI Children's Test Fund 31\n"
)
# Real shape (2026-07-12, UTI Banking & PSU Fund): banner carries the SAME
# trailing footnote marker the annexure footer defines ("@" = a co/assistant
# fund manager footnote), glued directly to the name with no space.
_UTI_SCHEME_E_MARKER_BANNER = (
    "50\n"
    "Debt objective text.\n"
    "27th January, 2014\n"
    "Nifty Banking & PSU Debt Index A-II\n"
    "Mr. Test Epsilon, Bcom, MSc, CA\n"
    "Managing the scheme since Dec 2021\n"
    "Total Exp: 15 Yrs\n"
    "Fund Manager\n"
    "UTI TEST EPSILON FUND@ (Erstwhile UTI Test Epsilon Debt Fund)\n"
    "An open ended debt scheme.\n"
)
# Real shape (2026-07-12, UTI Children's Equity Fund): banner uses the CURLY
# apostrophe (U+2019) while the annexure summary above uses a straight one
# for the SAME scheme (a real PDF-export inconsistency) — AND "Since" is
# capitalized with a real month/year (UTI Multi Cap Fund's real shape),
# exercising the scoped-case-insensitive "since" literal.
_UTI_SCHEME_F_CURLY_APOSTROPHE = (
    "31\n"
    "Equity objective text.\n"
    "30th January, 2008\n"
    "Nifty 500 TRI\n"
    "Mr. Test Zeta, B.Com, CFA\n"
    "Managing the scheme Since May 2025\n"
    "Total Exp: 24 Yrs\n"
    "Fund Manager\n"
    "UTI CHILDREN’S TEST FUND\n"
    "An open ended fund for investment for children.\n"
)


class TestFactsheetCompilationUti:
    def test_splits_two_schemes_using_summary_lookup_and_fallback(self) -> None:
        from dhanradar.mf import disclosure_parsers as dp

        reader = _FakeReader(
            [_UTI_SUMMARY_PAGE, _UTI_NON_SCHEME_PAGE, _UTI_SCHEME_A, _UTI_SCHEME_B_MULTI]
        )
        result = dp._parse_uti_compilation(reader)
        assert len(result) == 2
        alpha, beta = result
        # Summary-table Title Case wins over the ALL-CAPS banner (and the
        # "(Erstwhile ...)" annotation is stripped) — case matters because
        # pg_trgm similarity is case-sensitive (2026-07-11 finding).
        assert alpha["scheme_name"] == "UTI Test Alpha Fund"
        assert alpha["manager_pairs"] == [("Test Manager", date(2022, 9, 1))]
        # Not in the summary lookup — falls back to "UTI " + title-cased rest.
        assert beta["scheme_name"] == "UTI Test Beta Fund"
        assert beta["manager_pairs"] == [
            ("Alpha One", date(2020, 1, 1)),
            ("Beta Two", date(2021, 2, 1)),
        ]

    def test_disclaimer_prose_never_leaks_a_bogus_manager(self) -> None:
        from dhanradar.mf import disclosure_parsers as dp

        result = dp._parse_uti_compilation(_FakeReader([_UTI_SCHEME_B_MULTI]))
        assert len(result[0]["manager_pairs"]) == 2  # not 3+ from the trailing prose

    def test_summary_lookup_handles_manager_prefixed_first_line(self) -> None:
        # The FIRST scheme under each manager shares its physical line with
        # "<Sr No> Mr. <Manager Name>" — `.startswith("UTI")` silently
        # dropped it (real 2026-07-11 bug); `.find("UTI")` must not.
        from dhanradar.mf.disclosure_parsers import _uti_manager_summary_lookup

        lookup = _uti_manager_summary_lookup(_UTI_SUMMARY_PAGE)
        assert lookup["UTI TEST ALPHA FUND"] == "UTI Test Alpha Fund"
        assert lookup["UTI TEST BETA FUND"] == "UTI Test Beta Fund"

    def test_manager_qualifier_in_parens_with_no_separator_still_matches(self) -> None:
        # Real bug (2026-07-12, UTI Balanced Advantage Fund): the manager's
        # "(Equity Portion)"/"(Debt Portion)" role qualifier sometimes sits
        # directly after the name with NO comma/hyphen at all — the old
        # regex required an immediate comma/hyphen and silently dropped the
        # whole scheme (manager_pairs stayed empty). "(" is now a valid
        # separator too.
        from dhanradar.mf import disclosure_parsers as dp

        result = dp._parse_uti_compilation(_FakeReader([_UTI_SCHEME_C_PAREN_QUALIFIER]))
        assert len(result) == 1
        assert result[0]["scheme_name"] == "UTI Test Gamma Fund"
        assert result[0]["manager_pairs"] == [
            ("Test Gamma", date(2023, 8, 1)),
            ("Test Delta", date(2023, 8, 1)),
        ]

    def test_managing_this_scheme_and_bare_managing_since_both_match(self) -> None:
        # Real wording variants (2026-07-12): most pages say "Managing the
        # scheme since <date>", but real pages also say "Managing this
        # scheme since <date>" (Balanced Advantage) and bare "Managing since
        # <date>" with neither "this" nor "the scheme" (ULIP co-managers).
        # Reuses the same fixture as the paren-separator test for "this
        # scheme"; a dedicated bare-"Managing since" fixture below.
        from dhanradar.mf import disclosure_parsers as dp

        result = dp._parse_uti_compilation(_FakeReader([_UTI_SCHEME_G_BARE_MANAGING_SINCE]))
        assert len(result) == 1
        assert result[0]["scheme_name"] == "UTI Test Eta Plan"
        assert result[0]["manager_pairs"] == [("Test Eta", date(2025, 11, 1))]

    def test_banner_footnote_marker_no_longer_blocks_the_whole_scheme(self) -> None:
        # Real bug (2026-07-12, UTI Banking & PSU Fund / UTI Unit Linked
        # Insurance Plan): the ALL-CAPS banner sometimes carries a trailing
        # footnote marker glued directly to the name ("UTI BANKING & PSU
        # FUND@", "...PLAN*") — the SAME markers the FUND MANAGER SUMMARY
        # footer defines. The old banner regex's char class excluded these,
        # so the banner was never recognized and the scheme (even though its
        # manager line matched fine) was silently skipped entirely. The
        # marker must not leak into the final scheme_name either.
        from dhanradar.mf import disclosure_parsers as dp

        reader = _FakeReader([_UTI_SUMMARY_PAGE_MARKERS, _UTI_SCHEME_E_MARKER_BANNER])
        result = dp._parse_uti_compilation(reader)
        assert len(result) == 1
        assert result[0]["scheme_name"] == "UTI Test Epsilon Fund"
        assert "@" not in result[0]["scheme_name"]
        assert result[0]["manager_pairs"] == [("Test Epsilon", date(2021, 12, 1))]

    def test_curly_apostrophe_banner_resolves_via_lookup_not_mangled_title_case(self) -> None:
        # Real bug (2026-07-12, UTI Children's Equity/Hybrid Fund): the same
        # scheme name is printed with a STRAIGHT apostrophe in one place and
        # a CURLY apostrophe (U+2019) in another (a PDF-export
        # inconsistency, not a code bug) — the old banner regex's char class
        # excluded the curly glyph entirely (dropped the scheme), and even
        # once allowed through, an exact-string lookup miss falls back to
        # `.title()`, which mangles "Children's" -> "Children'S". The lookup
        # key is now apostrophe-normalized so the annexure's own correct
        # casing always wins over the fallback.
        from dhanradar.mf import disclosure_parsers as dp

        reader = _FakeReader([_UTI_SUMMARY_PAGE_MARKERS, _UTI_SCHEME_F_CURLY_APOSTROPHE])
        result = dp._parse_uti_compilation(reader)
        assert len(result) == 1
        assert result[0]["scheme_name"] == "UTI Children's Test Fund"
        # "Since" (capital) with a real month/year must also match — the
        # literal is scoped case-insensitive.
        assert result[0]["manager_pairs"] == [("Test Zeta", date(2025, 5, 1))]


_AXIS_SCHEME_A_MULTI = (
    "AXIS TEST ALPHA FUND\n"
    "INVESTMENT OBJECTIVE: To achieve long term capital appreciation.\n"
    "Axis Test Alpha Fund - Regular Plan - Growth Option -4.48% 9,553\n"
    "Past performance may or may not be sustained in future . Different "
    "plans have different expense structure . Test One is managing the "
    "scheme since 23rd November 2016 and he manages 6 schemes of Axis "
    "Mutual Fund & Test Two is managing the scheme since 4th November 2024 "
    "and he manages 8 schemes of Axis Mutual Fund . Please refer to "
    "annexure on Page 102 for performance of all schemes managed by the "
    "fund manager ."
)
# No "Regular Plan - Growth Option" row (e.g. a single-NAV ETF) — falls back
# to the ALL-CAPS banner, Title-Cased.
_AXIS_SCHEME_B_ETF = (
    "AXIS TEST GAMMA FUND\n"
    "Some objective text.\n"
    "Past performance may or may not be sustained in future. Different "
    "plans have different expense structure. Test Three is managing the "
    "scheme since 1st March 2024 and she manages 4 schemes of Axis Mutual "
    "Fund. Please refer to annexure on Page 102 for performance of all "
    "schemes managed by the fund manager."
)
_AXIS_NON_SCHEME_PAGE = "EQUITY REVIEW\nMarket commentary with no manager information at all."


class TestFactsheetCompilationAxis:
    def test_splits_two_schemes_and_skips_non_scheme_page(self) -> None:
        from dhanradar.mf import disclosure_parsers as dp

        reader = _FakeReader([_AXIS_NON_SCHEME_PAGE, _AXIS_SCHEME_A_MULTI, _AXIS_SCHEME_B_ETF])
        result = dp._parse_axis_compilation(reader)
        assert len(result) == 2
        alpha, beta = result
        assert alpha["scheme_name"] == "Axis Test Alpha Fund"
        assert alpha["manager_pairs"] == [
            ("Test One", date(2016, 11, 1)),
            ("Test Two", date(2024, 11, 1)),
        ]
        assert beta["scheme_name"] == "Axis Test Gamma Fund"
        assert beta["manager_pairs"] == [("Test Three", date(2024, 3, 1))]

    def test_disclaimer_prefix_never_leaks_into_first_manager_name(self) -> None:
        # Regression for the real 2026-07-11 bug: an unbounded name char
        # class consumed the ENTIRE preceding disclaimer sentence
        # ("Past performance ... expense structure . <Name>") as part of
        # the first manager's captured name.
        from dhanradar.mf import disclosure_parsers as dp

        result = dp._parse_axis_compilation(_FakeReader([_AXIS_SCHEME_A_MULTI]))
        first_name = result[0]["manager_pairs"][0][0]
        assert first_name == "Test One"
        assert "Past performance" not in first_name


# KOTAK: real 2026-07-11 shape — the manager-tenure annexure is organised BY
# MANAGER, and the reliable per-scheme anchor is the compound "Scheme
# Inception date is <date>. Mr./Ms. <Name>[...] has/have been managing the
# fund since <date>" sentence, with the scheme-name banner sitting a short
# distance BEFORE it (sometimes with unrelated benchmark-table noise in
# between, verified against real target-maturity/index scheme pages).
_KOTAK_SCHEME_A = (
    "163\n"
    "Fund Manager*: Mr. Test One\n"  # cover-page-style label the sniff keys on
    "Kotak Test Alpha Fund\n"
    "Kotak Test Alpha\n"
    "Index TRI#\n"
    "Index TRI#\n"
    "Scheme Inception date is 10/4/2023. Mr. Test One & Mr. Test Two have "
    "been managing the fund since 10/4/2023, Mr. Test Three have been "
    "managing the fund since 09/03/2026.\n"
    "Different plans have different expense structure. The performance "
    "details provided herein are of Regular Plan - Growth Option\n"
)
# Ms. honorific (real 2026-07-11 shape, Kotak Banking & Financial Services
# Fund uses "Ms. Shibani Kurian" in this exact sentence structure).
_KOTAK_SCHEME_B_MS_HONORIFIC = (
    "163\n"
    "Kotak Test Banking Fund\n"
    "Nifty Banking TRI#\n"
    "Scheme Inception date is 27/02/2023. Ms. Test Four has been managing "
    "the fund since 27/02/2023\n"
    "Different plans have different expense structure.\n"
)
# Regression for the real 2026-07-11 bug: a lazy "letters + spaces" name
# char class with no stopping literal bridged across TWO unrelated "Kotak
# ... Fund" mentions ("Kotak Silver ETF Domestic Prices of physical Silver
# Kotak Pioneer Fund") when the first one has no "Fund/ETF/FOF" terminal
# word of its own before the second "Kotak" starts. The nearest, cleanly-
# bounded name ("Kotak Test Pioneer Fund") must win, never the bridged one.
_KOTAK_SCHEME_C_BRIDGE_GUARD = (
    "165\n"
    "Kotak Test Silver ETF\n"
    "Domestic Prices of physical Silver\n"
    "Kotak Test Pioneer Fund\n"
    "Nifty Pioneer TRI#\n"
    "Scheme Inception date is 20/10/2023. Mr. Test Five has been managing "
    "the fund since 20/10/2023\n"
    "Different plans have different expense structure.\n"
)
_KOTAK_NON_SCHEME_PAGE = "Kotak Asset Allocation View\nMarket Cap Stance\nLarge Cap (Equal Weight)"


class TestFactsheetCompilationKotak:
    def test_splits_two_schemes_and_skips_non_scheme_page(self) -> None:
        from dhanradar.mf import disclosure_parsers as dp

        reader = _FakeReader(
            [_KOTAK_NON_SCHEME_PAGE, _KOTAK_SCHEME_A, _KOTAK_SCHEME_B_MS_HONORIFIC]
        )
        result = dp._parse_kotak_compilation(reader)
        assert len(result) == 2
        alpha, beta = result
        assert alpha["scheme_name"] == "Kotak Test Alpha Fund"
        assert alpha["manager_pairs"] == [
            ("Test One", date(2023, 4, 10)),
            ("Test Two", date(2023, 4, 10)),
            ("Test Three", date(2026, 3, 9)),
        ]
        # Ms. honorific must resolve just like Mr./Mrs.
        assert beta["scheme_name"] == "Kotak Test Banking Fund"
        assert beta["manager_pairs"] == [("Test Four", date(2023, 2, 27))]

    def test_never_bridges_across_unrelated_kotak_fund_mentions(self) -> None:
        from dhanradar.mf import disclosure_parsers as dp

        result = dp._parse_kotak_compilation(_FakeReader([_KOTAK_SCHEME_C_BRIDGE_GUARD]))
        assert len(result) == 1
        assert result[0]["scheme_name"] == "Kotak Test Pioneer Fund"
        assert "Silver ETF" not in result[0]["scheme_name"]
        assert result[0]["manager_pairs"] == [("Test Five", date(2023, 10, 20))]

    def test_non_scheme_page_yields_nothing(self) -> None:
        from dhanradar.mf import disclosure_parsers as dp

        assert dp._parse_kotak_compilation(_FakeReader([_KOTAK_NON_SCHEME_PAGE])) == []


# EDELWEISS: real 2026-07-11 shape — manager + tenure sit directly on the
# scheme's own page ("Fund Manager\nFund Managers Experience Managing
# Since\nMr. <Name> <N> years <DD-Mon-YY>"); the scheme banner sometimes
# wraps onto a 2nd/3rd line before the SEBI "An open ended..." description.
_EDEL_SCHEME_A = (
    "Some portfolio holdings text.\n"
    "Fund Manager\n"
    "Fund Managers Experience Managing Since\n"
    "Mr. Test Alpha 18 years 01-Oct-21\n"
    "Mr. Test Beta 17 years 02-May-17\n"
    "Minimum Investment Amount\n"
    "Rs. 100/- per application & in multiples of Re. 1/- thereafter\n"
    "Exit Load\n"
    "Edelweiss Test Alpha Fund\n"
    "An open ended equity scheme predominantly investing in large cap stocks\n"
)
# Real bug (2026-07-11): the scheme name itself wraps onto a 2nd line before
# the description ("Edelweiss Large & Mid Cap" / "Fund" / "An open
# ended..."), which the old single-line-only check silently dropped.
_EDEL_SCHEME_B_WRAPPED_NAME = (
    "Some portfolio holdings text.\n"
    "Fund Manager\n"
    "Fund Managers Experience Managing Since\n"
    "Ms. Test Gamma 18 years 01-Apr-24\n"
    "Minimum Investment Amount\n"
    "Edelweiss Large & Mid Test\n"
    "Fund\n"
    "An open ended equity scheme investing in both large cap and mid cap stocks\n"
)
# BHARAT Bond ETF/FOF series: real, sponsor-shared products with NO
# "Edelweiss" prefix and no "An open ended..." pairing at all — falls back
# to the literal "BHARAT Bond ... : <price>" NAV line.
_EDEL_SCHEME_C_BHARAT_BOND = (
    "About the Scheme\n"
    "NAV\n"
    "BHARAT Bond Test ETF - April 2030 : ₹1569.92\n"
    "Fund Manager\n"
    "Fund Managers Experience Managing Since\n"
    "Mr. Test Delta 26 Years 26-Dec-19\n"
    "Exit Load\n"
    "Nil\n"
)
_EDEL_NON_SCHEME_PAGE = "Expert Speaks\nJune quarter most challenging market commentary."


class TestFactsheetCompilationEdelweiss:
    def test_splits_two_schemes_and_skips_non_scheme_page(self) -> None:
        from dhanradar.mf import disclosure_parsers as dp

        reader = _FakeReader([_EDEL_NON_SCHEME_PAGE, _EDEL_SCHEME_A, _EDEL_SCHEME_B_WRAPPED_NAME])
        result = dp._parse_edelweiss_compilation(reader)
        assert len(result) == 2
        alpha, beta = result
        assert alpha["scheme_name"] == "Edelweiss Test Alpha Fund"
        assert alpha["manager_pairs"] == [
            ("Test Alpha", date(2021, 10, 1)),
            ("Test Beta", date(2017, 5, 2)),
        ]
        # Wrapped 2-line name + Ms. honorific
        assert beta["scheme_name"] == "Edelweiss Large & Mid Test Fund"
        assert beta["manager_pairs"] == [("Test Gamma", date(2024, 4, 1))]

    def test_bharat_bond_series_falls_back_to_nav_line(self) -> None:
        from dhanradar.mf import disclosure_parsers as dp

        result = dp._parse_edelweiss_compilation(_FakeReader([_EDEL_SCHEME_C_BHARAT_BOND]))
        assert len(result) == 1
        assert result[0]["scheme_name"] == "BHARAT Bond Test ETF - April 2030"
        assert result[0]["manager_pairs"] == [("Test Delta", date(2019, 12, 26))]

    def test_non_scheme_page_yields_nothing(self) -> None:
        from dhanradar.mf import disclosure_parsers as dp

        assert dp._parse_edelweiss_compilation(_FakeReader([_EDEL_NON_SCHEME_PAGE])) == []


# ABSL ("Empower Factsheet"): real 2026-07-11 shape — "Fund Manager - Mr.
# <Name>\nManaging the Fund Since: <Month DD, YYYY>" repeated per co-manager,
# scheme name recovered from the Investment Performance table's own row
# (Title-Case name immediately followed by a return percentage).
_ABSL_SCHEME_A = (
    "Fund Snapshot\n"
    "Date of Allotment : Aug 30, 2002\n"
    "Fund Manager - Mr. Test Alpha\n"
    "Managing the Fund Since: January 07, 2026\n"
    "Experience in Managing the Fund: 0.4 Years\n"
    "Investment Performance NAV as on May 29, 2026 : ₹500.21\n"
    "Aditya Birla Sun Life Test Alpha Fund 17.90% 11.73% 10.87% 11.34% -3.59%\n"
)
_ABSL_SCHEME_B_MULTI = (
    "Fund Snapshot\n"
    "Fund Manager - Mr. Test Beta\n"
    "Managing the Fund Since: November 03, 2023\n"
    "Experience in Managing the Fund: 2.6 Years\n"
    "Fund Manager - Ms. Test Gamma\n"
    "Managing the Fund Since: November 21, 2022\n"
    "Experience in Managing the Fund: 3.5 Years\n"
    "Aditya Birla Sun Life Test Beta Fund 20.67% 14.11% 12.70% 16.38% 4.93%\n"
)
_ABSL_NON_SCHEME_PAGE = "Tax Reckoner Tax Year 2026-27\nWithholding tax rate table."


class TestFactsheetCompilationAbsl:
    def test_splits_two_schemes_and_skips_non_scheme_page(self) -> None:
        from dhanradar.mf import disclosure_parsers as dp

        reader = _FakeReader([_ABSL_NON_SCHEME_PAGE, _ABSL_SCHEME_A, _ABSL_SCHEME_B_MULTI])
        result = dp._parse_absl_compilation(reader)
        assert len(result) == 2
        alpha, beta = result
        assert alpha["scheme_name"] == "Aditya Birla Sun Life Test Alpha Fund"
        assert alpha["manager_pairs"] == [("Test Alpha", date(2026, 1, 7))]
        # Second manager uses the Ms. honorific.
        assert beta["scheme_name"] == "Aditya Birla Sun Life Test Beta Fund"
        assert beta["manager_pairs"] == [
            ("Test Beta", date(2023, 11, 3)),
            ("Test Gamma", date(2022, 11, 21)),
        ]

    def test_non_scheme_page_yields_nothing(self) -> None:
        from dhanradar.mf import disclosure_parsers as dp

        assert dp._parse_absl_compilation(_FakeReader([_ABSL_NON_SCHEME_PAGE])) == []


class TestFactsheetCompilationDispatch:
    def test_unrecognized_amc_fails_closed(self, monkeypatch) -> None:
        import pypdf

        from dhanradar.mf import disclosure_parsers as dp

        monkeypatch.setattr(pypdf, "PdfReader", lambda _buf: _FakeReader([_AXIS_SCHEME_A_MULTI]))
        assert dp.parse_factsheet_compilation(b"%PDF-fake", "SOME_OTHER_AMC") == []
        assert dp.parse_factsheet_compilation(b"%PDF-fake", None) == []

    def test_unreadable_pdf_fails_closed(self) -> None:
        from dhanradar.mf.disclosure_parsers import parse_factsheet_compilation

        assert parse_factsheet_compilation(b"not a pdf at all", "HDFC") == []

    def test_dispatches_to_the_right_amc_extractor(self, monkeypatch) -> None:
        import pypdf

        from dhanradar.mf import disclosure_parsers as dp

        monkeypatch.setattr(pypdf, "PdfReader", lambda _buf: _FakeReader([_HDFC_SCHEME_A]))
        result = dp.parse_factsheet_compilation(b"%PDF-fake", "HDFC")
        assert result and result[0]["scheme_name"] == "HDFC Test Alpha Fund"

    def test_dispatches_kotak_edelweiss_absl_to_their_own_extractors(self, monkeypatch) -> None:
        import pypdf

        from dhanradar.mf import disclosure_parsers as dp

        cases = [
            ("KOTAK", _KOTAK_SCHEME_A, "Kotak Test Alpha Fund"),
            ("EDELWEISS", _EDEL_SCHEME_A, "Edelweiss Test Alpha Fund"),
            ("ABSL", _ABSL_SCHEME_A, "Aditya Birla Sun Life Test Alpha Fund"),
        ]
        for amc, page_text, expected_name in cases:
            monkeypatch.setattr(pypdf, "PdfReader", lambda _buf, t=page_text: _FakeReader([t]))
            result = dp.parse_factsheet_compilation(b"%PDF-fake", amc)
            assert result and result[0]["scheme_name"] == expected_name

    def test_sniff_scans_first_pages_and_recognizes_each_layout(self, monkeypatch) -> None:
        import pypdf

        from dhanradar.mf.disclosure_parsers import looks_like_factsheet_compilation

        for page_text in (
            _HDFC_SCHEME_A,
            _UTI_SCHEME_A,
            _AXIS_SCHEME_A_MULTI,
            _KOTAK_SCHEME_A,
            _EDEL_SCHEME_A,
            _ABSL_SCHEME_A,
        ):
            monkeypatch.setattr(pypdf, "PdfReader", lambda _buf, t=page_text: _FakeReader([t]))
            assert looks_like_factsheet_compilation(b"%PDF-fake")

    def test_sniff_rejects_non_compilations(self) -> None:
        from dhanradar.mf.disclosure_parsers import looks_like_factsheet_compilation

        assert not looks_like_factsheet_compilation(b"PK\x03\x04 xlsx")
        assert not looks_like_factsheet_compilation(b"%PDF-1.7 garbage without markers")


# ---------------------------------------------------------------------------
# scheme_summary_xls (2026-07-11) — TATA's per-scheme "SCHEME SUMMARY
# DOCUMENT" workbook. Real files are saved with a ".xls" extension but are,
# byte-for-byte, valid xlsx zips; fixtures below reproduce the real ~54-row
# label/value layout (trimmed to the fields this class extracts).
# ---------------------------------------------------------------------------


def _build_scheme_summary(
    *,
    fund_name: str = "Tata Large Cap Fund",
    manager_name: str = "FM-1 Abhinav Sharma, FM-2 Hasmukh Vishariya",
    manager_from: str = "FM-1 05/04/2023, FM-2 01/03/2025",
    exit_load: str = (
        "1. On or before 30 days from the date of allotment: 0.50%. "
        "2. After 30 days from the date of allotment: NIL."
    ),
    min_amount: str = "5000",
    isins: str = "INF277K01931,INF277K01923,INF277K01QY0,INF277K01QZ7",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["Fields", "SCHEME SUMMARY DOCUMENT", None])
    ws.append([1, "Fund Name", fund_name])
    ws.append([16, "Benchmark (Tier 1)", "Nifty 100 TRI"])
    ws.append([18, "Fund Manager Name", manager_name])
    ws.append([20, "Fund Manager From Date", manager_from])
    ws.append([21, "Annual Expense (Stated maximum)", "Regular 1.96, Direct 0.98"])
    ws.append([22, "Exit Load (if applicable)", exit_load])
    ws.append([28, "ISINs", isins])
    ws.append([31, "Minimum Application Amount", min_amount])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _break_dimension(data: bytes) -> bytes:
    """Force the same broken `<dimension>` bug TATA's real files carry
    (declares 'A1' while data spans far further) — patches the zip's
    `xl/worksheets/sheet1.xml` in place, same technique
    `_repair_corrupted_stylesheet_xlsx` uses for the styles.xml repair."""
    import re
    import zipfile

    with zipfile.ZipFile(io.BytesIO(data)) as zin:
        sheet_xml = zin.read("xl/worksheets/sheet1.xml").decode("utf-8")
        patched = re.sub(r'<dimension ref="[^"]+"\s*/>', '<dimension ref="A1"/>', sheet_xml)
        assert patched != sheet_xml, "fixture's sheet1.xml has no <dimension> to break"
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                content = (
                    patched.encode("utf-8")
                    if item.filename == "xl/worksheets/sheet1.xml"
                    else zin.read(item.filename)
                )
                zout.writestr(item, content)
        return buf.getvalue()


class TestSchemeSummaryXls:
    def test_sniff_and_classify(self) -> None:
        from dhanradar.mf.disclosure_parsers import looks_like_scheme_summary_xls

        data = _build_scheme_summary()
        assert looks_like_scheme_summary_xls(data)
        # Filename is a plain per-scheme name, indistinguishable from any
        # other AMC's portfolio disclosure — only the bytes decide.
        assert classify_file_class("Tata-Large-Cap-Fund.xls", data) == "scheme_summary_xls"

    def test_non_summary_workbook_yields_nothing_for_this_class(self) -> None:
        # must-not: an unrelated xlsx (a real portfolio-holdings sheet) must
        # never be sniffed or parsed as a scheme summary.
        from dhanradar.mf.disclosure_parsers import (
            looks_like_scheme_summary_xls,
            parse_scheme_summary_xls,
        )

        wb = Workbook()
        wb.active.append(["Name of the Instrument", "ISIN", "Quantity"])
        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()
        assert not looks_like_scheme_summary_xls(data)
        assert classify_file_class("Tata Large Cap Fund - 31-May-2026.xlsx", data) == "portfolio"
        assert parse_scheme_summary_xls(data) == {}

    def test_parses_real_layout_managers_exit_load_min_amount_isins(self) -> None:
        from dhanradar.mf.disclosure_parsers import parse_scheme_summary_xls

        parsed = parse_scheme_summary_xls(_build_scheme_summary())
        assert parsed["scheme_name"] == "Tata Large Cap Fund"
        assert parsed["manager_pairs"] == [
            ("Abhinav Sharma", date(2023, 4, 5)),
            ("Hasmukh Vishariya", date(2025, 3, 1)),
        ]
        assert parsed["exit_load_pct"] == 0.50
        assert parsed["exit_load_days"] == 30
        assert parsed["min_lumpsum_amount"] == 5000.0
        assert parsed["isins"] == [
            "INF277K01931",
            "INF277K01923",
            "INF277K01QY0",
            "INF277K01QZ7",
        ]

    def test_positional_pairing_when_neither_field_is_indexed(self) -> None:
        # Real file shape (verified): no 'FM-N'/'N.' marker at all, plain
        # comma lists that line up 1:1 by position.
        from dhanradar.mf.disclosure_parsers import parse_scheme_summary_xls

        data = _build_scheme_summary(
            manager_name="Tapan Patel, Nitin Bharat Sharma, Rakesh Prajapati",
            manager_from="19 Jan 2024, 09 March 2026, 09 March 2026",
        )
        parsed = parse_scheme_summary_xls(data)
        assert parsed["manager_pairs"] == [
            ("Tapan Patel", date(2024, 1, 19)),
            ("Nitin Bharat Sharma", date(2026, 3, 9)),
            ("Rakesh Prajapati", date(2026, 3, 9)),
        ]

    def test_fm_count_mismatch_writes_managers_without_dates_never_fabricated(self) -> None:
        # Real file shape (verified): 2 managers, ONE shared date with no
        # index to attribute it to either manager — never guess which one.
        from dhanradar.mf.disclosure_parsers import parse_scheme_summary_xls

        data = _build_scheme_summary(
            manager_name="Nitin Sharma, Rakesh Prajapati",
            manager_from="29-Dec-2025",
        )
        parsed = parse_scheme_summary_xls(data)
        assert parsed["manager_pairs"] == [
            ("Nitin Sharma", None),
            ("Rakesh Prajapati", None),
        ]

    def test_broken_dimension_regression(self) -> None:
        # Bug A: TATA's real files declare `<dimension ref="A1"/>` while data
        # spans A1:C60ish — read-only openpyxl trusts it and iter_rows()
        # truncates to one cell. `_iter_sheets_xlsx` must reset it before
        # iterating, or NEITHER the sniff NOR the parse can ever see past
        # the "Fields" banner cell.
        from dhanradar.mf.disclosure_parsers import (
            looks_like_scheme_summary_xls,
            parse_scheme_summary_xls,
        )

        broken = _break_dimension(_build_scheme_summary())
        assert looks_like_scheme_summary_xls(broken)
        parsed = parse_scheme_summary_xls(broken)
        assert parsed["scheme_name"] == "Tata Large Cap Fund"
        assert parsed["isins"] == [
            "INF277K01931",
            "INF277K01923",
            "INF277K01QY0",
            "INF277K01QZ7",
        ]

    def test_iter_sheets_xlsx_reset_dimensions_no_behavior_change_on_normal_file(self) -> None:
        # A CORRECTLY-dimensioned file must read identically before/after
        # the root fix — reset_dimensions() only drops an optimization hint,
        # it must never change which rows/cells come back.
        from dhanradar.mf.disclosure_parsers import _iter_sheets_xlsx

        data = _fund_performance_xlsx()
        _name, rows = next(iter(_iter_sheets_xlsx(data)))
        assert rows[3] == ["Fund Performance"]
        assert rows[4][:3] == ["Scheme Name", "Benchmark", "Riskometer Scheme"]
        assert len(rows) == 8  # 5 header/title rows + 3 data rows
