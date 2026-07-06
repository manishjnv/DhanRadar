"""
Unit tests for the SBI TER (expense ratio) xlsx provider — no DB, no network.

Fixture row tuples below are REAL data captured 2026-07-06 from SBI's live
"current-year-ter.xlsx" (docs/default-source/ter_allschemes/current-year-ter.xlsx,
verified reachable via real-browser-UA curl, no bot-block encountered) — trimmed
to a few rows, not invented. See amc_expense_sbi.py's module docstring for the
full capture context and header layout.

Covers:
  parse_sbi_ter_rows:
    - real header + 2 real data rows -> 4 RawExpenseRatioRow (2 plans x 2 schemes)
    - plan_suffix correctly distinguishes Regular vs Direct TER values
    - effective_date coerced from a datetime cell to a plain date
    - a row with ter_pct out of range (0, 10] for one plan is dropped for
      THAT plan only; the other plan (if valid) is kept
    - a row with an empty scheme_name is skipped entirely
    - header missing a required column -> [] (fail closed, never guessed)
    - empty input -> []

  fetch_sbi_expense_ratios:
    - HTTP non-200 -> unreachable
    - network error -> unreachable
    - HTTP 200 but corrupt/unparseable xlsx bytes -> format_mismatch
    - HTTP 200 valid xlsx with 0 usable rows -> format_mismatch
    - happy path -> ok, rows returned
"""

from __future__ import annotations

import datetime
from unittest.mock import AsyncMock, MagicMock

import httpx

from dhanradar.market_data.amc_expense_sbi import (
    RawExpenseRatioRow,
    _load_sbi_ter_rows,
    fetch_sbi_expense_ratios,
    parse_sbi_ter_rows,
)

# ---------------------------------------------------------------------------
# Real captured fixture — SBI current-year-ter.xlsx, 2026-07-06.
# Columns: NSDL Scheme Code | Scheme Name | TER Date | Regular Plan - Base
# Expense Ratio (BER) (%) | Regular Plan - Brokerage cost (%) | Regular Plan -
# Transaction Cost ... (%) | Regular Plan - Statutory Levies ... (%) |
# Regular Plan - Total TER (%) | Direct Plan - Base Expense Ratio (BER) (%) |
# Direct Plan - Brokerage cost (%) | Direct Plan - Transaction Cost ... (%) |
# Direct Plan - Statutory Levies ... (%) | Direct Plan - Total TER (%)
# ---------------------------------------------------------------------------
_REAL_HEADER = (
    "NSDL Scheme Code",
    "Scheme Name",
    "TER Date\n(DD/MM/\nYYYY)",
    "Regular Plan - Base Expense Ratio (BER) (%)",
    "Regular Plan - Brokerage cost (%)",
    "Regular Plan - Transaction Cost incurred for the purpose of execution of trade (%)",
    "Regular Plan - Statutory Levies (including GST) (%)",
    "Regular Plan - Total TER (%)",
    "Direct Plan - Base Expense Ratio (BER) (%)",
    "Direct Plan - Brokerage cost (%)",
    "Direct Plan - Transaction Cost incurred for the purpose of execution of trade (%)",
    "Direct Plan - Statutory Levies (including GST) (%)",
    "Direct Plan - Total TER (%)",
)

_REAL_ROW_1 = (
    "SBIM/O/E/THE/97/01/0001",
    "SBI ESG Exclusionary Strategy Fund",
    datetime.datetime(2026, 4, 1, 0, 0),
    1.61,
    0,
    0,
    0.28,
    1.89,
    1.1,
    0,
    0,
    0.19,
    1.29,
)
_REAL_ROW_2 = (
    "SBIM/O/E/LMF/98/04/0003",
    "SBI LARGE & MIDCAP FUND",
    datetime.datetime(2026, 4, 1, 0, 0),
    1.32,
    0,
    0,
    0.22,
    1.54,
    0.59,
    0,
    0,
    0.1,
    0.69,
)

VALID_FIXTURE_ROWS = [_REAL_HEADER, _REAL_ROW_1, _REAL_ROW_2]


def _make_fake_client(content: bytes, status_code: int = 200) -> MagicMock:
    resp = MagicMock()
    resp.status_code = status_code
    resp.content = content
    client = MagicMock()
    client.get = AsyncMock(return_value=resp)
    return client


# ===========================================================================
# parse_sbi_ter_rows — PURE parser tests
# ===========================================================================


class TestParseSbiTerRows:
    def test_real_fixture_yields_four_rows(self):
        """2 real data rows x 2 plans (Regular + Direct) = 4 RawExpenseRatioRow."""
        rows = parse_sbi_ter_rows(VALID_FIXTURE_ROWS)
        assert len(rows) == 4
        assert all(isinstance(r, RawExpenseRatioRow) for r in rows)

    def test_regular_and_direct_ter_values_distinguished(self):
        rows = parse_sbi_ter_rows(VALID_FIXTURE_ROWS)
        esg_regular = next(
            r
            for r in rows
            if r.scheme_name == "SBI ESG Exclusionary Strategy Fund"
            and r.plan_suffix == " - Regular Plan"
        )
        esg_direct = next(
            r
            for r in rows
            if r.scheme_name == "SBI ESG Exclusionary Strategy Fund"
            and r.plan_suffix == " - Direct Plan"
        )
        assert esg_regular.ter_pct == 1.89
        assert esg_direct.ter_pct == 1.29

    def test_effective_date_coerced_from_datetime_cell(self):
        rows = parse_sbi_ter_rows(VALID_FIXTURE_ROWS)
        assert rows[0].effective_date == datetime.date(2026, 4, 1)

    def test_out_of_range_ter_dropped_for_that_plan_only(self):
        """Direct-plan TER = 12.0 (>10, invalid) — Regular plan (1.89, valid) is
        still kept for the same row."""
        bad_row = (
            "SBIM/O/E/THE/97/01/0002",
            "SBI Bad TER Fund",
            datetime.datetime(2026, 4, 1, 0, 0),
            1.61,
            0,
            0,
            0.28,
            1.89,
            1.1,
            0,
            0,
            0.19,
            12.0,  # Direct Plan - Total TER (%) — out of range
        )
        rows = parse_sbi_ter_rows([_REAL_HEADER, bad_row])
        assert len(rows) == 1
        assert rows[0].plan_suffix == " - Regular Plan"
        assert rows[0].ter_pct == 1.89

    def test_zero_ter_dropped(self):
        zero_row = (
            "SBIM/O/E/THE/97/01/0003",
            "SBI Zero TER Fund",
            datetime.datetime(2026, 4, 1, 0, 0),
            0,
            0,
            0,
            0,
            0.0,
            0,
            0,
            0,
            0,
            0.0,
        )
        rows = parse_sbi_ter_rows([_REAL_HEADER, zero_row])
        assert rows == []

    def test_empty_scheme_name_row_skipped(self):
        blank_row = (
            "SBIM/O/E/THE/97/01/0004",
            "",
            datetime.datetime(2026, 4, 1, 0, 0),
            1.0,
            0,
            0,
            0.1,
            1.1,
            0.5,
            0,
            0,
            0.1,
            0.6,
        )
        rows = parse_sbi_ter_rows([_REAL_HEADER, blank_row])
        assert rows == []

    def test_unparseable_date_row_skipped(self):
        bad_date_row = (
            "SBIM/O/E/THE/97/01/0005",
            "SBI Bad Date Fund",
            "not-a-date",
            1.0,
            0,
            0,
            0.1,
            1.1,
            0.5,
            0,
            0,
            0.1,
            0.6,
        )
        rows = parse_sbi_ter_rows([_REAL_HEADER, bad_date_row])
        assert rows == []

    def test_header_missing_required_column_returns_empty(self):
        """Header without a 'Direct Plan - Total TER' column -> [] (fail closed,
        never guesses a substitute column)."""
        broken_header = _REAL_HEADER[:-1]  # drop the last column
        broken_row = _REAL_ROW_1[:-1]
        assert parse_sbi_ter_rows([broken_header, broken_row]) == []

    def test_empty_input_returns_empty_list(self):
        assert parse_sbi_ter_rows([]) == []


# ===========================================================================
# _load_sbi_ter_rows — thin openpyxl I/O wrapper (real xlsx round-trip)
# ===========================================================================


class TestLoadSbiTerRows:
    def test_round_trips_a_real_xlsx(self):
        """Builds a minimal real xlsx in-memory (via openpyxl) with the fixture
        rows and confirms _load_sbi_ter_rows extracts them back out unchanged —
        exercises the actual openpyxl read path, not just the pure parser."""
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        for row in VALID_FIXTURE_ROWS:
            ws.append(list(row))
        buf = io.BytesIO()
        wb.save(buf)

        loaded = _load_sbi_ter_rows(buf.getvalue())
        assert len(loaded) == 3
        assert loaded[0][1] == "Scheme Name"
        assert loaded[1][1] == "SBI ESG Exclusionary Strategy Fund"


# ===========================================================================
# fetch_sbi_expense_ratios — async fetcher, injected httpx client
# ===========================================================================


def _real_xlsx_bytes() -> bytes:
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    for row in VALID_FIXTURE_ROWS:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestFetchSbiExpenseRatios:
    async def test_happy_path_real_fixture(self):
        client = _make_fake_client(_real_xlsx_bytes())
        rows, status = await fetch_sbi_expense_ratios(client)
        assert len(rows) == 4
        assert status["ok"] == ["SBI"]
        assert status["unreachable"] == []
        assert status["format_mismatch"] == []
        assert status["bot_blocked"] == []

    async def test_http_non_200_marks_unreachable(self):
        client = _make_fake_client(b"", status_code=404)
        rows, status = await fetch_sbi_expense_ratios(client)
        assert rows == []
        assert status["unreachable"] == ["SBI"]

    async def test_network_error_marks_unreachable(self):
        client = MagicMock()
        client.get = AsyncMock(side_effect=httpx.ConnectTimeout("timed out"))
        rows, status = await fetch_sbi_expense_ratios(client)
        assert rows == []
        assert status["unreachable"] == ["SBI"]

    async def test_corrupt_xlsx_bytes_marks_format_mismatch(self):
        client = _make_fake_client(b"not a real xlsx file")
        rows, status = await fetch_sbi_expense_ratios(client)
        assert rows == []
        assert status["format_mismatch"] == ["SBI"]

    async def test_valid_xlsx_zero_usable_rows_marks_format_mismatch(self):
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Some Other Header"])  # no recognizable columns at all
        buf = io.BytesIO()
        wb.save(buf)

        client = _make_fake_client(buf.getvalue())
        rows, status = await fetch_sbi_expense_ratios(client)
        assert rows == []
        assert status["format_mismatch"] == ["SBI"]
