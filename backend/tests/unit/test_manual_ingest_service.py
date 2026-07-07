"""
Unit tests for the manual disclosure inbox's shared intake helpers
(dhanradar/mf/manual_ingest.py) — pure logic only, no DB/Redis/network.

`intake_file()` itself (dedup + persist + enqueue) needs a real DB session
(TaskSessionLocal) and is covered as an integration test instead —
tests/integration/test_manual_ingest_intake.py.
"""

from __future__ import annotations

import io
import zipfile
from datetime import date

from openpyxl import Workbook

from dhanradar.mf.manual_ingest import (
    ALLOWED_EXTENSIONS,
    MAX_BYTES,
    MAX_ZIP_MEMBERS,
    MAX_ZIP_TOTAL_BYTES,
    detect_amc,
    detect_amc_and_parse,
    detect_period_from_filename,
    expand_zip,
    sha256_bytes,
    stored_path_for,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------


def test_allowed_extensions_is_xls_xlsx_pdf():
    """Contract §2 — PDFs are accepted + archived (never parsed) this wave."""
    assert set(ALLOWED_EXTENSIONS) == {".xls", ".xlsx", ".pdf"}


def test_max_bytes_is_25mb():
    assert MAX_BYTES == 25 * 1024 * 1024


# ---------------------------------------------------------------------------
# sha256_bytes
# ---------------------------------------------------------------------------


def test_sha256_bytes_deterministic_and_distinct():
    a = sha256_bytes(b"hello world")
    b = sha256_bytes(b"hello world")
    c = sha256_bytes(b"hello world!")
    assert a == b
    assert a != c
    assert len(a) == 64  # hex digest


# ---------------------------------------------------------------------------
# stored_path_for — uuid-named on disk, extension from original_filename
# ---------------------------------------------------------------------------


def test_stored_path_for_uses_uuid_name_never_original_filename():
    path = stored_path_for("abc-123", "some user supplied name; rm -rf .xlsx")
    assert path.name == "abc-123.xlsx"
    assert "rm -rf" not in str(path)


def test_stored_path_for_xls_extension():
    path = stored_path_for("abc-123", "HDFC_disclosure.xls")
    assert path.suffix == ".xls"


# ---------------------------------------------------------------------------
# detect_amc — keyword matching (filename OR scheme name)
# ---------------------------------------------------------------------------


def test_detect_amc_matches_each_bot_blocked_amc():
    assert detect_amc("HDFC_Portfolio_May2026.xlsx") == "HDFC"
    assert detect_amc("SBI Mutual Fund disclosure") == "SBI"
    assert detect_amc("ICICI Prudential Bluechip Fund") == "ICICI_PRU"
    assert detect_amc("kotak-disclosure-2026-05.xlsx") == "KOTAK"
    assert detect_amc("AXIS Long Term Equity Fund") == "AXIS"


def test_detect_amc_returns_none_when_no_keyword_matches():
    assert detect_amc("generic_disclosure_file.xlsx") is None


def test_detect_amc_matches_edelweiss_filename_prefix_and_folder_hint():
    """Edelweiss's own filename convention abbreviates to "EDEL_" (not the full
    "EDELWEISS_"), while the watched-subfolder hint is the full "Edelweiss" name —
    one keyword ("edel") must match both forms."""
    assert detect_amc("EDEL_Monthly_Portfolio_May2026.xlsx") == "EDELWEISS"
    assert detect_amc("Edelweiss") == "EDELWEISS"
    assert detect_amc("Edelweiss Mutual Fund disclosure") == "EDELWEISS"


def test_detect_amc_matches_ppfas_filename_and_scheme_name():
    """PPFAS files say 'PPFAS' or 'Monthly-Portfolio' on amc.ppfas.com, but the
    SCHEMES are named 'Parag Parikh ...' — both forms must resolve (2026-07-07,
    PPFAS added to the automated scraper roots)."""
    assert detect_amc("PPFAS_Monthly-Portfolio-june-2026.xls") == "PPFAS"
    assert detect_amc("Parag Parikh Flexi Cap Fund") == "PPFAS"


def test_detect_amc_matches_absl_filename_and_scheme_name():
    """ABSL files say 'ABSLMF'; schemes say 'Aditya Birla Sun Life ...' (2026-07-07,
    founder's manual batch)."""
    assert detect_amc("31052026_ABSLMF_Monthly-Portfolio.zip") == "ABSL"
    assert detect_amc("Aditya Birla Sun Life Frontline Equity Fund") == "ABSL"


def test_detect_amc_matches_hsbc_filename_and_folder_hint():
    """HSBC is manual-only (site unreachable from KVM4); founder files are
    'hsbc-<scheme>-31-may-2026.xlsx' in an incoming/HSBC/ folder (2026-07-07)."""
    assert detect_amc("hsbc-multi-cap-fund-31-may-2026.xlsx") == "HSBC"
    assert detect_amc("HSBC") == "HSBC"


def test_ppfas_registered_in_scraper_roots_with_resolver_prefix():
    """The scraper root + the resolver's name-prefix override must stay in sync —
    a root without the 'Parag Parikh%' prefix would resolve zero schemes."""
    import inspect

    from dhanradar.tasks.mf import _AMC_DISCLOSURE_ROOTS, _resolve_scheme_isins  # noqa: F401

    roots = {r["name"] for r in _AMC_DISCLOSURE_ROOTS}
    assert "PPFAS" in roots
    src = inspect.getsource(_resolve_scheme_isins)
    assert "Parag Parikh%" in src


# ---------------------------------------------------------------------------
# detect_amc_and_parse — real .xlsx bytes through the EXISTING SEBI parser
# (_parse_sebi_xlsx, dhanradar/tasks/mf.py) — never a second parser.
# ---------------------------------------------------------------------------


def _build_disclosure_xlsx(scheme_name: str) -> bytes:
    """A trimmed but structurally real SEBI-format monthly disclosure: an
    'as on' date header, one scheme-name row, a column header row, two holding
    rows, and a 'Net Assets' total row (the AUM-extraction row, §8.4)."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Portfolio as on 31-May-2026"])
    ws.append([scheme_name])
    ws.append(["Name of Instrument", "ISIN", "% to NAV", "Market Value"])
    ws.append(["HDFC Bank Ltd", "INE040A01034", "5.00", "1234.56"])
    ws.append(["ICICI Bank Ltd", "INE090A01021", "4.50", "1000.00"])
    ws.append(["Net Assets", "", "100.00", "24000.00"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_detect_amc_and_parse_filename_detection():
    data = _build_disclosure_xlsx("HDFC Flexi Cap Fund")
    amc_name, period, rows = detect_amc_and_parse(data, "HDFC_May2026_Portfolio.xlsx")
    assert amc_name == "HDFC"
    assert period == date(2026, 5, 1)
    assert len(rows) == 3  # 2 holdings + 1 total row
    assert any(r["constituent_name"] == "HDFC Bank Ltd" for r in rows)
    total_row = next(r for r in rows if r.get("is_total_row"))
    assert total_row["market_value_cr"] == 240.0  # 24000.00 lakh -> crore


def test_detect_amc_and_parse_edelweiss_abbreviated_filename_prefix():
    """Edelweiss files arrive as "EDEL_..." (abbreviated), not "EDELWEISS_..." —
    filename detection must still resolve the full AMC name."""
    data = _build_disclosure_xlsx("Edelweiss Large Cap Fund")
    amc_name, period, rows = detect_amc_and_parse(data, "EDEL_Monthly_Portfolio_May2026.xlsx")
    assert amc_name == "EDELWEISS"
    assert period == date(2026, 5, 1)
    assert len(rows) == 3


def test_detect_amc_and_parse_falls_back_to_scheme_name_when_filename_uninformative():
    """A generically-named upload ('disclosure.xlsx') still resolves the AMC from
    the parsed scheme name — filename detection is a shortcut, not the only path."""
    data = _build_disclosure_xlsx("HDFC Flexi Cap Fund")
    amc_name, period, rows = detect_amc_and_parse(data, "disclosure.xlsx")
    assert amc_name == "HDFC"
    assert period == date(2026, 5, 1)
    assert len(rows) == 3


def test_detect_amc_and_parse_undetectable_amc_returns_none():
    """Neither the filename nor the scheme name names a known AMC — the caller
    (tasks/manual_ingest.py::parse_manual_disclosure_file) treats this as
    status='unsupported', never a guess."""
    data = _build_disclosure_xlsx("XYZ Multi Cap Fund")
    amc_name, _period, rows = detect_amc_and_parse(data, "disclosure_file.xlsx")
    assert amc_name is None
    assert len(rows) == 3  # still parsed — only AMC detection failed


# ---------------------------------------------------------------------------
# detect_amc_and_parse — amc_hint fallback (contract §3: per-AMC subfolders)
# ---------------------------------------------------------------------------


def test_detect_amc_and_parse_uses_hint_when_filename_and_scheme_undetectable():
    """Neither filename nor scheme name names a known AMC — the subfolder hint
    is the LAST-resort fallback."""
    data = _build_disclosure_xlsx("XYZ Multi Cap Fund")
    amc_name, _period, _rows = detect_amc_and_parse(data, "disclosure.xlsx", amc_hint="HDFC")
    assert amc_name == "HDFC"


def test_detect_amc_and_parse_filename_wins_over_hint():
    """Filename/content detection runs FIRST — a hint from the wrong subfolder
    must never override a real filename match."""
    data = _build_disclosure_xlsx("HDFC Flexi Cap Fund")
    amc_name, _period, _rows = detect_amc_and_parse(
        data, "HDFC_May2026_Portfolio.xlsx", amc_hint="SBI"
    )
    assert amc_name == "HDFC"


def test_detect_amc_and_parse_ignores_unknown_hint():
    """An unrecognized folder name (not in the known AMC keyword set) is
    ignored — amc_name stays None, never a guessed value."""
    data = _build_disclosure_xlsx("XYZ Multi Cap Fund")
    amc_name, _period, _rows = detect_amc_and_parse(
        data, "disclosure.xlsx", amc_hint="SomeUnknownFundHouse"
    )
    assert amc_name is None


# ---------------------------------------------------------------------------
# expand_zip — zip-bomb / abuse guards (contract §1)
# ---------------------------------------------------------------------------


def _zip_bytes(members: dict[str, bytes], *, dirs: list[str] | None = None) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        for name in dirs or []:
            zf.writestr(name + "/", "")
        for name, data in members.items():
            zf.writestr(name, data)
    return buf.getvalue()


def test_expand_zip_happy_path_multi_member():
    data = _zip_bytes(
        {
            "HDFC_June2026.xlsx": b"member-a-bytes",
            "SBI_June2026.xls": b"member-b-bytes",
            "notice.pdf": b"%PDF-1.4 fake",
        }
    )
    result = expand_zip(data, "bundle.zip")
    assert result.skipped == []
    names = {name for name, _data in result.eligible}
    assert names == {"HDFC_June2026.xlsx", "SBI_June2026.xls", "notice.pdf"}
    by_name = dict(result.eligible)
    assert by_name["HDFC_June2026.xlsx"] == b"member-a-bytes"


def test_expand_zip_skips_directories():
    data = _zip_bytes({"a.xlsx": b"bytes"}, dirs=["some_dir"])
    result = expand_zip(data, "bundle.zip")
    assert len(result.eligible) == 1
    assert result.skipped == []


def test_expand_zip_rejects_encrypted_zip_via_flag_bit(monkeypatch):
    """Stdlib zipfile can't WRITE a real encrypted zip (`writestr` always resets
    flag_bits), so fake the PKWARE traditional-encryption bit (bit 0 of
    ZipInfo.flag_bits — the same signal a real encrypted member carries) by
    patching infolist() to return it set, exactly what expand_zip() reads."""
    data = _zip_bytes({"secret.xlsx": b"irrelevant-if-encrypted"})

    real_infolist = zipfile.ZipFile.infolist

    def _fake_infolist(self):
        infos = real_infolist(self)
        for info in infos:
            info.flag_bits |= 0x1
        return infos

    monkeypatch.setattr(zipfile.ZipFile, "infolist", _fake_infolist)

    result = expand_zip(data, "bundle.zip")
    assert result.eligible == []
    assert result.skipped == [("bundle.zip", "encrypted_zip")]


def test_expand_zip_rejects_too_many_members():
    members = {f"file_{i}.xlsx": b"x" for i in range(MAX_ZIP_MEMBERS + 1)}
    data = _zip_bytes(members)
    result = expand_zip(data, "bundle.zip")
    assert result.eligible == []
    assert result.skipped == [("bundle.zip", "too_many_members")]


def test_expand_zip_skips_oversized_member():
    data = _zip_bytes({"huge.xlsx": b"0" * (MAX_BYTES + 1), "ok.xlsx": b"fine"})
    result = expand_zip(data, "bundle.zip")
    names = {name for name, _data in result.eligible}
    assert names == {"ok.xlsx"}
    assert ("huge.xlsx", "file_too_large") in result.skipped


def test_expand_zip_skips_when_total_uncompressed_exceeds_cap():
    # Members each well under the 25MB per-file cap, but summing past the
    # total cap — each is all-zero bytes so the ACTUAL zip stays tiny (this is
    # the zip-bomb shape the total-cap guard exists for). Member size is pinned
    # under MAX_BYTES and the count derived from the budget so the test tracks
    # the constants (300MB budget as of 2026-07-07, not the original 100MB).
    member_size = 20 * 1024 * 1024  # 20MB — under the per-member cap
    count = MAX_ZIP_TOTAL_BYTES // member_size + 2  # cumulative > total cap
    members = {f"member_{i}.xlsx": b"0" * member_size for i in range(count)}
    data = _zip_bytes(members)
    result = expand_zip(data, "bundle.zip")
    assert 0 < len(result.eligible) < count  # some fit the budget, the rest don't
    reasons = {reason for _name, reason in result.skipped}
    assert "zip_total_size_exceeded" in reasons


def test_expand_zip_skips_nested_zip_never_recursed():
    inner = _zip_bytes({"HDFC.xlsx": b"inner-bytes"})
    data = _zip_bytes({"inner.zip": inner, "outer.xlsx": b"outer-bytes"})
    result = expand_zip(data, "bundle.zip")
    names = {name for name, _data in result.eligible}
    assert names == {"outer.xlsx"}
    assert ("inner.zip", "nested_zip") in result.skipped


def test_expand_zip_rejects_corrupt_zip():
    result = expand_zip(b"not-a-real-zip-file-at-all", "bundle.zip")
    assert result.eligible == []
    assert result.skipped == [("bundle.zip", "corrupt_zip")]


def test_expand_zip_skips_unsupported_extension_member():
    data = _zip_bytes({"readme.txt": b"hello", "ok.xlsx": b"fine"})
    result = expand_zip(data, "bundle.zip")
    names = {name for name, _data in result.eligible}
    assert names == {"ok.xlsx"}
    assert ("readme.txt", "unsupported_extension:.txt") in result.skipped


# ---------------------------------------------------------------------------
# Parser-tolerance fixes (2026-07-07) — real-file evidence from the founder's
# first 120-file manual-ingest run: KOTAK consolidated, EDELWEISS monthly, and
# filename-only periods all parsed to 0 rows before these fixes.
# ---------------------------------------------------------------------------


def test_detect_period_from_filename_variants():
    # KOTAK consolidated: month name + year, no separator.
    assert detect_period_from_filename("KOTAKConsolidatedSEBIPortfolioMay2026.xlsx") == date(
        2026, 5, 1
    )
    # KOTAK fortnightly: month name + day + year, no separator.
    assert detect_period_from_filename("KotakFortnightlyPortfolioJune302026.xlsx") == date(
        2026, 6, 1
    )
    # EDELWEISS: DD-Mon-YYYY with trailing upload timestamp noise.
    assert detect_period_from_filename(
        "EDEL_Portfolio Monthly Notes 31-May-2026_10062026_021941_PM.xlsx"
    ) == date(2026, 5, 1)
    # Underscore-separated month + year.
    assert detect_period_from_filename("Monthly AAUM_May 2026_09062026.xlsx") == date(2026, 5, 1)
    # No period present — fail closed.
    assert detect_period_from_filename("SBIcurrent-year-ter.xlsx") is None


def _build_edel_style_xlsx() -> bytes:
    """EDELWEISS layout: banner title, header row, then BLANK rows before the
    first holding — the blank rows must NOT wipe the freshly-built col_map."""
    wb = Workbook()
    ws = wb.active
    ws.append(["PORTFOLIO STATEMENT OF EDELWEISS BANKING AND PSU DEBT FUND"])
    ws.append(["(An open-ended debt scheme)"])
    ws.append(
        [
            "Name of the Instrument",
            "ISIN",
            "Rating/Industry",
            "Quantity",
            "Market/Fair Value(Rs. In Lacs)",
            "% to Net Assets",
        ]
    )
    ws.append([])
    ws.append([])
    ws.append(["7.18% GOI 2033", "IN0020230085", "Sovereign", "500000", "512.34", "8.10"])
    ws.append(["HDFC Bank Ltd", "INE040A01034", "Banks", "1000", "234.56", "3.70"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_blank_rows_after_header_do_not_reset_col_map():
    amc, period, rows = detect_amc_and_parse(
        _build_edel_style_xlsx(),
        "EDEL_Portfolio Monthly Notes 31-May-2026_10062026_021941_PM.xlsx",
    )
    assert amc == "EDELWEISS"
    assert len(rows) == 2
    # Banner prefix stripped from the scheme name.
    assert rows[0]["scheme_name"] == "EDELWEISS BANKING AND PSU DEBT FUND"
    # Period recovered from the filename (sheet has no as-on line) and stamped on rows.
    assert period == date(2026, 5, 1)
    assert all(r["as_of_month"] == date(2026, 5, 1) for r in rows)


def _build_kotak_style_xlsx() -> bytes:
    """KOTAK layout: name header label in col 0 but every VALUE in col 2 (merged
    name block), ISIN labeled "ISIN Code" in col 3."""
    wb = Workbook()
    ws = wb.active
    ws.append([None, None, "Portfolio of Kotak Nifty India Tourism Index Fund"])
    ws.append(["Name of Instrument", None, None, "ISIN Code", "Industry", "Yield", "Quantity"])
    ws.append(["Equity & Equity related"])
    ws.append(
        [None, " ", "INDIAN HOTELS CO LTD", "INE053A01029", "Leisure Services", None, "103736"]
    )
    ws.append(
        [
            None,
            " ",
            "GMR AIRPORTS LIMITED",
            "INE776C01039",
            "Transport Infrastructure",
            None,
            "570973",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_offset_name_column_recovered_via_isin_and_banner_stripped():
    amc, period, rows = detect_amc_and_parse(
        _build_kotak_style_xlsx(), "KOTAKConsolidatedSEBIPortfolioMay2026.xlsx"
    )
    assert amc == "KOTAK"
    assert period == date(2026, 5, 1)
    names = {r["constituent_name"] for r in rows}
    assert "INDIAN HOTELS CO LTD" in names
    assert "GMR AIRPORTS LIMITED" in names
    # "Portfolio of " banner prefix stripped for scheme-name resolution.
    assert all(r["scheme_name"] == "Kotak Nifty India Tourism Index Fund" for r in rows)


def test_zip_member_cap_admits_real_sbi_sized_bundles():
    """SBI's real zips carry 441/57/156 members — the cap must admit them; the
    bomb guard is the total-uncompressed budget, not the member count."""
    assert MAX_ZIP_MEMBERS >= 500
    members = {f"scheme_{i}.xlsx": b"x" for i in range(441)}
    result = expand_zip(_zip_bytes(members), "SBIopen_ended_schemes.zip")
    assert len(result.eligible) == 441
    assert result.skipped == []
