"""
Integration test for the manual disclosure inbox parse task
(dhanradar/tasks/manual_ingest.py::parse_manual_disclosure_file) on a real
trimmed SEBI-format disclosure fixture — exercises the ACTUAL byte-parsing
path (openpyxl.load_workbook via _parse_sebi_xlsx, the SAME parser
mf_constituents_fetch uses), not a synthetic dict.

`_upsert_constituents` (dhanradar/tasks/mf.py) is monkeypatched: it resolves
scheme names to ISINs via a live pg_trgm similarity query against mf.mf_funds,
which this suite has no existing fixture-builder for (no test anywhere
exercises `_upsert_constituents` end-to-end yet — a pre-existing gap, not
introduced here). This test instead proves the manual-ingest ORCHESTRATION —
detect AMC/period from real bytes, route through the shared upsert function
with the RIGHT arguments (rows, amc_name, run_id), and record the outcome on
the mf.manual_ingest_files row — end to end.
"""

from __future__ import annotations

import io
import uuid

import pytest
from openpyxl import Workbook

from dhanradar.mf.manual_ingest import stored_path_for
from dhanradar.models.mf import MfManualIngestFile
from dhanradar.tasks import manual_ingest as mi

pytestmark = pytest.mark.integration


def _build_hdfc_disclosure_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["Portfolio as on 30-June-2026"])
    ws.append(["HDFC Flexi Cap Fund"])
    ws.append(["Name of Instrument", "ISIN", "% to NAV", "Market Value"])
    ws.append(["HDFC Bank Ltd", "INE040A01034", "5.00", "1234.56"])
    ws.append(["Reliance Industries Ltd", "INE002A01018", "4.50", "1000.00"])
    ws.append(["Net Assets", "", "100.00", "24000.00"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _insert_pending_row(
    db_session, *, filename: str, channel: str = "folder"
) -> tuple[str, bytes]:
    from dhanradar.mf.manual_ingest import sha256_bytes

    data = _build_hdfc_disclosure_xlsx()
    file_id = uuid.uuid4()
    db_session.add(
        MfManualIngestFile(
            id=file_id,
            sha256=sha256_bytes(data),
            original_filename=filename,
            channel=channel,
            status="pending",
        )
    )
    await db_session.commit()
    return str(file_id), data


@pytest.fixture(autouse=True)
def _tmp_store(monkeypatch: pytest.MonkeyPatch, tmp_path):
    from dhanradar.config import settings

    monkeypatch.setattr(settings, "MANUAL_INGEST_DIR", str(tmp_path))


async def test_parse_pipeline_parses_real_xlsx_and_marks_row_parsed(
    db_session, monkeypatch: pytest.MonkeyPatch
):
    file_id, data = await _insert_pending_row(db_session, filename="HDFC_June2026.xlsx")
    stored_path_for(file_id, "HDFC_June2026.xlsx").write_bytes(data)

    captured: dict = {}

    async def _fake_upsert(rows, amc_name, run_id=None):
        captured["rows"] = rows
        captured["amc_name"] = amc_name
        captured["run_id"] = run_id
        return len(rows), 1  # (rows_upserted, aum_updates)

    monkeypatch.setattr("dhanradar.tasks.mf._upsert_constituents", _fake_upsert)

    result = await mi._parse_pipeline(file_id)

    assert result.startswith("parsed:")
    assert captured["amc_name"] == "HDFC"
    assert len(captured["rows"]) == 3  # 2 holdings + 1 total row — real parser output
    assert captured["run_id"] is not None  # ingestion_run() provenance threaded through

    # _parse_pipeline wrote via a DIFFERENT session (TaskSessionLocal) — this
    # session's identity map still holds the pre-parse object (expire_on_commit=
    # False), so it must be told to re-read from the DB rather than serve a cache hit.
    db_session.expire_all()
    row = await db_session.get(MfManualIngestFile, uuid.UUID(file_id))
    assert row.status == "parsed"
    assert row.amc_detected == "HDFC"
    assert row.rows_ingested == 3
    assert row.period_detected is not None
    assert row.parsed_at is not None


async def test_parse_pipeline_amc_undetectable_marks_unsupported(
    db_session, monkeypatch: pytest.MonkeyPatch
):
    """A scheme name that names no known AMC → 'unsupported', never a guess, and
    _upsert_constituents is never called (nothing to route through)."""
    from dhanradar.mf.manual_ingest import sha256_bytes

    wb = Workbook()
    ws = wb.active
    ws.append(["Portfolio as on 30-June-2026"])
    ws.append(["XYZ Multi Cap Fund"])
    ws.append(["Name of Instrument", "ISIN", "% to NAV", "Market Value"])
    ws.append(["Some Stock Ltd", "INE999Z99999", "5.00", "100.00"])
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    file_id = uuid.uuid4()
    db_session.add(
        MfManualIngestFile(
            id=file_id,
            sha256=sha256_bytes(data),
            original_filename="unknown_amc.xlsx",
            channel="folder",
            status="pending",
        )
    )
    await db_session.commit()
    stored_path_for(str(file_id), "unknown_amc.xlsx").write_bytes(data)

    def _boom(*_a, **_kw):
        raise AssertionError("_upsert_constituents must never run for an undetectable AMC")

    monkeypatch.setattr("dhanradar.tasks.mf._upsert_constituents", _boom)

    result = await mi._parse_pipeline(str(file_id))

    assert result.startswith("unsupported:")
    db_session.expire_all()
    row = await db_session.get(MfManualIngestFile, file_id)
    assert row.status == "unsupported"
    assert row.error == "amc_or_period_undetectable"
    assert row.amc_detected is None


async def test_parse_pipeline_missing_stored_file_marks_failed(db_session):
    """The DB row exists but the on-disk file is gone (e.g. a manual delete) —
    fail closed, never crash silently."""
    file_id = uuid.uuid4()
    db_session.add(
        MfManualIngestFile(
            id=file_id,
            sha256="deadbeef" * 8,
            original_filename="missing.xlsx",
            channel="folder",
            status="pending",
        )
    )
    await db_session.commit()

    result = await mi._parse_pipeline(str(file_id))

    assert result == "failed: stored_file_missing"
    db_session.expire_all()
    row = await db_session.get(MfManualIngestFile, file_id)
    assert row.status == "failed"
    assert row.error == "stored_file_missing"


async def test_parse_pipeline_pdf_marks_archived_never_parsed(
    db_session, monkeypatch: pytest.MonkeyPatch
):
    """Contract §2 — the SEBI parser is xlsx-only; a PDF is stored + row kept,
    marked 'archived' immediately, and NEVER routed through detect_amc_and_parse
    / _upsert_constituents (no fake parsing, no OCR)."""
    from dhanradar.mf.manual_ingest import sha256_bytes

    data = b"%PDF-1.4 fake factsheet bytes"
    file_id = uuid.uuid4()
    db_session.add(
        MfManualIngestFile(
            id=file_id,
            sha256=sha256_bytes(data),
            original_filename="HDFC_factsheet_June2026.pdf",
            channel="upload",
            status="pending",
        )
    )
    await db_session.commit()
    stored_path_for(str(file_id), "HDFC_factsheet_June2026.pdf").write_bytes(data)

    def _boom(*_a, **_kw):
        raise AssertionError("a PDF must never be routed through detect_amc_and_parse")

    monkeypatch.setattr("dhanradar.tasks.manual_ingest.detect_amc_and_parse", _boom)
    monkeypatch.setattr(
        "dhanradar.tasks.mf._upsert_constituents",
        lambda *_a, **_kw: (_ for _ in ()).throw(AssertionError("must never upsert a PDF")),
    )

    result = await mi._parse_pipeline(str(file_id))

    assert result == "archived: pdf_saved_for_later"
    db_session.expire_all()
    row = await db_session.get(MfManualIngestFile, file_id)
    assert row.status == "archived"
    assert row.amc_detected is None
    assert row.rows_ingested is None


async def test_parse_pipeline_threads_amc_hint_into_detection(
    db_session, monkeypatch: pytest.MonkeyPatch
):
    """The optional amc_hint argument reaches detect_amc_and_parse as its
    3rd positional/keyword arg — the fallback plumbing contract §3 relies on."""
    file_id, data = await _insert_pending_row(db_session, filename="generic.xlsx")
    stored_path_for(file_id, "generic.xlsx").write_bytes(data)

    captured: dict = {}

    def _fake_detect(data, filename, amc_hint=None):
        captured["amc_hint"] = amc_hint
        return "HDFC", None, []

    monkeypatch.setattr("dhanradar.tasks.manual_ingest.detect_amc_and_parse", _fake_detect)

    result = await mi._parse_pipeline(file_id, amc_hint="HDFC")

    assert captured["amc_hint"] == "HDFC"
    # rows == [] -> amc_or_period_undetectable path (not the point of this test,
    # only that the hint was threaded through).
    assert result == "unsupported: amc_or_period_undetectable"


async def test_parse_pipeline_skips_already_terminal_row(db_session):
    """Idempotency guard: a re-enqueued/duplicate task run must never re-parse a
    row that already reached a terminal status."""
    file_id = uuid.uuid4()
    db_session.add(
        MfManualIngestFile(
            id=file_id,
            sha256="alreadydone" * 6,
            original_filename="already_parsed.xlsx",
            channel="folder",
            status="parsed",
            rows_ingested=5,
        )
    )
    await db_session.commit()

    result = await mi._parse_pipeline(str(file_id))
    assert result == "skip:parsed"


# ---------------------------------------------------------------------------
# File-class dispatcher (2026-07-07) — aaum / riskometer / performance files
# route to their own parsers + mf_funds writers, never the constituents path.
# Real DB writes asserted (no false positives): the actual mf_funds columns
# must change value.
# ---------------------------------------------------------------------------


def _build_riskometer_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "SR No.",
            "Scheme name",
            "Risk-o-meter level as on March 31, 2025",
            "Risk-o-meter level as on March 31, 2026",
            "Changes",
        ]
    )
    ws.append(["1", "ICICI Prudential Multicap Fund", "High", "Very High", "1"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_aaum_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "AAUM disclosure"
    ws.append([])
    ws.append(
        [
            "Sl. No.",
            "Scheme Category/ Scheme Name",
            "ICICI Mutual Fund : Net AAUM as on  2026-05-31",
            None,
            "GRAND TOTAL",
        ]
    )
    ws.append(["A)", "EQUITY SCHEMES", None, None, None])
    ws.append([None, "ICICI Prudential Multicap Fund", 1.0, 2.0, 15432.10])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _seed_fund(db_session, isin: str, name: str) -> None:
    from dhanradar.models.mf import MfFund

    db_session.add(MfFund(isin=isin, scheme_name=name, amc_name="ICICI Prudential Mutual Fund"))
    await db_session.commit()


async def _insert_pending_bytes(db_session, *, filename: str, data: bytes) -> str:
    from dhanradar.mf.manual_ingest import sha256_bytes

    file_id = uuid.uuid4()
    db_session.add(
        MfManualIngestFile(
            id=file_id,
            sha256=sha256_bytes(data),
            original_filename=filename,
            channel="folder",
            status="pending",
        )
    )
    await db_session.commit()
    stored_path_for(str(file_id), filename).write_bytes(data)
    return str(file_id)


async def test_dispatcher_riskometer_writes_mf_funds_band(db_session, monkeypatch):
    isin = "INF109KTEST1"
    await _seed_fund(db_session, isin, "ICICI Prudential Multicap Fund")
    data = _build_riskometer_xlsx()
    file_id = await _insert_pending_bytes(
        db_session,
        filename="ICICIAnnual Disclosure of Scheme Riskometer_FY 2025-26.xlsx",
        data=data,
    )

    async def _fake_resolve(names, amc_name):
        assert amc_name == "ICICI_PRU"
        return {"ICICI Prudential Multicap Fund": isin}

    monkeypatch.setattr("dhanradar.tasks.mf._resolve_scheme_isins", _fake_resolve)

    result = await mi._parse_pipeline(file_id)
    assert result.startswith("parsed: riskometer")

    db_session.expire_all()
    from dhanradar.models.mf import MfFund

    fund = await db_session.get(MfFund, isin)
    assert fund.risk_o_meter == "Very High"  # LATEST FY column, verbatim regulatory band
    row = await db_session.get(MfManualIngestFile, uuid.UUID(file_id))
    assert row.status == "parsed"
    assert row.rows_ingested == 1


async def test_dispatcher_aaum_fills_null_aum_and_history(db_session, monkeypatch):
    from datetime import date as _date

    isin = "INF109KTEST2"
    await _seed_fund(db_session, isin, "ICICI Prudential Multicap Fund")
    data = _build_aaum_xlsx()
    file_id = await _insert_pending_bytes(
        db_session, filename="ICICI Monthly AAUM Disclosure - May 2026.xlsx", data=data
    )

    async def _fake_resolve(names, amc_name):
        return {"ICICI Prudential Multicap Fund": isin}

    monkeypatch.setattr("dhanradar.tasks.mf._resolve_scheme_isins", _fake_resolve)

    result = await mi._parse_pipeline(file_id)
    assert result.startswith("parsed: aaum")

    db_session.expire_all()
    from sqlalchemy import select

    from dhanradar.models.mf import MfAumHistory, MfFund

    fund = await db_session.get(MfFund, isin)
    assert float(fund.aum_crore) == 15432.10
    assert fund.aum_as_of == _date(2026, 5, 1)
    hist = (
        (await db_session.execute(select(MfAumHistory).where(MfAumHistory.isin == isin)))
        .scalars()
        .all()
    )
    assert len(hist) == 1 and float(hist[0].aum_crore) == 15432.10


async def test_dispatcher_aaum_never_clobbers_fresher_net_assets(db_session, monkeypatch):
    """The portfolio disclosure's stated net-assets (same field, fresher month)
    must win over an AAUM figure — fill-if-null-or-older only."""
    from datetime import date as _date

    from dhanradar.models.mf import MfFund

    isin = "INF109KTEST3"
    db_session.add(
        MfFund(
            isin=isin,
            scheme_name="ICICI Prudential Multicap Fund",
            amc_name="ICICI Prudential Mutual Fund",
            aum_crore=16000.00,
            aum_as_of=_date(2026, 6, 1),  # fresher than the AAUM file's May
        )
    )
    await db_session.commit()
    data = _build_aaum_xlsx()
    file_id = await _insert_pending_bytes(
        db_session, filename="ICICI Monthly AAUM Disclosure - May 2026 v2.xlsx", data=data
    )

    async def _fake_resolve(names, amc_name):
        return {"ICICI Prudential Multicap Fund": isin}

    monkeypatch.setattr("dhanradar.tasks.mf._resolve_scheme_isins", _fake_resolve)

    result = await mi._parse_pipeline(file_id)
    # Schemes RESOLVED but zero rows eligible to write (fresher figure already
    # in place) -> a legitimate parsed no-op, never a failure.
    assert result.startswith("parsed: aaum updates=0")

    db_session.expire_all()
    fund = await db_session.get(MfFund, isin)
    assert float(fund.aum_crore) == 16000.00  # untouched
    assert fund.aum_as_of == _date(2026, 6, 1)


# ---------------------------------------------------------------------------
# amfi_aaum (2026-07-10) — AMFI's consolidated quarterly scheme-wise AAUM
# workbook: exact amfi_code join (zero fuzzy resolution), multi-AMC (no AMC
# gate), same fill-if-null-or-older + history rules as 'aaum'.
# ---------------------------------------------------------------------------


def _build_amfi_aaum_xlsx(
    title: str = (
        "Average Assets under Management (AAUM) for the quarter of April - June 2026 (Rs in Lakhs)"
    ),
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append([title])
    ws.append(
        [
            "AMFI Code",
            "Scheme NAV Name",
            "Excluding Fund of Funds - Domestic but including Fund of Funds - Overseas",
            "Fund Of Funds - Domestic",
        ]
    )
    ws.append(["360 ONE Mutual Fund"])  # AMC section row
    ws.append(["Other Scheme - Other  ETFs"])  # category section row
    ws.append(["154366", "360 ONE MSCI India ETF", "277.38", "0"])
    ws.append(["888888", "Scheme Not In Our Master", "1000.0", "0"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _seed_fund_with_code(db_session, isin: str, name: str, amfi_code: str, **kwargs) -> None:
    from dhanradar.models.mf import MfFund

    db_session.add(
        MfFund(
            isin=isin,
            scheme_name=name,
            amc_name="360 ONE Mutual Fund",
            amfi_code=amfi_code,
            **kwargs,
        )
    )
    await db_session.commit()


async def test_dispatcher_amfi_aaum_joins_by_code_and_writes_every_isin(db_session):
    """The AMFI code joins exactly on mf_funds.amfi_code — BOTH plan-variant
    ISINs under one code get the value; no name resolution is involved; the
    multi-AMC file needs no detectable AMC in its filename."""
    from datetime import date as _date

    from sqlalchemy import select

    from dhanradar.models.mf import MfAumHistory, MfFund

    await _seed_fund_with_code(db_session, "INF579MTEST1", "360 ONE MSCI India ETF", "154366")
    await _seed_fund_with_code(
        db_session, "INF579MTEST2", "360 ONE MSCI India ETF - IDCW", "154366"
    )
    data = _build_amfi_aaum_xlsx()
    file_id = await _insert_pending_bytes(db_session, filename="average-aum2.xlsx", data=data)

    result = await mi._parse_pipeline(file_id)
    assert result.startswith("parsed: amfi_aaum updates=2")

    db_session.expire_all()
    for isin in ("INF579MTEST1", "INF579MTEST2"):
        fund = await db_session.get(MfFund, isin)
        assert float(fund.aum_crore) == 2.77  # 277.38 lakhs → 2.7738 cr (col is 2dp)
        assert fund.aum_as_of == _date(2026, 6, 1)  # quarter END month, day 1
        hist = (
            (await db_session.execute(select(MfAumHistory).where(MfAumHistory.isin == isin)))
            .scalars()
            .all()
        )
        assert len(hist) == 1
        assert hist[0].source == "AMFI"
        assert hist[0].run_id is not None
    row = await db_session.get(MfManualIngestFile, uuid.UUID(file_id))
    assert row.status == "parsed"


async def test_dispatcher_amfi_aaum_never_clobbers_fresher_net_assets(db_session):
    """A fund whose stated net-assets figure is FRESHER than the AAUM quarter
    is never overwritten — parsed no-op, exactly like the 'aaum' class."""
    from datetime import date as _date

    from dhanradar.models.mf import MfFund

    await _seed_fund_with_code(
        db_session,
        "INF579MTEST3",
        "360 ONE MSCI India ETF",
        "154366",
        aum_crore=999.99,
        aum_as_of=_date(2026, 7, 1),  # fresher than the file's June quarter-end
    )
    data = _build_amfi_aaum_xlsx()
    file_id = await _insert_pending_bytes(db_session, filename="average-aum2 (1).xlsx", data=data)

    result = await mi._parse_pipeline(file_id)
    assert result.startswith("parsed: amfi_aaum updates=0")

    db_session.expire_all()
    fund = await db_session.get(MfFund, "INF579MTEST3")
    assert float(fund.aum_crore) == 999.99  # untouched
    assert fund.aum_as_of == _date(2026, 7, 1)


async def test_dispatcher_amfi_aaum_periodless_writes_no_as_of_and_no_history(db_session):
    """Title AND filename both period-less → aum_crore fills only truly-empty
    funds; aum_as_of stays NULL and no history row is written (§8.4)."""
    from sqlalchemy import select

    from dhanradar.models.mf import MfAumHistory, MfFund

    await _seed_fund_with_code(db_session, "INF579MTEST4", "360 ONE MSCI India ETF", "154366")
    data = _build_amfi_aaum_xlsx(title="Average Assets under Management (AAUM) (Rs in Lakhs)")
    file_id = await _insert_pending_bytes(db_session, filename="average-aum-x.xlsx", data=data)

    result = await mi._parse_pipeline(file_id)
    assert result.startswith("parsed: amfi_aaum updates=1")

    db_session.expire_all()
    fund = await db_session.get(MfFund, "INF579MTEST4")
    assert float(fund.aum_crore) == 2.77
    assert fund.aum_as_of is None  # never fabricated
    hist = (
        (await db_session.execute(select(MfAumHistory).where(MfAumHistory.isin == "INF579MTEST4")))
        .scalars()
        .all()
    )
    assert hist == []  # a history row needs a real as_of_month


# ---------------------------------------------------------------------------
# AMFI consolidated TER (2026-07-10) — the multi-AMC 'ter' file runs with
# amc_name=None; the REAL plan resolver (no monkeypatch) routes Regular vs
# Direct values to the right seeded plan-variant ISINs via bare-prefix ILIKE.
# ---------------------------------------------------------------------------


def _build_amfi_ter_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "TER_Revised"
    ws.append(
        [
            "NSDL Scheme Code",
            "Scheme Name",
            "Scheme Type",
            "Scheme Category",
            "TER Date",
            "Regular Plan - Base Expense Ratio (BER) (%)",
            "Regular Plan - Brokerage cost (%)",
            "Regular Plan - Transaction Cost incurred for the purpose of execution of trade (%)",
            "Regular Plan - Statutory Levies (including GST) (%)",
            "Regular Plan - Total TER (%)",
            "Direct Plan - Base Expense Ratio (BER) (%)",
            "Direct Plan - Brokerage cost (%)",
            "Direct Plan - Transaction Cost incurred for the purpose of execution of trade (%)",
            "Direct Plan - Statutory Levies (including GST) (%)",
            "Direct Plan - Total TER (%)",
        ]
    )
    from datetime import datetime as _dt

    # Two days for the same scheme — the LATEST (07-08) must win. TER Date
    # cells are genuine Excel dates (as in the real portal file).
    for day, reg_total, dir_total in (
        (_dt(2026, 7, 1), "2.00", "0.50"),
        (_dt(2026, 7, 8), "2.19", "0.62"),
    ):
        ws.append(
            [
                "AMFT/O/H/ALP/23/07/0001",
                "AMFITEST Alpha Fund",
                "Open Ended",
                "Hybrid Scheme - Balanced Hybrid ",
                day,
                "1.78",
                "0.03",
                "0",
                "0.38",
                reg_total,
                "0.46",
                "0.03",
                "0",
                "0.14",
                dir_total,
            ]
        )
    # A scheme absent from our master — must write nothing, never guess.
    ws.append(
        [
            "AMFT/O/H/GHO/23/07/0002",
            "AMFITEST Ghost Fund",
            "Open Ended",
            "Equity Scheme",
            _dt(2026, 7, 8),
            "1.00",
            "0",
            "0",
            "0.18",
            "1.18",
            "0.50",
            "0",
            "0",
            "0.09",
            "0.59",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def test_dispatcher_amfi_ter_plan_routing_with_no_amc(db_session):
    """End-to-end on the REAL resolver: the AMC-less consolidated TER file
    writes the Regular value to every Regular plan-variant ISIN and the
    Direct value to the Direct one; latest TER Date wins; a scheme absent
    from the master writes nothing."""
    from dhanradar.models.mf import MfFund

    for isin, name in (
        ("INFAMFTTER01", "AMFITEST Alpha Fund - Regular Plan - Growth"),
        ("INFAMFTTER02", "AMFITEST Alpha Fund - Regular Plan - IDCW"),
        ("INFAMFTTER03", "AMFITEST Alpha Fund - Direct Plan - Growth"),
    ):
        db_session.add(MfFund(isin=isin, scheme_name=name, amc_name="AMFITEST Mutual Fund"))
    await db_session.commit()

    data = _build_amfi_ter_xlsx()
    # The real portal filename: matches \bter\b (class 'ter') but names no AMC.
    file_id = await _insert_pending_bytes(db_session, filename="ter-of-mf-schemes.xlsx", data=data)

    result = await mi._parse_pipeline(file_id)
    assert result.startswith("parsed: ter")

    db_session.expire_all()
    for isin in ("INFAMFTTER01", "INFAMFTTER02"):
        fund = await db_session.get(MfFund, isin)
        assert float(fund.expense_ratio_pct) == 2.19  # Regular Total TER, latest date
    direct = await db_session.get(MfFund, "INFAMFTTER03")
    assert float(direct.expense_ratio_pct) == 0.62  # Direct Total TER, latest date
    row = await db_session.get(MfManualIngestFile, uuid.UUID(file_id))
    assert row.status == "parsed"
    assert row.amc_detected is None  # honestly multi-AMC — never a guessed AMC


# ---------------------------------------------------------------------------
# scheme_not_in_master (2026-07-10) — closed-ended/matured schemes classify as
# an HONEST TERMINAL outcome ('unsupported'), never an endlessly retryable
# 'failed' row. Fixture names are real prod failures (93-row tail, 2026-07-10).
# ---------------------------------------------------------------------------


def _build_series_xlsx(banner: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["SBI MUTUAL FUND"])
    ws.append(["SCHEME NAME :", banner])
    ws.append(["PORTFOLIO STATEMENT AS ON :", "2018-08-31"])
    ws.append(["Name of the Instrument / Issuer", "ISIN", "Quantity", "Market value", "% to AUM"])
    ws.append(["Govt Stock", "IN0020180058", "1000", "230.25", "5.10"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def test_zero_rows_closed_ended_marks_scheme_not_in_master(
    db_session, monkeypatch: pytest.MonkeyPatch
):
    """Series banner + failed lookup (re-checked, not assumed) → terminal
    'unsupported: scheme_not_in_master'."""
    data = _build_series_xlsx("SBI Debt Fund Series C-1")
    file_id = await _insert_pending_bytes(
        db_session, filename="SBI Debt Fund Series C - 1.xlsx", data=data
    )

    async def _fake_upsert(rows, amc_name, run_id=None):
        return 0, 0

    async def _fake_resolve(names, amc_name):
        return {}

    monkeypatch.setattr("dhanradar.tasks.mf._upsert_constituents", _fake_upsert)
    monkeypatch.setattr("dhanradar.tasks.mf._resolve_scheme_isins", _fake_resolve)

    result = await mi._parse_pipeline(file_id)

    assert result == "unsupported: scheme_not_in_master"
    db_session.expire_all()
    row = await db_session.get(MfManualIngestFile, uuid.UUID(file_id))
    assert row.status == "unsupported"  # terminal — failed-row resets skip it
    assert row.error == "scheme_not_in_master"
    assert row.amc_detected == "SBI"


async def test_zero_rows_open_ended_stays_retryable_failed(
    db_session, monkeypatch: pytest.MonkeyPatch
):
    """Same zero-rows outcome for an OPEN-ended scheme → the pre-existing
    retryable failure, never silently reclassified."""
    data = _build_series_xlsx("SBI Corporate Bond Fund")
    file_id = await _insert_pending_bytes(
        db_session, filename="SBI Corporate Bond Fund.xlsx", data=data
    )

    async def _fake_upsert(rows, amc_name, run_id=None):
        return 0, 0

    async def _fake_resolve(names, amc_name):
        return {}

    monkeypatch.setattr("dhanradar.tasks.mf._upsert_constituents", _fake_upsert)
    monkeypatch.setattr("dhanradar.tasks.mf._resolve_scheme_isins", _fake_resolve)

    result = await mi._parse_pipeline(file_id)

    assert result == "failed: zero_rows_upserted"
    db_session.expire_all()
    row = await db_session.get(MfManualIngestFile, uuid.UUID(file_id))
    assert row.status == "failed"
    assert row.error == "zero_rows_upserted_scheme_unresolved"


async def test_zero_rows_with_resolvable_names_reports_period_missing(
    db_session, monkeypatch: pytest.MonkeyPatch
):
    """Names resolve but nothing landed → the honest diagnosis is a period
    problem, not 'scheme_unresolved' (real ICICI failure shape, 2026-07-07)."""
    data = _build_series_xlsx("SBI Corporate Bond Fund")
    file_id = await _insert_pending_bytes(
        db_session, filename="SBI Corporate Bond Fund v2.xlsx", data=data
    )

    async def _fake_upsert(rows, amc_name, run_id=None):
        return 0, 0

    async def _fake_resolve(names, amc_name):
        return {"SBI Corporate Bond Fund": "INF200TEST99"}

    monkeypatch.setattr("dhanradar.tasks.mf._upsert_constituents", _fake_upsert)
    monkeypatch.setattr("dhanradar.tasks.mf._resolve_scheme_isins", _fake_resolve)

    result = await mi._parse_pipeline(file_id)

    assert result == "failed: zero_rows_upserted"
    db_session.expire_all()
    row = await db_session.get(MfManualIngestFile, uuid.UUID(file_id))
    assert row.error == "zero_rows_upserted_period_missing"


async def test_no_rows_closed_ended_filename_marks_scheme_not_in_master(
    db_session, monkeypatch: pytest.MonkeyPatch
):
    """A recognizably closed-ended FILE that parses to zero rows (garbled
    legacy layout) is also terminal — detect_amc works off the filename."""
    wb = Workbook()
    ws = wb.active
    ws.append(["HDFC MUTUAL FUND"])  # letterhead only, no header/data rows
    buf = io.BytesIO()
    wb.save(buf)
    file_id = await _insert_pending_bytes(
        db_session, filename="HDFC FMP 1269D March 2023.xlsx", data=buf.getvalue()
    )

    async def _boom(*_a, **_kw):
        raise AssertionError("_upsert_constituents must never run for zero parsed rows")

    monkeypatch.setattr("dhanradar.tasks.mf._upsert_constituents", _boom)

    result = await mi._parse_pipeline(file_id)

    assert result == "unsupported: scheme_not_in_master"
    db_session.expire_all()
    row = await db_session.get(MfManualIngestFile, uuid.UUID(file_id))
    assert row.status == "unsupported"
    assert row.error == "scheme_not_in_master"


# ---------------------------------------------------------------------------
# scheme_master_pdf (2026-07-10) — SBI "Fund Details" PDFs parse instead of
# archiving; all other PDFs keep the contract-§2 archive behavior.
# ---------------------------------------------------------------------------


async def test_fund_details_pdf_routes_to_scheme_master_writer(db_session, monkeypatch):
    from dhanradar.mf import disclosure_parsers as dp
    from dhanradar.models.mf import MfFund

    isin = "INF200KPDF01"
    db_session.add(
        MfFund(
            isin=isin,
            scheme_name="SBI Automotive Opportunities Fund - Direct Plan - Growth",
            amc_name="SBI MUTUAL FUND",
        )
    )
    await db_session.commit()

    flat = (
        "Fund Details For SBI Automotive Opportunities Fund "
        "Fund Name SBI Automotive Opportunities Fund "
        "Riskometer As on Date VERY HIGH "
        "Benchmark(Tier 1) Nifty Auto Tri "
        "Fund Manager From Date Mr. Tanmaya Desai:05-Jun-2024 "
        "Annual Expense (Stated Maximum) Direct Plan - 0.65 "
        "Exit Load ( If Applicable) For exit within 30 days - 1%. After - Nil. "
        "Custodian X"
    )
    # Real pypdf extraction is exercised by unit sniff tests; here the PDF
    # byte-decoding seam is patched so the DISPATCH + WRITER path runs real.
    monkeypatch.setattr(dp, "looks_like_scheme_master_pdf", lambda data: True)
    monkeypatch.setattr(dp, "_flatten_pdf", lambda data: flat)

    async def _fake_resolve(names, amc_name):
        return {"SBI Automotive Opportunities Fund": isin}

    monkeypatch.setattr("dhanradar.tasks.mf._resolve_scheme_isins", _fake_resolve)

    file_id = await _insert_pending_bytes(
        db_session, filename="SBI Automotive Opportunities Fund.pdf", data=b"%PDF-fake"
    )
    result = await mi._parse_pipeline(file_id)

    assert result.startswith("parsed: scheme_master_pdf")
    db_session.expire_all()
    fund = await db_session.get(MfFund, isin)
    assert fund.risk_o_meter == "Very High"
    assert fund.benchmark_index == "Nifty Auto Tri"
    assert float(fund.expense_ratio_pct) == 0.65
    assert float(fund.exit_load_pct) == 1.0
    assert fund.exit_load_days == 30
    row = await db_session.get(MfManualIngestFile, uuid.UUID(file_id))
    assert row.status == "parsed"


async def test_other_pdfs_still_archive(db_session, monkeypatch):
    from dhanradar.mf import disclosure_parsers as dp

    monkeypatch.setattr(dp, "looks_like_scheme_master_pdf", lambda data: False)
    file_id = await _insert_pending_bytes(
        db_session, filename="Complete Factsheet May 2026.pdf", data=b"%PDF-fake2"
    )
    result = await mi._parse_pipeline(file_id)

    assert result == "archived: pdf_saved_for_later"
    db_session.expire_all()
    row = await db_session.get(MfManualIngestFile, uuid.UUID(file_id))
    assert row.status == "archived"


# ---------------------------------------------------------------------------
# fund_performance writer (2026-07-10) — benchmark + riskometer written to
# EVERY plan-variant ISIN of the bare scheme name; real DB fields asserted.
# ---------------------------------------------------------------------------


def _build_fund_performance_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["Crisil Intelligence"])
    ws.append(["Fund Performance"])
    ws.append(["Scheme Name", "Benchmark", "Riskometer Scheme", "NAV Regular"])
    ws.append(["FP Test Large Cap Fund", "Nifty 100 TRI", "Very High", "12.3"])
    ws.append(["FP Unknown Fund", "Nifty 50 TRI", "High", "9.9"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def test_fund_performance_writes_both_fields_to_all_plan_variants(db_session):
    from dhanradar.models.mf import MfFund

    for isin, name in (
        ("INF200KFP001", "FP Test Large Cap Fund - Regular Plan - Growth"),
        ("INF200KFP002", "FP Test Large Cap Fund - Direct Plan - Growth"),
    ):
        db_session.add(MfFund(isin=isin, scheme_name=name, amc_name="TESTAMC"))
    await db_session.commit()

    data = _build_fund_performance_xlsx()
    file_id = await _insert_pending_bytes(
        db_session, filename="Fund-Performance-10-Jul-2026--9999.xlsx", data=data
    )
    result = await mi._parse_pipeline(file_id)

    assert result.startswith("parsed: fund_performance")
    db_session.expire_all()
    for isin in ("INF200KFP001", "INF200KFP002"):
        fund = await db_session.get(MfFund, isin)
        assert fund.risk_o_meter == "Very High", isin
        assert fund.benchmark_index == "Nifty 100 TRI", isin
    row = await db_session.get(MfManualIngestFile, uuid.UUID(file_id))
    assert row.status == "parsed"
    # "FP Unknown Fund" resolved nothing — resolved=1 of 2 is still a parse.


# ---------------------------------------------------------------------------
# factsheet_compilation (2026-07-11) — whole-AMC factsheet PDFs (HDFC/UTI/
# AXIS): manager history only, looped per scheme the splitter found. Real
# resolver (pg_trgm) and real writer/dedup run end to end; only the PDF
# byte-decode seam (parse_factsheet_compilation itself, unit-tested against
# real-file-shaped fixtures in test_disclosure_parsers.py) is patched, same
# convention as the scheme_master_pdf tests above.
# ---------------------------------------------------------------------------


async def test_factsheet_compilation_writes_manager_history_and_dedups_on_reingest(
    db_session, monkeypatch
):
    from datetime import date

    from sqlalchemy import select as sa_select

    from dhanradar.mf import disclosure_parsers as dp
    from dhanradar.models.mf import MfFund, MfFundManagerHistory

    isin = "INF200KCOMP1"
    db_session.add(
        MfFund(isin=isin, scheme_name="HDFC Test Alpha Fund", amc_name="HDFC MUTUAL FUND")
    )
    await db_session.commit()

    records = [
        {
            "scheme_name": "HDFC Test Alpha Fund",
            "manager_pairs": [("Amit Ganatra", date(2026, 2, 1))],
        }
    ]
    monkeypatch.setattr(dp, "looks_like_factsheet_compilation", lambda data: True)
    monkeypatch.setattr(dp, "parse_factsheet_compilation", lambda data, amc: records)

    file_id = await _insert_pending_bytes(
        db_session, filename="HDFC MF Factsheet - May 2026.pdf", data=b"%PDF-fake"
    )
    result = await mi._parse_pipeline(file_id)

    assert result.startswith("parsed: factsheet_compilation")
    db_session.expire_all()
    rows = (
        (
            await db_session.execute(
                sa_select(MfFundManagerHistory).where(MfFundManagerHistory.scheme_uid == isin)
            )
        )
        .scalars()
        .all()
    )
    assert len(rows) == 1
    assert rows[0].manager_name == "Amit Ganatra"
    assert rows[0].start_date == date(2026, 2, 1)

    # Re-ingest the SAME file (reset to pending, as a real re-drop would be
    # re-enqueued after a status reset) — the manager row must NOT duplicate.
    row = await db_session.get(MfManualIngestFile, uuid.UUID(file_id))
    row.status = "pending"
    await db_session.commit()
    result2 = await mi._parse_pipeline(file_id)

    assert result2.startswith("parsed: factsheet_compilation")
    db_session.expire_all()
    rows2 = (
        (
            await db_session.execute(
                sa_select(MfFundManagerHistory).where(MfFundManagerHistory.scheme_uid == isin)
            )
        )
        .scalars()
        .all()
    )
    assert len(rows2) == 1  # still exactly one row — zero duplicates


async def test_factsheet_compilation_unrecognized_layout_marks_unsupported(db_session, monkeypatch):
    from dhanradar.mf import disclosure_parsers as dp

    monkeypatch.setattr(dp, "looks_like_factsheet_compilation", lambda data: True)
    monkeypatch.setattr(dp, "parse_factsheet_compilation", lambda data, amc: [])

    file_id = await _insert_pending_bytes(
        db_session, filename="HDFC MF Factsheet - May 2026.pdf", data=b"%PDF-fake"
    )
    result = await mi._parse_pipeline(file_id)

    assert result == "unsupported: factsheet_compilation_layout_unrecognized"
    db_session.expire_all()
    row = await db_session.get(MfManualIngestFile, uuid.UUID(file_id))
    assert row.status == "unsupported"
