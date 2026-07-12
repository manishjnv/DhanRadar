"""
DhanRadar — pure parsers for NON-portfolio disclosure file classes arriving
through the manual-ingest inbox (2026-07-07, built against the founder's real
first-batch files):

  - AAUM annexure  (SEBI "Annexure I/A1" monthly Average-AUM disclosure —
    per-scheme AAUM ₹Cr; AXIS/EDEL publish .xlsx, SBI/HDFC legacy .xls)
  - Annual riskometer disclosure (ICICI publishes a clean 1-sheet .xlsx:
    scheme · start-FY band · end-FY band · change count)
  - Scheme performance disclosure (two real layouts: ICICI = repeating blocks
    in one sheet, benchmark row follows the "Scheme" row; SBI = one sheet per
    scheme, explicit "SCHEME NAME :" and "Scheme Benchmark: <name>" cells) —
    parsed ONLY for the per-scheme BENCHMARK NAME (the `mf_funds.benchmark_index`
    unblock, plan §11); returns are NOT extracted (we compute our own from NAV).
  - Scheme master details (2026-07-08 — SBI's per-scheme "Fund Details" page
    from their own website, saved with a misleading `.xls` extension: it's
    plain HTML, not a spreadsheet at all. Confirmed identical label/value
    TEMPLATE across every real sample — riskometer band, primary benchmark
    name, fund-manager name(s) + tenure start, and the stated TER. Exit load
    is deliberately NOT extracted — free-form multi-tier text that would need
    lossy guessing to fit a single (pct, days) pair; a wrong compliance-
    adjacent fee figure is worse than none, per §8.4 verbatim-only.)
  - TER (Total Expense Ratio) disclosure (2026-07-09 — SBI/ABSL/Edelweiss/
    HDFC all publish a Regular-Plan-vs-Direct-Plan table, one row per scheme
    per date; header depth/labels vary but every layout shares one invariant:
    a "Regular"-labelled column group and a "Direct"-labelled group, each
    ending in its own "Total TER (%)" column). Only the LATEST date's row per
    scheme is kept — TER changes over time and only the current effective
    rate should ever be written.
  - AMFI consolidated scheme-wise AAUM (2026-07-10 — the quarterly all-AMC
    workbook from the AMFI portal: one row per scheme carrying its AMFI CODE,
    which joins exactly on mf_funds.amfi_code — the one disclosure class that
    needs ZERO name resolution. Multi-AMC by design, so its file class
    ('amfi_aaum') is content-sniffed and bypasses AMC detection.)
  - TATA per-scheme "SCHEME SUMMARY DOCUMENT" (2026-07-11 — TATA's own
    ~70-file per-scheme download, saved with a misleading `.xls` extension:
    it's a real xlsx zip. One sheet, label/value rows, byte-identical layout
    across every scheme type sampled. Unlike every other class here, the
    file GIVES the scheme's ISINs directly — resolved by exact ISIN match,
    never fuzzy name matching. Manager name(s) + tenure, exit load, minimum
    application amount.)

All parsers are PURE (bytes in, tuples out — no DB, no network) so they unit-
test without fixtures beyond in-memory workbooks. Writers live in
tasks/manual_ingest.py. Fail-closed everywhere: an unrecognizable layout
yields an empty result (the caller marks the file 'unsupported'), never a
guess. Values are taken VERBATIM from the file — riskometer bands validated
against the 6 regulatory words, AAUM never derived or imputed (ADR-0035).
"""

from __future__ import annotations

import html
import io
import logging
import re
from collections.abc import Iterator
from datetime import date, datetime
from typing import Any

logger = logging.getLogger(__name__)

# The 6 SEBI riskometer bands — a parsed value must match one VERBATIM
# (case-insensitive) or the row is skipped; we never normalize into a band.
RISKOMETER_BANDS: tuple[str, ...] = (
    "Low",
    "Low to Moderate",
    "Moderate",
    "Moderately High",
    "High",
    "Very High",
)
_BANDS_LOWER = {b.lower(): b for b in RISKOMETER_BANDS}


def classify_file_class(filename: str, data: bytes | None = None) -> str:
    """'aaum' | 'amfi_aaum' | 'riskometer' | 'ter' | 'performance' |
    'scheme_master' | 'portfolio' (default) — pure.

    Keyed off the disclosure-type words AMCs put in their own filenames
    (verified against the founder's real batch: "Average_Assets_Under_
    Management", "Monthly AAUM_May 2026", "monthly-average-asset-under-
    management", "Annual Disclosure of Scheme Riskometer", "Scheme
    Performance Disclosure", "scheme-performance---may-2026").

    `data`, when provided, is CONTENT-sniffed first for `scheme_master` — SBI's
    "Fund Details" page (confirmed 2026-07-08) is saved with a normal
    per-scheme filename indistinguishable from a real portfolio disclosure
    (e.g. "SBI Multicap Fund.xls"), so filename keywords alone can't
    recognize it; content sniffing is the only reliable signal.
    """
    if data is not None and _looks_like_scheme_master_html(data):
        return "scheme_master"
    # TATA's per-scheme "SCHEME SUMMARY DOCUMENT" workbook (2026-07-11): real
    # filenames are plain per-scheme names ("Tata-Large-Cap-Fund.xls")
    # indistinguishable from any other AMC's portfolio disclosure — content
    # is the only reliable signal, same reasoning as scheme_master above.
    if data is not None and looks_like_scheme_summary_xls(data):
        return "scheme_summary_xls"
    # AMFI's consolidated scheme-wise AAUM download has an arbitrary filename
    # ("average-aum2.xlsx" from the portal) that matches NO keyword below —
    # content is the only reliable signal, same reasoning as scheme_master.
    if data is not None and _looks_like_amfi_aaum(data):
        return "amfi_aaum"
    if data is not None and _looks_like_fund_performance(data):
        return "fund_performance"
    low = filename.lower()
    if "aaum" in low or "average asset" in low or "average_asset" in low or "average-asset" in low:
        return "aaum"
    if "riskometer" in low or "risk-o-meter" in low or "risk o meter" in low:
        return "riskometer"
    if (
        "expense ratio" in low
        or "expenseratio" in low
        or "_ter_" in low
        or "_ter." in low
        or re.search(r"\bter\b", low)
    ):
        # `_`/`-` are `\w` for `\b` purposes, so a bare word-boundary regex
        # alone can't catch an underscore-delimited "_ter_"/"_ter." (e.g.
        # "HDFCMF_SCHEMES_TER_02-06-2026_1.xls") — those need the explicit
        # checks above. `\bter\b` then covers every hyphen/space-delimited
        # form (e.g. HSBC's real "...HSBC MF Total Exp Ter Report_
        # 06072026.xlsx", confirmed 2026-07-09) in one pattern instead of
        # enumerating every separator combination — never matches inside a
        # real word like "water"/"after".
        return "ter"
    if "performance" in low:
        return "performance"
    return "portfolio"


def _repair_corrupted_stylesheet_xlsx(data: bytes) -> bytes:
    """Patch a malformed `xl/styles.xml` `rgb="..."` color value in-place and
    return corrected xlsx bytes, unchanged if nothing needed fixing.

    Confirmed 2026-07-09 (Edelweiss's real "TotalExpenseRatio-2026-2027.xlsx"
    and its siblings): openpyxl raises `ValueError: Unable to read workbook:
    could not read stylesheet ... Colors must be aRGB hex values` because one
    `<color rgb="0000000"/>` element has a 7-character hex value — a genuinely
    truncated 8-char aRGB value (should be e.g. "00000000"), not a different
    color format. Zero-padding any `rgb="..."` value that is neither 6 nor 8
    hex characters (the only two valid lengths — plain RGB or aRGB) to 8
    chars recovers the file without altering any VALID color, and this is a
    zip-level, in-memory patch — never touches the on-disk stored original.
    """
    import zipfile

    with zipfile.ZipFile(io.BytesIO(data)) as zin:
        if "xl/styles.xml" not in zin.namelist():
            return data
        styles = zin.read("xl/styles.xml").decode("utf-8")

        def _pad(m: re.Match[str]) -> str:
            hexval = m.group(1)
            return m.group(0) if len(hexval) in (6, 8) else f'rgb="{hexval.zfill(8)}"'

        fixed = re.sub(r'rgb="([0-9A-Fa-f]+)"', _pad, styles)
        if fixed == styles:
            return data

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                content = (
                    fixed.encode("utf-8")
                    if item.filename == "xl/styles.xml"
                    else zin.read(item.filename)
                )
                zout.writestr(item, content)
        return buf.getvalue()


def _iter_sheets(data: bytes, ext: str) -> Iterator[tuple[str, list[list[Any]]]]:
    """Yield (sheet_name, rows-of-raw-values) — openpyxl for .xlsx, xlrd for
    legacy OLE2 .xls (xlrd>=2.0 reads ONLY .xls; already a dependency via the
    CAMS legacy parser). Raw values (not str) — AAUM needs the floats.

    The file EXTENSION is a hint, never trusted blindly both ways (confirmed
    2026-07-09, HDFC's own "...TER_02-06-2026_1.xls" — the opposite mismatch
    of the earlier SBI "Fund Details" HTML-saved-as-.xls case: this one is a
    genuine, valid XLSX zip that someone saved with a legacy ".xls"
    extension). If the extension-suggested reader fails with the SPECIFIC
    "wrong container format" error, retry with the other reader once before
    giving up — never guess data, just don't let a mislabeled extension mask
    a file that's perfectly readable under its real format.
    """
    import xlrd

    if ext == ".xls":
        try:
            book = xlrd.open_workbook(file_contents=data)
        except xlrd.XLRDError as exc:
            if "not supported" not in str(exc):
                raise
            yield from _iter_sheets_xlsx(data)
            return
        for sheet in book.sheets():
            rows = [
                [sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)
            ]
            yield sheet.name, rows
    else:
        import zipfile

        try:
            yield from _iter_sheets_xlsx(data)
        except zipfile.BadZipFile:
            book = xlrd.open_workbook(file_contents=data)
            for sheet in book.sheets():
                rows = [
                    [sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)
                ]
                yield sheet.name, rows


def _iter_sheets_xlsx(data: bytes) -> Iterator[tuple[str, list[list[Any]]]]:
    """The .xlsx branch of `_iter_sheets`, split out so it can be retried with
    repaired bytes on a stylesheet-corruption ValueError without duplicating
    the openpyxl-open-and-iterate logic.

    Root-cause fix (2026-07-11, TATA's real "SCHEME SUMMARY DOCUMENT" per-
    scheme files): a worksheet's own `<dimension ref="A1"/>` can understate
    the real used range (TATA's declares just A1 while data sits in
    A1:C60ish) — openpyxl's READ-ONLY mode trusts that declared dimension
    and `iter_rows()` yields only the single cell it covers. `reset_dimensions()`
    (a real `ReadOnlyWorksheet` method) makes it scan the sheet's actual XML
    instead of trusting the stale bound. Called on EVERY worksheet here (not
    just TATA's) so every xlsx parser sharing this one reader is protected,
    not just the one class that surfaced the bug — verified no behavior
    change on a correctly-dimensioned file (reset just removes an
    optimization hint, it never changes which rows exist in the XML)."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except ValueError as exc:
        if "could not read stylesheet" not in str(exc):
            raise
        repaired = _repair_corrupted_stylesheet_xlsx(data)
        wb = openpyxl.load_workbook(io.BytesIO(repaired), read_only=True, data_only=True)
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            ws.reset_dimensions()
            yield name, [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _s(v: Any) -> str:
    return str(v).strip() if v is not None else ""


_AS_ON_RE = re.compile(r"as on\s+(\d{4})-(\d{2})-(\d{2})", re.IGNORECASE)


def parse_aaum_annexure(data: bytes, ext: str) -> tuple[date | None, list[tuple[str, float]]]:
    """SEBI AAUM annexure → (period, [(scheme_name, aaum_crore)]).

    Anchors on the standardized header pair: a 'Scheme Category/ Scheme Name'
    cell and a 'GRAND TOTAL' cell in the same row (verified identical across
    AXIS .xlsx and SBI .xls). Scheme rows = name in the scheme-name column +
    a positive number in the GRAND TOTAL column; SUB-TOTAL/TOTAL and category
    header rows carry no grand-total number or a 'total' name and are skipped.
    Values are ₹ Cr as published (AXIS LIQUID grand total 59,127.66 verified
    against the fund's real ~₹59k Cr AAUM); a header mentioning 'lakh' scales
    /100 — anything else unrecognized fails closed to an empty result.
    """
    for _name, rows in _iter_sheets(data, ext):
        header_idx = None
        name_ci = None
        total_ci = None
        for idx, row in enumerate(rows[:12]):
            cells = [_s(c) for c in row]
            for ci, cell in enumerate(cells):
                if cell.lower().startswith("scheme category"):
                    name_ci = ci
                if "grand total" in cell.lower():
                    total_ci = ci
            if name_ci is not None and total_ci is not None:
                header_idx = idx
                break
        if header_idx is None or name_ci is None or total_ci is None:
            continue  # not the annexure sheet — try the next one

        # Units: ₹ Cr unless the header block explicitly says lakhs.
        head_text = " ".join(_s(c) for row in rows[: header_idx + 1] for c in row).lower()
        scale = 0.01 if "lakh" in head_text else 1.0

        # Period: the header carries "as on YYYY-MM-DD" (AXIS verified);
        # the caller falls back to the filename when this is absent.
        period: date | None = None
        m = _AS_ON_RE.search(head_text)
        if m:
            try:
                period = date(int(m.group(1)), int(m.group(2)), 1)
            except ValueError:
                period = None

        out: list[tuple[str, float]] = []
        for row in rows[header_idx + 1 :]:
            if name_ci >= len(row) or total_ci >= len(row):
                continue
            name = _s(row[name_ci])
            value = row[total_ci]
            if not name or "total" in name.lower():
                continue
            if not isinstance(value, (int, float)) or value <= 0:
                continue  # category headers / blank rows have no grand total
            out.append((name, round(float(value) * scale, 2)))
        if out:
            return period, out
    return None, []


def _looks_like_fund_performance(data: bytes) -> bool:
    """CONTENT sniff for AMFI's "Fund Performance" export (Crisil-powered,
    one xlsx per SEBI category, ~41 files/quarter): the first rows carry a
    'Fund Performance' title and a header row with 'Scheme Name',
    'Benchmark' and 'Riskometer Scheme'. Filename is arbitrary
    ("Fund-Performance-10-Jul-2026--2221 (3).xlsx"), so only bytes decide.
    Anything unreadable is simply not this class, never an error."""
    if not data.startswith(b"PK\x03\x04"):
        return False
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        title_seen = False
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            joined = " | ".join(cells).lower()
            if "fund performance" in joined:
                title_seen = True
            if title_seen and "scheme name" in joined and "riskometer scheme" in joined:
                return True
            if i > 8:
                break
    except Exception:  # noqa: BLE001
        return False
    return False


def parse_fund_performance(data: bytes, ext: str) -> list[tuple[str, str | None, str | None]]:
    """AMFI "Fund Performance" export → [(scheme_name, benchmark, risk_band)].

    Extracts ONLY the two factual scheme attributes (primary benchmark name +
    scheme riskometer, validated VERBATIM against the 6 regulatory band words);
    every NAV/return column is deliberately never read (house rule: returns
    are computed from NAV in-house, third-party computed returns are never
    ingested — see DATA_SOURCES.md §12 entry). Pure, fail-closed [].
    """
    results: list[tuple[str, str | None, str | None]] = []
    for _sheet, rows in _iter_sheets(data, ext):
        name_ci = bench_ci = risk_ci = None
        header_idx = None
        for idx, row in enumerate(rows[:10]):
            cells = [_s(c) for c in row]
            for ci, cell in enumerate(cells):
                low = cell.lower()
                if low == "scheme name":
                    name_ci = ci
                elif low == "benchmark":
                    bench_ci = ci
                elif low == "riskometer scheme":
                    risk_ci = ci
            if name_ci is not None and bench_ci is not None and risk_ci is not None:
                header_idx = idx
                break
        if header_idx is None or name_ci is None or bench_ci is None or risk_ci is None:
            continue
        for row in rows[header_idx + 1 :]:
            cells = [_s(c) for c in row]
            name = cells[name_ci] if name_ci < len(cells) else ""
            if not name:
                continue
            bench_raw = cells[bench_ci] if bench_ci < len(cells) else ""
            bench = bench_raw if bench_raw.upper() not in ("", "NA", "N/A", "-") else None
            risk_raw = cells[risk_ci] if risk_ci < len(cells) else ""
            band = _BANDS_LOWER.get(risk_raw.strip().lower())
            if bench is None and band is None:
                continue  # nothing factual to write for this row
            results.append((name, bench, band))
    return results


# ---------------------------------------------------------------------------
# TATA per-scheme "SCHEME SUMMARY DOCUMENT" workbook (2026-07-11) — TATA's own
# website serves ~70 real per-scheme files as ".xls" but they are, byte-for-
# byte, valid xlsx zips (PK header) — a genuine extension mislabel. One
# sheet, 3 columns (row idx | field label | value), ~54 label rows verified
# byte-identical across every scheme TYPE sampled (large-cap equity, index,
# liquid/debt, gold ETF, retirement plan) — only the VALUES vary. Unlike
# every other class here, this file GIVES the scheme's ISINs directly (one
# file lists every plan/option variant of ONE scheme) — resolved by exact
# ISIN lookup in the writer, never fuzzy name matching.
#
# Extracted: fund name (informational only — resolution is by ISIN, this is
# never used to match), fund manager name(s) + tenure start date(s), exit
# load (handed VERBATIM, after trim, to the existing `_parse_exit_load_text`
# — never a second exit-load reader), minimum application amount, and the
# ISIN list.
#
# Deliberately NOT extracted: "Annual Expense (Stated maximum)" is a STATED
# ceiling, not the actual TER the AMFI consolidated TER class already
# covers — writing it would pollute expense_ratio_pct from a less-accurate
# source; SIP/STP/SWP mechanics and custodian/auditor/registrar are
# administrative, not analytics-relevant. Benchmark (Tier 1) / Riskometer
# ARE present in the file but skipped this pass — the only existing writer
# for those fields (fund_performance, PR #544) resolves by fuzzy NAME
# matching, not ISIN, so reusing it here would mean a second write pattern,
# not a reuse; noted as the extension point (a follow-up could add two more
# `UPDATE ... WHERE isin = :i` statements identical in shape to the
# exit-load/min-lumpsum ones below), not built in this pass.
# ---------------------------------------------------------------------------

_TATA_ISIN_RE = re.compile(r"[A-Z]{2}[A-Z0-9]{9}\d")
# Two DISTINCT marker shapes, deliberately not one permissive "leading digits"
# regex: a bare date's day-of-month ('19 Jan 2024', '27-Mar-19') also starts
# with digits, so an unqualified `(\d+)` would misread the day as an index
# (caught by the positional-pairing test — '19 Jan 2024' was silently read as
# index 19). Real manager-index markers are unambiguous: either the literal
# 'FM' prefix, or a bare digit followed IMMEDIATELY by '.' (never '-' —
# dates use '-' right after the day digit, e.g. '27-Mar-19'; no real manager
# index in any sampled file uses a bare digit + dash).
_TATA_FM_MARKER_RE = re.compile(r"^FM\s*-?\s*(\d+)\s*[-.:]*\s*(.*)$", re.IGNORECASE)
_TATA_BARE_MARKER_RE = re.compile(r"^(\d+)\.\s*(.*)$")
_TATA_DATE_NUM_RE = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{2,4})")
_TATA_DATE_MON_RE = re.compile(r"(\d{1,2})[\s\-]+([A-Za-z]{3,9})[.,\s\-]*(\d{2,4})")


def looks_like_scheme_summary_xls(data: bytes) -> bool:
    """CONTENT sniff — never filename (TATA's per-scheme filenames are
    indistinguishable from any other AMC's by pattern alone). The first
    sheet's first row carries 'Fields' + 'SCHEME SUMMARY DOCUMENT' as a
    banner pair (verified identical across ~70 real files). Goes through
    `_iter_sheets_xlsx` (not a second openpyxl open) so it inherits the
    read-only broken-`<dimension>` fix above for free — without it, the
    banner row itself reads back truncated to one cell and this sniff would
    never fire."""
    if not data.startswith(b"PK\x03\x04"):
        return False
    try:
        _name, rows = next(iter(_iter_sheets_xlsx(data)))
    except Exception:  # noqa: BLE001 — a sniff must never break classification
        return False
    if not rows:
        return False
    head = [_s(c).lower() for c in rows[0]]
    return "fields" in head and any("scheme summary document" in c for c in head)


def _tata_isins(field: str) -> list[str]:
    """ISIN list, verbatim-validated. Real files carry 3 distinct data-entry
    errors: a missing leading letter, two ISINs concatenated with no
    separator, and a bare '0' placeholder — a strict 12-char ISIN shape
    check silently drops all three rather than guessing where to split or
    what letter is missing. Separator varies (comma, comma+space, or a bare
    newline — all three seen in real files); order-preserving de-dup."""
    seen: list[str] = []
    for token in re.split(r"[,\n;]", field):
        token = token.strip()
        if token and _TATA_ISIN_RE.fullmatch(token) and token not in seen:
            seen.append(token)
    return seen


def _tata_split_indexed(field: str) -> list[tuple[int | None, str]]:
    """Split a comma/semicolon list into (index, rest) pairs. `index` is the
    manager's own 'FM-N'/'N.' ordinal when present (real separator styles
    seen: 'FM-1', 'FM 1 -', '1.'), else None — the caller correlates the
    name field against the date field by this index."""
    out: list[tuple[int | None, str]] = []
    for item in re.split(r"[,;]", field):
        item = item.strip()
        if not item:
            continue
        m = _TATA_FM_MARKER_RE.match(item) or _TATA_BARE_MARKER_RE.match(item)
        if m:
            out.append((int(m.group(1)), m.group(2).strip(" -")))
        else:
            out.append((None, item))
    return out


def _tata_extract_date(text: str) -> date | None:
    """First recognizable date anywhere in `text` — real files mix DD/MM/YYYY
    ('01/07/2025') with DD-Month-YYYY in several spacing variants ('01-July-
    2025', '20-Dec-24', '16-March- 2026'). Fail-closed None on free text
    ('From the date of allotment' — a real value, never guessed into a
    date)."""
    m = _TATA_DATE_NUM_RE.search(text)
    if m:
        day, month, year = m.groups()
        y = int(year) + 2000 if len(year) == 2 else int(year)
        try:
            return date(y, int(month), int(day))
        except ValueError:
            return None
    m = _TATA_DATE_MON_RE.search(text)
    if m:
        day, mon, year = m.groups()
        y = int(year) + 2000 if len(year) == 2 else int(year)
        try:
            return datetime.strptime(f"{int(day):02d}-{mon[:3]}-{y}", "%d-%b-%Y").date()
        except ValueError:
            return None
    return None


def _tata_manager_pairs(name_field: str, date_field: str) -> list[tuple[str, date | None]]:
    """Correlate the 'Fund Manager Name' and 'Fund Manager From Date' fields.

    Real files use THREE different pairing shapes (verified across ~70
    files): (1) both fields carry a matching 'FM-N'/'N.' index per manager —
    paired by index, robust even when the date field ALSO repeats the name
    ('2. Rakesh Prajapati - 20-Dec-2024'); (2) neither field has an index and
    the split counts match — paired POSITIONALLY; (3) counts genuinely
    mismatch (e.g. 2 managers, 1 date with no index to attribute it to
    either) — every manager in that file gets `None`, never a guessed
    pairing. `None` dates are returned here, not dropped — the caller
    decides whether to write a dateless manager (`fund_manager_history.
    start_date` is NOT NULL, so in practice they're skipped at write time,
    never fabricated — §8.4).
    """
    names = _tata_split_indexed(name_field)
    dates = _tata_split_indexed(date_field)

    date_by_idx: dict[int, date] = {}
    fallback_dates: list[date] = []
    for idx, raw in dates:
        parsed = _tata_extract_date(raw)
        if parsed is None:
            continue
        if idx is not None:
            date_by_idx.setdefault(idx, parsed)
        else:
            fallback_dates.append(parsed)

    use_positional = not date_by_idx and len(names) == len(fallback_dates)

    pairs: list[tuple[str, date | None]] = []
    for i, (idx, raw_name) in enumerate(names):
        clean_name = " ".join(raw_name.split()).strip(" ,-")
        if not clean_name:
            continue
        start = date_by_idx.get(idx) if idx is not None else None
        if start is None and use_positional:
            start = fallback_dates[i]
        pairs.append((clean_name, start))
    return pairs


def parse_scheme_summary_xls(data: bytes) -> dict[str, Any]:
    """TATA per-scheme "SCHEME SUMMARY DOCUMENT" workbook → one dict (ONE
    file = ONE scheme, same output shape as `parse_factsheet_pdf`). Fail-
    closed {} when the file carries no 'Fund Name' label at all (the sniff
    already gates the common case; this is the belt-and-braces check).

    Keys: scheme_name (informational only — resolution is by ISIN, never by
    this name), manager_pairs (list[(name, date|None)] — see
    `_tata_manager_pairs`), exit_load_pct/exit_load_days (verbatim-trimmed
    text handed to the existing `_parse_exit_load_text`, same parser
    `factsheet_pdf`/`scheme_master` already use — never a second exit-load
    reader), min_lumpsum_amount, isins (every plan/option ISIN the file
    lists for this one scheme).
    """
    try:
        _name, rows = next(iter(_iter_sheets_xlsx(data)))
    except Exception:  # noqa: BLE001 — unreadable workbook, never a guess
        return {}
    labels: dict[str, str] = {}
    for row in rows:
        if not row or len(row) < 2 or row[1] is None:
            continue
        label = _s(row[1])
        if label:
            labels.setdefault(label, _s(row[2]) if len(row) > 2 else "")

    scheme_name = labels.get("Fund Name", "")
    if not scheme_name:
        return {}

    manager_pairs = _tata_manager_pairs(
        labels.get("Fund Manager Name", ""), labels.get("Fund Manager From Date", "")
    )

    exit_load_pct, exit_load_days = _parse_exit_load_text(
        labels.get("Exit Load (if applicable)", "").strip()
    )

    min_lumpsum: float | None = None
    min_raw = labels.get("Minimum Application Amount", "")
    try:
        val = float(min_raw)
        min_lumpsum = val if val > 0 else None
    except ValueError:
        min_lumpsum = None

    isins = _tata_isins(labels.get("ISINs", ""))

    return {
        "scheme_name": scheme_name,
        "manager_pairs": manager_pairs,
        "exit_load_pct": exit_load_pct,
        "exit_load_days": exit_load_days,
        "min_lumpsum_amount": min_lumpsum,
        "isins": isins,
    }


# ---------------------------------------------------------------------------
# scheme_summary_pdf (2026-07-12) — the SAME SEBI "SCHEME SUMMARY DOCUMENT"
# per-scheme template as scheme_summary_xls above, published as a PDF by
# HDFC / ICICI Prudential / HSBC / Franklin Templeton / Mirae (top-5 AMC
# ingestion wave), either from the AMC's own site or the AMFI portal's mirror
# (`portal.amfiindia.com/spages/SSD_<id>.pdf`). Real samples verified
# 2026-07-12 (docs/Sample/amc-data/{HDFC,ICICI,HSBC,Franklin-Templeton,
# MIRAE}/scheme-summary/): the numbered 1-53 field template is IDENTICAL in
# content to the xlsx variant, but pypdf's PLAIN `extract_text()` scrambles
# reading order on several real files (ICICI's AMFI-portal mirror: every
# field's LABEL prints first in one block, then every VALUE in a second
# block — label/value pairing by line order is impossible there).
# `extraction_mode="layout"` (pypdf's x/y-position-aware mode) restores the
# correct row order on EVERY sample tried — a label and its own value sit on
# the same line again — so this class reads PDFs through THAT mode
# exclusively, never plain extract_text().
#
# Same output dict shape as `parse_scheme_summary_xls` (scheme_name/
# manager_pairs/exit_load_pct/exit_load_days/min_lumpsum_amount/isins) PLUS
# two fields the xlsx parser deliberately skipped: benchmark_tier1 and
# risk_band — the writer fills mf_funds.benchmark_index/risk_o_meter ONLY
# where currently NULL (never overwrites — the same B104 COALESCE discipline
# the nightly NAV upsert now enforces; AMFI's Fund-Performance names already
# cover ~8.2k schemes and must not be clobbered by a lower-confidence SSD
# read).
#
# Field NUMBERS are not a safe anchor across AMCs (Mirae renumbers every
# field past 20 because each of its up-to-4 fund managers gets 3 OWN
# numbered rows instead of one shared comma-list row) — every anchor below
# is a LABEL text match, sequential/bounded (search only after the previous
# anchor), never a number.
# ---------------------------------------------------------------------------

_SSD_BANNER_RE = re.compile(r"SCHEME SUMMARY DOCUMENT", re.IGNORECASE)
_SSD_FUND_NAME_RE = re.compile(r"\bFund Name\b")
# HDFC's real files drop the "ti" ligature glyph ENTIRELY in several labels
# (a font/encoding defect confirmed 2026-07-12 against both real HDFC
# samples — `_flatten_pdf_layout` already normalizes the leaked NUL byte to
# a space, so `\s*` alone absorbs the gap here): "Option" -> "Op on",
# "Listing" -> "Lis ng", "Application" -> "Applica on" — 2 letters gone, not
# a misrender. `(?:ti)?` makes the pair optional everywhere it's
# structurally expected; the same `\s*` also absorbs the SEPARATE plain
# stray-space artifact the same files carry elsewhere (e.g. "R egular"),
# which loses no letters.
_SSD_OPTION_NAMES_RE = re.compile(r"Op(?:ti)?\s*on\s*Names?\s*\(\s*R\s*egular", re.IGNORECASE)
_SSD_RISK_RE = re.compile(r"Riskometer\s*\(as on[^)]*\)", re.IGNORECASE)
_SSD_CATEGORY_RE = re.compile(r"Category\s*as\s*Per\s*SEBI", re.IGNORECASE)
# Mirae's real files misspell this label "Benchmarch" on EVERY sample seen —
# the `k|ch` alternation matches both the correct and the misspelled form
# with one pattern (verified 2026-07-12 against all 5 target AMCs).
_SSD_BENCH1_RE = re.compile(r"Benchmar(?:k|ch)\s*\(Tier\s*1\)", re.IGNORECASE)
_SSD_BENCH2_RE = re.compile(r"Benchmar(?:k|ch)\s*\(Tier\s*2\)", re.IGNORECASE)
_SSD_EXPENSE_RE = re.compile(r"(?:Annual|Actual)\s+[Ee]xpenses?", re.IGNORECASE)
# Franklin's real file typos the closing bracket: "(if applicable}".
_SSD_EXIT_LOAD_RE = re.compile(r"Exit Load\s*\(if applicable[)\}]", re.IGNORECASE)
_SSD_CUSTODIAN_RE = re.compile(r"\bCustodian\b")
# ISINs' own label is used as a BOUNDARY (stop for Listing Details, in
# `_SSD_LISTING_DETAILS_RE` below) but never as the ISIN block's own START —
# see that constant's docstring for why. Both this and the min-application
# labels carry the same "ti"-drop defect `_SSD_OPTION_NAMES_RE` above
# documents ("Listing" -> "Lis ng", "Application" -> "Applica on").
_SSD_LISTING_DETAILS_RE = re.compile(r"Lis(?:ti)?\s*ng\s*De\s*tails", re.IGNORECASE)
_SSD_AMFI_CODE_RE = re.compile(r"AMFI Codes?\s*\(To be phased out\)", re.IGNORECASE)
_SSD_MIN_APP_RE = re.compile(r"Minimum\s*Applica(?:ti)?\s*on\s*Amoun\s*t(?!\s*in\b)", re.IGNORECASE)
_SSD_MIN_APP_MULT_RE = re.compile(
    r"Minimum\s*Applica(?:ti)?\s*on\s*Amoun\s*t\s+in\b", re.IGNORECASE
)
# The manager block: BOTH real row shapes seen — one shared comma-list row
# ("Fund Manager Name" / "... Type" / "... From Date", HDFC/HSBC/Franklin/
# DSP-family) and one row PER manager ("Fund Manager 1 - Name" / "Fund
# Manager 2 - Name" ..., ICICI/Mirae) — the optional `\d*` matches both.
# "Type" is tracked ONLY as a boundary (its own value is never needed) —
# without it, a Name row's value runs all the way to the NEXT Name/Date row,
# swallowing the whole Type row's text (and, for a comma-list Type value,
# splitting it into bogus extra "managers" — caught 2026-07-12 against every
# real sample).
_SSD_MGR_NAME_RE = re.compile(r"Fund Manager\s*\d*\s*[-–]?\s*Name\b", re.IGNORECASE)
_SSD_MGR_TYPE_RE = re.compile(r"Fund Manager\s*\d*\s*[-–]?\s*Type\b", re.IGNORECASE)
_SSD_MGR_DATE_RE = re.compile(r"Fund Manager\s*\d*\s*[-–]?\s*(?:From\s*)?Date\b", re.IGNORECASE)
_SSD_MONTH_FIRST_DATE_RE = re.compile(r"\b([A-Za-z]{3,9})\s+(\d{1,2}),?\s*(\d{4})\b")
_SSD_ISIN_RE = re.compile(r"\b[A-Z]{2}[A-Z0-9]{9}\d\b")
_SSD_FM_MARKER_RE = re.compile(r"\bFM\s*-?\s*(\d+)\s*[-.:]*\s*", re.IGNORECASE)
_SSD_NUM_RE = re.compile(r"(\d[\d,]*(?:\.\d+)?)")
# A field NUMBER that lands mid-row (real Mirae/ICICI artifact: the number
# column is vertically centered against a multi-manager block, so it prints
# next to a DIFFERENT physical line than the field it belongs to — e.g.
# ICICI's "Fund Manager 2 - Name" line carries manager 1's stray "18") gets
# swept into whichever value slice sits next to it. Stripped as a trailing
# short bare-digit token — never inside a real name/date, always a leaked
# row number.
_SSD_TRAILING_FIELDNUM_RE = re.compile(r"\s+\d{1,3}\s*$")


def looks_like_ssd_pdf(data: bytes) -> bool:
    """CONTENT sniff for the PDF "SCHEME SUMMARY DOCUMENT" template — the
    literal banner phrase every real sample (HDFC/ICICI/HSBC/Franklin/Mirae,
    both the AMC's own site and the AMFI-portal mirror) prints on its first
    page, next to the "Fields" column header. A cheap plain-text check is
    enough for a sniff — only the full parse needs layout mode."""
    if not data.startswith(b"%PDF"):
        return False
    try:
        import io as _io

        from pypdf import PdfReader

        first = PdfReader(_io.BytesIO(data)).pages[0].extract_text() or ""
    except Exception:  # noqa: BLE001 — unreadable PDF is simply not this class
        return False
    return bool(_SSD_BANNER_RE.search(first)) and "fields" in first.lower()


def _flatten_pdf_layout(data: bytes) -> str:
    """All pages' LAYOUT-mode text, joined and whitespace-collapsed to one
    string — see the module note above for why layout mode (not
    `_flatten_pdf`'s plain mode) is required for this class.

    Real HDFC files drop the "ti" ligature glyph entirely wherever it
    appears (a font/encoding defect confirmed 2026-07-12 against both real
    HDFC samples), and pypdf renders the gap as a literal NUL byte, not a
    space ("Op\\x00on" for "Option", byte-level). NUL bytes are stripped to
    spaces here — once, at the source — rather than special-cased in every
    downstream label regex, and this doubles as a safety net: a NUL byte in
    a manager name would make Postgres reject the whole write
    ("A string literal cannot contain NUL (0x00) characters")."""
    import io as _io

    from pypdf import PdfReader

    reader = PdfReader(_io.BytesIO(data))
    text = " ".join(page.extract_text(extraction_mode="layout") or "" for page in reader.pages)
    return re.sub(r"[\s\x00]+", " ", text).strip()


def _ssd_value_between(
    text: str,
    start_re: re.Pattern[str],
    stop_re: re.Pattern[str] | None,
    search_from: int = 0,
) -> str | None:
    """Text between the END of `start_re`'s first match (at/after
    `search_from`) and the START of `stop_re`'s next match — `None` if
    `start_re` never matches. Mirrors `_extract_all_labels`'s "next known
    label ends this one's value" contract, generalized to regex anchors
    since the SSD's own field NUMBERS are not a safe anchor (see module
    note)."""
    m = start_re.search(text, search_from)
    if not m:
        return None
    end = len(text)
    if stop_re is not None:
        stop = stop_re.search(text, m.end())
        if stop:
            end = stop.start()
    value = text[m.end() : end].strip(" :-")
    # The NEXT field's own row NUMBER routinely leaks onto the tail of THIS
    # field's captured value (real artifact: the row-number column is
    # vertically centered against a multi-line value cell, so it prints
    # beside whichever physical line ends up in the middle — often the
    # line right before the next field's own label; e.g. a Riskometer value
    # of "Very High" followed a line later by "6 Category as Per..." reads
    # back as "Very High 6"). Stripped as a trailing short bare-digit token
    # — never part of a real value in any field this parser reads.
    return _SSD_TRAILING_FIELDNUM_RE.sub("", value).strip()


def _ssd_extract_date(text: str) -> date | None:
    """First recognizable date in `text`, trying every real order seen across
    the 5 target AMCs' SSD PDFs: numeric DD/MM/YYYY and DD-Month-YYYY
    (reuses `_tata_extract_date`, the same generic parser TATA's xlsx class
    uses) PLUS "Month DD, YYYY" (HDFC "July 29, 2022", Franklin "May 02,
    2016") — a US-style order the xlsx samples never used, so it is added
    here rather than loosening the shared TATA regex. Fail-closed None."""
    found = _tata_extract_date(text)
    if found is not None:
        return found
    m = _SSD_MONTH_FIRST_DATE_RE.search(text)
    if m:
        mon, day, year = m.groups()
        try:
            return datetime.strptime(f"{int(day):02d}-{mon[:3]}-{year}", "%d-%b-%Y").date()
        except ValueError:
            return None
    return None


def _ssd_clean_manager_name(raw: str) -> str:
    """Whitespace-normalize a captured manager-name value, drop a trailing
    role/portfolio-segment qualifier in parens (real Mirae multi-asset
    scheme: "Mr. Harshad Borawake (Equity Portion)" -> "Mr. Harshad
    Borawake" — every other manager-name source on this platform stores the
    bare name), and strip a leaked field-number token (see
    `_SSD_TRAILING_FIELDNUM_RE`)."""
    name = " ".join(raw.split())
    name = re.split(r"\s*\(", name)[0]
    name = _SSD_TRAILING_FIELDNUM_RE.sub("", name)
    return name.strip(" ,-")


def _ssd_manager_rows(text: str, start: int, end: int) -> tuple[list[str], list[str]]:
    """Every 'Name'-labeled and 'From Date'/'Date'-labeled row's raw value
    text within [start, end), in document order. A row's own value can
    itself be a comma list (the shared-row shape) or a single manager (the
    per-manager-row shape) — `_ssd_manager_pairs` below tells them apart by
    how many rows came back. 'Type' rows are located too but ONLY used as a
    boundary — a Name row otherwise has no way to know where its own value
    ends and the following Type row's begins."""
    anchors: list[tuple[int, int, str]] = (
        [(m.start(), m.end(), "name") for m in _SSD_MGR_NAME_RE.finditer(text, start, end)]
        + [(m.start(), m.end(), "type") for m in _SSD_MGR_TYPE_RE.finditer(text, start, end)]
        + [(m.start(), m.end(), "date") for m in _SSD_MGR_DATE_RE.finditer(text, start, end)]
    )
    anchors.sort(key=lambda a: a[0])
    names: list[str] = []
    dates: list[str] = []
    for i, (_a_start, a_end, kind) in enumerate(anchors):
        if kind == "type":
            continue
        value_end = anchors[i + 1][0] if i + 1 < len(anchors) else end
        value = _SSD_TRAILING_FIELDNUM_RE.sub("", text[a_end:value_end].strip(" :-")).strip()
        # A "-" placeholder row (real Mirae shape: unused manager 2/3/4
        # slots) strips down to nothing — UNLESS a leaked field number (see
        # `_SSD_TRAILING_FIELDNUM_RE`'s own note) is the only thing left
        # after the dash itself is gone, leaving a bare digit string that
        # would otherwise be mistaken for a real name/date. Neither a real
        # manager name nor a real tenure date is ever pure digits.
        if not value or value.isdigit():
            continue
        (names if kind == "name" else dates).append(value)
    return names, dates


def _ssd_manager_pairs(names_raw: list[str], dates_raw: list[str]) -> list[tuple[str, date | None]]:
    """Correlate manager names to tenure-start dates across BOTH real row
    shapes (see `_ssd_manager_rows`).

    Every Name row is first expanded through `_tata_split_indexed` (the SAME
    comma/FM-marker splitter TATA's xlsx parser uses) — a row can itself
    carry a comma list (the shared-row shape) or already be exactly one
    manager (the per-manager-row shape, OR a lone manager on a single
    indexed row — Franklin's real "Fund Manager 1- Name" single-manager
    file uses this shape too). Concatenating every row's split in document
    order keeps the overall manager order intact either way.

    Per-manager DATE rows (ICICI/Mirae, or a lone Franklin-style row):
    len(dates_raw) already equals the manager count — paired POSITIONALLY,
    only when the counts agree (never a guessed pairing on a mismatch).

    One shared comma-list DATE row (HDFC/HSBC/Franklin-ELSS/DSP-family):
    two correlation strategies, tried in order, per manager:
      1. an 'FM-N' marker's POSITION in the date text (HDFC/HSBC-Liquid:
         "FM 1 - Managing Since July 29, 2022 and FM 2 - ..." — split by
         marker position, never by comma, because the date itself contains
         a comma ("July 29, 2022") a naive comma-split would misread as a
         field separator).
      2. the manager's own name repeated verbatim inside the date text
         (DSP/Franklin-ELSS/HSBC-SmallCap: "Mr. Anil Ghelani - 26/12/2023,
         ...") — find that substring, read the first date found right
         after it.
    Neither strategy finding a date is a legitimate, fail-closed outcome —
    `None`, never fabricated (§8.4); the writer already skips dateless
    pairs.
    """
    split_names: list[tuple[int | None, str]] = []
    for row in names_raw:
        split_names.extend(_tata_split_indexed(row))

    if dates_raw and len(dates_raw) == len(split_names):
        pairs: list[tuple[str, date | None]] = []
        for (_idx, raw_name), date_text in zip(split_names, dates_raw):
            clean = _ssd_clean_manager_name(raw_name)
            if clean:
                pairs.append((clean, _ssd_extract_date(date_text)))
        return pairs

    date_field = dates_raw[0] if dates_raw else ""
    fm_segments: dict[int, str] = {}
    hits = list(_SSD_FM_MARKER_RE.finditer(date_field))
    for i, hit in enumerate(hits):
        seg_end = hits[i + 1].start() if i + 1 < len(hits) else len(date_field)
        fm_segments[int(hit.group(1))] = date_field[hit.end() : seg_end]

    def _find_pos(name: str) -> int:
        pos = date_field.find(name)
        if pos < 0:
            bare = re.sub(r"^(?:Mr|Ms|Mrs)\.?\s+", "", name)
            pos = date_field.find(bare) if bare != name else -1
        return pos

    # Pre-locate every manager's own name-substring position in date_field,
    # in TEXT order — needed to bound each manager's own search window to
    # the NEXT located manager's position, never a fixed length. A fixed
    # window can swallow a NEIGHBOR's date when that neighbor's date uses a
    # "Month DD, YYYY" order the day-first regex doesn't recognize as a
    # match starting there but DOES accidentally match further along in the
    # window (caught 2026-07-12 against HSBC Small Cap's real file:
    # "Venugopal Manghat - Dec 17, 2019, Sonal Gupta - 05-July -21" —
    # "Dec 17, 2019" needs the month-first fallback, but a wide-enough
    # window lets the day-first regex match "05-July-21" first instead).
    located = sorted(
        (pos, i)
        for i, (_idx, raw_name) in enumerate(split_names)
        if (pos := _find_pos(_ssd_clean_manager_name(raw_name))) >= 0
    )
    window_end_by_index: dict[int, int] = {
        i: (located[j + 1][0] if j + 1 < len(located) else len(date_field))
        for j, (_pos, i) in enumerate(located)
    }

    pairs = []
    for i, (idx, raw_name) in enumerate(split_names):
        clean_name = _ssd_clean_manager_name(raw_name)
        if not clean_name:
            continue
        start_date: date | None = None
        if idx is not None and idx in fm_segments:
            start_date = _ssd_extract_date(fm_segments[idx])
        if start_date is None and i in window_end_by_index:
            pos = _find_pos(clean_name)
            start_date = _ssd_extract_date(date_field[pos : window_end_by_index[i]])
        pairs.append((clean_name, start_date))
    return pairs


def _ssd_match_risk_band(value: str) -> str | None:
    """Riskometer value -> one of the 6 regulatory bands, matched as a
    PREFIX rather than requiring an exact match (every other risk-band
    reader on this platform uses an exact `_BANDS_LOWER` lookup, but real
    HDFC files glue the SEBI disclaimer sentence directly onto the band word
    with no separator at all: "Very High RiskInvestors understand that
    their principal will be at Very High Risk" — an exact match would miss
    it entirely). Bands are checked longest-first so "Low to Moderate" is
    never mistaken for a "Low" prefix match. Fail-closed None on anything
    that isn't one of the 6 words verbatim at the start."""
    low = value.strip().lower()
    for band in sorted(RISKOMETER_BANDS, key=len, reverse=True):
        if low.startswith(band.lower()):
            return band
    return None


def _ssd_isins(text: str) -> list[str]:
    """ISIN list, order-preserving de-dup, strict 12-char shape (same
    verbatim-only rule TATA's `_tata_isins` enforces — a malformed token,
    e.g. a real DSP file dropping the leading letter off one ISIN, is simply
    never matched by the fixed-length pattern rather than guessed/repaired).
    `findall` over the raw blob (not a comma/newline pre-split) because real
    files glue the ISIN directly onto the preceding plan description with no
    separator at all (HDFC: "...OPTION- INF179K01VM3 HDFC FLEXI...")."""
    seen: list[str] = []
    for tok in _SSD_ISIN_RE.findall(text):
        if tok not in seen:
            seen.append(tok)
    return seen


def parse_ssd_pdf(data: bytes) -> dict[str, Any]:
    """PDF "SCHEME SUMMARY DOCUMENT" -> the SAME dict shape
    `parse_scheme_summary_xls` returns (scheme_name/manager_pairs/
    exit_load_pct/exit_load_days/min_lumpsum_amount/isins), plus
    benchmark_tier1/risk_band (see module note). Fail-closed {} when no
    'Fund Name' is found at all.
    """
    try:
        text = _flatten_pdf_layout(data)
    except Exception:  # noqa: BLE001 — unreadable PDF, never a guess
        return {}

    scheme_name = (_ssd_value_between(text, _SSD_FUND_NAME_RE, _SSD_OPTION_NAMES_RE) or "").strip()
    if not scheme_name:
        return {}

    risk_raw = _ssd_value_between(text, _SSD_RISK_RE, _SSD_CATEGORY_RE) or ""
    risk_band = _ssd_match_risk_band(risk_raw)

    bench_raw = (_ssd_value_between(text, _SSD_BENCH1_RE, _SSD_BENCH2_RE) or "").strip()
    benchmark_tier1 = (
        bench_raw if bench_raw.upper() not in ("", "NA", "N/A", "-", "NOT APPLICABLE") else None
    )

    bench2_match = _SSD_BENCH2_RE.search(text)
    mgr_start = bench2_match.end() if bench2_match else 0
    expense_match = _SSD_EXPENSE_RE.search(text, mgr_start)
    mgr_end = expense_match.start() if expense_match else len(text)
    names_raw, dates_raw = _ssd_manager_rows(text, mgr_start, mgr_end)
    manager_pairs = _ssd_manager_pairs(names_raw, dates_raw)

    exit_raw = _ssd_value_between(text, _SSD_EXIT_LOAD_RE, _SSD_CUSTODIAN_RE) or ""
    exit_load_pct, exit_load_days = _parse_exit_load_text(exit_raw)

    # Widened to start from "Listing Details" (the field immediately BEFORE
    # ISINs), not the "ISINs" label's own end: on every AMC with 5+ ISINs the
    # block wraps across enough lines that the "ISINs" label itself centers
    # PAST the block's first 1-2 wrapped lines — those lines still print
    # (right after the always-short, single-line Listing Details value) but
    # sit textually BEFORE the "ISINs" label match, so anchoring on that
    # label's own end silently drops the leading ISIN(s) (caught 2026-07-12:
    # HDFC Flexi Cap Fund's real file lost its first of 6). ISIN pattern-
    # matching is content-strict (`_ssd_isins`), so including Listing
    # Details' own short "Not Applicable"/"NA" value in the window is safe —
    # it never contains a 12-char ISIN-shaped token.
    isins_raw = _ssd_value_between(text, _SSD_LISTING_DETAILS_RE, _SSD_AMFI_CODE_RE) or ""
    isins = _ssd_isins(isins_raw)

    min_raw = _ssd_value_between(text, _SSD_MIN_APP_RE, _SSD_MIN_APP_MULT_RE) or ""
    min_lumpsum: float | None = None
    num_m = _SSD_NUM_RE.search(min_raw)
    if num_m:
        try:
            val = float(num_m.group(1).replace(",", ""))
            min_lumpsum = val if val > 0 else None
        except ValueError:
            min_lumpsum = None

    return {
        "scheme_name": scheme_name,
        "manager_pairs": manager_pairs,
        "exit_load_pct": exit_load_pct,
        "exit_load_days": exit_load_days,
        "min_lumpsum_amount": min_lumpsum,
        "isins": isins,
        "benchmark_tier1": benchmark_tier1,
        "risk_band": risk_band,
    }


_MANAGER_SINCE_RE = re.compile(
    r"([A-Z][A-Za-z.'\u2019 ]{2,50}?)\s*\n?\(Managing this fund since\s+([A-Za-z]+),?\s*(\d{4})",
)
_CLOSING_AUM_RE = re.compile(
    r"Closing AUM as on\s+(\d{1,2}-[A-Za-z]{3}-\d{2,4})\s*:\s*Rs\.?\s*([\d,]+(?:\.\d+)?)\s*crores",
    re.IGNORECASE,
)
_FRESH_SUBSCRIPTION_RE = re.compile(
    r"Application Amount for fresh Subscription\s*:?\s*\n?\s*Rs\.?\s*([\d,]+)",
    re.IGNORECASE,
)
_HEREIN_ARE_OF_RE = re.compile(
    r"performance details provided herein are of\s+([^.\n]+?)\s*(?:\.|\n)", re.IGNORECASE
)
_EXIT_LOAD_SECTION_RE = re.compile(
    r"Exit load for Redemption\s*/\s*Switch\s*out\s*:?-?(.{0,400})", re.IGNORECASE | re.DOTALL
)


def looks_like_factsheet_pdf(data: bytes) -> bool:
    """CONTENT sniff for an AMC's PER-SCHEME factsheet PDF (ICICI layout,
    2026-07-11: 2 pages, "Fund Managers** :" + "(Managing this fund since
    <Month>, <Year>)" + "Closing AUM as on <date>"). First page only — the
    whole-AMC "Complete Factsheet" compilations have a cover page and stay
    archived (per-scheme files are the reliable unit)."""
    if not data.startswith(b"%PDF"):
        return False
    try:
        import io as _io

        from pypdf import PdfReader

        first = PdfReader(_io.BytesIO(data)).pages[0].extract_text() or ""
    except Exception:  # noqa: BLE001 — unreadable PDF is simply not this class
        return False
    return "managing this fund since" in first.lower() and "aum as on" in first.lower()


_COMPILATION_SNIFF_PAGES = 20


def looks_like_factsheet_compilation(data: bytes) -> bool:
    """CONTENT sniff for a whole-AMC factsheet COMPILATION PDF (one file,
    every scheme in the AMC) — verified 2026-07-11 against the real HDFC/
    UTI/AXIS/KOTAK/EDELWEISS/ABSL May-June 2026 files, each with a distinct
    but stable per-scheme manager anchor: HDFC's "Name Since Total Exp"
    table header, UTI's "FUND MANAGER SUMMARY" annexure, AXIS's "...is
    managing the scheme since...schemes of Axis Mutual Fund" prose, Kotak's
    per-scheme "Fund Manager*:" label, Edelweiss's "Fund Managers Experience
    Managing Since" table header, ABSL's "Managing the Fund Since:" label.
    Extended 2026-07-12 for the AMCs' separate PASSIVE/INDEX-fund factsheet
    siblings (same AMC, a different sub-brochure — verified against the real
    HDFC/AXIS/MIRAE passive files): AXIS's passive layout carries no "...is
    managing the scheme since..." prose at all, but every scheme page has its
    own "Work experience:" label right after the manager's name; MIRAE (a
    brand-new AMC for this class) prints "Fund Managers :" followed by a
    "(since <Month> <Day>, <Year>)" parenthetical per manager — checked together
    (both substrings on the same page) since "Fund Managers :" alone is too
    generic a label to anchor on by itself. HDFC's own passive/index
    factsheet reuses the EXACT SAME "Name Since Total Exp" table anchor as
    its regular factsheet (real per-scheme layout, zero new anchor needed).
    Scans only the first _COMPILATION_SNIFF_PAGES pages — these run 100+
    pages, and scanning the whole document on every intake would be
    wasteful and, on the biggest files, memory-risky on the 640 MB worker —
    and returns on the first anchor found. NOTE: Kotak's own manager-tenure
    annexure only starts around page 146 (far past this window) but its
    per-scheme "Fund Manager*:" label already appears on every scheme's
    cover page from page 6 onward, so the early-page sniff still holds."""
    if not data.startswith(b"%PDF"):
        return False
    try:
        import io as _io

        from pypdf import PdfReader

        reader = PdfReader(_io.BytesIO(data))
        n = min(_COMPILATION_SNIFF_PAGES, len(reader.pages))
        for i in range(n):
            text = reader.pages[i].extract_text() or ""
            if (
                "Name Since Total Exp" in text
                or "FUND MANAGER SUMMARY" in text
                or "Managing the scheme since" in text
                or "schemes of Axis Mutual Fund" in text
                or "Fund Manager*:" in text
                or "Fund Managers Experience Managing Since" in text
                or "Managing the Fund Since:" in text
                or "Work experience:" in text
                or ("Fund Managers :" in text and "(since " in text)
            ):
                return True
    except Exception:  # noqa: BLE001 — unreadable PDF is simply not this class
        return False
    return False


def parse_factsheet_pdf(data: bytes) -> dict[str, Any]:
    """Per-scheme factsheet PDF → the scheme_master-style dict (ONE file =
    ONE scheme): scheme_name, manager_pairs, aum_crore/aum_as_of (CLOSING
    AUM — the stated net-assets figure, same semantics as the portfolio
    grand-total row), exit_load_pct/days, min_lumpsum_amount. Facts only;
    every NAV/return figure is deliberately never read. Fail-closed {}.
    """
    try:
        flat_lines = _flatten_pdf_lines(data)
    except Exception:  # noqa: BLE001
        return {}
    text = "\n".join(flat_lines)

    # Scheme name — two strategies (pypdf's page-text ordering varies between
    # files of the SAME layout, confirmed 2026-07-11):
    #   1. The standard performance disclaimer names the fund exactly:
    #      "...performance details provided herein are of <Fund Name>."
    #   2. Fallback: the first fund-title-looking line at the top of page 1
    #      (directly under "Portfolio as on <date>").
    scheme_name = ""
    m = _HEREIN_ARE_OF_RE.search(text)
    if m:
        scheme_name = " ".join(m.group(1).split()).strip(" .")
    if not scheme_name:
        for line in flat_lines[:8]:
            clean = line.strip()
            if (
                len(clean) > 8
                and any(kw in clean.lower() for kw in ("fund", "etf", "plan"))
                and "portfolio as on" not in clean.lower()
                and not clean[0].isdigit()
            ):
                scheme_name = clean
                break
    if not scheme_name:
        return {}

    manager_pairs: list[tuple[str, date]] = []
    for name, month, year in _MANAGER_SINCE_RE.findall(text):
        try:
            start = datetime.strptime(f"01-{month[:3]}-{year}", "%d-%b-%Y").date()
        except ValueError:
            continue
        clean_name = " ".join(name.split()).strip(" ,")
        if clean_name and (clean_name, start) not in manager_pairs:
            manager_pairs.append((clean_name, start))

    aum_crore: float | None = None
    aum_as_of: date | None = None
    m = _CLOSING_AUM_RE.search(text)
    if m:
        try:
            aum_crore = float(m.group(2).replace(",", ""))
            raw_date = m.group(1)
            fmt = "%d-%b-%y" if len(raw_date.split("-")[-1]) == 2 else "%d-%b-%Y"
            aum_as_of = datetime.strptime(raw_date, fmt).date()
        except ValueError:
            aum_crore, aum_as_of = None, None

    min_lumpsum: float | None = None
    m = _FRESH_SUBSCRIPTION_RE.search(text)
    if m:
        try:
            val = float(m.group(1).replace(",", ""))
            min_lumpsum = val if val > 0 else None
        except ValueError:
            min_lumpsum = None

    exit_load_pct: float | None = None
    exit_load_days: int | None = None
    m = _EXIT_LOAD_SECTION_RE.search(text)
    if m:
        exit_load_pct, exit_load_days = _parse_exit_load_text(m.group(1))

    return {
        "scheme_name": scheme_name,
        "manager_pairs": manager_pairs,
        "aum_crore": aum_crore,
        "aum_as_of": aum_as_of,
        "min_lumpsum_amount": min_lumpsum,
        "exit_load_pct": exit_load_pct,
        "exit_load_days": exit_load_days,
    }


def _flatten_pdf_lines(data: bytes) -> list[str]:
    """All pages' text as lines (whitespace-normalized per line, empties
    dropped) — the factsheet parser needs LINE structure (manager names sit
    on their own line above the tenure parenthetical), unlike the fully
    flattened stream `_flatten_pdf` produces for the label-pair pages."""
    import io as _io

    from pypdf import PdfReader

    reader = PdfReader(_io.BytesIO(data))
    lines: list[str] = []
    for page in reader.pages:
        for raw in (page.extract_text() or "").splitlines():
            cleaned = " ".join(raw.split())
            if cleaned:
                lines.append(cleaned)
    return lines


# ---------------------------------------------------------------------------
# Factsheet COMPILATION splitter (2026-07-11) — whole-AMC PDF, every scheme
# in one file. Reuses the SAME per-scheme output shape parse_factsheet_pdf
# returns (scheme_name + manager_pairs only — AUM/exit-load/min-lumpsum are
# deliberately NOT extracted here: those need their own per-layout verified
# anchors and manager is the metric this class targets, PR body has the
# per-file yield). Each AMC's layout is genuinely different (real files,
# verified 2026-07-11) — no shared regex, one small per-AMC extractor each,
# dispatched by the AMC the caller already detected from the filename.
# Every extractor walks reader.pages ONE PAGE AT A TIME (never accumulates
# every page's text before returning) — these run 100+ pages and the box has
# a 640 MB worker budget (see the memory-guard trap in the runbook).
# ---------------------------------------------------------------------------


def parse_factsheet_compilation(data: bytes, amc_name: str | None) -> list[dict[str, Any]]:
    """Whole-AMC factsheet compilation PDF -> [{"scheme_name", "manager_pairs"}],
    one dict per scheme found. `amc_name` is the AMC already detected from the
    filename by the caller (manual_ingest.py); an unrecognized/undetected AMC
    fails closed to [] rather than guessing a layout. A scheme's own detail
    page can legitimately be reprinted earlier in the document as a
    "featured fund" highlight (confirmed real, AXIS) — de-duplicated by
    scheme_name, first occurrence wins.
    """
    try:
        import io as _io

        from pypdf import PdfReader

        reader = PdfReader(_io.BytesIO(data))
    except Exception:  # noqa: BLE001 — unreadable PDF, never a guess
        return []

    if amc_name == "HDFC":
        return _parse_hdfc_compilation(reader)
    if amc_name == "UTI":
        return _parse_uti_compilation(reader)
    if amc_name == "AXIS":
        return _parse_axis_compilation(reader)
    if amc_name == "KOTAK":
        return _parse_kotak_compilation(reader)
    if amc_name == "EDELWEISS":
        return _parse_edelweiss_compilation(reader)
    if amc_name == "ABSL":
        return _parse_absl_compilation(reader)
    if amc_name == "MIRAE":
        return _parse_mirae_compilation(reader)
    return []  # unrecognized AMC for this class — fail closed, never guess


# Manager-name trailing-credential strip (2026-07-12), shared post-processing
# for EVERY factsheet_compilation extractor below. Root cause (live prod wart
# "Anurag Mittal Bcom", UTI Money Market Fund): the real line reads "Mr.
# Anurag Mittal Bcom, MSc, CA\nManaging the scheme since Dec 2021" — there is
# NO separator between the name and the FIRST credential token, so the
# manager-name regex's stopping literal (a comma) is the one AFTER "Bcom",
# not before it, and "Bcom" is swallowed into the captured name. Every OTHER
# credential in the same line ("MSc", "CA") already sits after ITS OWN comma,
# which the regex correctly stops at — they were never polluted.
# Verified by running every existing extractor against all 6 real baseline
# files (HDFC/UTI/AXIS/KOTAK/EDELWEISS/ABSL) and checking every captured name
# for a trailing token matching a candidate credential list (Bcom/B.Com/BCom/
# MSc/M.Sc/CA/CFA/CAIA/MBA/PGDM/FRM/ACA/B.E/BE/B.Tech) — "Bcom" on this ONE
# UTI line is the ONLY hit anywhere in the corpus. The allowlist below is
# intentionally this narrow — NOT a regex loosening (the name-capturing
# regexes are untouched); this is a fixed post-hoc token-membership check
# applied to a name AFTER the existing separator-bounded regex already
# captured it, so it can only ever REMOVE a known-bad trailing token, never
# widen what any regex matches.
_CREDENTIAL_SUFFIX_TOKENS = {"BCOM"}  # normalized: dots removed, upper-cased


def _strip_credential_suffix(name: str) -> str:
    """Drop a trailing academic-credential token the manager-name regex
    swallowed because the real file has no separator before it (see
    `_CREDENTIAL_SUFFIX_TOKENS`) — a bare string op on an already
    comma/whitespace-cleaned name, single-token names are left untouched
    (never strips a genuine one-word name)."""
    tokens = name.split()
    if len(tokens) > 1 and tokens[-1].upper().replace(".", "") in _CREDENTIAL_SUFFIX_TOKENS:
        return " ".join(tokens[:-1])
    return name


_HDFC_BANNER_RE = re.compile(r"^\d{1,3}\s*\|\s*[A-Za-z]+\s+\d{4}\s*$")
# HDFC's per-scheme "FUND MANAGER" table: "Name Since Total Exp" header, then
# repeating "<Name>[ (<Portfolio qualifier>)] <Month> <DD>, <YYYY> Over <N>
# years" blocks (verified 2026-07-11, incl. multi-manager schemes like HDFC
# Balanced Advantage Fund — 5 managers on one page). The name char class
# excludes ',' and '-' so it naturally stops at the credential/qualifier
# separator instead of over-consuming into the next word (no explicit month
# list needed — datetime.strptime is the real validator, same idiom as
# _MANAGER_SINCE_RE above).
_HDFC_MANAGER_BLOCK_RE = re.compile(
    r"([A-Z][A-Za-z.’' ]{1,40}?)\s*(?:\([^)]*\))?\s*([A-Za-z]+)\s+(\d{1,2}),\s*(\d{4})"
    r"\s*Over\s+\d+\s*years?",
)


def _hdfc_scheme_name_from_lines(lines: list[str]) -> str:
    """The scheme name sits on the line immediately after the running
    "<page> | <Month> <Year>" banner (verified: every scheme page repeats
    this banner + name at its own top)."""
    for i, line in enumerate(lines[:-1]):
        if _HDFC_BANNER_RE.match(line):
            cand = lines[i + 1]
            if len(cand) > 5 and not cand.lower().startswith("for product label"):
                return cand
    return ""


def _parse_hdfc_compilation(reader: Any) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for page in reader.pages:
        text = page.extract_text() or ""
        # Uppercase "FUND MANAGER" is the section header (never the lowercase
        # "...Dedicated Fund Manager for Overseas Investments" boilerplate
        # that also appears on continuation pages) — a real anchor, verified
        # against every one of the ~50 scheme pages sampled. Cheap raw-text
        # filter first (this literal never wraps across a line break).
        if "FUND MANAGER" not in text:
            continue
        lines = [" ".join(ln.split()) for ln in text.splitlines() if ln.strip()]
        flat = " ".join(lines)
        # "Name Since Total Exp" checked on the LINE-NORMALIZED flat text, not
        # raw text — HDFC Gold ETF's real page (Index Solutions passive
        # factsheet, verified 2026-07-12) wraps the header itself across a
        # column break ("Name Since Total \nExp"), so a raw-text substring
        # check silently skipped the whole scheme even though the manager
        # table is genuinely present. `flat` only ever ADDS matches versus
        # `text` (whitespace-only normalization), so this can never regress
        # a page that already matched on raw text.
        if "Name Since Total Exp" not in flat:
            continue
        scheme_name = _hdfc_scheme_name_from_lines(lines)
        if not scheme_name or scheme_name in seen:
            continue
        # Slice strictly AFTER "Name Since Total Exp" (never from "FUND
        # MANAGER" itself) — some real pages omit the "¥" footnote marker
        # that would otherwise act as an accidental separator, and without
        # this the permissive name char class (letters + spaces, needed for
        # multi-word names) swallows "FUND MANAGER" itself into the first
        # captured name (caught 2026-07-11 against the real file: "FUND
        # MANAGER Rakesh Sethia" instead of "Rakesh Sethia").
        header_pos = flat.find("Name Since Total Exp")
        end = flat.find("DATE OF ALLOTMENT", header_pos)
        if header_pos < 0 or end < 0:
            continue
        block = flat[header_pos + len("Name Since Total Exp") : end]
        manager_pairs: list[tuple[str, date]] = []
        for name, month, _day, year in _HDFC_MANAGER_BLOCK_RE.findall(block):
            clean_name = _strip_credential_suffix(" ".join(name.split()).strip(" ,"))
            if not clean_name:
                continue
            try:
                start_date = datetime.strptime(f"01-{month[:3]}-{year}", "%d-%b-%Y").date()
            except ValueError:
                continue
            if (clean_name, start_date) not in manager_pairs:
                manager_pairs.append((clean_name, start_date))
        if not manager_pairs:
            continue
        out.append({"scheme_name": scheme_name, "manager_pairs": manager_pairs})
        seen.add(scheme_name)
    return out


# UTI: one scheme per page. Scheme name comes from the ALL-CAPS banner line
# ("UTI LARGE CAP FUND"), but pg_trgm similarity is case-sensitive and the
# master's scheme_name is Title Case ("UTI Large Cap Fund") — an ALL-CAPS
# query would share almost no trigrams with the true row and silently fail
# to resolve. The "FUND MANAGER SUMMARY" annexure (verified 2026-07-11,
# pages ~79-80 printed) lists every scheme in the AMC's own Title Case, so
# it is parsed FIRST into a lookup and used to recover the correct casing;
# `.title()` is only a fallback for a scheme the summary doesn't cover.
# Char class also allows the SAME footnote markers (@^*) the summary annexure
# uses ("UTI BANKING & PSU FUND@", "UTI UNIT LINKED INSURANCE PLAN*") and the
# curly apostrophe U+2019 ("UTI CHILDREN’S EQUITY FUND") — both verified
# 2026-07-12 against the real file: without them the banner never matched at
# all, silently dropping the whole scheme (9 of 45 real sections missed).
_UTI_BANNER_LINE_RE = re.compile(r"^UTI\b[A-Z0-9 &\-',.@^*’]*$")
_UTI_SUMMARY_TRAILER_RE = re.compile(r"[@^*\s]*(?:-|\d{1,3})\s*$")
_UTI_PAREN_RE = re.compile(r"\s*\([^)]*\)")
# "Mr./Ms./Mrs. <Name>[,-(]<credentials...>Managing [this/the scheme ]since
# <Month> <Year>" — repeated per co-manager (verified 2026-07-11, incl.
# 3-manager schemes like UTI Flexi Cap Fund). The name char class excludes
# ',' '-' '(' so it stops at the credential separator instead of
# over-consuming; '(' is a valid separator too (verified 2026-07-12: some
# real pages put the "(Equity Portion)"/"(Debt Portion)" qualifier directly
# after the name with no comma/hyphen at all, e.g. "Mr. Sachin Trivedi
# (Equity Portion) B.Com..." on UTI Balanced Advantage Fund — silently
# dropped the whole scheme under the old comma/hyphen-only separator). The
# "since" literal is scoped case-insensitive (real pages mix "Managing the
# scheme since Dec 2021" and "Managing the scheme Since May 2025" for the
# exact same phrase) and "this scheme"/"the scheme" is optional (real pages
# also just say "Managing since Nov 2025" with neither word, e.g. UTI Unit
# Linked Insurance Plan's co-managers) — all 3 wording variants verified
# against the real file, never a guess.
_UTI_MANAGER_RE = re.compile(
    r"Mr\.?\s+([A-Z][A-Za-z.’' ]{1,40}?)\s*[,\-(]\s*.{0,120}?"
    r"Managing\s+(?:(?:this|the)\s+scheme\s+)?(?i:since)\s+([A-Za-z]+)[\s-]+(\d{4})",
)
_UTI_DISCLAIMER_START = "Past performance may or may not be sustained"

# Rare UTI shape (verified 2026-07-12, real file
# uti_fund_watch_active_june_2026_rv1.pdf, pages 28/29 — UTI Quant Fund + UTI
# Income Plus Arbitrage Active Fund Of Fund, the last 2 of 45 real sections
# _UTI_MANAGER_RE still missed): the manager block reads "Mr. <Name>[,-]
# <credentials>\nManaging the scheme Since Inception\nTotal Exp: <N> Yrs" —
# there is NO month/year at all, so _UTI_MANAGER_RE's `([A-Za-z]+)[\s-]+
# (\d{4})` capture never matches (nothing to capture: "Inception" is not
# followed by a 4-digit year). Name-only anchor, stops at the SAME first
# comma/hyphen separator _UTI_MANAGER_RE uses.
_UTI_MANAGER_NAME_ONLY_RE = re.compile(r"^Mr\.?\s+([A-Z][A-Za-z.’' ]{1,40}?)\s*[,\-]")
_UTI_SINCE_INCEPTION_RE = re.compile(
    r"Managing\s+(?:(?:this|the)\s+scheme\s+)?since\s+inception", re.IGNORECASE
)
# The bare "<Day><ordinal> <Month>, <Year>" value line (the page's own "Date
# of inception/allotment" field) — verified: on both real pages this is the
# ONLY line matching this shape start-to-end; every OTHER date on the page
# ("NAV per unit as on 29th May, 2026", etc.) is embedded mid-sentence in a
# longer label, never a bare line, so the `^...$` full-line anchor can't
# false-match those.
_UTI_BARE_DATE_RE = re.compile(r"^(\d{1,2})(?:st|nd|rd|th)\s+([A-Za-z]+),?\s*(\d{4})$")


def _uti_inception_fallback(lines: list[str]) -> list[tuple[str, date]]:
    """Recover the manager+date pair for the "Managing the scheme Since
    Inception" shape `_UTI_MANAGER_RE` can't match (see
    `_UTI_MANAGER_NAME_ONLY_RE` above). The scheme's actual inception date
    sits a few lines ABOVE the manager name (verified 2026-07-12: same
    "Date of inception/allotment" value the page prints near its top,
    directly before the Benchmark Index / Fund Manager values) — searched
    backward in a tight 6-line window so it can only ever find that one
    value, never a random later-page date."""
    pairs: list[tuple[str, date]] = []
    for i, line in enumerate(lines):
        m = _UTI_MANAGER_NAME_ONLY_RE.match(line)
        if not m:
            continue
        if not _UTI_SINCE_INCEPTION_RE.search(" ".join(lines[i : i + 3])):
            continue
        clean = _strip_credential_suffix(m.group(1).strip(" ,-"))
        if not clean:
            continue
        start: date | None = None
        for back in lines[max(0, i - 6) : i]:
            dm = _UTI_BARE_DATE_RE.match(back)
            if dm:
                day, month, year = dm.groups()
                try:
                    start = datetime.strptime(f"{day}-{month[:3]}-{year}", "%d-%b-%Y").date()
                except ValueError:
                    start = None
        if start and (clean, start) not in pairs:
            pairs.append((clean, start))
    return pairs


def _uti_manager_summary_lookup(text: str) -> dict[str, str]:
    """Parse the "FUND MANAGER SUMMARY" page(s) into {UPPER(clean name):
    clean Title-Case name} — the AMC's own scheme-name spelling, ground
    truth for the whole document. A summary line is "<Sr No> Mr. <Manager
    Name> <Scheme Name> <page ref>" for the FIRST scheme under a manager and
    just "<Scheme Name> <page ref>" for every scheme after — `line.find`
    (not `startswith`) handles both shapes (caught 2026-07-11: `startswith`
    silently dropped every manager's FIRST scheme, e.g. UTI Flexi Cap Fund)."""
    lookup: dict[str, str] = {}
    idx = text.find("FUND MANAGER SUMMARY")
    if idx < 0:
        return lookup
    for raw_line in text[idx:].splitlines():
        line = raw_line.strip()
        uti_pos = line.find("UTI")
        if uti_pos < 0:
            continue
        line = line[uti_pos:]
        line = _UTI_SUMMARY_TRAILER_RE.sub("", line)
        name = _UTI_PAREN_RE.sub("", line).strip(" @^*")
        if name:
            # Straight vs curly apostrophe (U+2019) key normalization: the
            # real annexure prints the SAME scheme with different apostrophe
            # glyphs across entries ("Children's" here, "Children’s" a few
            # lines later — verified 2026-07-12, a PDF-export inconsistency,
            # not a code bug) — normalize the lookup KEY only, so whichever
            # glyph the per-scheme banner later uses still resolves to this
            # annexure's own (correctly cased) spelling instead of falling
            # through to the `.title()` fallback, which mangles apostrophes
            # ("Children's" -> "Children'S").
            lookup[name.upper().replace("’", "'")] = name
    return lookup


def _uti_banner_line(lines: list[str]) -> str:
    """The ALL-CAPS scheme banner, followed within a few lines by the
    SEBI-mandated "An open ended..." description (verified on every real
    scheme page; front-matter/media pages that also mention "Fund Manager"
    never carry this pair). Many banners carry a mixed-case "(Erstwhile <old
    name>)" annotation — sometimes inline on the same line ("UTI LARGE CAP
    FUND (Erstwhile UTI Mastershare Unit Scheme)"), sometimes its OWN line
    right after the banner ("UTI OVERNIGHT FUND" / "(Maturity of 1 day)" /
    "An open ended..." — 3 separate lines, caught 2026-07-11) — the ALL-CAPS
    check runs only on the part before any inline annotation, and the
    description search allows a short gap for a possible annotation line."""
    for i, line in enumerate(lines[:-1]):
        head = line.split("(")[0].strip()
        if not _UTI_BANNER_LINE_RE.match(head):
            continue
        # "open-ended" (hyphen) and "open ended" (space) both appear across
        # real scheme pages (caught 2026-07-11: hyphenated form on UTI
        # Focused Fund was silently missed) — normalize before checking.
        if any("open ended" in nxt.lower().replace("-", " ") for nxt in lines[i + 1 : i + 4]):
            return line
    return ""


def _uti_manager_pairs(lines: list[str]) -> list[tuple[str, date]]:
    flat = " ".join(lines)
    # Bound the search to BEFORE the trailing disclaimer paragraph — it
    # restates "Mr. <Name> since <date>, Mr. <Name2> ... Managing the scheme
    # since <date2>" in prose closely enough to false-match the same regex
    # (caught 2026-07-11: produced a bogus 4th "manager" named "Ajay Tyagi
    # since Jan" on UTI Flexi Cap Fund). The real structured manager block
    # always precedes this paragraph.
    boundary = flat.find(_UTI_DISCLAIMER_START)
    search_text = flat[:boundary] if boundary > 0 else flat
    pairs: list[tuple[str, date]] = []
    for name, month, year in _UTI_MANAGER_RE.findall(search_text):
        clean = _strip_credential_suffix(" ".join(name.split()).strip(" ,-"))
        if not clean:
            continue
        try:
            start = datetime.strptime(f"01-{month[:3]}-{year}", "%d-%b-%Y").date()
        except ValueError:
            continue
        if (clean, start) not in pairs:
            pairs.append((clean, start))
    if pairs:
        return pairs
    # No dated block on this page at all — try the "Since Inception" shape
    # (see `_uti_inception_fallback`) rather than reporting the scheme as
    # having no manager.
    return _uti_inception_fallback(lines)


def _parse_uti_compilation(reader: Any) -> list[dict[str, Any]]:
    name_lookup: dict[str, str] = {}
    pages_text: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        pages_text.append(text)
        if "FUND MANAGER SUMMARY" in text:
            name_lookup.update(_uti_manager_summary_lookup(text))

    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for text in pages_text:
        if "Fund Manager" not in text:
            continue
        lines = [" ".join(ln.split()) for ln in text.splitlines() if ln.strip()]
        banner = _uti_banner_line(lines)
        if not banner:
            continue
        # Strip the SAME footnote markers (@^*) the annexure lookup keys
        # already strip (2026-07-12: banner char class now allows them
        # through so the ALL-CAPS match succeeds, e.g. "UTI BANKING & PSU
        # FUND@" / "UTI UNIT LINKED INSURANCE PLAN*" — but the trailing
        # marker itself is never part of the real scheme name).
        clean_banner = _UTI_PAREN_RE.sub("", banner).strip(" @^*")
        # Fallback only (the summary lookup misses a genuine scheme-name
        # spelling difference between the annexure and the banner, e.g.
        # "and" vs "&"): plain `.title()` mangles "UTI" -> "Uti" since it
        # treats the whole ALL-CAPS run as one word (caught 2026-07-11), and
        # mangles a mid-name apostrophe too ("Children's" -> "Children'S",
        # caught 2026-07-12). The banner is guaranteed (by
        # _UTI_BANNER_LINE_RE) to start with the literal "UTI" token, so keep
        # it and title-case only the rest. The lookup key is apostrophe-
        # normalized (see `_uti_manager_summary_lookup`) so the same
        # normalization is applied here before querying.
        scheme_name = name_lookup.get(clean_banner.upper().replace("’", "'")) or (
            "UTI " + clean_banner[3:].strip().title()
        )
        if scheme_name in seen:
            continue
        manager_pairs = _uti_manager_pairs(lines)
        if not manager_pairs:
            continue
        out.append({"scheme_name": scheme_name, "manager_pairs": manager_pairs})
        seen.add(scheme_name)
    return out


# AXIS: one scheme per page. Preferred scheme-name source is the performance
# table's own Title-Case row ("Axis Large Cap Fund - Regular Plan - Growth
# Option" — this file is the "Regular Fund Factsheet", so every scheme's
# Regular-Plan row is present); the ALL-CAPS banner + .title() is only a
# fallback for a layout the row pattern doesn't cover (e.g. a single-NAV ETF
# with no plan structure) — same case-sensitivity reasoning as UTI above.
_AXIS_BANNER_LINE_RE = re.compile(r"^AXIS\b[A-Z0-9 &\-',.]*$")
_AXIS_SCHEME_NAME_RE = re.compile(
    r"(Axis\s[A-Za-z0-9&.,’' \-]*?)\s*-\s*Regular Plan\s*-\s*Growth Option"
)
# "<Name> is managing the scheme since <Day><st/nd/rd/th> <Month> <Year> and
# he/she manages <N> schemes of Axis Mutual Fund" — repeated per co-manager,
# joined by '&' (verified 2026-07-11, incl. 3-manager schemes like Axis
# Large Cap Fund). The digit in "<N> schemes" between entries prevents the
# name group from ever bridging into the next manager's name.
_AXIS_MANAGER_RE = re.compile(
    r"([A-Z][A-Za-z.’' ]{1,40}?)\s+is managing the scheme since\s+"
    r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)\s+(\d{4})",
)


def _axis_scheme_name(lines: list[str], flat: str) -> str:
    m = _AXIS_SCHEME_NAME_RE.search(flat)
    if m:
        return " ".join(m.group(1).split()).strip(" -")
    for line in lines:
        if _AXIS_BANNER_LINE_RE.match(line):
            return line.title()
    return ""


def _axis_manager_pairs(flat: str) -> list[tuple[str, date]]:
    pairs: list[tuple[str, date]] = []
    for name, _day, month, year in _AXIS_MANAGER_RE.findall(flat):
        clean = _strip_credential_suffix(" ".join(name.split()).strip(" ,"))
        if not clean:
            continue
        try:
            start = datetime.strptime(f"01-{month[:3]}-{year}", "%d-%b-%Y").date()
        except ValueError:
            continue
        if (clean, start) not in pairs:
            pairs.append((clean, start))
    return pairs


# AXIS PASSIVE (2026-07-12): a separate real sibling file ("Axis Passive
# Factsheet"), one scheme per page, genuinely different layout from the
# regular factsheet above — no "...is managing the scheme since..." prose and
# no "- Regular Plan - Growth Option" performance row anywhere on a scheme's
# own page (verified: those phrases only occur in this file's Scheme-Returns
# annexure, pages 44-57, never on a per-scheme page). Scheme name is the
# banner line immediately before the "MONTHLY FACTSHEET" tagline (sometimes
# 1-2 lines earlier if the name wraps, or if a parenthetical NSE/BSE-code or
# "(Formerly known as ...)" annotation line sits between them — same
# skip-parenthetical idiom as Edelweiss/Kotak above). The tagline itself
# sometimes prints as "MONTHL Y FACTSHEET" (a stray kerning space INSIDE
# "MONTHLY", verified 2026-07-12 on 3 real pages: NIFTY IT/Healthcare/India
# Consumption ETF — a raw "MONTHLY FACTSHEET" substring check silently
# skipped the whole scheme) — matched via a tolerant regex instead of a
# literal. Manager block is "Mr./Ms./Mrs. <Name>Work experience: <N>
# years.<He/She> <has been|is> managing <this|the> <fund|scheme> since
# <Day><ordinal> <Month> <Year>" — pypdf's extraction glues several of these
# word boundaries with NO space at all (verified: "MalikWork experience:",
# "15 years.He", "hasbeen managing", "6thMarch 2026" all appear with zero
# spaces on real pages, while other real pages spell the SAME phrase with
# normal spaces) — every joint the codebase has seen either glued or spaced
# uses `\s*` (zero-or-more), never `\s+`. Name length is bounded ({1,40}?)
# the same as every other AMC's manager regex — "Work experience:" is a
# real, specific literal that never appears in ordinary prose, so a
# mid-sentence false start still can't reach it within the bound.
_AXIS_PASSIVE_FACTSHEET_TAG_RE = re.compile(r"MONTHL\s*Y\s*FACTSHEET")
_AXIS_PASSIVE_MANAGER_RE = re.compile(
    r"(?:Mr\.|Ms\.|Mrs\.)\s*([A-Z][A-Za-z.’' ]{1,40}?)Work experience:\s*\d+\s*years\.\s*"
    r"(?:He|She)\s*(?:has\s*been|is)\s*managing\s*(?:this|the)\s*(?:fund|scheme)\s*since\s*"
    r"(\d{1,2})(?:st|nd|rd|th)?\s*([A-Za-z]+)\s+(\d{4})"
)


def _axis_passive_scheme_name(lines: list[str]) -> str:
    parts: list[str] = []
    for line in lines:
        m = _AXIS_PASSIVE_FACTSHEET_TAG_RE.search(line)
        if m:
            head = line[: m.start()].strip()
            if head:
                parts.append(head)
            break
        if line.startswith("("):
            continue  # parenthetical annotation, e.g. "(NSE Symbol: AXISNIFTY)"
        parts.append(line)
    name = " ".join(parts).strip()
    return name if name.startswith("Axis") else ""


def _axis_passive_manager_pairs(flat: str) -> list[tuple[str, date]]:
    pairs: list[tuple[str, date]] = []
    for name, _day, month, year in _AXIS_PASSIVE_MANAGER_RE.findall(flat):
        clean = _strip_credential_suffix(" ".join(name.split()).strip(" ,"))
        if not clean:
            continue
        try:
            start = datetime.strptime(f"01-{month[:3]}-{year}", "%d-%b-%Y").date()
        except ValueError:
            continue
        if (clean, start) not in pairs:
            pairs.append((clean, start))
    return pairs


def _parse_axis_compilation(reader: Any) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for page in reader.pages:
        text = page.extract_text() or ""
        lines = [" ".join(ln.split()) for ln in text.splitlines() if ln.strip()]
        flat = " ".join(lines)
        if "is managing the scheme since" in text:
            scheme_name = _axis_scheme_name(lines, flat)
            if not scheme_name or scheme_name in seen:
                continue
            manager_pairs = _axis_manager_pairs(flat)
        elif "Work experience:" in text and _AXIS_PASSIVE_FACTSHEET_TAG_RE.search(text):
            # Regular-layout pages never reach here (they match the branch
            # above first — every regular scheme page carries "is managing
            # the scheme since" on the SAME page as its own "Work experience:"
            # annexure block, verified 2026-07-12), so this branch only ever
            # fires on the genuinely different passive-file layout.
            scheme_name = _axis_passive_scheme_name(lines)
            if not scheme_name or scheme_name in seen:
                continue
            manager_pairs = _axis_passive_manager_pairs(flat)
        else:
            continue
        if not manager_pairs:
            continue
        out.append({"scheme_name": scheme_name, "manager_pairs": manager_pairs})
        seen.add(scheme_name)
    return out


# KOTAK: the manager-tenure annexure ("ABOUT OUR FUND MANAGERS - REGULAR/
# DIRECT PLAN", verified 2026-07-11 pages ~146-190 of the real 190-page
# June-2026 file) is organised BY MANAGER, not by scheme, and repeats the
# SAME schemes across a Regular-Plan pass and a Direct-Plan pass — dedup by
# scheme_name (first occurrence wins) handles both passes for free. Every
# scheme's own cover page names the manager but explicitly defers the
# tenure date to this annexure ("For Fund Manager experience, please refer
# page 162-166 & 181-185") — so the annexure, not the cover page, is the
# only real tenure source. The one reliable per-scheme anchor inside it is a
# compound sentence that appears once per scheme wherever its performance
# table sits: "Scheme Inception date is <DD/MM/YYYY>[, <optional clause>].
# Mr./Ms./Mrs. <Name>[, & Mr. <Name2>...] has/have been managing the fund
# since <DD/MM/YYYY>" (repeated per co-manager clause via `finditer` — real
# schemes have up to 4 managers, each clause its own separate match, joined
# by "&"/"and"/". " indifferently since each is found independently;
# verified against Kotak Balanced Advantage Fund's 3 managers on 3 distinct
# dates). pypdf's table-cell reading order sometimes separates the scheme-
# name banner from this sentence with other reordered cells in between
# (verified: passive/target-maturity index & debt-index schemes print their
# NAV/benchmark table BETWEEN the banner and the sentence) — the scheme
# name is therefore recovered by searching BACKWARD from the anchor for the
# closest "Kotak ... Fund/ETF/FOF" run within a bounded window, never by
# requiring direct adjacency.
_KOTAK_INCEPTION_RE = re.compile(r"Scheme Inception date\s+is\s+\d{1,2}/\d{1,2}/\d{2,4}")
# The name char class blocks re-entering a SECOND "Kotak" occurrence via a
# per-character `(?!Kotak)` lookahead — without this, a lazy match bridges
# across unrelated intervening benchmark/table text that has no stopping
# literal of its own, swallowing TWO unrelated fund names into one bogus
# scheme_name (caught 2026-07-11 against the real file: "Kotak Silver ETF
# Domestic Prices of physical Silver Kotak Pioneer Fund"). Same over-capture
# failure class as the 2026-07-11 HDFC/UTI/AXIS manager-name RCA — bound or
# fail closed, never trust "the next literal will stop it."
_KOTAK_SCHEME_NAME_RE = re.compile(
    r"Kotak\s+(?:(?!Kotak)[A-Za-z0-9&.,’'\- ]){2,80}?\s+(?:Fund|ETF|FOF)\b"
)
_KOTAK_MANAGER_CLAUSE_RE = re.compile(
    r"((?:Mr\.|Ms\.|Mrs\.)\s+[A-Z][A-Za-z.’' ]{1,40}?"
    r"(?:\s*(?:,|&|and)\s*(?:Mr\.|Ms\.|Mrs\.)\s+[A-Z][A-Za-z.’' ]{1,40}?)*)"
    r"\s+(?:has|have)\s+been\s+managing\s+the\s+fund\s+since\s+(\d{1,2}/\d{1,2}/\d{2,4})"
)
_KOTAK_NAME_ONLY_RE = re.compile(
    r"(?:Mr\.|Ms\.|Mrs\.)\s+([A-Z][A-Za-z.’' ]{1,40}?)(?=\s*(?:,|&|and\b|$))"
)


def _parse_kotak_compilation(reader: Any) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for page in reader.pages:
        text = page.extract_text() or ""
        if "Scheme Inception date" not in text or "managing the fund since" not in text:
            continue
        flat = " ".join(text.split())
        for m in _KOTAK_INCEPTION_RE.finditer(flat):
            window_start = max(0, m.start() - 250)
            names_before = list(_KOTAK_SCHEME_NAME_RE.finditer(flat, window_start, m.start()))
            if not names_before:
                continue
            scheme_name = " ".join(names_before[-1].group(0).split())
            if scheme_name in seen:
                continue
            next_anchor = _KOTAK_INCEPTION_RE.search(flat, m.end())
            block_end = next_anchor.start() if next_anchor else min(len(flat), m.end() + 500)
            block = flat[m.end() : block_end]
            manager_pairs: list[tuple[str, date]] = []
            for names_blob, raw_date in _KOTAK_MANAGER_CLAUSE_RE.findall(block):
                try:
                    start_date = datetime.strptime(raw_date, "%d/%m/%Y").date()
                except ValueError:
                    continue
                for name in _KOTAK_NAME_ONLY_RE.findall(names_blob):
                    clean = _strip_credential_suffix(" ".join(name.split()).strip(" ,"))
                    if clean and (clean, start_date) not in manager_pairs:
                        manager_pairs.append((clean, start_date))
            if not manager_pairs:
                continue
            out.append({"scheme_name": scheme_name, "manager_pairs": manager_pairs})
            seen.add(scheme_name)
    return out


# EDELWEISS: one or two pages per scheme (verified 2026-07-11 against the
# real ~193-page May-2026 file). Unlike Kotak, manager + tenure sit directly
# on the scheme's own page, no cross-page annexure needed: "Fund Manager\n
# Fund Managers Experience Managing Since\nMr. <Name> <N> years <DD-Mon-YY>"
# repeated per co-manager. The scheme banner ("Edelweiss <Name>", or the
# BHARAT Bond target-maturity series which carries no "Edelweiss" prefix at
# all — a real, sponsor-shared product line, verified against the TOC) sits
# on its own line(s) immediately before the SEBI "An open ended..."
# description line; the name itself sometimes WRAPS onto a 2nd/3rd line
# before that description (e.g. "Edelweiss Large & Mid Cap" / "Fund" / "An
# open ended...", caught 2026-07-11 — a single-line-only check silently
# dropped 43 of 71 real sections).
_EDEL_MANAGER_ROW_RE = re.compile(
    r"(?:Mr\.|Ms\.|Mrs\.)\s+([A-Z][A-Za-z.’' ]{1,40}?)\s+\d+\s*[Yy]ears?\s+"
    r"(\d{1,2}-[A-Za-z]{3}-\d{2,4})"
)
_EDEL_BLOCK_END_RE = re.compile(
    r"Minimum Investment Amount|Exit Load|Minimum Creation Unit Size|Minimum Additional"
)


def _edel_scheme_name_from_lines(lines: list[str]) -> str:
    for i, line in enumerate(lines[:-1]):
        s = line.strip()
        if not (s.startswith("Edelweiss ") or s.startswith("BHARAT Bond")):
            continue
        for span in (1, 2, 3):
            if i + span >= len(lines):
                break
            nxt = lines[i + span].strip()
            if nxt.lower().startswith("an open"):
                return " ".join(lines[i : i + span]).strip()
    # BHARAT Bond ETF/FOF series: single-NAV product, no "An open ended..."
    # description line pairs with the banner — fall back to the literal
    # "BHARAT Bond ... : <price>" NAV line (verified 2026-07-11).
    for line in lines:
        s = line.strip()
        if s.startswith("BHARAT Bond") and ":" in s:
            return s.split(":")[0].strip()
    return ""


def _parse_edelweiss_compilation(reader: Any) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for page in reader.pages:
        text = page.extract_text() or ""
        if "Managing Since" not in text:
            continue
        lines = [" ".join(ln.split()) for ln in text.splitlines() if ln.strip()]
        scheme_name = _edel_scheme_name_from_lines(lines)
        if not scheme_name or scheme_name in seen:
            continue
        flat = " ".join(lines)
        anchor = flat.find("Managing Since")
        if anchor < 0:
            continue
        # Slice strictly to the manager-table block (between the column
        # header and the next known section literal) — same bounded-slice
        # idiom as HDFC's "Name Since Total Exp" -> "DATE OF ALLOTMENT"
        # window, so the permissive name char class can never bridge into
        # unrelated later page text.
        end_m = _EDEL_BLOCK_END_RE.search(flat, anchor)
        block_end = end_m.start() if end_m else min(len(flat), anchor + 600)
        block = flat[anchor + len("Managing Since") : block_end]
        manager_pairs: list[tuple[str, date]] = []
        for name, raw_date in _EDEL_MANAGER_ROW_RE.findall(block):
            clean = _strip_credential_suffix(" ".join(name.split()).strip(" ,"))
            if not clean:
                continue
            try:
                start_date = datetime.strptime(raw_date, "%d-%b-%y").date()
            except ValueError:
                continue
            if (clean, start_date) not in manager_pairs:
                manager_pairs.append((clean, start_date))
        if not manager_pairs:
            continue
        out.append({"scheme_name": scheme_name, "manager_pairs": manager_pairs})
        seen.add(scheme_name)
    return out


# ABSL ("Empower Factsheet"): one or two pages per scheme (verified
# 2026-07-11 against the real ~254-page June-2026 file). Manager + tenure
# sit directly on the scheme's own page: "Fund Manager - Mr. <Name>\n
# Managing the Fund Since: <Month DD, YYYY>" repeated per co-manager, name
# and date immediately adjacent with nothing else between them (verified
# incl. 4-manager schemes, e.g. Aditya Birla Sun Life Balanced Advantage
# Fund). Scheme name comes from the Investment Performance table's own row,
# which always starts with the scheme's Title-Case name followed by a
# return percentage — a real, bounded stopping literal (mirrors HDFC's
# "Over N years" idiom: the digit-percent pattern is the actual validator,
# not a hand-picked char class boundary).
_ABSL_MANAGER_ROW_RE = re.compile(
    r"Fund Manager\s*-\s*(?:Mr\.|Ms\.|Mrs\.)\s+([A-Z][A-Za-z.’' ]{1,40}?)\s+"
    r"Managing the Fund Since:\s*([A-Za-z]+\s+\d{1,2},\s*\d{4})"
)
_ABSL_SCHEME_NAME_RE = re.compile(
    r"(Aditya Birla Sun Life\s+[A-Za-z0-9&.,’'\-—/ ]{2,60}?)\s+-?\d+\.\d+\s*%"
)


def _parse_absl_compilation(reader: Any) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for page in reader.pages:
        text = page.extract_text() or ""
        if "Managing the Fund Since" not in text or "Fund Manager" not in text:
            continue
        flat = " ".join(text.split())
        m = _ABSL_SCHEME_NAME_RE.search(flat)
        if not m:
            continue
        scheme_name = " ".join(m.group(1).split()).strip(" -")
        if not scheme_name or scheme_name in seen:
            continue
        manager_pairs: list[tuple[str, date]] = []
        for name, raw_date in _ABSL_MANAGER_ROW_RE.findall(flat):
            clean = _strip_credential_suffix(" ".join(name.split()).strip(" ,"))
            if not clean:
                continue
            try:
                start_date = datetime.strptime(raw_date, "%B %d, %Y").date()
            except ValueError:
                continue
            if (clean, start_date) not in manager_pairs:
                manager_pairs.append((clean, start_date))
        if not manager_pairs:
            continue
        out.append({"scheme_name": scheme_name, "manager_pairs": manager_pairs})
        seen.add(scheme_name)
    return out


# MIRAE (2026-07-12, brand-new AMC for this class): one scheme per page,
# verified against the real ~135-page June-2026 passive factsheet. Every
# scheme's Performance Report section repeats a single clean, well-punctuated
# summary line — "Fund Managers : Mr./Ms. <Name> (since <Month> <Day>,
# <Year> )[, Mr./Ms. <Name2> (since <Date2>)]" — the ONLY occurrence of the
# manager's name paired with a tenure date (an earlier "Fund Managers :"
# label at the top of the page lists names alone, no date, so anchoring on
# the "(since " parenthetical naturally skips it). A trailing space before
# the closing paren is real on some pages ("(since May 09, 2025 )", verified
# 2026-07-12 on Nifty50 Equal Weight ETF) — `\s*` before `\)` tolerates it.
# ~9 of the file's 65 real sections (verified: newly-launched schemes under
# 6 months old, e.g. Nifty 500 Value 50 ETF, BSE India Defence ETF) carry
# NO Performance Report at all — the page states outright "the scheme is in
# existence for less than 6 months, hence performance shall not be
# provided" (SEBI master circular clause 14.2.2) — only the dateless name
# list survives, so no (name, start_date) pair can be honestly written for
# them; §8.4 forbids fabricating a date. Scheme name is recovered from the
# SAME Performance Report table's own header row — "Period <Scheme Name>
# Scheme Benchmark* Additional Benchmark**" — which prints the AMC's own
# Title-Case spelling (the page's OTHER banner is ALL-CAPS, and this
# codebase's pg_trgm resolver is case-sensitive: an ALL-CAPS query shares
# almost no trigrams with the Title-Case master row, same root cause as the
# UTI banner-casing fix above) and may wrap across 1-3 lines for a long
# scheme name — collected up to the bounded, always-present "Scheme
# Benchmark" stopping literal, never scanned past it. A wrap can also split
# mid-word at a hyphen ("...G-\nSec ETF", verified on the 8-13 yr G-Sec ETF)
# — joined WITHOUT an inserted space when the prior fragment ends in "-",
# unlike every other line join which uses a space.
_MIRAE_MANAGER_SINCE_RE = re.compile(
    r"(?:Mr\.|Ms\.|Mrs\.)\s+([A-Z][A-Za-z.’' ]{1,40}?)\s*\(since\s+"
    r"([A-Za-z]+)\s+(\d{1,2}),\s*(\d{4})\s*\)"
)


def _mirae_join(name: str, seg: str) -> str:
    if not seg:
        return name
    if not name:
        return seg
    return name + seg if name.endswith("-") else f"{name} {seg}"


def _mirae_scheme_name_from_lines(lines: list[str]) -> str:
    name = ""
    collecting = False
    for line in lines:
        if not collecting:
            if not line.startswith("Period "):
                continue
            collecting = True
            line = line[len("Period ") :]
        if "Scheme Benchmark" in line:
            name = _mirae_join(name, line.split("Scheme Benchmark")[0].strip(" *"))
            break
        name = _mirae_join(name, line.strip(" *"))
    name = name.strip(" *")
    return name if name.startswith("Mirae Asset") else ""


def _mirae_manager_pairs(flat: str) -> list[tuple[str, date]]:
    pairs: list[tuple[str, date]] = []
    # Unlike HDFC/UTI/AXIS (Month+Year only, day forced to 1), MIRAE's real
    # "(since <Month> <Day>, <Year>)" line states the exact day — keep it
    # verbatim (§8.4), same convention as Kotak/Edelweiss/ABSL below.
    for name, month, day, year in _MIRAE_MANAGER_SINCE_RE.findall(flat):
        clean = _strip_credential_suffix(" ".join(name.split()).strip(" ,"))
        if not clean:
            continue
        try:
            start = datetime.strptime(f"{day}-{month[:3]}-{year}", "%d-%b-%Y").date()
        except ValueError:
            continue
        if (clean, start) not in pairs:
            pairs.append((clean, start))
    return pairs


def _parse_mirae_compilation(reader: Any) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for page in reader.pages:
        text = page.extract_text() or ""
        if "Fund Managers :" not in text or "(since " not in text:
            continue
        lines = [" ".join(ln.split()) for ln in text.splitlines() if ln.strip()]
        flat = " ".join(lines)
        scheme_name = _mirae_scheme_name_from_lines(lines)
        if not scheme_name or scheme_name in seen:
            continue
        manager_pairs = _mirae_manager_pairs(flat)
        if not manager_pairs:
            continue
        out.append({"scheme_name": scheme_name, "manager_pairs": manager_pairs})
        seen.add(scheme_name)
    return out


def _looks_like_amfi_aaum(data: bytes) -> bool:
    """CONTENT sniff for the AMFI consolidated scheme-wise AAUM workbook:
    a header row carrying BOTH 'AMFI Code' and 'Scheme NAV Name' plus an
    'Average Assets' title within the first few rows. xlsx-only (the portal
    serves xlsx; the check needs to open the zip) — anything unreadable is
    simply not this class, never an error.
    """
    if not data.startswith(b"PK\x03\x04"):
        return False
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            head: list[str] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 6:
                    break
                head.extend(_s(c).lower() for c in row if c is not None)
        finally:
            wb.close()
    except Exception:  # noqa: BLE001 — a sniff must never break classification
        return False
    return (
        any("average assets" in c for c in head)
        and any(c == "amfi code" for c in head)
        and any("scheme nav name" in c for c in head)
    )


_MONTH_NUMBERS = {
    name: num
    for num, name in enumerate(
        (
            "january",
            "february",
            "march",
            "april",
            "may",
            "june",
            "july",
            "august",
            "september",
            "october",
            "november",
            "december",
        ),
        start=1,
    )
}
_AMFI_AAUM_QUARTER_RE = re.compile(
    r"quarter of\s+[a-z]+\s*[-–]\s*([a-z]+)\s+(\d{4})", re.IGNORECASE
)


def parse_amfi_aaum(data: bytes, ext: str) -> tuple[date | None, list[tuple[str, str, float]]]:
    """AMFI consolidated scheme-wise AAUM workbook →
    (period, [(amfi_code, scheme_name, aaum_crore)]).

    Layout (verified against the real portal download 2026-07-10 — 8,545
    scheme rows): r0 title "Average Assets under Management (AAUM) for the
    quarter of April - June 2026 (Rs in Lakhs)"; r1 header ['AMFI Code',
    'Scheme NAV Name', 'Excluding Fund of Funds - Domestic but including
    Fund of Funds - Overseas', 'Fund Of Funds - Domestic']; below that, AMC
    section rows and category section rows (single-cell, no numeric code)
    interleave with scheme rows (numeric AMFI code + name + a value in
    exactly ONE of the two value columns — verified: no scheme carries
    both). Scheme AAUM = sum of the value columns.

    Units MUST be stated in the title block: 'lakh' → /100, 'crore' → as-is;
    an unstated unit FAILS CLOSED (a silent crore default would 100×-inflate
    every fund if AMFI ever dropped the suffix). Period comes from the
    "quarter of <M1> - <M2> <year>" title as month-start of the quarter's
    END month (the as_of_month convention used everywhere else); the caller
    may fall back to the filename, and a period-less result must never
    invent one (§8.4).
    """
    for _name, rows in _iter_sheets(data, ext):
        header_idx = None
        code_ci = None
        name_ci = None
        val_cis: list[int] = []
        for idx, row in enumerate(rows[:8]):
            lows = [_s(c).lower() for c in row]
            if "amfi code" in lows and any("scheme nav name" in c for c in lows):
                header_idx = idx
                code_ci = lows.index("amfi code")
                name_ci = next(ci for ci, c in enumerate(lows) if "scheme nav name" in c)
                val_cis = [ci for ci, c in enumerate(lows) if c and ci not in (code_ci, name_ci)]
                break
        if header_idx is None or code_ci is None or name_ci is None or not val_cis:
            continue

        head_text = " ".join(_s(c) for row in rows[: header_idx + 1] for c in row).lower()
        if "average assets" not in head_text:
            continue
        if "lakh" in head_text:
            scale = 0.01
        elif "crore" in head_text or "cr." in head_text:
            scale = 1.0
        else:
            continue  # unstated units → fail closed, never guess a 100× factor

        period: date | None = None
        m = _AMFI_AAUM_QUARTER_RE.search(head_text)
        if m:
            month = _MONTH_NUMBERS.get(m.group(1).lower())
            if month:
                try:
                    period = date(int(m.group(2)), month, 1)
                except ValueError:
                    period = None

        out: list[tuple[str, str, float]] = []
        for row in rows[header_idx + 1 :]:
            if code_ci >= len(row) or name_ci >= len(row):
                continue
            code = _s(row[code_ci])
            if code.endswith(".0"):
                code = code[:-2]  # xlrd renders integer cells as floats
            if not code.isdigit():
                continue  # AMC / category section rows carry no numeric code
            name = _s(row[name_ci])
            if not name:
                continue
            total = 0.0
            seen = False
            for ci in val_cis:
                if ci >= len(row):
                    continue
                v = row[ci]
                if isinstance(v, bool):
                    continue
                if isinstance(v, (int, float)):
                    total += float(v)
                    seen = True
                elif isinstance(v, str):
                    try:
                        total += float(v.replace(",", ""))
                        seen = True
                    except ValueError:
                        pass
            if not seen or total <= 0:
                continue  # a scheme with no/zero AAUM is skipped, never written as 0
            out.append((code, name, round(total * scale, 4)))
        if out:
            return period, out
    return None, []


def parse_riskometer_annual(data: bytes, ext: str) -> list[tuple[str, str]]:
    """Annual riskometer disclosure → [(scheme_name, band)] using the LATEST
    'Risk-o-meter level as on <date>' column (files carry start-of-FY and
    end-of-FY columns; the last one is current). Bands validated VERBATIM
    against the 6 regulatory words — 'NA'/unknown rows skipped, never coerced.
    """
    for _name, rows in _iter_sheets(data, ext):
        header_idx = None
        name_ci = None
        band_ci = None
        for idx, row in enumerate(rows[:10]):
            cells = [_s(c).lower() for c in row]
            if any("scheme name" in c for c in cells) and any("risk-o-meter" in c for c in cells):
                header_idx = idx
                name_ci = next(ci for ci, c in enumerate(cells) if "scheme name" in c)
                band_cols = [ci for ci, c in enumerate(cells) if "risk-o-meter level" in c]
                band_ci = band_cols[-1] if band_cols else None
                break
        if header_idx is None or name_ci is None or band_ci is None:
            continue

        out: list[tuple[str, str]] = []
        for row in rows[header_idx + 1 :]:
            if name_ci >= len(row) or band_ci >= len(row):
                continue
            name = _s(row[name_ci])
            band_raw = _s(row[band_ci])
            band = _BANDS_LOWER.get(band_raw.lower())
            if name and band:
                out.append((name, band))
        if out:
            return out
    return []


_SBI_BENCH_RE = re.compile(r"^scheme benchmark\s*:\s*(.+)$", re.IGNORECASE)
_TRAILING_BENCH_RE = re.compile(r"\s*\(benchmark\)?\s*$", re.IGNORECASE)


def parse_scheme_performance(data: bytes, ext: str) -> list[tuple[str, str]]:
    """Scheme-performance disclosure → [(scheme_name, benchmark_name)].

    Handles THREE real layouts (2026-07-07/08):
      - SBI: one sheet per scheme; a 'SCHEME NAME :' label row (name in the
        next non-empty cell) + an explicit 'Scheme Benchmark: <name>' cell.
      - ICICI: repeating blocks in one sheet; block header = the scheme name
        repeated across >=2 cells, then a 'Scheme' particulars row, then the
        benchmark row whose first cell is '<index name> (Benchmark)'.
      - Kotak: a plain table — one header row literally naming its own
        columns ("Scheme name" / "Benchmark Name (Tier 1)"), one scheme per
        row after that. Found inside Kotak's OWN "riskometer" disclosure
        file (its risk-o-meter LEVEL column is genuinely blank in every row —
        a change-COUNT disclosure, not a current-level one — but the same
        file carries this usable benchmark table instead); the caller tries
        this parser on any `riskometer`-classified file too, not only ones
        named/classified `performance`.
    Additional benchmarks ('Additional Benchmark: ...') are deliberately
    ignored — `benchmark_index` holds the scheme's PRIMARY benchmark.
    """
    out: list[tuple[str, str]] = []
    seen: set[str] = set()

    for _name, rows in _iter_sheets(data, ext):
        sheet_scheme: str | None = None
        block_scheme: str | None = None
        after_scheme_row = False
        table_name_ci: int | None = None
        table_bench_ci: int | None = None

        for row in rows:
            cells = [_s(c) for c in row]
            non_empty = [c for c in cells if c]
            if not non_empty:
                continue

            # Kotak-style plain table: a header row literally naming its own
            # columns. Once found, every later row in THIS sheet is read by
            # column position directly — this shape never satisfies the
            # SBI/ICICI heuristics below, so skip straight past them.
            if table_name_ci is None:
                low_cells = [c.lower() for c in cells]
                name_hits = [ci for ci, c in enumerate(low_cells) if "scheme name" in c]
                bench_hits = [ci for ci, c in enumerate(low_cells) if "benchmark name" in c]
                if name_hits and bench_hits:
                    table_name_ci, table_bench_ci = name_hits[0], bench_hits[0]
                    continue
            if table_name_ci is not None and table_bench_ci is not None:
                if table_name_ci < len(cells) and table_bench_ci < len(cells):
                    # Strip a trailing footnote marker ("Kotak Active Momentum
                    # Fund^") — Kotak (and several AMCs) mark ~1 in 7 scheme
                    # names this way in this exact table; never touches a real
                    # scheme name, which always ends alphanumeric.
                    name = cells[table_name_ci].rstrip("^*#†‡ ")
                    bench = cells[table_bench_ci]
                    if name and bench and name not in seen:
                        out.append((name, bench))
                        seen.add(name)
                continue

            # SBI: "SCHEME NAME :" label — the scheme is the next non-empty cell.
            # The captured value must LOOK like a scheme name: the per-column
            # table header row also carries a literal 'Scheme Name' cell whose
            # neighbor is 'CAGR %' — without this guard it overwrites the real
            # scheme captured from the true label row above it.
            for ci, cell in enumerate(cells):
                if cell.upper().startswith("SCHEME NAME"):
                    rest = next((c for c in cells[ci + 1 :] if c), "")
                    if rest and any(kw in rest.lower() for kw in ("fund", "etf", "scheme", "plan")):
                        sheet_scheme = rest
                    break

            # SBI: explicit benchmark label anywhere in the row.
            for cell in non_empty:
                m = _SBI_BENCH_RE.match(cell)
                if m and sheet_scheme and sheet_scheme not in seen:
                    out.append((sheet_scheme, m.group(1).strip()))
                    seen.add(sheet_scheme)
                    break

            # ICICI: block header = same value repeated across >=2 cells — and it
            # must LOOK like a scheme name (repeated 'CAGR (%)' / 'Returns' header
            # rows otherwise satisfy the repetition check and steal the block).
            if len(non_empty) >= 2 and len(set(non_empty)) == 1:
                cand = non_empty[0]
                cand_low = cand.lower()
                if any(kw in cand_low for kw in ("fund", "etf", "scheme", "plan")) and not any(
                    kw in cand_low for kw in ("cagr", "return", "%")
                ):
                    block_scheme = cand
                    after_scheme_row = False
                continue

            first = cells[0] if cells else ""
            if first.lower() == "scheme":
                after_scheme_row = True
                continue
            if after_scheme_row and first and first.lower() not in ("particulars",):
                if block_scheme and block_scheme not in seen:
                    bench = _TRAILING_BENCH_RE.sub("", first).strip()
                    # A benchmark cell names an index — require an index-ish word
                    # (every real benchmark carries one), never a bare number.
                    if any(
                        kw in bench.lower()
                        for kw in ("index", "tri", "nifty", "sensex", "crisil", "bse")
                    ):
                        out.append((block_scheme, bench))
                        seen.add(block_scheme)
                after_scheme_row = False

    return out


def _parse_ter_date(v: Any) -> date | None:
    """A TER-date cell is either a real `datetime` (openpyxl auto-converts a
    genuine Excel date cell) or a string in one of the two conventions AMCs
    actually use in these files (DD-MON-YYYY / DD/MM/YYYY — Indian date
    order, never US MM/DD)."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _s(v)
    if not s:
        return None
    for fmt in ("%d-%b-%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_ter_number(v: Any) -> float | None:
    """A TER percentage cell can be a real float/int OR a numeric-looking
    string (confirmed 2026-07-09, Edelweiss: every value cell is text like
    "0.0000", not a genuine number) — accept either, never guess on anything
    else."""
    if isinstance(v, (int, float)):
        return float(v)
    s = _s(v)
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_ter_disclosure(data: bytes, ext: str) -> list[tuple[str, date, float, float]]:
    """Total Expense Ratio (TER) disclosure -> [(scheme_name, ter_date,
    regular_total_ter_pct, direct_total_ter_pct)], ONE row per scheme (the
    LATEST date only — TER changes over time and only the current effective
    rate should ever be written).

    Confirmed 2026-07-09 across 3 real AMC layouts (SBI, ABSL, Edelweiss) that
    all share one structural invariant despite differing header shapes: a
    "Regular"-labelled column GROUP and a "Direct"-labelled column group
    (Regular always precedes Direct left-to-right), each ending in its own
    "Total TER (%)" sub-column — the all-in number; the BER/brokerage/
    transaction-cost/GST components that precede it are deliberately NOT
    extracted (§8.4 verbatim-only — we want the one number the AMC itself
    calls the total, never a recomputation). Header depth varies (SBI: one
    row, every column literally prefixed "Regular Plan - " / "Direct Plan -";
    Edelweiss: a group-title row ["Regular Plan", "Direct Plan"] followed by
    a shared 5-label sub-header row; ABSL: the same 2-row shape with bare
    "Regular"/"Direct" group titles) — detected generically by locating the
    "regular"/"direct" group-start columns and every "total ter" column,
    rather than hardcoding 3 separate branches.

    A scheme's holdings are IDENTICAL across Regular and Direct plans (TER is
    a plan-level, not option-level, fee) — the resolver in manual_ingest.py
    applies these two values to ALL option-variant ISINs (Growth/IDCW/Bonus/
    etc.) under each plan, never a single top-1 fuzzy match.

    AMFI's consolidated all-AMC TER workbook (2026-07-10, 'TER_Revised'
    sheet, one row per scheme per DAY, 15 columns "Regular Plan - ..." /
    "Direct Plan - ..." each ending in "Total TER (%)") is covered by this
    SAME generic detection with no extra branch — verified against the real
    portal download: 1,813 schemes, latest-date dedupe correct. Only the
    per-plan Total TER (all-in incl. GST) is extracted, matching every other
    layout. That file is multi-AMC: its writer path runs with amc_name=None
    (see _resolve_scheme_isins_by_plan's None handling).
    """
    latest: dict[str, tuple[date, float, float]] = {}

    for _name, rows in _iter_sheets(data, ext):
        header_rows = rows[:4]
        has_regular = False
        has_direct = False
        name_ci: int | None = None
        date_ci: int | None = None
        total_ter_cols: list[int] = []
        header_end = 0

        for ri, hr in enumerate(header_rows):
            non_empty_count = sum(1 for c in hr if _s(c))
            row_has_total_ter = False
            for ci, cell in enumerate(hr):
                s = _s(cell).lower()
                if not s:
                    continue
                if "regular" in s:
                    has_regular = True
                if "direct" in s:
                    has_direct = True
                if "name" in s and name_ci is None:
                    name_ci = ci
                if "date" in s and date_ci is None:
                    date_ci = ci
                # A free-text title banner ("Total Expense Ratio (TER) for
                # Mutual Fund Schemes") can spell out both "total" and "ter"
                # in prose — confirmed 2026-07-09, HDFC's TER file — and is
                # never a real multi-column header row (always <=2 non-empty
                # cells: either the single banner cell itself, or the
                # 2-cell "Regular Plan"/"Direct Plan" group-title row, which
                # never contains "total"/"ter" anyway). Require >2 non-empty
                # cells so only a genuine tabular header row can contribute.
                if non_empty_count > 2 and "total" in s and "ter" in s:
                    total_ter_cols.append(ci)
                    row_has_total_ter = True
            if row_has_total_ter:
                header_end = ri

        # Every real layout has EXACTLY one "Total TER" column per plan — the
        # group-TITLE cells ("Regular Plan" / "Direct Plan") are not reliably
        # positioned at the START of their own column block (confirmed
        # 2026-07-09, Edelweiss: title cells sit at columns 0/1 while the
        # actual Regular/Direct sub-column blocks are 2-6/7-11), so bucketing
        # by title-column proximity is unreliable. Left-to-right POSITION
        # order is: every real AMC lists the Regular block before the Direct
        # block, so with exactly 2 "Total TER" columns the smaller index is
        # always Regular's and the larger is always Direct's — simpler and
        # correct across all 3 known layouts (SBI/ABSL/Edelweiss). More or
        # fewer than 2 is an unrecognized layout — skip the sheet, never
        # guess which column is which.
        if (
            not has_regular
            or not has_direct
            or name_ci is None
            or date_ci is None
            or len(total_ter_cols) != 2
        ):
            continue

        regular_total_ci, direct_total_ci = sorted(total_ter_cols)

        for row in rows[header_end + 1 :]:
            if max(name_ci, date_ci, regular_total_ci, direct_total_ci) >= len(row):
                continue
            name = _s(row[name_ci])
            if not name:
                continue
            ter_date = _parse_ter_date(row[date_ci])
            reg_val = _parse_ter_number(row[regular_total_ci])
            dir_val = _parse_ter_number(row[direct_total_ci])
            if ter_date is None or reg_val is None or dir_val is None:
                continue
            prior = latest.get(name)
            if prior is None or ter_date > prior[0]:
                latest[name] = (ter_date, round(float(reg_val), 4), round(float(dir_val), 4))

    return [(name, d, reg, dirv) for name, (d, reg, dirv) in latest.items()]


# ---------------------------------------------------------------------------
# Scheme master details — SBI's per-scheme "Fund Details" HTML page
# (confirmed 2026-07-08, ~52 files — saved with a misleading ".xls"
# extension, plain HTML, not a spreadsheet at all).
# ---------------------------------------------------------------------------

# The AMC's OWN template schema — every real sample carries this identical
# label list in this identical order. This is a generic template shared by
# EVERY scheme (like RISKOMETER_BANDS' 6 fixed words), never a per-scheme
# lookup table: used only to find where one label's value ENDS (the position
# of whichever label comes next), not to hardcode any scheme's own data.
_FUND_DETAILS_LABELS: tuple[str, ...] = (
    "Fund Name",
    "Options Names",
    "Fund Type",
    "Riskometer At Launch",
    "Riskometer As on Date",
    "Category as per SEBI categorization Circular",
    "Potential Risk Class",
    "Description",
    "Stated Asset Allocation",
    "Face Value",
    "NFO Open Date",
    "NFO Close Date",
    "Allotment Date",
    "Reopen Date",
    "Maturity Date",
    "Benchmark(Tier 1)",
    "Benchmark(Tier 2)",
    "Fund Manager",
    "Fund Manager Type",
    "Fund Manager From Date",
    "Annual Expense (Stated Maximum)",
    "Exit Load",
    "Custodian",
    "Auditor",
    "Registrar",
    "RTA Code",
    "Swing Pricing",
    "Side-Pocketing",
    "Listing Details",
    "ISIN",
)

_TAG_RE = re.compile(r"<[^>]+>")
_MANAGER_DATE_RE = re.compile(r"((?:Mr|Ms|Mrs)\.?[^:]+?):\s*(\d{1,2}-[A-Za-z]{3}-\d{4})")


def _looks_like_scheme_master_html(data: bytes) -> bool:
    """Content sniff — the file's own filename looks like a normal per-scheme
    portfolio file (e.g. "SBI Multicap Fund.xls"), so only the BYTES reveal
    this is actually an AMC website page, not a spreadsheet."""
    head = data[:4000].decode("utf-8", errors="ignore").lower()
    return "<html" in head and "fund details for" in head


def _flatten_html(data: bytes) -> str:
    """Strip every tag to a single space (collapsing nested tables' cell
    values into the surrounding text stream) and unescape entities — turns
    the label/value table into one flat, whitespace-normalized string that
    `_value_after` can slice by label position."""
    text = data.decode("utf-8", errors="replace")
    text = _TAG_RE.sub(" ", text)
    text = html.unescape(text)
    return re.sub(r"\s+", " ", text).strip()


def _extract_all_labels(flat_text: str) -> dict[str, str]:
    """ONE sequential pass through the known label order (each label
    searched for starting only AFTER the previous label's own match ends),
    extracting every label's value as the text up to the NEXT label found in
    sequence. Sequential, monotonically-advancing search means an EARLIER
    label's name reappearing INSIDE a LATER value can never be mistaken for
    a boundary (confirmed 2026-07-08: a co-manager designation "(Co-Fund
    Manager)" literally contains the substring "Fund Manager", which a
    whole-list re-scan from position 0 wrongly matched as the next section
    start, truncating "Fund Manager From Date"'s real value)."""
    positions: list[tuple[str, int]] = []
    search_from = 0
    for label in _FUND_DETAILS_LABELS:
        idx = flat_text.find(label, search_from)
        if idx < 0:
            continue
        positions.append((label, idx))
        search_from = idx + len(label)

    values: dict[str, str] = {}
    for i, (label, idx) in enumerate(positions):
        start = idx + len(label)
        end = positions[i + 1][1] if i + 1 < len(positions) else len(flat_text)
        values[label] = flat_text[start:end].strip(" :")
    return values


def parse_scheme_master_details(data: bytes) -> dict[str, Any]:
    """SBI's per-scheme "Fund Details" page → one dict (ONE file = ONE
    scheme). Returns {} if the file doesn't even carry a Fund Name (fail-
    closed — caller marks 'unsupported', never guesses).

    Keys (any may be absent/None if the file omits a label):
      scheme_name: str
      risk_band: str | None — validated VERBATIM against the 6 regulatory
        words (RISKOMETER_BANDS), same rule as the dedicated riskometer parser.
      benchmark_tier1: str | None — the scheme's PRIMARY benchmark name.
      ter_pct: float | None — "Annual Expense (Stated Maximum)"'s number.
      manager_pairs: list[tuple[str, date]] — (manager_name, tenure start),
        parsed from "Fund Manager From Date"'s own "Name:DD-Mon-YYYY" pairs
        (NOT the separate "Fund Manager" list — that field has no per-name
        date, so pairing off this field is the only way to attribute the
        right start date to the right manager when a scheme has more than one).

    Exit load is deliberately NOT extracted (see module docstring).
    """
    return _scheme_master_from_flat(_flatten_html(data))


def _flatten_pdf(data: bytes) -> str:
    """pypdf text extraction across all pages, whitespace-normalized into the
    same flat "Label Value Label Value" stream `_extract_all_labels` slices.
    pypdf is already a declared dependency (NIPPON factsheet parser)."""
    import io

    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    text = " ".join(page.extract_text() or "" for page in reader.pages)
    return re.sub(r"\s+", " ", text).strip()


def looks_like_scheme_master_pdf(data: bytes) -> bool:
    """Content sniff for SBI's per-scheme "Fund Details" PDF (2026-07-10:
    the founder's 52 SBI uploads are this page rendered to PDF — they sat
    'archived' while carrying CURRENT riskometer/TER/benchmark/manager/exit
    load). First page only — cheap enough for the intake dispatch path."""
    if not data.startswith(b"%PDF"):
        return False
    try:
        import io

        from pypdf import PdfReader

        first = PdfReader(io.BytesIO(data)).pages[0].extract_text() or ""
    except Exception:  # noqa: BLE001 — a corrupt PDF is simply not this class
        return False
    return "fund details for" in first.lower()


def parse_scheme_master_pdf(data: bytes) -> dict[str, Any]:
    """PDF twin of `parse_scheme_master_details` — same labels, same output
    dict (the AMC renders the same Fund-Details page to both HTML and PDF).
    Fail-closed {} when no Fund Name is found."""
    try:
        flat = _flatten_pdf(data)
    except Exception:  # noqa: BLE001 — unreadable PDF → unsupported, never a guess
        return {}
    return _scheme_master_from_flat(flat)


def _scheme_master_from_flat(flat: str) -> dict[str, Any]:
    labels = _extract_all_labels(flat)

    scheme_name = labels.get("Fund Name", "")
    if not scheme_name:
        return {}

    risk_raw = labels.get("Riskometer As on Date", "")
    risk_band = _BANDS_LOWER.get(risk_raw.strip().lower())

    benchmark_raw = labels.get("Benchmark(Tier 1)", "").strip()
    benchmark: str | None = (
        benchmark_raw if benchmark_raw.upper() not in ("", "NA", "N/A", "-") else None
    )

    expense_raw = labels.get("Annual Expense (Stated Maximum)", "")
    ter_pct: float | None = None
    ter_m = re.search(r"(\d+(?:\.\d+)?)", expense_raw)
    if ter_m:
        try:
            ter_pct = float(ter_m.group(1))
        except ValueError:
            ter_pct = None

    since_raw = labels.get("Fund Manager From Date", "")
    manager_pairs: list[tuple[str, date]] = []
    for name_part, date_str in _MANAGER_DATE_RE.findall(since_raw):
        try:
            manager_pairs.append(
                (name_part.strip(), datetime.strptime(date_str, "%d-%b-%Y").date())
            )
        except ValueError:
            continue

    exit_load_pct, exit_load_days = _parse_exit_load_text(labels.get("Exit Load", ""))

    return {
        "scheme_name": scheme_name,
        "risk_band": risk_band,
        "benchmark_tier1": benchmark,
        "ter_pct": ter_pct,
        "manager_pairs": manager_pairs,
        "exit_load_pct": exit_load_pct,
        "exit_load_days": exit_load_days,
    }


_EXIT_PCT_RE = re.compile(r"(\d+(?:\.\d+)?)\s*%")
_EXIT_PERIOD_RES = (
    (re.compile(r"(\d+)\s*day", re.IGNORECASE), 1),
    (re.compile(r"(\d+)\s*month", re.IGNORECASE), 30),
    (re.compile(r"(\d+)\s*year", re.IGNORECASE), 365),
)


def _parse_exit_load_text(text: str) -> tuple[float | None, int | None]:
    """Exit-load prose → (pct, days) — fail-closed (None, None) when unstated.

    Real SBI Fund-Details form (2026-07-10): "For exit within 30 days from
    the date of allotment - 1%. For exit after 30 days - Nil." → (1.0, 30).
    A bare "Nil"/"NIL" is a REAL fact (no exit load) → (0.0, None). The pct
    is the FIRST percentage stated (the binding near-term tier); days is the
    longest period mentioned. Historically this label was deliberately
    skipped; extracting it now (founder-directed, 2026-07-10) because the
    BSE StAR source stays dormant until prod creds."""
    cleaned = text.strip()
    if not cleaned:
        return None, None
    m = _EXIT_PCT_RE.search(cleaned)
    if m is None:
        if re.fullmatch(r"nil\.?", cleaned, re.IGNORECASE):
            return 0.0, None
        return None, None
    try:
        pct = float(m.group(1))
    except ValueError:
        return None, None
    if not 0 <= pct <= 20:
        return None, None
    periods = [
        int(pm.group(1)) * mult
        for pattern, mult in _EXIT_PERIOD_RES
        for pm in pattern.finditer(cleaned)
        if 0 < int(pm.group(1)) * mult <= 3650
    ]
    return pct, (max(periods) if periods else None)
