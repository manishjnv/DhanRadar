"""
DhanRadar — AMFI half-yearly Large/Mid/Small Cap stock classification parser.

Source (Phase A verified 2026-07-05): AMFI publishes a half-yearly consolidated
list at a predictable URL:

    https://portal.amfiindia.com/spages/AverageMarketCapitalization{DD}{Mon}{YYYY}.xlsx

e.g. AverageMarketCapitalization30Jun2026.xlsx (period ended 30 Jun 2026, i.e.
the Jan-Jun 2026 half). The Jul-Dec half is expected to land at the 31 Dec
stem. Same lightly-protected portal.amfiindia.com domain as NAVAll — a plain
browser User-Agent is sufficient, no Playwright needed.

File layout (single sheet "FINAL"):
    row 0: title ("Average Market Capitalization of listed companies ...")
    row 1: header (Sr. No. | Company name | ISIN | BSE Symbol | BSE 6m Avg Mkt
           Cap (Rs Cr) | NSE Symbol | NSE 6m Avg Mkt Cap (Rs Cr) | MSEI Symbol |
           MSEI 6m Avg Mkt Cap (Rs Cr) | Average of All Exchanges (Rs Cr) |
           Categorization as per SEBI Circular dated Oct 6, 2017)
    row 2+: one row per stock. ISIN is the EQUITY ISIN (INE... prefix) — this
           is a stock-level dataset, distinct from MF ISINs (INF... prefix).

Pure module: `parse_cap_classification_rows` takes already-extracted
(row-tuple) data — no I/O, no DB — so it is unit-testable with plain Python
tuples mirroring the real layout above. `fetch_cap_classification` is the thin
I/O wrapper (httpx GET + openpyxl parse) used by the ingestion task.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import date

import httpx

from dhanradar.market_data.exceptions import ProviderError

logger = logging.getLogger(__name__)

_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)
_TIMEOUT = httpx.Timeout(30.0, connect=10.0)

_VALID_CAP_CLASSES = frozenset({"Large Cap", "Mid Cap", "Small Cap"})

# Half-year period ends: (month, day) tuples in calendar order.
_PERIOD_ENDS = ((6, 30), (12, 31))


@dataclass(frozen=True)
class StockCapRow:
    """One parsed row: a stock's SEBI large/mid/small-cap categorization."""

    stock_isin: str
    stock_name: str
    cap_class: str  # one of _VALID_CAP_CLASSES
    avg_market_cap_cr: float | None
    effective_period: str  # e.g. "2026H1"


def period_label(period_end: date) -> str:
    """Public label for a half-year period end, e.g. date(2026,6,30) -> '2026H1'."""
    half = "H1" if period_end.month == 6 else "H2"
    return f"{period_end.year}{half}"


def candidate_period_ends(today: date) -> list[date]:
    """Most-recent-first half-year period ends to try (current, then previous).

    AMFI publishes with a lag after each half closes, so the just-closed half
    may not be up yet — the caller tries these in order and stops at the first
    URL that resolves.
    """
    candidates: list[date] = []
    year = today.year
    for _ in range(3):
        for month, day in reversed(_PERIOD_ENDS):
            period_end = date(year, month, day)
            if period_end <= today:
                candidates.append(period_end)
        year -= 1
    # Dedup while preserving order, keep the 3 most recent.
    seen: set[date] = set()
    ordered: list[date] = []
    for c in sorted(candidates, reverse=True):
        if c not in seen:
            seen.add(c)
            ordered.append(c)
    return ordered[:3]


def build_url(period_end: date) -> str:
    """Predictable AMFI URL for a half-year period end date."""
    return (
        "https://portal.amfiindia.com/spages/AverageMarketCapitalization"
        f"{period_end.day:02d}{period_end.strftime('%b')}{period_end.year}.xlsx"
    )


def parse_cap_classification_rows(rows: list[tuple], *, effective_period: str) -> list[StockCapRow]:
    """PURE parser — takes already-extracted (openpyxl) row tuples.

    Skips the title row, header row, and any row missing a valid ISIN or an
    unrecognized cap_class label (never guessed/coerced — §8.4 no-fabrication).
    Column order (0-indexed): 0=Sr, 1=Company name, 2=ISIN, 3=BSE Symbol,
    4=BSE avg cap, 5=NSE Symbol, 6=NSE avg cap, 7=MSEI Symbol, 8=MSEI avg cap,
    9=Average of All Exchanges, 10=Categorization.
    """
    out: list[StockCapRow] = []
    for row in rows:
        if not row or len(row) < 11:
            continue
        isin = row[2]
        name = row[1]
        cap_class = row[10]
        if not isinstance(isin, str) or not isin.strip():
            continue
        if not isinstance(cap_class, str) or cap_class.strip() not in _VALID_CAP_CLASSES:
            continue
        if not isinstance(name, str) or not name.strip():
            continue
        avg_cap = row[9]
        avg_cap_val: float | None = None
        if isinstance(avg_cap, (int, float)):
            avg_cap_val = float(avg_cap)
        out.append(
            StockCapRow(
                stock_isin=isin.strip(),
                stock_name=name.strip(),
                cap_class=cap_class.strip(),
                avg_market_cap_cr=avg_cap_val,
                effective_period=effective_period,
            )
        )
    return out


async def fetch_cap_classification(
    client: httpx.AsyncClient | None, today: date | None = None
) -> tuple[list[StockCapRow], str, str]:
    """Fetch + parse the most-recent resolvable half-year cap classification file.

    Returns (rows, effective_period, source_url). Raises ProviderError if no
    candidate period resolves (never guesses/fabricates data).
    """
    import openpyxl

    today = today or date.today()
    headers = {"User-Agent": _USER_AGENT}

    for period_end in candidate_period_ends(today):
        url = build_url(period_end)
        try:
            if client is not None:
                resp = await client.get(url, headers=headers, timeout=_TIMEOUT)
            else:
                async with httpx.AsyncClient() as new_client:
                    resp = await new_client.get(url, headers=headers, timeout=_TIMEOUT)
        except httpx.HTTPError as exc:
            logger.warning("fetch_cap_classification: %s network error: %s", url, exc)
            continue

        if resp.status_code != 200:
            logger.debug(
                "fetch_cap_classification: %s returned HTTP %s, trying earlier period",
                url,
                resp.status_code,
            )
            continue

        import io

        try:
            wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            raw_rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        except Exception as exc:  # noqa: BLE001
            logger.warning("fetch_cap_classification: %s unparseable: %s", url, exc)
            continue

        # Data rows start after the title (row 0) and header (row 1).
        data_rows = raw_rows[2:]
        period_lbl = period_label(period_end)
        parsed = parse_cap_classification_rows(data_rows, effective_period=period_lbl)
        if not parsed:
            logger.warning("fetch_cap_classification: %s returned 0 parseable rows", url)
            continue
        return parsed, period_lbl, url

    raise ProviderError(
        "fetch_cap_classification: no candidate half-year period resolved "
        f"(tried {[build_url(p) for p in candidate_period_ends(today)]})"
    )
