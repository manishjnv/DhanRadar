"""
DhanRadar — SBI TER (expense ratio) xlsx provider (Block 0.5, 2026-07-06).

SBI Mutual Fund's factsheet-listing pages 404 on plain HTTP (portfolio-related
content is JS/AJAX-rendered — see amc_registry.py's SBI comment), BUT its
"Total Expense Ratio of Mutual Fund Schemes" page (`/total-expense-ratio`) is a
plain server-rendered HTML page that links directly to two STATIC xlsx files
(verified reachable 2026-07-06, real browser UA, no bot-block/challenge page
encountered on the whole sbimf.com domain):

    https://www.sbimf.com/docs/default-source/ter_allschemes/current-year-ter.xlsx
    https://www.sbimf.com/docs/default-source/ter_allschemes/historical-ter.xlsx

This module fetches + parses the CURRENT-YEAR file only (11 863 real data
rows verified 2026-07-06). Real header row (single sheet "Sheet1"):

    NSDL Scheme Code | Scheme Name | TER Date (DD/MM/YYYY) |
    Regular Plan - Base Expense Ratio (BER) (%) | Regular Plan - Brokerage cost (%) |
    Regular Plan - Transaction Cost ... (%) | Regular Plan - Statutory Levies ... (%) |
    Regular Plan - Total TER (%) |
    Direct Plan - Base Expense Ratio (BER) (%) | Direct Plan - Brokerage cost (%) |
    Direct Plan - Transaction Cost ... (%) | Direct Plan - Statutory Levies ... (%) |
    Direct Plan - Total TER (%)

There is NO ISIN column — only an NSDL scheme code (e.g.
"SBIM/O/E/THE/97/01/0001") and a plan-agnostic Scheme Name. Each row yields UP
TO TWO `RawExpenseRatioRow`s (Regular Plan Total TER + Direct Plan Total TER),
keyed by scheme name + a plan-name suffix appended before fuzzy resolution —
the SAME shared pg_trgm matcher the constituents/UTI/NIPPON pipelines use
(`dhanradar.tasks.mf._resolve_scheme_isins`), never a second matcher (see
`tasks/mf_expense_ratio.py` for where that resolution happens).

Known coverage limit (honest, not fabricated): the fuzzy matcher resolves to
its SINGLE top-similarity ISIN per (scheme name + plan) — a scheme with
multiple options (Growth vs IDCW) under the same plan only has its
top-matching option's ISIN written; sibling option ISINs are not
back-filled. This mirrors the existing UTI/NIPPON fund-manager fetchers'
resolution behavior exactly — no new gap introduced.
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from datetime import date, datetime

import httpx

logger = logging.getLogger(__name__)

_TER_URL = "https://www.sbimf.com/docs/default-source/ter_allschemes/current-year-ter.xlsx"

# Header keywords matched case-insensitively, substring, against the normalized
# (lowercased, whitespace-collapsed) header cell — real header text wraps
# "TER Date" across newlines ("TER Date\n(DD/MM/\nYYYY)"), so match on the
# short unambiguous keyword, not the full cell text.
_COL_SCHEME_NAME = "scheme name"
_COL_TER_DATE = "ter date"
_COL_REGULAR_TOTAL_TER = "regular plan - total ter"
_COL_DIRECT_TOTAL_TER = "direct plan - total ter"


@dataclass
class RawExpenseRatioRow:
    """A TER data point keyed by scheme NAME + plan suffix (not ISIN).

    ``plan_suffix`` (" - Direct Plan" or " - Regular Plan") is appended to
    ``scheme_name`` by the caller before fuzzy ISIN resolution — SBI's
    mf_funds scheme_name rows carry this exact AMFI convention, so appending
    it materially improves trigram similarity vs matching the bare scheme
    name alone (which is plan-ambiguous in this source file).
    """

    scheme_name: str
    plan_suffix: str
    ter_pct: float
    effective_date: date


def _normalize_header_cell(cell: object) -> str:
    return " ".join(str(cell or "").lower().split())


def _find_col(header: tuple, keyword: str) -> int | None:
    for i, cell in enumerate(header):
        if keyword in _normalize_header_cell(cell):
            return i
    return None


def _load_sbi_ter_rows(file_bytes: bytes) -> list[tuple]:
    """CPU-bound (no network) — extracts raw (row-tuple) data from the xlsx's
    first sheet via openpyxl. Isolated from `parse_sbi_ter_rows` so the parser
    itself stays pure/testable against captured fixture tuples (mirrors
    amc_managers_nippon.extract_pdf_page_texts)."""
    import openpyxl  # lazily imported — matches amc_expense.py/tasks/mf.py convention

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return list(ws.iter_rows(values_only=True))


def parse_sbi_ter_rows(rows: list[tuple]) -> list[RawExpenseRatioRow]:
    """PURE parser — no I/O. Takes the raw (header + data) row tuples already
    extracted from the xlsx (see `_load_sbi_ter_rows`) and returns
    high-confidence TER rows.

    Rules (fail-closed, mirrors amc_expense.parse_expense_ratios):
    - The header row is located by column-name match, not by position (SBI
      has changed column order/wording before per this file's own history).
      A header missing ANY of the 4 required columns → return [] (the caller
      records this as format_mismatch, never a partial/guessed row).
    - A data row with an empty scheme_name, unparseable TER date, or ter_pct
      outside (0, 10] for a given plan is SKIPPED for that plan only — the
      row's OTHER plan (if valid) is still kept.
    - Does NOT fabricate, interpolate, or impute any missing value.
    """
    if not rows:
        return []

    header = rows[0]
    name_idx = _find_col(header, _COL_SCHEME_NAME)
    date_idx = _find_col(header, _COL_TER_DATE)
    regular_idx = _find_col(header, _COL_REGULAR_TOTAL_TER)
    direct_idx = _find_col(header, _COL_DIRECT_TOTAL_TER)

    if name_idx is None or date_idx is None or regular_idx is None or direct_idx is None:
        logger.warning(
            "parse_sbi_ter_rows: header missing required column(s) "
            "(name=%s date=%s regular=%s direct=%s) — refusing to guess",
            name_idx,
            date_idx,
            regular_idx,
            direct_idx,
        )
        return []

    out: list[RawExpenseRatioRow] = []
    for row in rows[1:]:
        if row is None or len(row) <= max(name_idx, date_idx, regular_idx, direct_idx):
            continue

        scheme_name = str(row[name_idx] or "").strip()
        if not scheme_name:
            continue

        effective_date = _coerce_date(row[date_idx])
        if effective_date is None:
            logger.debug(
                "parse_sbi_ter_rows: unparseable TER date %r for scheme %r — skipping row",
                row[date_idx],
                scheme_name,
            )
            continue

        for plan_suffix, idx in ((" - Regular Plan", regular_idx), (" - Direct Plan", direct_idx)):
            ter_pct = _coerce_ter_pct(row[idx])
            if ter_pct is None:
                continue
            out.append(
                RawExpenseRatioRow(
                    scheme_name=scheme_name,
                    plan_suffix=plan_suffix,
                    ter_pct=ter_pct,
                    effective_date=effective_date,
                )
            )

    return out


def _coerce_date(raw: object) -> date | None:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, str) and raw.strip():
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(raw.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _coerce_ter_pct(raw: object) -> float | None:
    if not isinstance(raw, (int, float, str)) or isinstance(raw, bool):
        return None
    try:
        ter_pct = float(raw)
    except (TypeError, ValueError):
        return None
    # Same boundary as amc_expense.parse_expense_ratios / the DB CHECK constraint.
    if ter_pct <= 0 or ter_pct > 10:
        return None
    return ter_pct


async def fetch_sbi_expense_ratios(
    client: httpx.AsyncClient,
) -> tuple[list[RawExpenseRatioRow], dict]:
    """Fetch + parse SBI's current-year TER xlsx.

    Returns (rows, status) with the SAME status-bucket contract as
    `amc_expense.fetch_expense_ratios`: "bot_blocked" (always [] — SBI is not
    bot-blocked, see amc_registry.py), "unreachable" (network/HTTP failure),
    "format_mismatch" (HTTP 200 but the xlsx shape changed / 0 usable rows),
    "ok" (["SBI"] if >= 1 row extracted). Never raises.
    """
    empty_unreachable = {"bot_blocked": [], "unreachable": ["SBI"], "format_mismatch": [], "ok": []}
    empty_format_mismatch = {
        "bot_blocked": [],
        "unreachable": [],
        "format_mismatch": ["SBI"],
        "ok": [],
    }

    try:
        resp = await client.get(_TER_URL, timeout=30.0, follow_redirects=True)
        if resp.status_code != 200:
            logger.warning("fetch_sbi_expense_ratios: HTTP %s", resp.status_code)
            return [], empty_unreachable
        file_bytes = resp.content
    except (httpx.TimeoutException, httpx.TransportError, httpx.HTTPError) as exc:
        logger.warning("fetch_sbi_expense_ratios: network error: %s", exc)
        return [], empty_unreachable

    try:
        raw_rows = _load_sbi_ter_rows(file_bytes)
    except Exception as exc:  # noqa: BLE001 — a corrupt/changed xlsx must not crash the run
        logger.warning("fetch_sbi_expense_ratios: xlsx parse failed: %s", exc)
        return [], empty_format_mismatch

    rows = parse_sbi_ter_rows(raw_rows)
    if not rows:
        logger.info(
            "fetch_sbi_expense_ratios: 0 high-confidence rows in %d raw rows (format_mismatch)",
            len(raw_rows),
        )
        return [], empty_format_mismatch

    logger.info("fetch_sbi_expense_ratios: %d rows parsed", len(rows))
    return rows, {"bot_blocked": [], "unreachable": [], "format_mismatch": [], "ok": ["SBI"]}
