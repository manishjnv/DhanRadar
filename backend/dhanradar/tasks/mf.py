"""
DhanRadar — MF Celery tasks (Phase 5).

`parse_cas_job` is the CAS→report worker (architecture pipeline steps 3–8):
parse → upsert holdings → snapshot → score (via the engine interface) → cache the
report → done. The raw CAS file is deleted immediately after parse; a daily
`purge_cas_files` sweep is the 24h backstop (anti-pattern guard).

`nav_daily_fetch` pulls the current NAVAll.txt from AMFI, bulk-upserts
mf_nav_history and mf_funds metadata (amfi_code, scheme_name, category only —
other metadata columns such as expense_ratio/aum are not overwritten).

`nav_backfill` bootstraps multi-year historical NAV data by iterating over
<=90-day windows and upserting into mf_nav_history.  It is NOT in the beat
schedule — invoke manually.

All Celery tasks are sync; async DB/Redis/network pipelines run under asyncio.run.
Pure mapping helpers are factored out for unit testing without a worker.
"""

from __future__ import annotations

import asyncio
import csv
import io
import json
import logging
import math
import os
import re
import time
import zipfile
from collections import Counter
from collections.abc import AsyncIterator
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta, timezone
from typing import Any
from uuid import uuid4

import httpx
from structlog.contextvars import bind_contextvars

from dhanradar.celery_app import celery_app
from dhanradar.core.logging import get_logger, hash_user_ref
from dhanradar.mf import service
from dhanradar.mf.benchmark_mapping import map_index_fund_benchmark
from dhanradar.mf.cas import (
    CasParseError,
    ParsedHolding,
    classify_cas_failure,
    detect_and_parse,
    filter_foreign_pan_folios,
    normalize_folio,
    parser_version_for,
    resolve_blank_folios,
    split_ledger_eligible,
    suppress_placeholder_restatements,
)
from dhanradar.mf.cohort import CohortBenchmark, FundStats
from dhanradar.mf.scoring_bridge import score_fund, upsert_user_fund_score
from dhanradar.mf.signals import CategoryRelative, compute_fund_signals
from dhanradar.mf.snapshot import CashFlow, Holding, build_snapshot
from dhanradar.mf.taxonomy import (
    canonical_for,
    derive_short_name,
    parse_idcw_frequency,
    parse_plan_option,
)
from dhanradar.mf.taxonomy import summarize as taxonomy_summarize

logger = logging.getLogger(__name__)
_slog = get_logger(__name__)

_UPLOAD_TTL_SECONDS = 24 * 3600

# Batch size for bulk-upsert statements — bounds memory and statement size.
_UPSERT_CHUNK = 2000

# Top-10 AMCs by AUM — SEBI monthly portfolio disclosure discovery roots.
# These are the SEBI-mandated scheme-portfolio disclosure landing pages.
# Each AMC publishes monthly disclosure XLSX/CSV under these paths.
# Format: name (for source_amc provenance) + discovery URL.
_AMC_DISCLOSURE_ROOTS: list[dict] = [
    # json_api_url_template: Drupal/CMS JSON API returning rows with ZIP download URLs.
    # {year} = 4-digit year.  zip_xlsx_member_pattern = filename substring to match inside the ZIP.
    {
        "name": "UTI",
        "json_api_url_template": "https://www.utimf.com/api/get-consolidate-portfolio-disclosure?year={year}",
        "zip_xlsx_member_pattern": "Sebi Exposure",
    },
    # direct_url_template: predictable static paths discovered via HTTP (no Playwright needed).
    # {month_full} = full month name (e.g. "May"), {year} = 4-digit year.
    # Remaining AMCs use Playwright discovery (requires chromium in container).
    # HDFC, ICICI_PRU, KOTAK: known bot-protection in place (Akamai/Radware); Playwright
    # attempts are kept in the schedule — future-proof for when protections change.
    {"name": "HDFC", "url": "https://www.hdfcfund.com/investor-relations/portfolio-disclosure"},
    {"name": "SBI", "url": "https://www.sbimf.com/portfolios"},
    {"name": "ICICI_PRU", "url": "https://www.icicipruamc.com/portfolio-disclosure"},
    # Nippon publishes .xls (legacy Excel 97-2004) via its download centre.
    {
        "name": "NIPPON",
        "url": "https://mf.nipponindiaim.com/investor-service/downloads/factsheet-portfolio-and-other-disclosures",
    },
    {"name": "KOTAK", "url": "https://www.kotakmf.com/portfolio-disclosure"},
    # Axis: correct path is /downloads/portfolio-disclosure (not /portfolio-disclosure which 404s).
    {"name": "AXIS", "url": "https://www.axismf.com/downloads/portfolio-disclosure"},
    # Mirae: static HTML page; one XLSX per scheme — discover all links via plain HTTP.
    {
        "name": "MIRAE",
        "url": "https://www.miraeassetmf.co.in/downloads/portfolio",
        "static_multi": True,
    },
    # PPFAS (added 2026-07-07, verified from KVM4): static page with direct
    # per-month .xls links ("Monthly-Portfolio-<month>-<year>.xls"); no WAF on
    # homepage or this page; ONE consolidated file per month — the month/year
    # filter in the static discovery picks the current one. Resolver override
    # "Parag Parikh%" required (scheme names don't carry the AMC brand).
    {
        "name": "PPFAS",
        "url": "https://amc.ppfas.com/downloads/portfolio-disclosure/",
        "static_multi": True,
    },
    # Franklin: Angular SPA; domain corrected from franklintempletonmutualfund.com (blocked/parked).
    {
        "name": "FRANKLIN",
        "url": "https://www.franklintempletonindia.com/investor/portfolio-disclosure",
    },
    # DSP: domain moved from dspmf.com (GoDaddy) to dspim.com; disclosure page is JS-rendered.
    {"name": "DSP", "url": "https://www.dspim.com/downloads"},
]


def parsed_to_snapshot_holdings(
    parsed: list[ParsedHolding],
    nav_map: dict[str, float] | None = None,
    category_map: dict[str, str] | None = None,
    invested_map: dict[tuple[str, str], float] | None = None,
) -> list[Holding]:
    """Map parsed CAS holdings → snapshot.Holding, applying the latest NAV when
    available (else falling back to the CAS-reported valuation). Pure + testable.

    ``category_map`` (isin → SEBI-canonical or raw AMFI category from mf_funds) fills
    each holding's category so the portfolio category-allocation is real; a holding
    whose ISIN is not in the master stays ``"uncategorized"`` (honest)."""
    nav_map = nav_map or {}
    category_map = category_map or {}
    invested_map = invested_map or {}
    out: list[Holding] = []
    for p in parsed:
        nav = nav_map.get(p.isin, p.nav)
        current_value = (p.units * nav) if (nav is not None) else (p.value or 0.0)
        # B86: one invested definition — the ledger net-invested written to the holdings table; the CAS
        # cost is the fallback only for holdings the ledger can't fully reconstruct (matches the table).
        # invested_map is keyed by the CANONICAL folio (B82 hardening) — normalize the lookup key too.
        invested = invested_map.get(
            (p.isin, normalize_folio(p.folio_number)), p.cost if p.cost is not None else 0.0
        )
        cashflows = [CashFlow(when=t.when, amount=t.amount) for t in p.txns if t.amount]
        if current_value:
            cashflows.append(CashFlow(when=p.as_of_date or date.today(), amount=current_value))
        out.append(
            Holding(
                isin=p.isin,
                units=p.units,
                invested_amount=invested,
                current_value=current_value,
                category=(category_map.get(p.isin) or "uncategorized"),
                cashflows=cashflows,
            )
        )
    return out


# ---------------------------------------------------------------------------
# Pure mapping helpers — unit-testable without DB
# ---------------------------------------------------------------------------


def _navrows_to_nav_upserts(rows: Any) -> list[dict]:
    """
    Map a list of NavRow → list of dicts ready for mf_nav_history upsert.

    Keying strategy: prefer isin_growth; fall back to isin_reinvest.
    Rows where BOTH ISINs are None are silently skipped (cannot key the row).
    Returns dicts with keys: isin, nav_date, nav, source.

    Deduplication: the AMFI feed sometimes emits duplicate (isin, nav_date) pairs
    within a single batch, which triggers asyncpg.CardinalityViolationError on the
    ON CONFLICT DO UPDATE.  Last-seen row wins (dict keyed by (isin, nav_date)).
    """
    out: dict[tuple[str, object], dict] = {}
    for row in rows:
        isin = row.isin_growth or row.isin_reinvest
        if isin is None:
            continue
        out[(isin, row.nav_date)] = {
            "isin": isin,
            "nav_date": row.nav_date,
            "nav": row.nav,
            "source": "amfi",
        }
    return list(out.values())


async def _batch_nav_upserts(
    rows: Any, chunk_size: int = _UPSERT_CHUNK
) -> AsyncIterator[list[dict]]:
    """
    Buffer an ASYNC stream of NavRow (e.g. ``amfi.stream_nav_history``) into
    upsert-dict batches of at most ``chunk_size``, applying
    ``_navrows_to_nav_upserts`` dedup PER BATCH.

    Memory-flat backfill helper: a 90-day all-funds window is ~1.1M NAV
    lines — materializing the whole window as NavRow objects, then a second
    full list of upsert dicts, doubles peak memory past the 640MiB
    celery-batch cap. Flushing at ``chunk_size`` keeps peak memory O(batch),
    not O(window). Dedup scoped to the batch (not the whole window) is
    sufficient: it only needs to prevent a duplicate (isin, nav_date) key
    inside one multi-row INSERT (asyncpg.CardinalityViolationError on
    ON CONFLICT DO UPDATE) — that only requires no duplicates WITHIN a
    chunk, not across the whole window.
    """
    buf: list = []
    async for row in rows:
        buf.append(row)
        if len(buf) >= chunk_size:
            batch = _navrows_to_nav_upserts(buf)
            if batch:
                yield batch
            buf = []
    if buf:
        batch = _navrows_to_nav_upserts(buf)
        if batch:
            yield batch


def _navrows_to_fund_upserts(rows: Any) -> list[dict]:
    """
    Map a list of NavRow → list of dicts ready for mf_funds upsert.

    Columns this feed owns: amfi_code, scheme_name, category, sebi_category,
    plan_type, option_type, fund_name_short, idcw_frequency, benchmark_index.
    isin is the PK (isin_growth preferred, else isin_reinvest).  Rows without
    a keyable ISIN are skipped.

    Deduplication: last-seen row wins for duplicate ISINs in one batch (dict keyed
    by isin), preventing ON CONFLICT DO UPDATE cardinality errors.

    benchmark_index (Block 0.7): populated ONLY for index funds with a
    high-confidence name match against a BENCHMARK_REGISTRY entry
    (map_index_fund_benchmark, gated on sebi_category — see mf/
    benchmark_mapping.py). The registry membership is re-checked here (not
    just trusted from the mapping function) so a future registry key rename
    can never silently write a dangling benchmark_index value. Every other
    fund (including all active/non-index funds) gets None — never a guessed
    peer index (architecture plan §19).
    """
    out: dict[str, dict] = {}
    for row in rows:
        isin = row.isin_growth or row.isin_reinvest
        if isin is None:
            continue
        plan_type, option_type = parse_plan_option(row.scheme_name)
        is_segregated = "segregated portfolio" in row.scheme_name.lower()
        sebi_category = canonical_for(row.category)
        benchmark_key = map_index_fund_benchmark(row.scheme_name, sebi_category)
        out[isin] = {
            "isin": isin,
            "amfi_code": row.amfi_code,
            "scheme_name": row.scheme_name,
            "category": row.category,
            "sebi_category": sebi_category,
            "plan_type": plan_type,
            "option_type": option_type,
            # Display-only clean name + IDCW cadence, derived from the same scheme
            # name (single source of truth: taxonomy). scheme_name stays official.
            "fund_name_short": derive_short_name(row.scheme_name, isin),
            "idcw_frequency": parse_idcw_frequency(row.scheme_name),
            "is_segregated": is_segregated,
            "launch_date": row.nav_date,
            "benchmark_index": benchmark_key if benchmark_key in BENCHMARK_REGISTRY else None,
        }
    return list(out.values())


# ---------------------------------------------------------------------------
# CAS pipeline helpers (unchanged)
# ---------------------------------------------------------------------------


@celery_app.task(name="dhanradar.tasks.mf.parse_cas_job", bind=True, max_retries=2)
def parse_cas_job(
    self,
    job_id: str,
    path: str,
    user_id: str,
    portfolio_id: str,
    request_id: str | None = None,
) -> str:
    """CAS→report worker. Always purges the raw file; marks the job failed with an
    OPAQUE code (never the raw exception, which could carry a path/PII)."""
    try:
        return asyncio.run(_run_pipeline(job_id, path, user_id, portfolio_id, request_id))
    except CasParseError as exc:
        # Log the underlying reason SERVER-SIDE for diagnosis — it is never sent
        # to the client. casparser raises short reason strings (incorrect
        # password / unsupported format / header parse error), not PII, and the
        # CAS password is never part of these messages. Without this, every CAS
        # failure is an undiagnosable opaque "parse_failed".
        logger.warning("CAS parse failed job=%s reason=%s", job_id, exc)
        # error_message is served to the client (CasJobStatus) — keep it OPAQUE: a fixed,
        # closed-enum code derived from the exception CLASS (classify_cas_failure), NEVER
        # exc / str(exc) verbatim. The FE maps the code to plain-language copy.
        code = classify_cas_failure(exc)
        asyncio.run(_mark_failed(job_id, code))
        return f"failed: {code}"
    except Exception:  # noqa: BLE001 — record opaque code + purge, never leak detail
        logger.exception("CAS pipeline error job=%s", job_id)
        asyncio.run(_mark_failed(job_id, "internal_error"))
        return "failed: internal_error"
    finally:
        _purge(path)


async def _fetch_fund_categories(db: Any, isins: list[str]) -> dict[str, str]:
    """Map isin → validated SEBI category (sebi_category preferred; falls back to raw
    AMFI category) from the mf_funds master (read-only). Fills holding categories so
    the portfolio category-allocation is real; an ISIN absent from the master (or with
    no category data) stays 'uncategorized' (honest)."""
    if not isins:
        return {}
    from sqlalchemy import select

    from dhanradar.models.mf import MfFund

    rows = (
        await db.execute(
            select(MfFund.isin, MfFund.sebi_category, MfFund.category).where(MfFund.isin.in_(isins))
        )
    ).all()
    return {i: (sc or c) for i, sc, c in rows if (sc or c)}


async def _store_or_validate_identity(user_id: str, identity: Any) -> str | None:
    """Store investor PAN + name on the user after a CAS upload; returns the user's AUTHORITATIVE
    `investor_pan` after this call (existing value if unchanged, freshly stored on a first upload,
    or the pre-existing value on a mismatch — never overwritten). The caller uses this return value
    for the per-folio ownership guard (family-merge protection, 2026-07-04): a folio-level PAN that
    disagrees with this authoritative PAN gets excluded from the upload entirely.

    First upload (investor_pan IS NULL): store both PAN and full_name.
    Subsequent uploads: validate PAN matches; log a warning if it doesn't.
    full_name is set only if still NULL (preserves any manually-set name).
    Uses a separate admin session — runs outside the RLS user session so it
    can read AND update auth.users without the MF-scope GUC interfering.
    Fails silently: a storage error must never break the CAS pipeline (returns None).
    """
    import uuid as _uuid

    from sqlalchemy import select, update

    from dhanradar.db import admin_task_session
    from dhanradar.models.auth import User

    try:
        async with admin_task_session() as db:
            uid = _uuid.UUID(user_id)
            row = (
                await db.execute(select(User.investor_pan, User.full_name).where(User.id == uid))
            ).one_or_none()
            if row is None:
                return None  # user not found — shouldn't happen, but safe fail

            existing_pan, existing_name = row

            if not identity.pan and not identity.investor_name:
                return existing_pan  # nothing new to store; existing PAN (if any) still governs

            if existing_pan and identity.pan and existing_pan != identity.pan:
                # PAN mismatch — log prominently; do NOT overwrite (the stored PAN
                # is the authoritative one from the user's own first upload).
                _slog.warning(
                    "cas.identity.pan_mismatch",
                    user_ref=hash_user_ref(user_id),
                    stored_pan_prefix=existing_pan[:5],  # first 5 chars only (DPDP log discipline)
                    new_pan_prefix=identity.pan[:5],
                )
                return (
                    existing_pan  # don't update anything on a mismatch; existing PAN still governs
                )

            update_vals: dict[str, Any] = {}
            if not existing_pan and identity.pan:
                update_vals["investor_pan"] = identity.pan
            if not existing_name and identity.investor_name:
                update_vals["full_name"] = identity.investor_name

            if update_vals:
                await db.execute(update(User).where(User.id == uid).values(**update_vals))
                await db.commit()
                _slog.info(
                    "cas.identity.stored",
                    user_ref=hash_user_ref(user_id),
                    fields=list(update_vals.keys()),
                )
            return update_vals.get("investor_pan", existing_pan)
    except Exception:  # noqa: BLE001 — identity storage must never break the CAS pipeline
        logger.exception("cas.identity: storage failed for user_id=%s (non-fatal)", user_id)
        return None


#: Cap the replay window at ~10 years — mirrors portfolio_read._MAX_VALUATION_DAYS (kept as a
#: separate local constant; tasks/mf.py does not import from the read-model module).
_MAX_REPLAY_DAYS = 3650


async def _reset_valuation_series(db: Any, user_id: str, portfolio_id: str) -> None:
    """Rebuild the daily-valuation series after a CAS (re-)upload (§39.5 — retires the old
    delete-and-reseed-today patch; a composition change is just more ledger rows, and the FULL
    replay stays continuous and TRUE across it — no more "series restarts" on every upload).

    Ledger non-empty (S2/S8/S11 — detailed CAS, superset/subset re-upload, full composition
    change): delete the stored rows, then REPLAY the whole ledger-covered window (earliest
    txn_date → today, capped at _MAX_REPLAY_DAYS) against a single batched NAV query. Ledger
    EMPTY for this portfolio (S3 — CDSL/summary CAS, no transaction section at all): keep the
    existing forward-only fallback — seed exactly today's row from the just-written holdings +
    latest NAVs (the 04:00 task idempotently re-upserts today's row later either way).

    Runs in the pipeline's owner-RLS session — mf_portfolio_daily_values carries the owner's
    user_id (0056 RLS).
    """
    from sqlalchemy import delete as sa_delete
    from sqlalchemy import select as sa_select
    from sqlalchemy import text as sa_text

    from dhanradar.mf.valuation import compute_daily_value, replay_valuation_series
    from dhanradar.models.mf import MfPortfolioDailyValue, MfPortfolioTransaction, MfUserHolding

    # S20: same per-portfolio lock the ledger-append transaction takes — this function now does a
    # real delete + bulk rebuild of mf_portfolio_daily_values, the SAME table the nightly
    # compute_portfolio_daily_valuations task writes, so it needs the same serialization.
    await db.execute(sa_text("SELECT pg_advisory_xact_lock(hashtext(:pid))"), {"pid": portfolio_id})

    await db.execute(
        sa_delete(MfPortfolioDailyValue).where(MfPortfolioDailyValue.portfolio_id == portfolio_id)
    )

    ledger_rows = (
        (
            await db.execute(
                sa_select(
                    MfPortfolioTransaction.instrument_id,
                    MfPortfolioTransaction.units,
                    MfPortfolioTransaction.amount,
                    MfPortfolioTransaction.txn_type,
                    MfPortfolioTransaction.txn_date,
                    # nav_or_price feeds the replay's synthetic price seeding (the +212% RCA fix:
                    # units are valued at their txn price until real NAV coverage begins).
                    MfPortfolioTransaction.nav_or_price,
                ).where(MfPortfolioTransaction.portfolio_id == portfolio_id)
            )
        )
        .mappings()
        .all()
    )

    if ledger_rows:
        today = date.today()
        earliest = min(r["txn_date"] for r in ledger_rows)
        from_date = max(earliest, today - timedelta(days=_MAX_REPLAY_DAYS))
        isins = sorted({r["instrument_id"] for r in ledger_rows})

        # ONE batched query for every ISIN x the whole window (never per-day, never per-ISIN).
        nav_rows = await db.execute(
            sa_text(
                "SELECT isin, nav_date, nav FROM mf.mf_nav_history"
                " WHERE isin = ANY(:isins) AND nav_date >= :from_date AND nav_date <= :to_date"
                " ORDER BY isin, nav_date"
            ),
            # Fetch NAVs from 7 days BEFORE the replay window so carry-forward is seeded at
            # from_date (the NAV in effect on day 1 was usually published a day or two earlier —
            # weekend/holiday). replay_valuation_series applies any nav_date <= d correctly.
            {"isins": isins, "from_date": from_date - timedelta(days=7), "to_date": today},
        )
        nav_by_isin: dict[str, list[tuple[date, float]]] = {}
        for r in nav_rows:
            nav_by_isin.setdefault(r.isin, []).append((r.nav_date, float(r.nav)))

        points = replay_valuation_series(ledger_rows, nav_by_isin, from_date, today)
        if points:
            # UUID-typed ids are REQUIRED here: add_all uses SQLAlchemy insertmanyvalues, whose
            # sentinel matching needs the Python value to equal the driver's returned type
            # exactly — a str portfolio_id raises "Can't match sentinel values in result set".
            from uuid import UUID as _UUID

            pid_u, uid_u = _UUID(portfolio_id), _UUID(user_id)
            db.add_all(
                [
                    MfPortfolioDailyValue(
                        portfolio_id=pid_u,
                        user_id=uid_u,
                        valuation_date=p.valuation_date,
                        total_value=p.total_value,
                        total_invested=p.total_invested,
                    )
                    for p in points
                ]
            )
    else:
        # No ledger for this portfolio (S3 — txn-less CDSL/summary CAS): forward-only fallback,
        # unchanged from the pre-replay behaviour — seed exactly today's row.
        holdings = (
            await db.execute(
                sa_select(
                    MfUserHolding.isin, MfUserHolding.units, MfUserHolding.invested_amount
                ).where(MfUserHolding.portfolio_id == portfolio_id)
            )
        ).all()
        if holdings:
            isins = [h.isin for h in holdings]
            nav_rows = await db.execute(
                sa_text(
                    "SELECT DISTINCT ON (isin) isin, nav FROM mf.mf_nav_history"
                    " WHERE isin = ANY(:isins) ORDER BY isin, nav_date DESC"
                ),
                {"isins": isins},
            )
            nav_map = {r.isin: float(r.nav) for r in nav_rows}
            point = compute_daily_value(
                [(float(h.units or 0), nav_map.get(h.isin, 0.0)) for h in holdings],
                sum(float(h.invested_amount or 0) for h in holdings),
                date.today(),
            )
            db.add(
                MfPortfolioDailyValue(
                    portfolio_id=portfolio_id,
                    user_id=user_id,
                    valuation_date=point.valuation_date,
                    total_value=point.total_value,
                    total_invested=point.total_invested,
                )
            )
    await db.commit()


async def _run_pipeline(
    job_id: str,
    path: str,
    user_id: str,
    portfolio_id: str,
    request_id: str | None = None,
) -> str:
    from sqlalchemy import select, text, update

    # B81 PR-2: CAS writes the uploader's personal tables (cas_job/holdings/sip/scores/history/
    # snapshot), so it runs RLS WITH-CHECK enforced AS THE OWNER, not on the bypass admin engine.
    # rls_user_session re-applies the app.user_id GUC on every transaction begin, so the owner scope
    # survives the pipeline's 3+ progress commits AND the per-fund commit inside append_score_history
    # (a plain SET LOCAL would be cleared by each). user_id is the authenticated uploader (task arg),
    # never from the parsed file.
    from dhanradar.db_security import rls_user_session
    from dhanradar.models.mf import MfCasJob
    from dhanradar.redis_client import get_redis
    from dhanradar.scoring.engine import RatingEngine
    from dhanradar.scoring.engine.schemas import DISCLAIMER_VERSION

    # Bind correlation context for all structured log lines in this pipeline run.
    bind_contextvars(job_id=job_id, request_id=request_id, user_ref=hash_user_ref(user_id))
    _slog.info("cas_pipeline_start", job_id=job_id)

    redis = get_redis()

    # Consume-and-delete the CAS password from its short-lived Redis key (it was
    # never put on the broker). Missing → None (an unprotected CAS still parses).
    pw_key = f"mf:cas:pw:{job_id}"
    password = await redis.get(pw_key)
    if password is not None:
        await redis.delete(pw_key)

    parsed, identity = detect_and_parse(path, password)  # routes to CAS PDF or CAMS TDS parser

    # Store investor identity on the user (first upload sets it; subsequent uploads
    # validate against it so a wrong-investor CAS is caught early). Returns the user's
    # AUTHORITATIVE PAN (post-store) for the per-folio ownership guard below.
    user_pan = await _store_or_validate_identity(user_id, identity)

    # Per-folio ownership guard (family-merge incident 2026-07-04): a consolidated statement can
    # carry a DIFFERENT investor's folios (e.g. a household member sharing one RTA email) — a
    # folio whose OWN PAN is present and disagrees with the uploader's authoritative PAN is
    # excluded ENTIRELY, before it can produce a single ledger row, holding, or checkpoint. A folio
    # with no PAN info (CAMS TDS predates this on some exports; CDSL account-level rows) is assumed
    # the owner's — status quo. No stored PAN yet (a first-ever upload whose CAS itself carried no
    # identity at all) → nothing to compare against, so nothing is excluded.
    owned_holdings, foreign_holdings = filter_foreign_pan_folios(parsed, user_pan)
    excluded_folios = len(foreign_holdings)
    if foreign_holdings:
        _slog.warning(
            "cas.ownership.folio_excluded",
            excluded_count=excluded_folios,
            folios=[normalize_folio(p.folio_number) for p in foreign_holdings],
        )
    parsed = owned_holdings

    # Resolve CAMS TDS placeholders (isin="CAMS:<product_code>") to real ISINs via the
    # B94-hardened resolver (AMC token gate + similarity floor/ambiguity margin +
    # txn-price vs NAV-history validation — see resolve_cams_isins). Unresolved holdings
    # keep the placeholder ISIN and score as insufficient_data (honest fail-safe).
    cams_placeholders = [p for p in parsed if p.isin.startswith("CAMS:")]
    if cams_placeholders:
        from dataclasses import replace as _dc_replace

        from dhanradar.db import task_session
        from dhanradar.mf.cas import resolve_cams_isins

        async with task_session() as _db:
            resolved_map = await resolve_cams_isins(_db, cams_placeholders)
        if resolved_map:
            parsed = [
                _dc_replace(p, isin=resolved_map[p.isin]) if p.isin in resolved_map else p
                for p in parsed
            ]

    # Fix 2 (2026-07-04 plan-variant double-count): rewrite any holding parsed under an AMFI
    # secondary/reinvest-plan ISIN to its primary mf_funds.isin BEFORE ledger fingerprinting, so
    # the same real position printed by two formats (e.g. a CAMS-resolved primary ISIN + a
    # consolidated PDF printing the plan's other ISIN) collides instead of becoming two holdings.
    from dhanradar.db import task_session as _alias_task_session
    from dhanradar.mf.cas import alias_secondary_isins

    async with _alias_task_session() as _db:
        alias_map = await alias_secondary_isins(_db, parsed)
    if alias_map:
        from dataclasses import replace as _dc_replace

        for secondary, primary in alias_map.items():
            _slog.info("cas.isin.aliased", from_isin=secondary, to_isin=primary)
        parsed = [
            _dc_replace(p, isin=alias_map[p.isin]) if p.isin in alias_map else p for p in parsed
        ]

    rengine = RatingEngine()

    async with rls_user_session(user_id) as db:
        # The after_begin GUC re-apply scopes every transaction below to the uploader, so the first
        # MfCasJob UPDATE matches the row (without it the policy denies it and the UPDATE no-ops) and
        # every holdings/sip/score/snapshot write is RLS WITH-CHECK enforced for this owner.
        await db.execute(
            update(MfCasJob)
            .where(MfCasJob.job_id == job_id)
            .values(
                status="parsing",
                progress_pct=40,
                stmt_from=identity.stmt_from,
                stmt_to=identity.stmt_to,
                excluded_folios=excluded_folios,
            )
        )
        await db.commit()

        # B2/B3: append the CAS transactions to the append-only ledger (the spine) FIRST, then project
        # holdings FROM the ledger. Idempotent diff-and-append (ON CONFLICT DO NOTHING). Runs under
        # rls_user_session (after_begin re-applies the owner GUC), so RLS WITH CHECK enforces the
        # uploader as owner. HARD-FAIL (B3): holdings now DERIVE from the ledger, so a dropped ledger
        # write would mean wrong holdings — let the failure propagate and fail the job (the outer
        # handler marks it failed + purges) rather than completing with stale holdings.
        from dhanradar.mf.cas import build_cas_ledger_rows
        from dhanradar.mf.ledger import append_transactions

        # S20 (§39.4): serialize every writer of THIS portfolio (double-click upload, upload racing
        # the 04:00 valuation task, a future sync) — first statement of the write transaction so a
        # concurrent racer blocks here rather than corrupting the append/project step. xact-scoped:
        # released at this transaction's commit/rollback (NullPool hands the next transaction a fresh
        # connection anyway, mirroring why rls_user_session re-applies its GUC per-transaction too) —
        # _project_and_write_holdings and the nightly compute_portfolio_daily_valuations task each
        # re-acquire the SAME lock at the start of their own write transaction.
        await db.execute(
            text("SELECT pg_advisory_xact_lock(hashtext(:pid))"), {"pid": portfolio_id}
        )

        # Fix 3 (2026-07-04 blank-folio double-count): resolve any blank-folio holding (a
        # holdings-only valuation file, or a demat/CDSL section within a consolidated CAS) against
        # real folios for the SAME isin — from THIS batch and the portfolio's already-stored
        # holdings — before it can duplicate a position that already has a folio.
        existing_folios_by_isin: dict[str, set[str]] = {}
        for p in parsed:
            fn = normalize_folio(p.folio_number)
            if fn:
                existing_folios_by_isin.setdefault(p.isin, set()).add(fn)
        blank_isins = {p.isin for p in parsed if not normalize_folio(p.folio_number)}
        if blank_isins:
            from dhanradar.models.mf import MfUserHolding

            stored_folios = (
                await db.execute(
                    select(MfUserHolding.isin, MfUserHolding.folio_number).where(
                        MfUserHolding.portfolio_id == portfolio_id,
                        MfUserHolding.isin.in_(blank_isins),
                    )
                )
            ).all()
            for isin, folio in stored_folios:
                if folio:
                    existing_folios_by_isin.setdefault(isin, set()).add(folio)

        parsed, blank_folio_skipped = resolve_blank_folios(parsed, existing_folios_by_isin)
        if blank_folio_skipped:
            _slog.warning(
                "cas.blank_folio.ambiguous",
                count=len(blank_folio_skipped),
                isins=sorted({p.isin for p in blank_folio_skipped}),
            )

        # Placeholder-restatement suppression (2026-07-04 hand-cleaned incident: 3 active + 5 closed
        # dupes on prod): a holdings-only file can print an ABBREVIATED fund name the resolver
        # correctly refuses -- the SAME position then shows up twice, once resolved and once under
        # the unresolvable "CAMS:<code>" placeholder, both carrying the identical canonical folio +
        # units. Suppress the placeholder twin when a resolved holding with that (folio, units) pair
        # exists in THIS batch or the portfolio's already-stored holdings. Only queries the DB when
        # this batch actually carries a placeholder (cheap no-op on every ordinary upload).
        if any(p.isin.startswith("CAMS:") for p in parsed):
            resolved_positions = {
                (normalize_folio(p.folio_number), p.units)
                for p in parsed
                if not p.isin.startswith("CAMS:")
            }
            from dhanradar.models.mf import MfUserHolding

            stored_rows = (
                await db.execute(
                    select(MfUserHolding.folio_number, MfUserHolding.units).where(
                        MfUserHolding.portfolio_id == portfolio_id,
                        ~MfUserHolding.isin.like("CAMS:%"),
                    )
                )
            ).all()
            resolved_positions |= {
                (normalize_folio(folio), float(units)) for folio, units in stored_rows
            }
            parsed, restatements_suppressed = suppress_placeholder_restatements(
                parsed, resolved_positions
            )
            if restatements_suppressed:
                _slog.info("cas.placeholder.restatement_suppressed", count=restatements_suppressed)

        # Fix 1 (2026-07-04 placeholder-ISIN ledger leak): a holding still keyed on an unresolved
        # CAMS:<code> placeholder must never produce a ledger row — it still gets a holdings row
        # (B3 snapshot-fallback below) + a statement checkpoint, both keyed off (isin, folio).
        ledger_input, placeholder_holdings = split_ledger_eligible(parsed)
        if placeholder_holdings:
            _slog.info("cas.placeholder.ledger_barred", count=len(placeholder_holdings))

        # Format-specific parser_version: the ledger's natural-key dedup scopes itself to
        # CROSS-format matches only (same-format rows can be legitimate same-day twins).
        ledger_rows = build_cas_ledger_rows(
            ledger_input,
            user_id=user_id,
            portfolio_id=portfolio_id,
            parser_version=parser_version_for(path),
        )
        ledger_inserted, ledger_skipped = await append_transactions(db, ledger_rows)
        await db.commit()
        _slog.info(
            "mf.ledger.appended",
            job_id=job_id,
            inserted=ledger_inserted,
            skipped=ledger_skipped,
            total=len(ledger_rows),
        )

        # B3: holdings are now a PROJECTION of the ledger (units + net-invested), not a direct copy of
        # the parsed file. B86: capture the net-invested map so the fresh report/snapshot use the SAME
        # invested as the holdings table (one invested definition everywhere).
        invested_map, projected = await _project_and_write_holdings(
            db, user_id, parsed, portfolio_id
        )

        # §39.4 — persist this upload's statement-checkpoint evidence (stated vs ledger units).
        # Never mutates the ledger (I12); a mismatch is flagged + logged, not corrected.
        await _write_statement_checkpoints(db, user_id, portfolio_id, job_id, parsed, projected)

        # Persist SIP transactions so get_sip_day() can infer the user's SIP date
        from dhanradar.models.mf import MfSipTransaction

        sip_rows = [
            MfSipTransaction(
                portfolio_id=portfolio_id,
                user_id=user_id,
                txn_date=h_txn.when,
                amount=abs(h_txn.amount),
            )
            for h in parsed
            for h_txn in h.txns
            if h_txn.is_sip
        ]
        if sip_rows:
            # replace all rows for this portfolio so re-uploads stay idempotent
            from sqlalchemy import delete

            from dhanradar.models.mf import MfSipTransaction as _Sip

            await db.execute(delete(_Sip).where(_Sip.portfolio_id == portfolio_id))
            db.add_all(sip_rows)
            await db.flush()

        await db.execute(
            update(MfCasJob)
            .where(MfCasJob.job_id == job_id)
            .values(status="scoring", progress_pct=70)
        )
        await db.commit()

        # The (re-)upload may have replaced the holdings composition — restart the daily
        # valuation series so day-change/charts never span two different portfolios
        # (RCA 2026-07-02). Non-fatal: a reset failure must never break the upload.
        try:
            await _reset_valuation_series(db, user_id, portfolio_id)
        except Exception:  # noqa: BLE001
            _slog.warning("mf.valuation_series.reset_failed", job_id=job_id, exc_info=True)

        # Resolve Plus status once — controls whether history rows are written.
        from dhanradar.deps import is_plus
        from dhanradar.mf import history as mf_history

        plus = await is_plus(user_id, db)

        # Run NAV history, peer-cohort, and category reads concurrently — all three
        # are independent reads on public MF data (no user-personal data, no RLS).
        # Each opens its own task_session so asyncio.gather can interleave them safely.
        isins_to_score = [p.isin for p in parsed]
        from dhanradar.db import task_session as _task_session

        async def _par_nav() -> tuple[dict, dict]:
            async with _task_session() as _db:
                return await _load_nav_series(_db, isins_to_score)

        async def _par_cohort() -> dict:
            async with _task_session() as _db:
                return await _compute_cohort(_db, isins_to_score)

        async def _par_cats() -> dict:
            async with _task_session() as _db:
                return await _fetch_fund_categories(_db, isins_to_score)

        (nav_series, latest_nav), cohort, category_map = await asyncio.gather(
            _par_nav(), _par_cohort(), _par_cats()
        )

        snapshot_holdings = parsed_to_snapshot_holdings(
            parsed, nav_map=latest_nav, category_map=category_map, invested_map=invested_map
        )
        snap = build_snapshot(snapshot_holdings)
        # B65 observability: an uncomputable XIRR must be visible in worker logs,
        # not discovered by eyeball in a report. Log the boolean only — a user's
        # XIRR value is personal financial data (DPDP log discipline).
        _slog.info(
            "mf.snapshot.built",
            funds=len(parsed),
            cashflows=sum(len(h.cashflows) for h in snapshot_holdings),
            xirr_computed=snap.xirr_pct is not None,
        )

        from dhanradar.compliance import service as compliance_service

        funds_payload: list[dict] = []
        today = date.today()
        # Fetch prior labels for all funds in one query before the loop writes new ones.
        # get_prior_label filters snapshot_date < today, so we can read all at once.
        prior_labels: dict[str, str | None] = {
            p.isin: await mf_history.get_prior_label(db, portfolio_id, p.isin, today)
            for p in parsed
        }
        for p in parsed:
            # Signals are computed from the fund's own NAV series (momentum/risk);
            # fundamentals-backed axes stay None → partial_coverage (≤ medium).
            # category_relative carries the peer-cohort comparison (B58).
            signals = compute_fund_signals(
                p.isin, nav_series.get(p.isin, []), category_relative=cohort.get(p.isin)
            )
            result = await score_fund(rengine, signals)
            # commit=False: accumulate all fund score + history writes; one commit at end.
            await upsert_user_fund_score(db, user_id, result, portfolio_id, commit=False)
            # Write label history for ALL users (not just Plus) so the delta feature
            # (Feature 3: ↑/↓ arrow on the report) works for free users too.
            # The full history READ endpoint stays Plus-gated (router.py).
            await mf_history.append_score_history(
                db,
                user_id=user_id,
                result=result,
                snapshot_date=today,
                source="cas_upload",
                portfolio_id=portfolio_id,
                commit=False,
            )
            # B26 — persist (label, model_used, disclaimer_version) for this served
            # label at GENERATION (once, with full provenance). Fire-and-forget: a
            # failure is logged and never breaks the report pipeline.
            await compliance_service.record_served_label(
                surface="mf_report",
                label=result.verb_label.value,
                model=rengine.model_version,
                disclaimer_version=result.disclaimer_version,
                user_id=user_id,
                identifier=p.isin,
                confidence_band=result.confidence_band.value,
                request_id=request_id,
            )
            funds_payload.append(
                {
                    "isin": p.isin,
                    "scheme_name": p.scheme_name,
                    "folio_number": p.folio_number,
                    # Display-only clean name from the CAS scheme name (same derivation
                    # as the master). scheme_name (official) is still carried + shown.
                    "fund_name_short": derive_short_name(p.scheme_name, p.isin),
                    "idcw_frequency": parse_idcw_frequency(p.scheme_name),
                    "category": category_map.get(p.isin),
                    "units": p.units,
                    "invested_amount": invested_map.get(
                        (p.isin, normalize_folio(p.folio_number)), p.cost
                    ),
                    "current_value": p.value,
                    "verb_label": result.verb_label.value,
                    "confidence_band": result.confidence_band.value,
                    "contributing_signals": result.contributing_signals,
                    "contradicting_signals": result.contradicting_signals,
                    "previous_label": prior_labels.get(p.isin),
                    "confidence_factors": dict(result.confidence_factors),
                }
            )
        # Commit all fund scores + history rows in one transaction.
        await db.commit()

        # Plus-only: persist the portfolio-level snapshot (numbers stay server-side).
        if plus:
            await mf_history.persist_portfolio_snapshot(
                db,
                user_id=user_id,
                snapshot_date=date.today(),
                snap=snap,
                portfolio_id=portfolio_id,
            )

        report_payload = {
            "job_id": job_id,
            "status": "done",
            "snapshot": {
                "total_invested": snap.total_invested,
                "current_value": snap.current_value,
                "xirr_pct": snap.xirr_pct,
                "category_allocation": snap.category_allocation,
                "overlap_matrix": snap.overlap_matrix,
            },
            "funds": funds_payload,
            "model_version": rengine.model_version,
            "generated_at": datetime.now(UTC).isoformat(),
            # Stamp the in-force disclaimer version on the served + cached report so
            # it matches the audit rows written above (B26 tie-to-version).
            "disclaimer_version": DISCLAIMER_VERSION,
        }
        # First AI consumer — governed portfolio commentary (B20/B21/B22 gated).
        # Best-effort: a commentary failure NEVER breaks the report (mirrors the
        # fire-and-forget audit pattern above).
        try:
            from dhanradar.ai_gateway.gateway import OpenRouterGateway
            from dhanradar.mf.commentary import generate_commentary, is_commentary_entitled

            if await is_commentary_entitled(user_id, db):
                # Cap AI commentary at 12 s — prevents a slow model from blocking the whole pipeline.
                report_payload["commentary"] = await asyncio.wait_for(
                    generate_commentary(
                        OpenRouterGateway(),
                        user_id=user_id,
                        db=db,
                        snapshot=snap,
                        funds=funds_payload,
                        request_id=request_id,
                    ),
                    timeout=12.0,
                )
            else:
                report_payload["commentary"] = {
                    "state": "upgrade_required",
                    "reason": "plus_feature",
                }
        except TimeoutError:
            logger.warning("AI commentary timed out after 12 s job=%s", job_id)
            report_payload["commentary"] = {"state": "unavailable", "reason": "timeout"}
        except Exception:  # noqa: BLE001 — commentary is best-effort; report still completes
            logger.exception("AI commentary failed job=%s", job_id)
            report_payload["commentary"] = {"state": "unavailable", "reason": "internal_error"}

        await redis.set(
            f"{service._REPORT_PREFIX}{job_id}", json.dumps(report_payload), ex=service._REPORT_TTL
        )
        await db.execute(
            update(MfCasJob)
            .where(MfCasJob.job_id == job_id)
            .values(status="done", progress_pct=100, completed_at=datetime.now(UTC))
        )
        # Stamp latest_job_id on the portfolio so GET /mf/portfolio/latest works
        # and the daily refresh task knows which job to rebuild. Portfolio lifecycle fix.
        import uuid as _uuid

        from dhanradar.models.mf import MfPortfolio as _MfPortfolio

        await db.execute(
            update(_MfPortfolio)
            .where(_MfPortfolio.id == _uuid.UUID(portfolio_id))
            .values(latest_job_id=_uuid.UUID(job_id))
        )
        await db.commit()
    return f"done: {len(parsed)} schemes"


async def _project_and_write_holdings(
    db: Any, user_id: str, parsed: list[ParsedHolding], portfolio_id: str
) -> tuple[dict[tuple[str, str], float], dict[tuple[str, str], dict[str, Any]]]:
    """B3 cutover: holdings are a PROJECTION of the ledger, not a direct copy of the parsed file.

    Read the portfolio's full ledger (just appended), project current holdings (units = Σ unit deltas;
    invested = Σ net capital invested, plan §13), and UPSERT each parsed holding using the projected
    values WHERE the ledger has txns for it. A holding without ledger txns (txn-less CDSL / summary CAS,
    or a KFin holdings-only source with no transaction section — S3) stays snapshot-derived: its
    `invested_amount` is the STATED basis from the statement itself (`ParsedHolding.cost`, when the
    source printed one) rather than NULL, so `portfolio.summary`'s invested/cost_value totals cover the
    WHOLE portfolio, not just its ledger-backed holdings (Fix 2a, 2026-07-04). A ledger-backed holding
    keeps the ledger's net-CASH basis (Σ purchases − Σ redemptions) instead — both are the user's own
    statement facts, just from different evidence; neither is invented. UPSERT, never truncate, so a
    partial/older CAS can't erase funds from a prior upload (§22).

    Runs under the caller's rls_user_session → RLS scopes the ledger read and WITH-CHECK scopes the
    holding write to the uploader (user_id is the uploader, never the file). avg_cost_nav stays the
    CAS-reported per-holding NAV (market context, not a ledger fact). The holding folio is written
    CANONICAL (normalize_folio, B82 hardened 2026-07-04) — matching the ledger's own normalized
    folio — so the SAME real folio printed with different spacing/case across two source formats
    always collides to ONE `uq_mf_holding` row and ONE key for XIRR/day-change joins downstream,
    closing the double-map hazard a raw-folio write used to leave open.

    SAFETY: if a holding HAS ledger txns but Σ units != the AMC close balance (an un-captured txn type
    or a deduped identical txn), the AMC close is authoritative → fall back to the parsed snapshot and
    warn, rather than ship wrong units. So the cutover can never regress units below today's behaviour;
    it only upgrades to ledger-derived values where the ledger fully reconstructs the position.

    Returns (invested_map, projected) — both keyed by (isin, CANONICAL folio); `projected` is exposed
    so the caller can persist statement checkpoints (§39.4) without re-querying the ledger again."""
    import uuid as _uuid
    from decimal import Decimal

    from sqlalchemy import func, select, text
    from sqlalchemy.dialects.postgresql import insert

    from dhanradar.mf.projection import (
        ENGINE_VERSION,
        UNITS_GAP_TOLERANCE,
        project_holdings_from_ledger,
    )
    from dhanradar.models.mf import MfPortfolioTransaction, MfUserHolding

    # S20 (§39.4): re-acquire the per-portfolio lock — this is a NEW transaction (NullPool hands out a
    # fresh connection per commit), so the lock taken before the ledger append does not carry over here.
    await db.execute(text("SELECT pg_advisory_xact_lock(hashtext(:pid))"), {"pid": portfolio_id})

    pid = _uuid.UUID(portfolio_id)
    ledger_rows = (
        (
            await db.execute(
                select(
                    MfPortfolioTransaction.instrument_id,
                    MfPortfolioTransaction.folio_number,
                    MfPortfolioTransaction.units,
                    MfPortfolioTransaction.amount,
                    MfPortfolioTransaction.txn_type,
                    MfPortfolioTransaction.txn_date,
                ).where(MfPortfolioTransaction.portfolio_id == pid)
            )
        )
        .mappings()
        .all()
    )
    projected = project_holdings_from_ledger(ledger_rows)

    # B86: the FINAL invested written per (isin, CANONICAL folio) — net-invested where the ledger
    # fully reconstructs the holding, else the AMC-cost fallback — returned so the fresh
    # report/snapshot use the SAME invested as the holdings table (one invested definition everywhere).
    invested_map: dict[tuple[str, str], float] = {}
    projected_n = 0
    for p in parsed:
        folio_norm = normalize_folio(p.folio_number)
        proj = projected.get((p.isin, folio_norm))
        if proj is not None and abs(proj["units"] - Decimal(str(p.units))) <= UNITS_GAP_TOLERANCE:
            # Ledger COMPLETE for this holding (Σ units == the AMC close balance) → use the projection.
            units, invested, as_of = proj["units"], proj["invested_amount"], proj["as_of"]
            projected_n += 1
        else:
            # No ledger txns (txn-less CDSL/summary) OR the ledger is INCOMPLETE (Σ units != the AMC
            # close — an un-captured txn type / a deduped identical txn). The AMC close is authoritative,
            # so fall back to the parsed snapshot rather than ship wrong units; warn on a partial gap so
            # it drives a cas.py txn-mapping extension (the B3 gap detector).
            if proj is not None:
                _slog.warning(
                    "mf.holdings.units_gap",
                    isin=p.isin,
                    folio=folio_norm,
                    projected=str(proj["units"]),
                    cas_close=p.units,
                )
            units, invested, as_of = p.units, p.cost, p.as_of_date
        stmt = (
            insert(MfUserHolding)
            .values(
                user_id=user_id,
                portfolio_id=portfolio_id,
                isin=p.isin,
                folio_number=folio_norm,
                units=units,
                avg_cost_nav=p.nav,
                invested_amount=invested,
                source="cas",
                as_of_date=as_of,
            )
            .on_conflict_do_update(
                constraint="uq_mf_holding",
                set_={
                    "units": units,
                    "invested_amount": invested,
                    "source": "cas",
                    "as_of_date": as_of,
                    "updated_at": func.now(),
                },
            )
        )
        await db.execute(stmt)
        # float for the report/snapshot; 0.0 for a None-cost holding (every reader coerces
        # invested_amount or 0 → 0.0, so the report stays consistent with the table). Never float(None).
        invested_map[(p.isin, folio_norm)] = float(invested) if invested is not None else 0.0
    _slog.info(
        "mf.holdings.projected", projected=projected_n, total=len(parsed), engine=ENGINE_VERSION
    )
    await db.commit()
    return invested_map, projected


async def _write_statement_checkpoints(
    db: Any,
    user_id: str,
    portfolio_id: str,
    upload_ref: str,
    parsed: list[ParsedHolding],
    projected: dict[tuple[str, str], dict[str, Any]],
) -> None:
    """§39.4 — persist ONE checkpoint row per parsed holding: the statement's stated units/cost vs the
    ledger's replayed units for that (instrument, folio). This is evidence, never a ledger mutation
    (I12) — a disagreement (S10) is flagged `reconciliation_status='mismatch'` and logged, nothing is
    corrected or overwritten. Also the only data source in HOLDINGS_ONLY state (S3: a txn-less CDSL /
    summary CAS has no `projected` entry at all, so every checkpoint there is trivially 'ok' — there is
    nothing in the ledger to disagree with). `folio_number` is written CANONICAL (normalize_folio,
    B82 hardened 2026-07-04) so this table's own folio key agrees with the ledger and the holdings row.

    Also the cheap post-append reconciliation tripwire: when the ledger's replayed units EXCEED what
    THIS statement states for a folio it covers (the opposite direction from a generic mismatch — the
    ledger has MORE than the RTA says it should), that specifically smells like duplicate/double-counted
    rows (the 2026-07-04 incident's symptom) rather than an under-captured txn type, so it gets its own
    `ledger.inflation_suspected` warning alongside the ordinary mismatch flag.

    Runs on the caller's rls_user_session db (same owner-scoped write as the holdings upsert); commits
    are the caller's responsibility (mirrors the sip_rows add_all pattern above)."""
    from decimal import Decimal

    from dhanradar.mf.projection import UNITS_GAP_TOLERANCE
    from dhanradar.models.mf import MfPortfolioStatementCheckpoint

    rows: list[MfPortfolioStatementCheckpoint] = []
    for p in parsed:
        folio_norm = normalize_folio(p.folio_number)
        proj = projected.get((p.isin, folio_norm))
        status = "ok"
        if proj is not None:
            gap = proj["units"] - Decimal(str(p.units))
            if abs(gap) > UNITS_GAP_TOLERANCE:
                status = "mismatch"
                _slog.warning(
                    "mf.checkpoint.reconciliation_mismatch",
                    isin=p.isin,
                    folio=folio_norm,
                    stated_units=p.units,
                    ledger_units=str(proj["units"]),
                )
            if gap > UNITS_GAP_TOLERANCE:
                # Ledger units > what THIS statement states for a folio it covers — possible
                # duplicate/double-counted rows (cross-format dedup miss), not a units gap.
                _slog.warning(
                    "ledger.inflation_suspected",
                    isin=p.isin,
                    folio=folio_norm,
                    ledger_units=str(proj["units"]),
                    stated_units=p.units,
                )
        rows.append(
            MfPortfolioStatementCheckpoint(
                user_id=user_id,
                portfolio_id=portfolio_id,
                upload_ref=upload_ref,
                instrument_id=p.isin,
                folio_number=folio_norm,
                stated_units=p.units,
                stated_cost=p.cost,
                stmt_date=p.as_of_date,
                reconciliation_status=status,
            )
        )
    if rows:
        db.add_all(rows)
        await db.flush()


async def _load_nav_series(
    db: Any, isins: list[str], lookback_days: int = 400
) -> tuple[dict[str, list[tuple[date, float]]], dict[str, float]]:
    """Load recent NAV history for the given ISINs from mf_nav_history.

    Returns (series, latest_nav) where series maps isin → [(nav_date, nav), …]
    ascending, and latest_nav maps isin → its most recent NAV.  Read-only on the
    MF module's own schema (no cross-module access).  Empty isins → ({}, {})."""
    if not isins:
        return {}, {}

    from sqlalchemy import select

    from dhanradar.models.mf import MfNavHistory

    cutoff = datetime.now(timezone.utc).date() - timedelta(days=lookback_days)  # noqa: UP017
    result = await db.execute(
        select(MfNavHistory.isin, MfNavHistory.nav_date, MfNavHistory.nav)
        .where(MfNavHistory.isin.in_(isins), MfNavHistory.nav_date >= cutoff)
        .order_by(MfNavHistory.isin, MfNavHistory.nav_date)
    )
    series: dict[str, list[tuple[date, float]]] = {}
    for isin, nav_date, nav in result.all():
        series.setdefault(isin, []).append((nav_date, float(nav)))
    latest_nav = {isin: pts[-1][1] for isin, pts in series.items()}
    return series, latest_nav


# Long-range lookback for the peer-cohort benchmark (B58): the 3Y comparison needs
# > 3 years of NAV, unlike the 400-day momentum/risk load.
_COHORT_LOOKBACK_DAYS = 1200

# Lookback for mf_metrics_refresh — needs 5Y window (vs 3Y for cohort builder).
_METRICS_LOOKBACK_DAYS = 1900

# B63: peers' NAV series are loaded in batches of this many ISINs. Loading every
# peer's 1200-day series at once OOM-killed (SIGKILL) the 640M batch worker the
# moment the NAV table became complete (5.9M rows; hundreds of peers per
# category). Peak memory is now one batch; per-peer stats are identical — the
# same long_horizon_stats runs on the same per-fund series either way.
_COHORT_PEER_CHUNK = 200

# Batch size for the nightly mf_metrics_refresh upsert (ISINs per iteration).
_METRICS_REFRESH_CHUNK = 500

# Batch size for compute_market_ranks' per-fund 400-day NAV load (RCA 2026-07-05).
# Same discipline as _METRICS_REFRESH_CHUNK/B63: peak memory is one batch, never
# all ~14k funds' series at once (22M-row NAV table post-backfill, 640MiB worker).
_RANKS_NAV_CHUNK = 500

# Minimum number of funds with a non-None value in a category before we write
# category-level percentiles (p25/p50/p75/p90) for that (category, metric_key).
# Fewer than 5 funds is too thin to produce meaningful distribution stats.
_MIN_CATEGORY_FUNDS = 5

# Peer-cohort GROUPING KEY — VERSIONED METHODOLOGY (B6/B28 two-person gate), B66-f1
# part 2. Which mf_funds column groups peers into a category cohort:
#   * "category"      — the RAW AMFI string (v1.1). Malformed variants (bare "ELSS",
#     double-space "Other  ETFs", curly-apostrophe "Children's") and pre-2017 legacy
#     umbrellas ("Income"/"Growth"/"Gilt") each form their own string-distinct cohort
#     → fragmentation + a few bogus mega-cohorts.
#   * "sebi_category" — the VALIDATED canonical SEBI leaf (taxonomy.canonical_for,
#     B66). Malformed variants collapse into the correct canonical cohort; legacy
#     umbrellas are sebi_category NULL → excluded by SQL `IN` (NULLs never match) and
#     by the `if c` target filter → those funds stay HONESTLY UNCOHORTED (on_track +
#     the COHORT_NO_CANONICAL_CATEGORY context, B71). NEVER auto-mapped.
# This is the grouping key ONLY; the raw `category` column is never mutated
# (taxonomy.py invariant). Mirrored in ranking_configs_v1.json
# labels.cohort_grouping_key — lockstep test-enforced.
#
# v1.2 ACTIVATED 2026-06-14 (B66-f1 pt2): flipped to "sebi_category" under the B6/B28
# two-person gate (founder = approved_by ≠ created_by) + founder deploy approval, per
# ADR-0034. Read-only prod backtest measured 196 funds (1.40%) changing label (within
# the 5% churn gate). Prereqs B71 + B58-f5 landed first (PR #131). To roll back: set
# this to "category" + manifest to v1.1 and redeploy (indexes stay).
_COHORT_GROUPING_KEY = "sebi_category"


def _grouping_column(grouping_key: str):
    """Resolve the methodology cohort grouping key to its mf_funds column.

    Pure mapping (no DB access); raises on an unknown key so a typo in the
    methodology manifest fails loud rather than silently regrouping cohorts."""
    from dhanradar.models.mf import MfFund

    cols = {"category": MfFund.category, "sebi_category": MfFund.sebi_category}
    if grouping_key not in cols:
        raise ValueError(f"unknown cohort grouping key: {grouping_key!r}")
    return cols[grouping_key]


@dataclass(frozen=True)
class _CohortContext:
    """Pre-built peer-cohort benchmarks + per-fund long-horizon stats (B58-f2).

    Built ONCE per task run by :func:`_build_cohort_context`; per-portfolio label
    inputs then come from :func:`_relative_from_context` as pure lookups — the
    expensive peer NAV loads never repeat inside a per-portfolio loop."""

    category_by_isin: dict[str, str]
    stats_by_isin: dict[str, FundStats]
    benchmarks: dict[str, CohortBenchmark]
    # Targets present in mf_funds but with NO usable cohort key (NULL/blank/
    # "uncategorized" grouping value) — uncohorted, but KNOWN-uncohorted, so the
    # label carries an honest "no canonical category" context (B71) rather than a
    # silent on_track. Empty under the active "category" key (no prod fund lacks a
    # raw category); populated by the legacy umbrellas under "sebi_category".
    uncategorized_isins: frozenset[str] = frozenset()


_EMPTY_COHORT_CONTEXT = _CohortContext({}, {}, {})


async def _build_cohort_context(
    db: Any,
    target_isins: list[str],
    *,
    as_of: date | None = None,
    grouping_key: str = _COHORT_GROUPING_KEY,
) -> _CohortContext:
    """Build category peer-cohort benchmarks for the given target funds (B58).

    Read-only on the mf schema (MfFund + mf_fund_metrics) — no cross-module access.
    The peer set for a category is every fund AMFI tags in that category; the fund
    itself is included (negligible self-bias on a ≥5-peer median).

    ``grouping_key`` (B66-f1 pt2) selects the mf_funds column peers are grouped by
    — "category" (raw, v1.1 active) or "sebi_category" (validated canonical leaf).
    See ``_COHORT_GROUPING_KEY``: a fund whose grouping column is NULL/blank is
    dropped at step 1 and never appears as a peer (SQL ``IN`` excludes NULL), so it
    stays uncohorted → on_track fail-safe. The raw ``category`` column is never
    mutated regardless of the key (taxonomy.py invariant).

    Step 3 reads precomputed ``mf_fund_metrics`` rows (refreshed nightly by
    ``mf_metrics_refresh``) instead of loading peer NAV series into memory (B63).
    The stats are numerically identical — ``mf_metrics_refresh`` stores exactly
    ``long_horizon_stats()`` output on the same NAV data.

    ``as_of`` is kept for caller compatibility; stats are now precomputed so the
    value is not forwarded to the DB read.
    """
    from sqlalchemy import select

    from dhanradar.mf.cohort import build_benchmark
    from dhanradar.models.mf import MfFund, MfFundMetrics

    if not target_isins:
        return _EMPTY_COHORT_CONTEXT
    # as_of retained for signature compatibility; stats are precomputed nightly.
    group_col = _grouping_column(grouping_key)

    # 1. Resolve each target's grouping value; keep only real cohort keys. A target
    #    present in mf_funds but with a NULL/blank/"uncategorized" grouping value is
    #    KNOWN-uncohorted → it carries an honest "no canonical category" context
    #    (B71), distinct from a genuine matching-category on_track.
    cat_rows = (
        await db.execute(select(MfFund.isin, group_col).where(MfFund.isin.in_(target_isins)))
    ).all()
    target_category: dict[str, str] = {i: c for i, c in cat_rows if c and c != "uncategorized"}
    uncategorized = frozenset(i for i, c in cat_rows if not (c and c != "uncategorized"))
    categories = set(target_category.values())
    if not categories:
        # No target has a cohort key — still surface the known-uncohorted ones (B71).
        return _CohortContext({}, {}, {}, uncategorized)

    # 2. All peers in those cohorts (SQL ``IN`` excludes NULL-keyed funds).
    peer_rows = (
        await db.execute(select(MfFund.isin, group_col).where(group_col.in_(categories)))
    ).all()
    peers_by_cat: dict[str, list[str]] = {}
    all_peer_isins: list[str] = []
    for i, c in peer_rows:
        peers_by_cat.setdefault(c, []).append(i)
        all_peer_isins.append(i)

    # 3. Read precomputed long-horizon stats from mf_fund_metrics (refreshed nightly
    #    after nav_daily_fetch) in bounded chunks to stay under bind-param limits.
    #    Peers with no row default to (None, None, None) — identical to the old path's
    #    long_horizon_stats([]) for a fund AMFI tags in a category but with no NAV
    #    (that mf_funds-vs-mf_nav_history asymmetry predates this change; both paths
    #    treat it the same way).
    unique_peers = sorted(set(all_peer_isins))
    stats_by_isin: dict[str, FundStats] = {}
    found_any = False
    seen_isins: set[str] = set()
    for start in range(0, len(unique_peers), _COHORT_PEER_CHUNK):
        batch = unique_peers[start : start + _COHORT_PEER_CHUNK]
        metric_rows = (
            await db.execute(
                select(
                    MfFundMetrics.isin,
                    MfFundMetrics.return_1y_pct,
                    MfFundMetrics.return_3y_pct,
                    MfFundMetrics.max_drawdown_pct,
                ).where(MfFundMetrics.isin.in_(batch))
            )
        ).all()
        if metric_rows:
            found_any = True
        row_map = {
            r.isin: (r.return_1y_pct, r.return_3y_pct, r.max_drawdown_pct) for r in metric_rows
        }
        seen_isins.update(row_map.keys())
        for i in batch:
            stats_by_isin[i] = row_map.get(i, (None, None, None))

    # Partial-miss: some peers have metrics, others don't (e.g. new funds ingested
    # today before mf_metrics_refresh ran). Absent peers default to (None, None, None)
    # and may slightly skew the category median. Log for observability.
    if found_any:
        no_row = [i for i in unique_peers if i not in seen_isins]
        if no_row:
            logger.warning(
                "_build_cohort_context: %d/%d peers have no mf_fund_metrics row "
                "(possibly new funds added since last nightly refresh). "
                "Category benchmark may be slightly skewed. Sample: %s",
                len(no_row),
                len(unique_peers),
                no_row[:5],
            )

    # Empty-table safety net: if mf_fund_metrics has NO row for ANY peer (a fresh
    # deploy before the first mf_metrics_refresh, or a wiped table), every benchmark
    # would silently withhold → on_track for everyone. Fall back to the live NAV
    # computation for THIS run (the exact pre-refactor math — equivalent, just
    # memory-heavier) and log loudly so the missed populate is observable, not silent.
    if unique_peers and not found_any:
        from dhanradar.mf.signals import long_horizon_stats

        logger.critical(
            "mf_fund_metrics empty for %d peers — falling back to live cohort "
            "computation; run mf_metrics_refresh to populate",
            len(unique_peers),
        )
        # NOTE: as_of mismatch vs precomputed path — mf_metrics_refresh computes
        # stats with as_of=date.today(); this fallback honors the caller's as_of.
        # long_horizon_stats currently ignores as_of (windows anchor on the latest
        # NAV point), so there is no live divergence. If as_of is ever made
        # window-sensitive for backtesting, reconcile both paths before enabling it.
        ref_as_of = as_of or date.today()
        stats_by_isin = {}
        for start in range(0, len(unique_peers), _COHORT_PEER_CHUNK):
            batch = unique_peers[start : start + _COHORT_PEER_CHUNK]
            series, _ = await _load_nav_series(db, batch, lookback_days=_COHORT_LOOKBACK_DAYS)
            for i in batch:
                stats_by_isin[i] = long_horizon_stats(series.get(i, []), as_of=ref_as_of)

    # 4. Per-category median benchmark — unchanged.
    benchmarks = {
        cat: build_benchmark(cat, [stats_by_isin[i] for i in cat_isins])
        for cat, cat_isins in peers_by_cat.items()
    }
    return _CohortContext(target_category, stats_by_isin, benchmarks, uncategorized)


def _relative_from_context(
    ctx: _CohortContext, target_isins: list[str]
) -> dict[str, CategoryRelative]:
    """Each target's own stats vs its category benchmark — pure lookup against a
    pre-built context.  A fund absent from the result (no category at build time,
    thin cohort, or no NAV) is scored with no category red flag → on_track, the
    honest fail-safe.  A KNOWN-uncohorted fund (its grouping value was NULL/blank —
    a legacy umbrella with no canonical SEBI category) gets an explicit
    ``COHORT_NO_CANONICAL_CATEGORY`` context (B71) so its on_track reads
    honest-not-positive rather than implying a peer comparison was made."""
    from dhanradar.mf.cohort import compare_to_cohort
    from dhanradar.scoring.engine.signal_names import SignalName, display

    out: dict[str, CategoryRelative] = {}
    for i in target_isins:
        c = ctx.category_by_isin.get(i)
        if c is None:
            if i in ctx.uncategorized_isins:
                out[i] = CategoryRelative(
                    contributing=[display(SignalName.COHORT_NO_CANONICAL_CATEGORY)]
                )
            continue
        out[i] = compare_to_cohort(
            ctx.stats_by_isin.get(i, (None, None, None)), ctx.benchmarks.get(c)
        )
    return out


async def _compute_cohort(
    db: Any,
    target_isins: list[str],
    *,
    as_of: date | None = None,
    grouping_key: str = _COHORT_GROUPING_KEY,
) -> dict[str, CategoryRelative]:
    """Build category peer-cohort benchmarks and return each target fund's
    category-relative LABEL inputs (B58) — build + lookup in one call, for the
    single-portfolio path (CAS upload).  The monthly rescore builds the context
    once and reuses it across portfolios instead (B58-f2)."""
    ctx = await _build_cohort_context(db, target_isins, as_of=as_of, grouping_key=grouping_key)
    return _relative_from_context(ctx, target_isins)


async def _mark_failed(job_id: str, message: str) -> None:
    from sqlalchemy import update

    from dhanradar.db import admin_task_session
    from dhanradar.models.mf import MfCasJob

    async with admin_task_session() as db:
        await db.execute(
            update(MfCasJob)
            .where(MfCasJob.job_id == job_id)
            .values(status="failed", error_message=message[:500])
        )
        await db.commit()


def _purge(path: str) -> None:
    try:
        if path and os.path.exists(path):
            os.remove(path)
    except OSError:
        pass


# ---------------------------------------------------------------------------
# NAV ingestion tasks
# ---------------------------------------------------------------------------


@celery_app.task(name="dhanradar.tasks.mf.nav_daily_fetch")
def nav_daily_fetch() -> str:
    """AMFI NAV daily refresh: fetch NAVAll.txt → bulk-upsert mf_nav_history +
    mf_funds metadata (amfi_code, scheme_name, category only).  Real network call
    via fetch_navall_rows_with_category.  Wired to the beat schedule."""
    try:
        return asyncio.run(_nav_daily_pipeline())
    except Exception:  # noqa: BLE001
        logger.exception("nav_daily_fetch pipeline error")
        return "nav_daily_fetch: failed — see worker logs"


async def _nav_daily_pipeline() -> str:
    from sqlalchemy import func
    from sqlalchemy.dialects.postgresql import insert

    from dhanradar.db import TaskSessionLocal
    from dhanradar.market_data import amfi
    from dhanradar.models.mf import MfFund, MfNavHistory

    logger.info("nav_daily_fetch: fetching NAVAll.txt from AMFI")
    rows = await amfi.fetch_navall_rows_with_category()
    logger.info("nav_daily_fetch: fetched %d rows", len(rows))

    # Taxonomy validation — best-effort; never let logging raise and break ingestion.
    try:
        summary = taxonomy_summarize(r.category for r in rows)
        logger.info(
            "nav_daily_fetch: taxonomy counts %s",
            summary.counts,
        )
        if summary.counts.get("unknown", 0) > 0 or summary.counts.get("legacy", 0) > 0:
            logger.warning(
                "nav_daily_fetch: taxonomy drift detected — unknown_samples=%r legacy_samples=%r",
                summary.unknown_samples,
                summary.legacy_samples,
            )
    except Exception:  # noqa: BLE001
        logger.warning("nav_daily_fetch: taxonomy summarize failed (non-fatal)", exc_info=True)

    nav_dicts = _navrows_to_nav_upserts(rows)
    fund_dicts = _navrows_to_fund_upserts(rows)

    async with TaskSessionLocal() as db:
        # -- mf_nav_history bulk upsert in chunks ---------------------------------
        n_nav = 0
        for i in range(0, len(nav_dicts), _UPSERT_CHUNK):
            chunk = nav_dicts[i : i + _UPSERT_CHUNK]
            if not chunk:
                continue
            stmt = (
                insert(MfNavHistory)
                .values(chunk)
                .on_conflict_do_update(
                    constraint="uq_mf_nav_isin_date",
                    set_={
                        "nav": insert(MfNavHistory).excluded.nav,
                        "source": "amfi",
                        "ingested_at": func.now(),
                    },
                )
            )
            await db.execute(stmt)
            n_nav += len(chunk)

        # -- mf_funds upsert (only the 3 AMFI-owned columns) ---------------------
        n_funds = 0
        for i in range(0, len(fund_dicts), _UPSERT_CHUNK):
            chunk = fund_dicts[i : i + _UPSERT_CHUNK]
            if not chunk:
                continue
            stmt = (
                insert(MfFund)
                .values(chunk)
                .on_conflict_do_update(
                    index_elements=["isin"],
                    set_={
                        "amfi_code": insert(MfFund).excluded.amfi_code,
                        "scheme_name": insert(MfFund).excluded.scheme_name,
                        "category": insert(MfFund).excluded.category,
                        "sebi_category": insert(MfFund).excluded.sebi_category,
                        "plan_type": insert(MfFund).excluded.plan_type,
                        "option_type": insert(MfFund).excluded.option_type,
                        "fund_name_short": insert(MfFund).excluded.fund_name_short,
                        "idcw_frequency": insert(MfFund).excluded.idcw_frequency,
                        "is_segregated": insert(MfFund).excluded.is_segregated,
                        "benchmark_index": insert(MfFund).excluded.benchmark_index,
                        # Keep the earliest date seen — LEAST ignores NULL so a NULL
                        # existing launch_date gets replaced by the incoming nav_date.
                        "launch_date": func.least(
                            MfFund.launch_date, insert(MfFund).excluded.launch_date
                        ),
                    },
                )
            )
            await db.execute(stmt)
            n_funds += len(chunk)

        await db.commit()

    summary = f"nav_daily_fetch: {n_nav} navs, {n_funds} funds"
    logger.info(summary)
    return summary


@celery_app.task(name="dhanradar.tasks.mf.nav_backfill")
def nav_backfill(years: int = 3) -> str:
    """Bootstrap multi-year historical NAV from AMFI history endpoint.

    Iterates backwards in <=90-day windows from (today - years*365) to today.
    Upserts nav rows only (no mf_funds update — history feed carries no category).
    NOT in the beat schedule — invoke manually.
    """
    try:
        return asyncio.run(_nav_backfill_pipeline(years))
    except Exception:  # noqa: BLE001
        logger.exception("nav_backfill pipeline error years=%d", years)
        return f"nav_backfill: failed after error — see worker logs (years={years})"


async def _nav_backfill_pipeline(years: int) -> str:
    import asyncio as _asyncio

    from sqlalchemy import func
    from sqlalchemy.dialects.postgresql import insert

    from dhanradar.db import TaskSessionLocal
    from dhanradar.market_data import amfi
    from dhanradar.market_data.exceptions import ProviderError
    from dhanradar.models.mf import MfNavHistory

    today = datetime.now(timezone.utc).date()  # noqa: UP017 — matches pre-existing style in this file
    start = today - timedelta(days=years * 365)

    # Build list of (frmdt, todt) windows of at most 90 days each, from
    # start → today, non-overlapping.
    _WINDOW_DAYS = 90
    windows: list[tuple[date, date]] = []
    cursor = start
    while cursor < today:
        end = min(cursor + timedelta(days=_WINDOW_DAYS - 1), today)
        windows.append((cursor, end))
        cursor = end + timedelta(days=1)

    logger.info(
        "nav_backfill: years=%d, windows=%d, start=%s, end=%s",
        years,
        len(windows),
        start,
        today,
    )

    total_rows = 0
    windows_fetched = 0

    for idx, (frmdt, todt) in enumerate(windows, start=1):
        window_rows = 0
        try:
            # Streamed + batched (not one full-window list): a 90-day
            # all-funds window is ~1.1M lines — materializing it whole (plus
            # a second upsert-dict list) is what OOM-killed the celery-batch
            # worker (640MiB cap). _batch_nav_upserts yields chunk_size-sized
            # upsert-dict batches as rows arrive, so peak memory is O(batch).
            async for nav_dicts in _batch_nav_upserts(amfi.stream_nav_history(frmdt, todt)):
                async with TaskSessionLocal() as db:
                    stmt = (
                        insert(MfNavHistory)
                        .values(nav_dicts)
                        .on_conflict_do_update(
                            constraint="uq_mf_nav_isin_date",
                            set_={
                                "nav": insert(MfNavHistory).excluded.nav,
                                "source": "amfi",
                                "ingested_at": func.now(),
                            },
                        )
                    )
                    await db.execute(stmt)
                    await db.commit()
                window_rows += len(nav_dicts)
        except ProviderError as exc:
            logger.warning(
                "nav_backfill: window %d/%d (%s–%s) fetch failed: %s",
                idx,
                len(windows),
                frmdt,
                todt,
                exc,
            )
            await _asyncio.sleep(1)
            continue

        total_rows += window_rows

        windows_fetched += 1
        logger.info(
            "nav_backfill: window %d/%d (%s–%s) → %d rows (total so far: %d)",
            idx,
            len(windows),
            frmdt,
            todt,
            window_rows,
            total_rows,
        )
        await _asyncio.sleep(1)

    summary = (
        f"nav_backfill: {total_rows} rows upserted across {windows_fetched}/{len(windows)} windows"
        f" (years={years})"
    )
    logger.info(summary)
    return summary


@celery_app.task(name="dhanradar.tasks.mf.mf_metrics_refresh")
def mf_metrics_refresh() -> str:
    """Nightly precompute of per-fund long-horizon stats into mf_fund_metrics.

    Runs after nav_daily_fetch so metrics reflect the day's fresh NAV.
    The cohort builder reads these rows instead of loading peer NAV series
    into worker memory (B63 memory cap).  Wired to the beat schedule.
    """
    try:
        return asyncio.run(_metrics_refresh_pipeline())
    except Exception:  # noqa: BLE001
        logger.exception("mf_metrics_refresh pipeline error")
        return "mf_metrics_refresh: failed — see worker logs"


async def _metrics_refresh_pipeline() -> str:
    from collections import defaultdict

    from sqlalchemy import func, select
    from sqlalchemy.dialects.postgresql import insert

    from dhanradar.config import settings as _s
    from dhanradar.db import TaskSessionLocal
    from dhanradar.mf.risk import (
        benchmark_relative_stats,
        calendar_year_returns,
        percentile,
        resolve_risk_free_rate,
        risk_adjusted_stats,
        rolling_3y_returns,
    )
    from dhanradar.mf.signals import extended_horizon_stats
    from dhanradar.models.mf import (
        MfBenchmarkDaily,
        MfCategoryStats,
        MfFund,
        MfFundMetrics,
        MfMacroIndicator,
        MfNavHistory,
    )

    today = date.today()
    run_id = str(uuid4())

    # Resolve the Sharpe/Sortino risk-free rate ONCE per run (tbill enrichment).
    # RBI DBIE is fragile (undocumented SPA, can silently zero-fill/stale) —
    # resolve_risk_free_rate() fails CLOSED to the existing placeholder unless
    # the latest ingested 91-day T-bill yield is fresh and sane.
    async with TaskSessionLocal() as db:
        tbill_row = (
            await db.execute(
                select(MfMacroIndicator.indicator_value, MfMacroIndicator.as_of_date)
                .where(MfMacroIndicator.indicator_key == "tbill_91d_yield_pct")
                .order_by(MfMacroIndicator.as_of_date.desc())
                .limit(1)
            )
        ).first()
    tbill_value, tbill_as_of = tbill_row if tbill_row else (None, None)

    rf_resolution = resolve_risk_free_rate(
        tbill_value_pct=float(tbill_value) if tbill_value is not None else None,
        tbill_as_of=tbill_as_of,
        today=today,
        placeholder_annual=_s.RISK_FREE_RATE_ANNUAL,
    )
    risk_free_annual = rf_resolution.rate_annual
    if rf_resolution.source == "placeholder":
        _slog.warning(
            "mf_metrics_refresh.risk_free_rate_fallback",
            reason=rf_resolution.rejected_reason,
            placeholder_annual=_s.RISK_FREE_RATE_ANNUAL,
        )
    else:
        logger.info(
            "mf_metrics_refresh: risk-free rate source=tbill@%s rate_annual=%.4f",
            rf_resolution.as_of_date,
            risk_free_annual,
        )

    async with TaskSessionLocal() as db:
        # Load all ISINs that have any NAV data.
        isin_rows = (await db.execute(select(MfNavHistory.isin).distinct())).all()
        all_isins = [r[0] for r in isin_rows]

    logger.info("mf_metrics_refresh: %d ISINs to process", len(all_isins))

    n_processed = 0
    # Per-fund calendar-year returns, accumulated across chunks for the category-stats
    # step below (W2 §10.5) — a small {year: pct} dict per ISIN, not the NAV series
    # itself, so this stays well within the B63 memory cap even at full fund count.
    cy_returns_by_isin: dict[str, dict[int, float]] = {}
    for start in range(0, len(all_isins), _METRICS_REFRESH_CHUNK):
        chunk = all_isins[start : start + _METRICS_REFRESH_CHUNK]
        if not chunk:
            continue

        async with TaskSessionLocal() as db:
            # Load long NAV series for this chunk (5Y window for extended metrics).
            series, _ = await _load_nav_series(db, chunk, lookback_days=_METRICS_LOOKBACK_DAYS)

            # Block 0.7 — benchmark_index map for this chunk (index funds only;
            # every other fund's benchmark_index is NULL, so this dict only ever
            # has entries for a subset of `chunk`). One extra read-only query per
            # chunk, scoped to the ISINs already loaded above.
            bm_rows = (
                await db.execute(
                    select(MfFund.isin, MfFund.benchmark_index).where(
                        MfFund.isin.in_(chunk), MfFund.benchmark_index.isnot(None)
                    )
                )
            ).all()
            benchmark_by_isin: dict[str, str] = {r.isin: r.benchmark_index for r in bm_rows}

            # Fetch each DISTINCT benchmark's daily-close series ONCE per chunk
            # (multiple index funds commonly share the same benchmark) — same
            # query shape as mf/router.py's _benchmark_returns (PR #483/block 0.4,
            # read-only reuse, no edits to that function/table).
            bench_series_cache: dict[str, list[tuple[date, float]]] = {}
            for registry_key in set(benchmark_by_isin.values()):
                spec = BENCHMARK_REGISTRY.get(registry_key)
                if spec is None:
                    # Registry drift guard — a stale benchmark_index value that no
                    # longer resolves to a registry entry; skip, don't crash the chunk.
                    continue
                bench_rows = (
                    await db.execute(
                        select(MfBenchmarkDaily.close_date, MfBenchmarkDaily.close_value)
                        .where(MfBenchmarkDaily.benchmark == spec.storage_key)
                        .order_by(MfBenchmarkDaily.close_date.asc())
                    )
                ).all()
                bench_series_cache[registry_key] = [
                    (r.close_date, float(r.close_value)) for r in bench_rows
                ]

            upsert_dicts: list[dict] = []
            for isin in chunk:
                # as_of is currently window-irrelevant (long_horizon_stats anchors
                # windows on the latest NAV point, not as_of). Stored for
                # forward-compatibility only. If as_of is ever wired to anchor
                # windows, the refresh must store per-fund as_of and the cohort
                # builder must filter by it to stay equivalent with the live path.
                r3m, r6m, r1, r3, r5, dd = extended_horizon_stats(series.get(isin, []))

                # Risk-adjusted metrics (Sharpe, Sortino, vol, rolling 1Y).
                rs = risk_adjusted_stats(
                    series.get(isin, []),
                    risk_free_annual=risk_free_annual,
                )
                # Rolling 3Y (migration 0065, W2 §10.5) — same algorithm as rolling
                # 1Y above, longer window; independent of risk_adjusted_stats/RiskStats
                # (which stays fund-NAV 1Y-only; also reused by the portfolio M2.3 path).
                r3y_avg, r3y_min, r3y_max, r3y_pct_pos = rolling_3y_returns(series.get(isin, []))

                # Benchmark-relative stats (alpha/beta/tracking error, Block 0.7) —
                # only for index funds with a registry-valid mapped benchmark_index
                # (bench_series_cache is empty for every other fund); everyone else
                # gets all-None here, never a fabricated/guessed comparison.
                bkey = benchmark_by_isin.get(isin)
                bench_points = bench_series_cache.get(bkey) if bkey else None
                if bench_points:
                    brs = benchmark_relative_stats(
                        series.get(isin, []),
                        bench_points,
                        risk_free_annual=risk_free_annual,
                    )
                    alpha_1y, beta_1y, te_pct = brs.alpha_1y, brs.beta_1y, brs.tracking_error_pct
                else:
                    alpha_1y = beta_1y = te_pct = None

                # ponytail: _METRICS_LOOKBACK_DAYS (1900d, ~5.2y) can fall a little short
                # of the ~6y span calendar_year_returns wants for a full 5-year strip —
                # the oldest year(s) just come back missing, same honest-partial-data
                # pattern as every other window here. Extend _METRICS_LOOKBACK_DAYS if a
                # deeper calendar-year history is ever needed.
                cy_returns_by_isin[isin] = calendar_year_returns(series.get(isin, []), as_of=today)

                upsert_dicts.append(
                    {
                        "isin": isin,
                        "return_3m_pct": r3m,
                        "return_6m_pct": r6m,
                        "return_1y_pct": r1,
                        "return_3y_pct": r3,
                        "return_5y_pct": r5,
                        "max_drawdown_pct": dd,
                        "nav_points": len(series.get(isin, [])),
                        "as_of_date": today,
                        "source_run_id": run_id,
                        # Risk-adjusted metrics (migration 0042).
                        "sharpe_ratio": rs.sharpe_ratio,
                        "sortino_ratio": rs.sortino_ratio,
                        "volatility_pct": rs.volatility_pct,
                        "rolling_1y_avg_pct": rs.rolling_1y_avg_pct,
                        "rolling_1y_min_pct": rs.rolling_1y_min_pct,
                        "rolling_1y_max_pct": rs.rolling_1y_max_pct,
                        "rolling_1y_pct_positive": rs.rolling_1y_pct_positive,
                        # Rolling 3Y stats (migration 0065).
                        "rolling_3y_avg_pct": r3y_avg,
                        "rolling_3y_min_pct": r3y_min,
                        "rolling_3y_max_pct": r3y_max,
                        "rolling_3y_pct_positive": r3y_pct_pos,
                        # Benchmark-relative stats (migration 0071, Block 0.7).
                        "alpha_1y": alpha_1y,
                        "beta_1y": beta_1y,
                        "tracking_error_pct": te_pct,
                    }
                )

            # Bulk upsert in sub-chunks to bound statement size.
            for i in range(0, len(upsert_dicts), _UPSERT_CHUNK):
                sub = upsert_dicts[i : i + _UPSERT_CHUNK]
                if not sub:
                    continue
                stmt = (
                    insert(MfFundMetrics)
                    .values(sub)
                    .on_conflict_do_update(
                        index_elements=["isin"],
                        set_={
                            "return_3m_pct": insert(MfFundMetrics).excluded.return_3m_pct,
                            "return_6m_pct": insert(MfFundMetrics).excluded.return_6m_pct,
                            "return_1y_pct": insert(MfFundMetrics).excluded.return_1y_pct,
                            "return_3y_pct": insert(MfFundMetrics).excluded.return_3y_pct,
                            "return_5y_pct": insert(MfFundMetrics).excluded.return_5y_pct,
                            "max_drawdown_pct": insert(MfFundMetrics).excluded.max_drawdown_pct,
                            "nav_points": insert(MfFundMetrics).excluded.nav_points,
                            "as_of_date": insert(MfFundMetrics).excluded.as_of_date,
                            "source_run_id": insert(MfFundMetrics).excluded.source_run_id,
                            "computed_at": func.now(),
                            # Risk-adjusted metrics (migration 0042).
                            "sharpe_ratio": insert(MfFundMetrics).excluded.sharpe_ratio,
                            "sortino_ratio": insert(MfFundMetrics).excluded.sortino_ratio,
                            "volatility_pct": insert(MfFundMetrics).excluded.volatility_pct,
                            "rolling_1y_avg_pct": insert(MfFundMetrics).excluded.rolling_1y_avg_pct,
                            "rolling_1y_min_pct": insert(MfFundMetrics).excluded.rolling_1y_min_pct,
                            "rolling_1y_max_pct": insert(MfFundMetrics).excluded.rolling_1y_max_pct,
                            "rolling_1y_pct_positive": insert(
                                MfFundMetrics
                            ).excluded.rolling_1y_pct_positive,
                            # Rolling 3Y stats (migration 0065).
                            "rolling_3y_avg_pct": insert(MfFundMetrics).excluded.rolling_3y_avg_pct,
                            "rolling_3y_min_pct": insert(MfFundMetrics).excluded.rolling_3y_min_pct,
                            "rolling_3y_max_pct": insert(MfFundMetrics).excluded.rolling_3y_max_pct,
                            "rolling_3y_pct_positive": insert(
                                MfFundMetrics
                            ).excluded.rolling_3y_pct_positive,
                            # Benchmark-relative stats (migration 0071, Block 0.7).
                            "alpha_1y": insert(MfFundMetrics).excluded.alpha_1y,
                            "beta_1y": insert(MfFundMetrics).excluded.beta_1y,
                            "tracking_error_pct": insert(MfFundMetrics).excluded.tracking_error_pct,
                        },
                    )
                )
                await db.execute(stmt)
            await db.commit()
            n_processed += len(chunk)

    # ------------------------------------------------------------------
    # Category-stats step: p25/p50/p75/p90 per (sebi_category, metric_key)
    # for return_1y_pct, return_3y_pct, max_drawdown_pct.
    # ------------------------------------------------------------------
    _CATEGORY_METRIC_KEYS = ("return_1y_pct", "return_3y_pct", "max_drawdown_pct")

    async with TaskSessionLocal() as db:
        rows = (
            await db.execute(
                select(
                    MfFund.isin,
                    MfFund.sebi_category,
                    MfFundMetrics.return_1y_pct,
                    MfFundMetrics.return_3y_pct,
                    MfFundMetrics.max_drawdown_pct,
                )
                .join(MfFundMetrics, MfFund.isin == MfFundMetrics.isin)
                .where(MfFund.sebi_category.isnot(None))
            )
        ).all()

    # Group per-metric values by category.
    # Structure: {category: {metric_key: [values]}}
    by_cat_metric: dict[str, dict[str, list[float | None]]] = defaultdict(
        lambda: {k: [] for k in _CATEGORY_METRIC_KEYS}
    )
    # Per-calendar-year cohort (W2 §10.5 consistency strip): {category: {year: [values]}}.
    # metric_key is written as f"return_cy_{year}" below — the composite PK
    # (sebi_category, metric_key, as_of) already supports an arbitrary key, no
    # migration needed.
    by_cat_cy: dict[str, dict[int, list[float]]] = defaultdict(lambda: defaultdict(list))
    for row in rows:
        cat = row.sebi_category
        by_cat_metric[cat]["return_1y_pct"].append(row.return_1y_pct)
        by_cat_metric[cat]["return_3y_pct"].append(row.return_3y_pct)
        by_cat_metric[cat]["max_drawdown_pct"].append(row.max_drawdown_pct)
        for year, ret in cy_returns_by_isin.get(row.isin, {}).items():
            by_cat_cy[cat][year].append(ret)

    cat_stat_upserts: list[dict] = []
    for cat, metric_map in by_cat_metric.items():
        for metric_key, values in metric_map.items():
            valid = sorted(v for v in values if v is not None)
            if len(valid) < _MIN_CATEGORY_FUNDS:
                # Too few funds — skip; noisy percentiles are worse than None.
                continue
            cat_stat_upserts.append(
                {
                    "sebi_category": cat,
                    "metric_key": metric_key,
                    "p25": percentile(valid, 25.0),
                    "p50": percentile(valid, 50.0),
                    "p75": percentile(valid, 75.0),
                    "p90": percentile(valid, 90.0),
                    "as_of": today,
                }
            )
    for cat, year_map in by_cat_cy.items():
        for year, cy_values in year_map.items():
            valid_cy = sorted(cy_values)
            if len(valid_cy) < _MIN_CATEGORY_FUNDS:
                continue
            cat_stat_upserts.append(
                {
                    "sebi_category": cat,
                    "metric_key": f"return_cy_{year}",
                    "p25": percentile(valid_cy, 25.0),
                    "p50": percentile(valid_cy, 50.0),
                    "p75": percentile(valid_cy, 75.0),
                    "p90": percentile(valid_cy, 90.0),
                    "as_of": today,
                }
            )

    n_cat_stats = 0
    if cat_stat_upserts:
        async with TaskSessionLocal() as db:
            for i in range(0, len(cat_stat_upserts), _UPSERT_CHUNK):
                chunk_cs = cat_stat_upserts[i : i + _UPSERT_CHUNK]
                if not chunk_cs:
                    continue
                stmt_cs = (
                    insert(MfCategoryStats)
                    .values(chunk_cs)
                    .on_conflict_do_update(
                        index_elements=["sebi_category", "metric_key", "as_of"],
                        set_={
                            "p25": insert(MfCategoryStats).excluded.p25,
                            "p50": insert(MfCategoryStats).excluded.p50,
                            "p75": insert(MfCategoryStats).excluded.p75,
                            "p90": insert(MfCategoryStats).excluded.p90,
                            "computed_at": func.now(),
                        },
                    )
                )
                await db.execute(stmt_cs)
            await db.commit()
            n_cat_stats = len(cat_stat_upserts)

    summary = (
        f"mf_metrics_refresh: {n_processed} funds, {n_cat_stats} category-stat rows, "
        f"rf_source={rf_resolution.source}"
    )
    logger.info(summary)
    return summary


@celery_app.task(name="dhanradar.tasks.mf.compute_market_ranks")
def compute_market_ranks() -> str:
    """Market-wide per-category rank computation.

    Runs nightly at 01:00 IST after mf_metrics_refresh (00:15). Reads precomputed
    mf_fund_metrics stats, builds per-sebi_category benchmarks, scores each fund
    category-relatively (no user context), and upserts ordinal ranks into
    mf_fund_ranks. unified_score is used ONLY internally for ordering — it is never
    written to mf_fund_ranks (non-neg #2). Idempotent: re-running on the same day
    upserts, never duplicates.

    W2 (§10.1, migration 0064): also persists confidence_band + confidence_factors +
    contributing/contradicting signals from the SAME `score_fund()` call — this was
    already being computed here and discarded; now every browsed (not just held)
    fund gets a band + signals. insufficient_data funds get a null band/factors
    (signals may still carry the honest reason) — never a fabricated read.
    """
    try:
        return asyncio.run(_compute_market_ranks_pipeline())
    except Exception:  # noqa: BLE001
        logger.exception("compute_market_ranks pipeline error")
        return "compute_market_ranks: failed — see worker logs"


async def _compute_market_ranks_pipeline() -> str:
    from collections import defaultdict

    from sqlalchemy import func, select
    from sqlalchemy.dialects.postgresql import insert

    from dhanradar.db import TaskSessionLocal
    from dhanradar.mf.cohort import build_benchmark, compare_to_cohort
    from dhanradar.mf.scoring_bridge import score_fund
    from dhanradar.mf.signals import compute_fund_signals
    from dhanradar.models.mf import MfFund, MfFundMetrics, MfFundRanks
    from dhanradar.scoring.engine import RatingEngine
    from dhanradar.scoring.engine.schemas import VerbLabel

    today = date.today()

    # Load all funds with a valid sebi_category + precomputed metrics in one query.
    async with TaskSessionLocal() as db:
        rows = (
            await db.execute(
                select(
                    MfFund.isin,
                    MfFund.sebi_category,
                    MfFundMetrics.return_1y_pct,
                    MfFundMetrics.return_3y_pct,
                    MfFundMetrics.max_drawdown_pct,
                )
                .join(MfFundMetrics, MfFund.isin == MfFundMetrics.isin)
                .where(MfFund.sebi_category.isnot(None))
            )
        ).all()

    if not rows:
        logger.warning(
            "compute_market_ranks: mf_fund_metrics has no rows with sebi_category "
            "— run mf_metrics_refresh first; skipping"
        )
        return "compute_market_ranks: skipped (mf_fund_metrics empty or no categorised funds)"

    # Group by sebi_category.
    by_cat: dict[str, list] = defaultdict(list)
    for row in rows:
        by_cat[row.sebi_category].append(row)

    rengine = RatingEngine()
    all_upserts: list[dict] = []

    for cat, cat_rows in by_cat.items():
        # Build the category benchmark from precomputed long-horizon stats (no NAV loads).
        stats_list = [(r.return_1y_pct, r.return_3y_pct, r.max_drawdown_pct) for r in cat_rows]
        benchmark = build_benchmark(cat, stats_list)

        # Score each fund on its OWN 400-day NAV window (same loader + default
        # window as the CAS report path) + the category-relative context.
        # RCA 2026-07-05: this used to pass an empty NAV list, which tripped
        # compute_fund_signals' _MIN_POINTS guard BEFORE category_relative was
        # applied — every market label degraded to insufficient_data since
        # 2026-06-15. NAV is loaded per _RANKS_NAV_CHUNK batch and discarded
        # after scoring, so peak memory stays one batch (B63 discipline).
        scored: list[tuple[str, int, str, str | None, dict | None, list, list]] = []
        for start in range(0, len(cat_rows), _RANKS_NAV_CHUNK):
            batch = cat_rows[start : start + _RANKS_NAV_CHUNK]
            async with TaskSessionLocal() as db:
                nav_series, _ = await _load_nav_series(db, [r.isin for r in batch])
            for r in batch:
                fund_stats = (r.return_1y_pct, r.return_3y_pct, r.max_drawdown_pct)
                cat_rel = compare_to_cohort(fund_stats, benchmark)
                signals = compute_fund_signals(
                    r.isin, nav_series.get(r.isin, []), category_relative=cat_rel
                )
                result = await score_fund(rengine, signals)
                # W2 (§10.1): persist the SAME result the label already came from —
                # never re-derive. insufficient_data has no rateable band/factors
                # (non-neg #4 fail-safe); signals may still carry the honest reason.
                refused = result.verb_label == VerbLabel.insufficient_data
                scored.append(
                    (
                        r.isin,
                        result.unified_score or 0,
                        result.verb_label.value,
                        None if refused else result.confidence_band.value,
                        None if refused else dict(result.confidence_factors),
                        list(result.contributing_signals),
                        list(result.contradicting_signals),
                    )
                )

        # Sort: highest unified_score first; isin alphabetically as deterministic tiebreaker.
        scored.sort(key=lambda x: (-x[1], x[0]))
        total = len(scored)
        for rank, (
            isin,
            _score,
            verb_label,
            band,
            factors,
            contributing,
            contradicting,
        ) in enumerate(scored, start=1):
            all_upserts.append(
                {
                    "isin": isin,
                    "sebi_category": cat,
                    "rank": rank,
                    "total_in_cat": total,
                    "verb_label": verb_label,
                    "confidence_band": band,
                    "confidence_factors": factors,
                    "contributing_signals": contributing,
                    "contradicting_signals": contradicting,
                    "as_of_date": today,
                }
            )

    # Bulk upsert — idempotent on (isin, as_of_date) PK.
    async with TaskSessionLocal() as db:
        for i in range(0, len(all_upserts), _UPSERT_CHUNK):
            chunk = all_upserts[i : i + _UPSERT_CHUNK]
            if not chunk:
                continue
            stmt = (
                insert(MfFundRanks)
                .values(chunk)
                .on_conflict_do_update(
                    index_elements=["isin", "as_of_date"],
                    set_={
                        "sebi_category": insert(MfFundRanks).excluded.sebi_category,
                        "rank": insert(MfFundRanks).excluded.rank,
                        "total_in_cat": insert(MfFundRanks).excluded.total_in_cat,
                        "verb_label": insert(MfFundRanks).excluded.verb_label,
                        "confidence_band": insert(MfFundRanks).excluded.confidence_band,
                        "confidence_factors": insert(MfFundRanks).excluded.confidence_factors,
                        "contributing_signals": insert(MfFundRanks).excluded.contributing_signals,
                        "contradicting_signals": insert(MfFundRanks).excluded.contradicting_signals,
                        "computed_at": func.now(),
                    },
                )
            )
            await db.execute(stmt)
        await db.commit()

    summary = f"compute_market_ranks: {len(all_upserts)} ranks across {len(by_cat)} categories"
    logger.info(summary)
    return summary


@celery_app.task(name="dhanradar.tasks.mf.fund_events_refresh")
def fund_events_refresh() -> str:
    """What-Changed diff engine (FUND_DETAIL_DATA_ARCHITECTURE_PLAN.md §10.6, §17 W2).

    Runs nightly at 01:15 IST, after compute_market_ranks (01:00). Diffs the two latest
    `mf_fund_ranks.as_of_date` rows (rank_change), the last two `expense_ratio_history`
    rows per isin (ter_change), and the two latest `mf_fund_constituents.as_of_month`
    snapshots (holding_change), upserting typed events into `mf_fund_events`. Idempotent:
    re-running upserts on (isin, event_type, as_of), never duplicates.
    """
    try:
        return asyncio.run(_fund_events_refresh_pipeline())
    except Exception:  # noqa: BLE001
        logger.exception("fund_events_refresh pipeline error")
        return "fund_events_refresh: failed — see worker logs"


async def _fund_events_refresh_pipeline() -> str:
    from collections import defaultdict

    from sqlalchemy import func, select
    from sqlalchemy.dialects.postgresql import insert

    from dhanradar.db import TaskSessionLocal
    from dhanradar.mf.fund_events import (
        cap_fund_events,
        detect_aum_change,
        detect_holding_change,
        detect_rank_change,
        detect_ter_change,
    )
    from dhanradar.models.mf import (
        MfAumHistory,
        MfExpenseRatioHistory,
        MfFundConstituent,
        MfFundEvent,
        MfFundRanks,
    )

    events_by_isin: dict[str, list[dict]] = defaultdict(list)

    async with TaskSessionLocal() as db:
        # --- rank_change: the two latest as_of_dates are shared across the whole
        # nightly compute_market_ranks run, so this is one pair of dates, not per-isin.
        rank_dates = (
            (
                await db.execute(
                    select(MfFundRanks.as_of_date)
                    .distinct()
                    .order_by(MfFundRanks.as_of_date.desc())
                    .limit(2)
                )
            )
            .scalars()
            .all()
        )
        if len(rank_dates) >= 2:
            new_date, old_date = rank_dates[0], rank_dates[1]
            new_ranks = (
                (await db.execute(select(MfFundRanks).where(MfFundRanks.as_of_date == new_date)))
                .scalars()
                .all()
            )
            old_ranks = {
                r.isin: r
                for r in (
                    await db.execute(select(MfFundRanks).where(MfFundRanks.as_of_date == old_date))
                )
                .scalars()
                .all()
            }
            for new_r in new_ranks:
                old_r = old_ranks.get(new_r.isin)
                if old_r is None:
                    continue
                ev = detect_rank_change(
                    old_rank=old_r.rank,
                    old_total=old_r.total_in_cat,
                    new_rank=new_r.rank,
                    new_total=new_r.total_in_cat,
                )
                if ev is not None:
                    events_by_isin[new_r.isin].append(
                        {
                            "isin": new_r.isin,
                            "event_type": "rank_change",
                            "as_of": new_date,
                            "payload": ev,
                        }
                    )

        # --- ter_change: per-isin last-two rows via a window function (one query,
        # no N+1 — design principle 1, §2).
        ter_subq = select(
            MfExpenseRatioHistory.isin,
            MfExpenseRatioHistory.ter_pct,
            MfExpenseRatioHistory.effective_date,
            func.row_number()
            .over(
                partition_by=MfExpenseRatioHistory.isin,
                order_by=MfExpenseRatioHistory.effective_date.desc(),
            )
            .label("rn"),
        ).subquery()
        ter_rows = (
            await db.execute(
                select(
                    ter_subq.c.isin, ter_subq.c.ter_pct, ter_subq.c.effective_date, ter_subq.c.rn
                )
                .where(ter_subq.c.rn <= 2)
                .order_by(ter_subq.c.isin, ter_subq.c.rn)
            )
        ).all()
        ter_by_isin: dict[str, list] = defaultdict(list)
        for row in ter_rows:
            ter_by_isin[row.isin].append(row)
        for isin, rows in ter_by_isin.items():
            if len(rows) < 2:
                continue
            newest = next(r for r in rows if r.rn == 1)
            older = next(r for r in rows if r.rn == 2)
            ev = detect_ter_change(
                old_ter=float(older.ter_pct),
                new_ter=float(newest.ter_pct),
                effective_date=newest.effective_date,
            )
            if ev is not None:
                events_by_isin[isin].append(
                    {
                        "isin": isin,
                        "event_type": "ter_change",
                        "as_of": newest.effective_date,
                        "payload": ev,
                    }
                )

        # --- aum_change: per-isin last-two rows via a window function, same pattern
        # as ter_change above, sourced from mf.aum_history instead of expense_ratio_history.
        aum_subq = select(
            MfAumHistory.isin,
            MfAumHistory.aum_crore,
            MfAumHistory.as_of_month,
            func.row_number()
            .over(
                partition_by=MfAumHistory.isin,
                order_by=MfAumHistory.as_of_month.desc(),
            )
            .label("rn"),
        ).subquery()
        aum_rows = (
            await db.execute(
                select(aum_subq.c.isin, aum_subq.c.aum_crore, aum_subq.c.as_of_month, aum_subq.c.rn)
                .where(aum_subq.c.rn <= 2)
                .order_by(aum_subq.c.isin, aum_subq.c.rn)
            )
        ).all()
        aum_by_isin: dict[str, list] = defaultdict(list)
        for row in aum_rows:
            aum_by_isin[row.isin].append(row)
        for isin, rows in aum_by_isin.items():
            if len(rows) < 2:
                continue
            newest = next(r for r in rows if r.rn == 1)
            older = next(r for r in rows if r.rn == 2)
            ev = detect_aum_change(
                old_aum_crore=float(older.aum_crore),
                new_aum_crore=float(newest.aum_crore),
                as_of_month=newest.as_of_month,
            )
            if ev is not None:
                events_by_isin[isin].append(
                    {
                        "isin": isin,
                        "event_type": "aum_change",
                        "as_of": newest.as_of_month,
                        "payload": ev,
                    }
                )

        # --- holding_change: the two latest as_of_months are shared across the whole
        # monthly disclosure ingestion run, same pattern as rank_change above.
        months = (
            (
                await db.execute(
                    select(MfFundConstituent.as_of_month)
                    .distinct()
                    .order_by(MfFundConstituent.as_of_month.desc())
                    .limit(2)
                )
            )
            .scalars()
            .all()
        )
        if len(months) >= 2:
            new_month, old_month = months[0], months[1]
            new_c_rows = (
                (
                    await db.execute(
                        select(MfFundConstituent).where(MfFundConstituent.as_of_month == new_month)
                    )
                )
                .scalars()
                .all()
            )
            old_c_rows = (
                (
                    await db.execute(
                        select(MfFundConstituent).where(MfFundConstituent.as_of_month == old_month)
                    )
                )
                .scalars()
                .all()
            )
            new_by_isin: dict[str, list[dict]] = defaultdict(list)
            for r in new_c_rows:
                new_by_isin[r.isin].append(
                    {
                        "name": r.constituent_name,
                        "weight_pct": float(r.weight_pct) if r.weight_pct is not None else None,
                    }
                )
            old_by_isin: dict[str, list[dict]] = defaultdict(list)
            for r in old_c_rows:
                old_by_isin[r.isin].append(
                    {
                        "name": r.constituent_name,
                        "weight_pct": float(r.weight_pct) if r.weight_pct is not None else None,
                    }
                )
            for isin, new_holdings in new_by_isin.items():
                old_holdings = old_by_isin.get(isin)
                if not old_holdings:
                    continue
                ev = detect_holding_change(old_holdings=old_holdings, new_holdings=new_holdings)
                if ev is not None:
                    events_by_isin[isin].append(
                        {
                            "isin": isin,
                            "event_type": "holding_change",
                            "as_of": new_month,
                            "payload": ev,
                        }
                    )

    # Cap per fund (<=3, one per type — §10.6) THEN flatten for the bulk upsert.
    all_events: list[dict] = []
    for isin, evs in events_by_isin.items():
        all_events.extend(cap_fund_events(evs))

    async with TaskSessionLocal() as db:
        for i in range(0, len(all_events), _UPSERT_CHUNK):
            chunk = all_events[i : i + _UPSERT_CHUNK]
            if not chunk:
                continue
            stmt = (
                insert(MfFundEvent)
                .values(chunk)
                .on_conflict_do_update(
                    index_elements=["isin", "event_type", "as_of"],
                    set_={"payload": insert(MfFundEvent).excluded.payload},
                )
            )
            await db.execute(stmt)
        await db.commit()

    summary = f"fund_events_refresh: {len(all_events)} events across {len(events_by_isin)} funds"
    logger.info(summary)
    return summary


@celery_app.task(name="dhanradar.tasks.mf.monthly_rescore_plus_users")
def monthly_rescore_plus_users() -> str:
    """Re-score every Plus user's current holdings from the latest NAV without
    requiring a re-upload.  Writes label history + portfolio snapshot (Plus only).
    Free users are skipped — no history rows, no snapshot.  Wired to the beat
    schedule (1st of month, 03:00 IST).
    """
    try:
        return asyncio.run(_monthly_rescore())
    except Exception:  # noqa: BLE001
        logger.exception("monthly_rescore_plus_users pipeline error")
        return "monthly_rescore: failed — see worker logs"


async def _monthly_rescore() -> str:
    from sqlalchemy import select

    from dhanradar.db import admin_task_session
    from dhanradar.deps import is_plus
    from dhanradar.mf import history as mf_history
    from dhanradar.mf.snapshot import CashFlow as _CashFlow
    from dhanradar.mf.snapshot import Holding as _Holding
    from dhanradar.models.mf import MfFund, MfPortfolio, MfUserHolding
    from dhanradar.notifications import service as notif_service
    from dhanradar.redis_client import get_redis
    from dhanradar.scoring.engine import RatingEngine
    from dhanradar.scoring.engine.schemas import DISCLAIMER_VERSION

    rengine = RatingEngine()
    today = date.today()
    redis = get_redis()

    async with admin_task_session() as db:
        # All distinct (portfolio, isin) pairs that currently have holdings.
        pid_isin_rows = (
            await db.execute(select(MfUserHolding.portfolio_id, MfUserHolding.isin).distinct())
        ).all()
        isins_by_pid: dict[str, set[str]] = {}
        raw_pids: list[Any] = []  # native UUIDs for the owner query — no str round-trip
        for pid_raw, isin_raw in pid_isin_rows:
            key = str(pid_raw)
            if key not in isins_by_pid:
                raw_pids.append(pid_raw)
            isins_by_pid.setdefault(key, set()).add(isin_raw)
        portfolio_ids = list(isins_by_pid)

        # B58-f2: build the peer-cohort context ONCE per run, over the union of
        # Plus portfolios' holdings — not per portfolio inside the loop (the
        # build re-fetched the same category peer NAV sets for every portfolio).
        # Peer loads stay chunked (B63).  A portfolio that turns Plus mid-run is
        # absent from the union and scores without category flags this month —
        # the same honest fail-safe as an uncategorized fund.
        owner_rows = (
            await db.execute(
                select(MfPortfolio.id, MfPortfolio.user_id).where(MfPortfolio.id.in_(raw_pids))
            )
        ).all()
        uid_by_pid = {str(r[0]): str(r[1]) for r in owner_rows}
        plus_by_uid: dict[str, bool] = {}
        for uid in set(uid_by_pid.values()):
            plus_by_uid[uid] = await is_plus(uid, db)
        plus_isins: set[str] = set()
        for pid, uid in uid_by_pid.items():
            if plus_by_uid.get(uid):
                plus_isins |= isins_by_pid[pid]
        cohort_ctx = (
            await _build_cohort_context(db, sorted(plus_isins), as_of=today)
            if plus_isins
            else _EMPTY_COHORT_CONTEXT
        )

    rescored = 0

    for pid in portfolio_ids:
        try:
            async with admin_task_session() as db:
                # Resolve the owning user_id + portfolio name for this portfolio.
                port_row = (
                    await db.execute(
                        select(MfPortfolio.user_id, MfPortfolio.name).where(
                            MfPortfolio.id == pid  # type: ignore[arg-type]
                        )
                    )
                ).first()
                if port_row is None:
                    continue
                uid = str(port_row[0])
                portfolio_name: str = port_row[1] or ""

                if not await is_plus(uid, db):
                    continue  # Free users — no history/snapshot written.

                # Load holding rows for this portfolio.
                holding_rows = (
                    (
                        await db.execute(
                            select(MfUserHolding).where(
                                MfUserHolding.portfolio_id == pid  # type: ignore[arg-type]
                            )
                        )
                    )
                    .scalars()
                    .all()
                )

                if not holding_rows:
                    continue

                isins = [h.isin for h in holding_rows]
                nav_series, latest_nav = await _load_nav_series(db, isins)
                # Category-relative label inputs — pure lookup against the
                # run-level cohort context built before the loop (B58-f2).
                cohort = _relative_from_context(cohort_ctx, isins)

                # Batch-fetch scheme names for alert copy (never scores or numerics).
                scheme_rows = await db.execute(
                    select(MfFund.isin, MfFund.scheme_name).where(MfFund.isin.in_(isins))
                )
                scheme_by_isin: dict[str, str] = {i: n for i, n in scheme_rows.all()}

                # Score each fund via the bridge (never recompute the engine directly).
                for h_row in holding_rows:
                    isin = h_row.isin
                    signals = compute_fund_signals(
                        isin, nav_series.get(isin, []), category_relative=cohort.get(isin)
                    )
                    result = await score_fund(rengine, signals)
                    await upsert_user_fund_score(db, uid, result, pid)

                    new_label = result.verb_label.value
                    prior_label = await mf_history.get_prior_label(db, pid, isin, today)

                    inserted = await mf_history.append_score_history(
                        db,
                        user_id=uid,
                        result=result,
                        snapshot_date=today,
                        source="monthly_rescore",
                        portfolio_id=pid,
                    )

                    # Enqueue educational label-change alert (best-effort).
                    # Guard: inserted=True (real new row) + prior exists + label changed.
                    # The inserted guard makes alerts idempotent: a beat re-run on the
                    # same day returns inserted=False and skips the enqueue entirely.
                    if inserted and prior_label is not None and prior_label != new_label:
                        payload = {
                            "scheme_name": scheme_by_isin.get(isin, "A fund in your portfolio"),
                            "portfolio_name": portfolio_name,
                            "prior_label": prior_label,
                            "new_label": new_label,
                            "isin": isin,
                            "confidence_band": result.confidence_band.value,
                            # Stamp the in-force disclaimer version at GENERATION so the
                            # B26 deliver-seam audit ties to what was served, not the live
                            # constant at drain time (Tier-B audit-integrity fix).
                            "disclaimer_version": DISCLAIMER_VERSION,
                        }
                        for ch in ("telegram", "email"):
                            try:
                                await notif_service.publish_notification(
                                    redis, uid, ch, "mf_label_change", payload
                                )
                            except Exception:  # noqa: BLE001 — alert is best-effort; never abort rescore
                                # user_ref is hashed (never the raw uid in logs).
                                _slog.warning(
                                    "label_change_alert_enqueue_failed",
                                    user_ref=hash_user_ref(str(uid)),
                                    isin=isin,
                                    ch=ch,
                                )

                # Build snapshot from current holdings + latest NAV.
                holdings: list[_Holding] = []
                for h_row in holding_rows:
                    nav = latest_nav.get(h_row.isin)
                    current_value = float(h_row.units) * nav if nav is not None else 0.0
                    invested = (
                        float(h_row.invested_amount) if h_row.invested_amount is not None else 0.0
                    )
                    holdings.append(
                        _Holding(
                            isin=h_row.isin,
                            units=float(h_row.units),
                            invested_amount=invested,
                            current_value=current_value,
                            category="uncategorized",
                            cashflows=[_CashFlow(when=today, amount=current_value)],
                        )
                    )

                snap = build_snapshot(holdings)

                # The rescore builds Holdings with a single cashflow
                # (current_value as of today).  xirr() requires ≥2 cashflows, so
                # snap.xirr_pct is always None here — the transaction ledger that
                # would supply the purchase/SIP flows is not yet built.
                # PRESERVE the most recently-computed xirr_pct from the DB rather
                # than writing NULL and clobbering the value computed at CAS-upload
                # time.  This is the correct carry-forward until the transaction
                # ledger exists (Path-B / ADR-0037).
                if snap.xirr_pct is None:
                    from dataclasses import replace as _dc_replace

                    from sqlalchemy import select as _select

                    from dhanradar.models.mf import MfPortfolioSnapshot as _MfSnap

                    prior_xirr_row = (
                        await db.execute(
                            _select(_MfSnap.xirr_pct)
                            .where(_MfSnap.portfolio_id == pid)  # type: ignore[arg-type]
                            .order_by(_MfSnap.snapshot_date.desc())
                            .limit(1)
                        )
                    ).scalar_one_or_none()
                    if prior_xirr_row is not None:
                        snap = _dc_replace(snap, xirr_pct=float(prior_xirr_row))

                await mf_history.persist_portfolio_snapshot(
                    db,
                    user_id=uid,
                    snapshot_date=today,
                    snap=snap,
                    portfolio_id=pid,
                )
                rescored += 1

        except Exception:  # noqa: BLE001 — one bad portfolio never aborts the beat
            logger.exception("monthly_rescore: failed for portfolio_id=%s", pid)
            continue

    summary = f"monthly_rescore: rescored {rescored} Plus users"
    logger.info(summary)
    return summary


@celery_app.task(name="dhanradar.tasks.mf.reap_stuck_cas_jobs")
def reap_stuck_cas_jobs() -> str:
    """Beat task: mark CAS jobs that have been in a non-terminal status for more
    than 10 minutes as failed with the opaque code 'stuck_timeout'.

    Safe to run every few minutes — idempotent (WHERE completed_at IS NULL guards
    against re-touching already-terminal rows).  Also clears each reaped job's Redis
    dedup key so a clean re-upload reprocesses (uses the existing service.dedup_clear
    helper, which expects (user_id, portfolio_id, source_hash) — all present on the
    MfCasJob row).
    """
    try:
        return asyncio.run(_reap_stuck_cas_jobs())
    except Exception:  # noqa: BLE001
        logger.exception("reap_stuck_cas_jobs pipeline error")
        return "reap_stuck_cas_jobs: failed — see worker logs"


async def _reap_stuck_cas_jobs() -> str:
    from sqlalchemy import select, update

    from dhanradar.db import admin_task_session
    from dhanradar.models.mf import MfCasJob
    from dhanradar.redis_client import get_redis

    cutoff = datetime.now(timezone.utc) - timedelta(minutes=10)  # noqa: UP017 — matches pre-existing style

    async with admin_task_session() as db:
        # SELECT the rows we are about to reap so we can clear their dedup keys.
        result = await db.execute(
            select(
                MfCasJob.job_id, MfCasJob.user_id, MfCasJob.portfolio_id, MfCasJob.source_hash
            ).where(
                MfCasJob.status.in_(["queued", "parsing", "scoring"]),
                MfCasJob.created_at < cutoff,
                MfCasJob.completed_at.is_(None),
            )
        )
        rows = result.all()

        if not rows:
            _slog.debug("reap_stuck_cas_jobs: nothing to reap")
            return "reaped 0 stuck jobs"

        job_ids = [str(r.job_id) for r in rows]

        # Bulk UPDATE to 'failed' / 'stuck_timeout' in one statement.
        await db.execute(
            update(MfCasJob)
            .where(MfCasJob.job_id.in_(job_ids))  # type: ignore[arg-type]
            .values(status="failed", error_message="stuck_timeout")
        )
        await db.commit()

    # Best-effort: clear Redis dedup keys so re-upload reprocesses cleanly.
    # Runs OUTSIDE the DB session (Redis is independent; a failure here is non-fatal).
    try:
        redis = get_redis()
        for r in rows:
            if r.portfolio_id is not None:
                await service.dedup_clear(
                    redis,
                    str(r.user_id),
                    str(r.portfolio_id),
                    r.source_hash,
                )
    except Exception:  # noqa: BLE001 — dedup clear is best-effort; reaper result is already committed
        logger.warning("reap_stuck_cas_jobs: Redis dedup_clear failed (non-fatal)", exc_info=True)

    n = len(rows)
    _slog.info("reap_stuck_cas_jobs: reaped", count=n, job_ids=job_ids)
    return f"reaped {n} stuck jobs"


@celery_app.task(name="dhanradar.tasks.mf.daily_portfolio_refresh")
def daily_portfolio_refresh() -> str:
    """Rebuild cached reports for every portfolio that has been uploaded at least once.

    Runs at 01:30 IST — after nav_daily_fetch (23:30) and mf_metrics_refresh (00:15)
    so current_value reflects today's NAV. Users who log in after this task has run
    see a fresh portfolio without re-uploading their CAS statement.
    """
    try:
        return asyncio.run(_daily_portfolio_refresh_pipeline())
    except Exception:
        logger.exception("daily_portfolio_refresh pipeline error")
        return "daily_portfolio_refresh: failed — see worker logs"


async def _daily_portfolio_refresh_pipeline() -> str:
    from sqlalchemy import select

    from dhanradar.db import admin_task_session
    from dhanradar.models.mf import MfPortfolio
    from dhanradar.redis_client import get_redis

    redis = get_redis()
    refreshed = 0
    failed = 0

    async with admin_task_session() as db:
        portfolios = (
            (await db.execute(select(MfPortfolio).where(MfPortfolio.latest_job_id.isnot(None))))
            .scalars()
            .all()
        )

        for portfolio in portfolios:
            try:
                job_id = str(portfolio.latest_job_id)
                portfolio_id = str(portfolio.id)
                # Invalidate the stale cache so rebuild_report_from_db writes fresh data.
                await redis.delete(f"{service._REPORT_PREFIX}{job_id}")
                await service.rebuild_report_from_db(
                    job_id=job_id, portfolio_id=portfolio_id, redis=redis, db=db
                )
                refreshed += 1
            except Exception:
                logger.exception(
                    "daily_portfolio_refresh: failed for portfolio_id=%s", portfolio.id
                )
                failed += 1

    return f"daily_portfolio_refresh: refreshed={refreshed} failed={failed}"


@celery_app.task(name="dhanradar.tasks.mf.purge_cas_files")
def purge_cas_files() -> str:
    """24h backstop: delete any raw CAS file older than the TTL (anti-pattern guard)."""
    import tempfile

    d = os.path.join(tempfile.gettempdir(), "dhanradar_cas")
    if not os.path.isdir(d):
        return "purge: no dir"
    now = time.time()
    removed = 0
    for name in os.listdir(d):
        fp = os.path.join(d, name)
        try:
            if os.path.isfile(fp) and (now - os.path.getmtime(fp)) > _UPLOAD_TTL_SECONDS:
                os.remove(fp)
                removed += 1
        except OSError:
            pass
    return f"purge: removed {removed}"


@celery_app.task(name="dhanradar.tasks.mf.mf_fund_metadata_backfill")
def mf_fund_metadata_backfill() -> str:
    """One-shot backfill: stamp plan_type, option_type, is_segregated, launch_date on all MfFund rows.

    NOT in the beat schedule — invoke manually via Celery CLI or admin panel.
    """
    try:
        return asyncio.run(_mf_fund_metadata_backfill_pipeline())
    except Exception:  # noqa: BLE001
        logger.exception("mf_fund_metadata_backfill pipeline error")
        return "mf_fund_metadata_backfill: failed — see worker logs"


async def _mf_fund_metadata_backfill_pipeline() -> str:
    from sqlalchemy import func as sa_func
    from sqlalchemy import select, text

    from dhanradar.db import TaskSessionLocal
    from dhanradar.models.mf import MfFund, MfNavHistory

    _CHUNK = 500

    # Core-level text() UPDATE avoids the "ORM Bulk UPDATE by Primary Key" path
    # that SQLAlchemy takes for update(Model) + list-of-params, which would require
    # the PK column name in every params dict instead of the bindparam alias.
    update_stmt = text(
        "UPDATE mf.mf_funds"
        " SET plan_type = :b_plan_type, option_type = :b_option_type,"
        " fund_name_short = :b_fund_name_short, idcw_frequency = :b_idcw_frequency,"
        " is_segregated = :b_is_segregated, launch_date = :b_launch_date"
        " WHERE isin = :b_isin"
    )

    async with TaskSessionLocal() as db:
        min_date_rows = (
            await db.execute(
                select(
                    MfNavHistory.isin, sa_func.min(MfNavHistory.nav_date).label("min_date")
                ).group_by(MfNavHistory.isin)
            )
        ).all()
        min_date_map: dict[str, Any] = {r.isin: r.min_date for r in min_date_rows}

        # Select only scalar columns — avoids loading ORM instances into the identity map,
        # which would conflict with the subsequent bulk update.
        fund_rows = (await db.execute(select(MfFund.isin, MfFund.scheme_name))).all()
        n = 0

        for i in range(0, len(fund_rows), _CHUNK):
            chunk = fund_rows[i : i + _CHUNK]
            params = []
            for fund in chunk:
                plan_type, option_type = parse_plan_option(fund.scheme_name)
                name = (fund.scheme_name or "").lower()
                params.append(
                    {
                        "b_isin": fund.isin,
                        "b_plan_type": plan_type,
                        "b_option_type": option_type,
                        "b_fund_name_short": derive_short_name(fund.scheme_name, fund.isin),
                        "b_idcw_frequency": parse_idcw_frequency(fund.scheme_name),
                        "b_is_segregated": "segregated portfolio" in name,
                        "b_launch_date": min_date_map.get(fund.isin),
                    }
                )
            await db.execute(update_stmt, params, execution_options={"synchronize_session": False})
            n += len(chunk)
        await db.commit()

    return f"mf_fund_metadata_backfill: updated {n} funds"


# ---------------------------------------------------------------------------
# mf_constituents_fetch — ADR-0033(a) SEBI Monthly Portfolio Disclosure Scraper
# ---------------------------------------------------------------------------
# Manual-only task (NOT in beat schedule).  Fetches SEBI-format monthly
# portfolio disclosure files from top-10 AMCs, upserts constituent rows into
# mf.mf_fund_constituents, and updates mf_funds.aum_crore from the net-assets
# column.  Coverage: top-10 AMCs (~75-80% market AUM); remainder is a logged
# gap (§8.4 — never imputed from AMC aggregate).
# ---------------------------------------------------------------------------


@celery_app.task(name="dhanradar.tasks.mf.mf_constituents_fetch")
def mf_constituents_fetch() -> str:
    try:
        return asyncio.run(_mf_constituents_pipeline())
    except Exception:  # noqa: BLE001
        logger.exception("mf_constituents_fetch pipeline error")
        return "mf_constituents_fetch: failed — see worker logs"


async def _mf_constituents_pipeline() -> str:
    """Fetch SEBI monthly disclosures for top-10 AMCs, upsert constituents."""
    total_rows = 0
    aum_updates = 0

    # Separate AMCs by resolution strategy.
    template_amcs = [a for a in _AMC_DISCLOSURE_ROOTS if a.get("direct_url_template")]
    json_api_amcs = [a for a in _AMC_DISCLOSURE_ROOTS if a.get("json_api_url_template")]
    static_multi_amcs = [a for a in _AMC_DISCLOSURE_ROOTS if a.get("static_multi")]
    playwright_amcs = [
        a
        for a in _AMC_DISCLOSURE_ROOTS
        if not a.get("direct_url_template")
        and not a.get("json_api_url_template")
        and not a.get("static_multi")
    ]

    async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
        # --- Direct-URL AMCs (no browser needed) ---
        for amc in template_amcs:
            amc_name: str = amc["name"]
            try:
                rows, aum_cnt = await _process_amc_direct(
                    client, amc_name, amc["direct_url_template"]
                )
                total_rows += rows
                aum_updates += aum_cnt
                logger.info(
                    "mf_constituents_fetch amc=%s rows=%d aum_updates=%d", amc_name, rows, aum_cnt
                )
            except Exception:  # noqa: BLE001
                logger.exception("mf_constituents_fetch amc=%s failed — skipping", amc_name)

        # --- JSON-API AMCs (Drupal/CMS API returns ZIP URL, no browser needed) ---
        for amc in json_api_amcs:
            amc_name = amc["name"]
            try:
                rows, aum_cnt = await _process_amc_json_api(
                    client,
                    amc_name,
                    amc["json_api_url_template"],
                    amc.get("zip_xlsx_member_pattern", ""),
                )
                total_rows += rows
                aum_updates += aum_cnt
                logger.info(
                    "mf_constituents_fetch amc=%s rows=%d aum_updates=%d", amc_name, rows, aum_cnt
                )
            except Exception:  # noqa: BLE001
                logger.exception("mf_constituents_fetch amc=%s failed — skipping", amc_name)

        # --- Static-multi AMCs (plain HTML page, one XLSX per scheme, no JS rendering) ---
        for amc in static_multi_amcs:
            amc_name = amc["name"]
            try:
                rows, aum_cnt = await _process_amc_static_multi(client, amc_name, amc["url"])
                total_rows += rows
                aum_updates += aum_cnt
                logger.info(
                    "mf_constituents_fetch amc=%s rows=%d aum_updates=%d", amc_name, rows, aum_cnt
                )
            except Exception:  # noqa: BLE001
                logger.exception("mf_constituents_fetch amc=%s failed — skipping", amc_name)

        # --- Playwright AMCs (JS SPA discovery) ---
        if playwright_amcs:
            import tracemalloc

            tracemalloc.start()
            try:
                from playwright.async_api import async_playwright

                async with async_playwright() as p:
                    browser = await p.chromium.launch(
                        headless=True,
                        args=[
                            "--no-sandbox",
                            "--disable-dev-shm-usage",
                            "--disable-gpu",
                            "--memory-pressure-off",
                            "--disable-extensions",
                        ],
                    )
                    try:
                        first_amc = True
                        for amc in playwright_amcs:
                            amc_name = amc["name"]
                            if not first_amc:
                                await asyncio.sleep(10)
                            first_amc = False
                            try:
                                rows, aum_cnt = await _process_amc(
                                    client, browser, amc_name, amc["url"]
                                )
                                total_rows += rows
                                aum_updates += aum_cnt
                                logger.info(
                                    "mf_constituents_fetch amc=%s rows=%d aum_updates=%d",
                                    amc_name,
                                    rows,
                                    aum_cnt,
                                )
                            except Exception:  # noqa: BLE001
                                logger.exception(
                                    "mf_constituents_fetch amc=%s failed — skipping", amc_name
                                )
                    finally:
                        await browser.close()
            except Exception as e:  # noqa: BLE001
                logger.warning(
                    "mf_constituents_fetch playwright unavailable (%s: %s) — skipping %d JS-SPA AMCs: %s",
                    type(e).__name__,
                    e,
                    len(playwright_amcs),
                    [a["name"] for a in playwright_amcs],
                )
            finally:
                _, peak = tracemalloc.get_traced_memory()
                tracemalloc.stop()
                logger.info(
                    "mf_constituents_fetch playwright_memory_peak peak_mb=%.1f",
                    peak / 1_048_576,
                )
                if peak > 500 * 1_048_576:
                    logger.warning(
                        "mf_constituents_fetch playwright_memory_ceiling_risk peak_mb=%.1f"
                        " — approaching 640MB celery-batch ceiling",
                        peak / 1_048_576,
                    )

    return (
        f"mf_constituents_fetch done: "
        f"total_rows={total_rows} aum_updates={aum_updates} "
        f"amcs={len(_AMC_DISCLOSURE_ROOTS)}"
    )


async def _process_amc_direct(
    client: httpx.AsyncClient, amc_name: str, url_template: str
) -> tuple[int, int]:
    """Download and parse a disclosure file whose URL is known via a predictable template.

    Tries the previous month first, then 2 months back (handles the 10-day publication lag).
    """
    now = datetime.now(UTC)
    for months_back in (1, 2):
        # Subtract months_back months.
        target = (now.replace(day=1) - timedelta(days=months_back * 28)).replace(day=1)
        file_url = url_template.format(month_full=target.strftime("%B"), year=target.strftime("%Y"))
        try:
            resp = await client.get(
                file_url,
                headers={"User-Agent": "DhanRadar/1.0 (research; contact@dhanradar.com)"},
            )
            if resp.status_code == 404:
                continue
            resp.raise_for_status()
        except httpx.HTTPStatusError:
            continue

        logger.info("mf_constituents_fetch amc=%s url=%s", amc_name, file_url)
        file_bytes = resp.content
        content_type = resp.headers.get("content-type", "")
        if "html" in content_type:
            logger.warning(
                "mf_constituents_fetch amc=%s url=%s returned text/html — CDN 404 or file not published yet",
                amc_name,
                file_url,
            )
            continue
        if "spreadsheetml" in content_type or file_url.lower().endswith(".xlsx"):
            parsed = _parse_sebi_xlsx(file_bytes, amc_name)
        else:
            parsed = _parse_sebi_csv(file_bytes.decode("utf-8", errors="replace"), amc_name)

        if not parsed:
            logger.warning("mf_constituents_fetch amc=%s parsed 0 rows from %s", amc_name, file_url)
            return 0, 0

        return await _upsert_constituents(parsed, amc_name)

    logger.warning(
        "mf_constituents_fetch amc=%s no disclosure file found (tried template)", amc_name
    )
    return 0, 0


async def _process_amc_json_api(
    client: httpx.AsyncClient,
    amc_name: str,
    api_url_template: str,
    zip_member_pattern: str,
) -> tuple[int, int]:
    """Fetch the SEBI disclosure for AMCs that publish a JSON API returning ZIP URLs.

    Strategy (confirmed for UTI Drupal CMS):
      1. GET ``api_url_template.format(year=YYYY)`` — returns a JSON array of rows,
         each row has a ``file`` field containing a CloudFront/CDN ZIP URL.
      2. Match the row whose ``month`` field equals the target month name.
      3. Download the ZIP; extract the member whose filename contains
         ``zip_member_pattern`` (e.g. "Sebi Exposure").
      4. Parse the extracted XLSX with ``_parse_sebi_xlsx()``.

    Tries 1 month back first, then 2 months back (handles publication lag).
    """
    now = datetime.now(UTC)
    for months_back in (1, 2):
        target = (now.replace(day=1) - timedelta(days=months_back * 28)).replace(day=1)
        target_month = target.strftime("%B")  # e.g. "May"
        target_year = target.strftime("%Y")  # e.g. "2026"

        api_url = api_url_template.format(year=target_year)
        try:
            resp = await client.get(
                api_url,
                headers={"User-Agent": "DhanRadar/1.0 (research; contact@dhanradar.com)"},
            )
            resp.raise_for_status()
        except Exception:  # noqa: BLE001
            logger.debug(
                "mf_constituents_fetch amc=%s json-api call failed url=%s",
                amc_name,
                api_url,
                exc_info=True,
            )
            continue

        try:
            data = resp.json()
        except Exception:  # noqa: BLE001
            logger.warning(
                "mf_constituents_fetch amc=%s json-api response is not JSON url=%s",
                amc_name,
                api_url,
            )
            continue

        # UTI (and possibly others) wrap the rows list under a top-level key.
        if isinstance(data, dict):
            data = data.get("rows", data.get("data", []))

        if not isinstance(data, list):
            logger.warning(
                "mf_constituents_fetch amc=%s json-api returned non-list type=%s",
                amc_name,
                type(data),
            )
            continue

        rows_json: list[dict] = data

        # Find the row for the target month (case-insensitive).
        zip_url: str | None = None
        for row in rows_json:
            row_month = str(row.get("month", row.get("Month", ""))).strip()
            if row_month.lower() == target_month.lower():
                zip_url = str(row.get("file", row.get("File", row.get("url", ""))))
                break

        if not zip_url:
            logger.debug(
                "mf_constituents_fetch amc=%s json-api: no row for month=%s year=%s (rows=%d)",
                amc_name,
                target_month,
                target_year,
                len(rows_json),
            )
            continue

        # Download the ZIP.
        logger.info("mf_constituents_fetch amc=%s downloading zip=%s", amc_name, zip_url)
        try:
            zip_resp = await client.get(
                zip_url,
                headers={"User-Agent": "DhanRadar/1.0 (research; contact@dhanradar.com)"},
            )
            zip_resp.raise_for_status()
        except Exception:  # noqa: BLE001
            logger.warning(
                "mf_constituents_fetch amc=%s zip download failed url=%s",
                amc_name,
                zip_url,
                exc_info=True,
            )
            continue

        # Extract the target XLSX member from the ZIP.
        try:
            with zipfile.ZipFile(io.BytesIO(zip_resp.content)) as zf:
                member_names = zf.namelist()
                target_member: str | None = None
                for name in member_names:
                    if zip_member_pattern.lower() in name.lower() and name.lower().endswith(
                        (".xlsx", ".xls")
                    ):
                        target_member = name
                        break
                if target_member is None:
                    # Fall back to first XLSX in archive.
                    for name in member_names:
                        if name.lower().endswith((".xlsx", ".xls")):
                            target_member = name
                            break
                if target_member is None:
                    logger.warning(
                        "mf_constituents_fetch amc=%s zip has no xlsx member (members=%s)",
                        amc_name,
                        member_names,
                    )
                    continue
                file_bytes = zf.read(target_member)
        except zipfile.BadZipFile:
            logger.warning(
                "mf_constituents_fetch amc=%s zip response is not a valid ZIP url=%s",
                amc_name,
                zip_url,
            )
            continue

        parsed = _parse_sebi_xlsx(file_bytes, amc_name)
        if not parsed:
            logger.warning(
                "mf_constituents_fetch amc=%s parsed 0 rows from zip member=%s",
                amc_name,
                target_member,
            )
            return 0, 0

        return await _upsert_constituents(parsed, amc_name)

    logger.warning(
        "mf_constituents_fetch amc=%s no disclosure file found (tried json-api)", amc_name
    )
    return 0, 0


async def _process_amc(
    client: httpx.AsyncClient, browser: Any, amc_name: str, discovery_url: str
) -> tuple[int, int]:
    """Discover and parse the latest monthly disclosure file for one AMC.

    Returns (rows_upserted, aum_updates).
    """
    # Step 1: Discover the latest disclosure file URL via Playwright.
    file_url = await _discover_url_playwright(browser, discovery_url, amc_name)
    if file_url is None:
        logger.warning("mf_constituents_fetch amc=%s no disclosure file found", amc_name)
        return 0, 0

    # Step 2: Download the file bytes.
    logger.info("mf_constituents_fetch amc=%s fetching %s", amc_name, file_url)
    resp = await client.get(
        file_url,
        headers={"User-Agent": "DhanRadar/1.0 (research; contact@dhanradar.com)"},
    )
    resp.raise_for_status()

    content_type = resp.headers.get("content-type", "")
    file_bytes = resp.content

    # Step 3: Parse the file into constituent rows.
    if "spreadsheetml" in content_type or file_url.lower().endswith((".xlsx", ".xls")):
        parsed = _parse_sebi_xlsx(file_bytes, amc_name)
    elif "csv" in content_type or file_url.lower().endswith(".csv"):
        parsed = _parse_sebi_csv(file_bytes.decode("utf-8", errors="replace"), amc_name)
    else:
        # Try XLSX first, fall back to CSV.
        try:
            parsed = _parse_sebi_xlsx(file_bytes, amc_name)
        except Exception:  # noqa: BLE001
            parsed = _parse_sebi_csv(file_bytes.decode("utf-8", errors="replace"), amc_name)

    if not parsed:
        logger.warning("mf_constituents_fetch amc=%s parsed 0 rows", amc_name)
        return 0, 0

    # Step 4: Resolve scheme names → ISINs and upsert.
    return await _upsert_constituents(parsed, amc_name)


async def _discover_url_playwright(browser: Any, discovery_url: str, amc_name: str) -> str | None:
    """Render the AMC disclosure SPA with Playwright and extract the latest XLSX/CSV URL.

    Caches the discovered URL in Redis for 25 days (key mf:disclosure_url:{amc}:{YYYY-MM})
    so Playwright runs once per month per AMC on cache miss only.
    """
    from dhanradar.redis_client import get_redis

    now = datetime.now(UTC)
    cache_key = f"mf:disclosure_url:{amc_name}:{now.strftime('%Y-%m')}"

    redis = get_redis()
    cached = await redis.get(cache_key)
    if cached:
        return cached.decode() if isinstance(cached, bytes) else str(cached)

    page = await browser.new_page()
    try:
        await page.goto(discovery_url, wait_until="networkidle", timeout=60_000)
        hrefs: list[str] = await page.evaluate(
            "() => Array.from(document.querySelectorAll('a[href]')).map(a => a.href)"
        )
        # Prefer links labelled portfolio/disclosure; fall back to any xlsx/xls/csv.
        candidates = [
            h
            for h in hrefs
            if h.lower().endswith((".xlsx", ".xls", ".csv"))
            and ("portfolio" in h.lower() or "disclosure" in h.lower())
        ]
        if not candidates:
            candidates = [h for h in hrefs if h.lower().endswith((".xlsx", ".xls", ".csv"))]
        if not candidates:
            logger.warning(
                "mf_constituents_fetch amc=%s playwright found no xlsx/xls/csv links at %s",
                amc_name,
                discovery_url,
            )
            return None
        url = candidates[0]
        await redis.set(cache_key, url, ex=25 * 86400)
        return url
    except Exception:  # noqa: BLE001
        logger.debug(
            "mf_constituents_fetch amc=%s playwright discovery failed url=%s",
            amc_name,
            discovery_url,
            exc_info=True,
        )
        return None
    finally:
        await page.close()


async def _discover_all_urls_playwright(
    browser: Any, discovery_url: str, amc_name: str
) -> list[str]:
    """Like _discover_url_playwright but returns ALL matching file URLs.

    Used for AMCs like MIRAE that publish one XLSX per scheme on a single SPA page.
    Caches the full list as a JSON string in Redis for 25 days.
    """
    import json

    from dhanradar.redis_client import get_redis

    now = datetime.now(UTC)
    cache_key = f"mf:disclosure_urls:{amc_name}:{now.strftime('%Y-%m')}"

    redis = get_redis()
    cached = await redis.get(cache_key)
    if cached:
        raw = cached.decode() if isinstance(cached, bytes) else str(cached)
        try:
            return json.loads(raw)
        except Exception:  # noqa: BLE001
            return [raw] if raw else []

    page = await browser.new_page()
    try:
        await page.goto(discovery_url, wait_until="networkidle", timeout=60_000)
        hrefs: list[str] = await page.evaluate(
            "() => Array.from(document.querySelectorAll('a[href]')).map(a => a.href)"
        )
        candidates = [
            h
            for h in hrefs
            if h.lower().endswith((".xlsx", ".xls", ".csv"))
            and ("portfolio" in h.lower() or "disclosure" in h.lower())
        ]
        if not candidates:
            candidates = [h for h in hrefs if h.lower().endswith((".xlsx", ".xls", ".csv"))]
        # Deduplicate while preserving order.
        candidates = list(dict.fromkeys(candidates))
        if not candidates:
            logger.warning(
                "mf_constituents_fetch amc=%s playwright found no xlsx/xls/csv links at %s",
                amc_name,
                discovery_url,
            )
            return []
        logger.info(
            "mf_constituents_fetch amc=%s playwright found %d file links", amc_name, len(candidates)
        )
        await redis.set(cache_key, json.dumps(candidates), ex=25 * 86400)
        return candidates
    except Exception:  # noqa: BLE001
        logger.debug(
            "mf_constituents_fetch amc=%s playwright multi-url discovery failed url=%s",
            amc_name,
            discovery_url,
            exc_info=True,
        )
        return []
    finally:
        await page.close()


async def _process_amc_multi(
    client: httpx.AsyncClient, browser: Any, amc_name: str, discovery_url: str
) -> tuple[int, int]:
    """Like _process_amc but processes ALL discovered files (for AMCs like MIRAE).

    Returns (total_rows_upserted, total_aum_updates).
    """
    file_urls = await _discover_all_urls_playwright(browser, discovery_url, amc_name)
    if not file_urls:
        logger.warning("mf_constituents_fetch amc=%s no files found (multi)", amc_name)
        return 0, 0

    total_rows = 0
    total_aum = 0
    for file_url in file_urls:
        logger.info("mf_constituents_fetch amc=%s multi fetching %s", amc_name, file_url)
        try:
            resp = await client.get(
                file_url,
                headers={"User-Agent": "DhanRadar/1.0 (research; contact@dhanradar.com)"},
            )
            resp.raise_for_status()
        except Exception:  # noqa: BLE001
            logger.warning(
                "mf_constituents_fetch amc=%s failed to fetch %s", amc_name, file_url, exc_info=True
            )
            continue

        content_type = resp.headers.get("content-type", "")
        file_bytes = resp.content

        if "spreadsheetml" in content_type or file_url.lower().endswith((".xlsx", ".xls")):
            parsed = _parse_sebi_xlsx(file_bytes, amc_name)
        elif "csv" in content_type or file_url.lower().endswith(".csv"):
            parsed = _parse_sebi_csv(file_bytes.decode("utf-8", errors="replace"), amc_name)
        else:
            try:
                parsed = _parse_sebi_xlsx(file_bytes, amc_name)
            except Exception:  # noqa: BLE001
                parsed = _parse_sebi_csv(file_bytes.decode("utf-8", errors="replace"), amc_name)

        if not parsed:
            logger.debug("mf_constituents_fetch amc=%s parsed 0 rows from %s", amc_name, file_url)
            continue

        rows, aum_cnt = await _upsert_constituents(parsed, amc_name)
        total_rows += rows
        total_aum += aum_cnt

    return total_rows, total_aum


def _normalize_col(name: str) -> str:
    """Lowercase + strip a column header for loose matching."""
    return name.lower().strip()


async def _discover_all_urls_static(
    client: httpx.AsyncClient,
    url: str,
    amc_name: str,
    target_month: date | None = None,
) -> list[str]:
    """Discover all XLSX download links from a plain HTML disclosure page.

    Used for AMCs like MIRAE whose landing page renders static HTML (no JS).
    If target_month is provided, links are filtered to those whose URL contains
    the abbreviated month name and year (e.g. "may" + "2026").
    """
    import re
    from urllib.parse import urljoin

    try:
        resp = await client.get(
            url,
            headers={"User-Agent": "DhanRadar/1.0 (research; contact@dhanradar.com)"},
        )
        resp.raise_for_status()
    except Exception:  # noqa: BLE001
        logger.warning(
            "mf_constituents_fetch amc=%s static discovery failed url=%s",
            amc_name,
            url,
            exc_info=True,
        )
        return []

    raw_hrefs = re.findall(r'href=["\']([^"\']+\.xlsx[^"\']*)["\']', resp.text, re.IGNORECASE)
    links: list[str] = list(dict.fromkeys(urljoin(url, h) for h in raw_hrefs))

    if target_month and links:
        month_abbr = target_month.strftime("%b").lower()  # e.g. "may"
        year_str = target_month.strftime("%Y")  # e.g. "2026"
        filtered = [lnk for lnk in links if month_abbr in lnk.lower() and year_str in lnk]
        if filtered:
            links = filtered

    logger.info(
        "mf_constituents_fetch amc=%s static discovery found %d links url=%s",
        amc_name,
        len(links),
        url,
    )
    return links


async def _process_amc_static_multi(
    client: httpx.AsyncClient, amc_name: str, discovery_url: str
) -> tuple[int, int]:
    """Download and parse all disclosure files for an AMC with a plain HTML index page.

    Tries the previous month first, then 2 months back (SEBI publication lag).
    Used for MIRAE, which publishes one XLSX per scheme on a static page.
    """
    now = datetime.now(UTC)
    for months_back in (1, 2):
        target = (now.replace(day=1) - timedelta(days=months_back * 28)).replace(day=1)
        file_urls = await _discover_all_urls_static(
            client, discovery_url, amc_name, target_month=target.date()
        )
        if not file_urls:
            logger.debug(
                "mf_constituents_fetch amc=%s static multi: no links for month=%s",
                amc_name,
                target.strftime("%B %Y"),
            )
            continue

        logger.info(
            "mf_constituents_fetch amc=%s static multi: %d scheme files for %s",
            amc_name,
            len(file_urls),
            target.strftime("%B %Y"),
        )
        total_rows = 0
        total_aum = 0
        parsed_files = 0

        for file_url in file_urls:
            try:
                resp = await client.get(
                    file_url,
                    headers={"User-Agent": "DhanRadar/1.0 (research; contact@dhanradar.com)"},
                )
                resp.raise_for_status()
            except Exception:  # noqa: BLE001
                logger.warning(
                    "mf_constituents_fetch amc=%s failed to fetch %s",
                    amc_name,
                    file_url,
                    exc_info=True,
                )
                continue

            content_type = resp.headers.get("content-type", "")
            file_bytes = resp.content

            if "spreadsheetml" in content_type or file_url.lower().endswith((".xlsx", ".xls")):
                parsed = _parse_sebi_xlsx(file_bytes, amc_name)
            elif "csv" in content_type or file_url.lower().endswith(".csv"):
                parsed = _parse_sebi_csv(file_bytes.decode("utf-8", errors="replace"), amc_name)
            else:
                try:
                    parsed = _parse_sebi_xlsx(file_bytes, amc_name)
                except Exception:  # noqa: BLE001
                    parsed = _parse_sebi_csv(file_bytes.decode("utf-8", errors="replace"), amc_name)

            if not parsed:
                logger.debug(
                    "mf_constituents_fetch amc=%s parsed 0 rows from %s", amc_name, file_url
                )
                continue

            rows, aum_cnt = await _upsert_constituents(parsed, amc_name)
            total_rows += rows
            total_aum += aum_cnt
            parsed_files += 1

        logger.info(
            "mf_constituents_fetch amc=%s static multi: parsed %d/%d files rows=%d",
            amc_name,
            parsed_files,
            len(file_urls),
            total_rows,
        )
        return total_rows, total_aum

    logger.warning("mf_constituents_fetch amc=%s static multi: no disclosure files found", amc_name)
    return 0, 0


class _XlrdSheetShim:
    """Wraps one xlrd `Sheet` to expose the single openpyxl worksheet method
    `_parse_sebi_xlsx` actually calls — `iter_rows(values_only=True)` — so the
    SAME row-parsing loop below runs unchanged for legacy binary .xls files.
    xlrd already returns "" (not None) for a blank cell and raw numbers as
    floats (not formatted strings); both convert identically through the
    loop's own `str(c).strip() if c is not None else ""` either way.
    """

    def __init__(self, sheet: Any) -> None:
        self._sheet = sheet

    def iter_rows(self, values_only: bool = True) -> Any:
        for r in range(self._sheet.nrows):
            yield tuple(self._sheet.row_values(r))


class _XlrdWorkbookShim:
    """Wraps an xlrd `Book` to expose the two openpyxl workbook surfaces
    `_parse_sebi_xlsx` actually uses — `.sheetnames` and `wb[name]` — so the
    xlrd (legacy binary .xls) and openpyxl (.xlsx) code paths share one parser."""

    def __init__(self, book: Any) -> None:
        self._book = book
        self.sheetnames = book.sheet_names()

    def __getitem__(self, name: str) -> _XlrdSheetShim:
        return _XlrdSheetShim(self._book.sheet_by_name(name))


def _parse_sebi_xlsx(file_bytes: bytes, amc_name: str) -> list[dict]:
    """Parse a SEBI-format monthly portfolio disclosure XLSX.

    SEBI circular SEBI/HO/IMD/IMD-II DOF3/P/CIR/2021/024 mandates a standard
    format. Column names vary slightly per AMC; we match loosely.

    Returns list of dicts with keys:
        scheme_name, constituent_name, constituent_isin,
        sector, rating, weight_pct, market_value_cr, as_of_month
    """
    import openpyxl  # lazily imported — not installed everywhere

    try:
        wb: Any = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except zipfile.BadZipFile:
        # openpyxl only reads the zip-based .xlsx container. A genuine legacy
        # binary .xls (OLE2/CFBF magic `D0 CF 11 E0` — e.g. ABSL's ~107-sheet
        # monthly portfolio, confirmed 2026-07-08) raises BadZipFile here.
        # Fall back to xlrd via the shim above so this SAME parsing loop runs
        # unchanged for both formats — never a second parser. A file that is
        # NEITHER real format (confirmed 2026-07-08: ~52 SBI files are an AMC
        # website "Fund Details" HTML page mislabeled ".xls", not a
        # spreadsheet at all — a DIFFERENT data source, tracked separately,
        # not a portfolio disclosure) fails xlrd too and re-raises, so the
        # caller marks the file failed with an honest parse error instead of
        # silently extracting zero rows.
        import xlrd

        book = xlrd.open_workbook(file_contents=file_bytes)
        wb = _XlrdWorkbookShim(book)

    result: list[dict] = []

    current_scheme: str | None = None
    as_of_month: date | None = None

    # NIPPON May-2026+ multi-sheet format: "Index" sheet maps 2-letter codes → scheme names.
    nippon_code_map: dict[str, str] = {}
    if amc_name == "NIPPON":
        for sn in wb.sheetnames:
            if sn.lower() == "index":
                import re as _re

                for irow in wb[sn].iter_rows(values_only=True):
                    cells = [str(c).strip() if c is not None else "" for c in irow]
                    non_empty = [c for c in cells if c and c.lower() not in ("none", "")]
                    if len(non_empty) >= 2 and _re.fullmatch(r"[A-Za-z]{2,5}", non_empty[0]):
                        nippon_code_map[non_empty[0].upper()] = non_empty[1]
                break
        if nippon_code_map:
            logger.info("mf_constituents_fetch NIPPON index map: %d entries", len(nippon_code_map))

    for sheet in wb.sheetnames:
        # Skip NIPPON's "Index" sheet — it's a code→name lookup, not portfolio data.
        if amc_name == "NIPPON" and sheet.lower() == "index":
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        col_map: dict[str, int] = {}
        rows_since_header = 0  # holdings extracted since col_map was (re)built
        current_scheme: str | None = None  # Reset per sheet for per-scheme files.

        # For per-scheme files (e.g. MIRAE), infer scheme name from sheet name
        # if the sheet name looks like a scheme (contains "fund", "plan", etc.).
        # Also accept "Portfolio" which is common in MIRAE files.
        sheet_scheme: str | None = None
        # NIPPON multi-sheet: map 2-letter abbreviation to full scheme name via index.
        if amc_name == "NIPPON" and nippon_code_map.get(sheet.upper()):
            sheet_scheme = nippon_code_map[sheet.upper()]
        elif sheet and any(
            kw in sheet.lower()
            for kw in (
                "fund",
                "scheme",
                "plan",
                "etf",
                "index",
                "growth",
                "idcw",
                "direct",
                "regular",
                "portfolio",
            )
        ):
            sheet_scheme = sheet

        for idx, row in enumerate(rows):  # noqa: B007
            row_strs = [str(c).strip() if c is not None else "" for c in row]

            # Detect as_of_month from header rows (e.g. "Portfolio as on 31-May-2025"
            # or "AS OF 31/05/2026").
            joined = " ".join(row_strs).lower()
            if as_of_month is None and (
                "portfolio as on" in joined
                or "as at" in joined
                or "month end" in joined
                or "as of" in joined
                or "as on" in joined
            ):
                import re

                # Try "DD-Mon-YYYY" or "DD Mon YYYY" format first.
                date_m = re.search(
                    r"(\d{1,2})[- ](\w+)[- ](\d{4})",
                    " ".join(row_strs),
                    re.IGNORECASE,
                )
                if date_m:
                    try:
                        from datetime import datetime as _dt

                        as_of_month = (
                            _dt.strptime(f"01-{date_m.group(2)[:3]}-{date_m.group(3)}", "%d-%b-%Y")
                            .date()
                            .replace(day=1)
                        )
                    except ValueError:
                        pass
                # Fallback: "DD/MM/YYYY" format (UTI style).
                if as_of_month is None:
                    slash_m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", " ".join(row_strs))
                    if slash_m:
                        try:
                            from datetime import date as _date

                            as_of_month = _date(int(slash_m.group(3)), int(slash_m.group(2)), 1)
                        except ValueError:
                            pass
                # Fallback: "Month DD, YYYY" format (MIRAE style: "May 31, 2026").
                # ICICI_PRU's own banners omit the space after the comma
                # ("Portfolio as on May 31,2026" / "Figures as on Mar 31,2026" —
                # confirmed 2026-07-07 across every sampled ICICI per-scheme
                # file), which the strict `,?\s+` used to reject outright —
                # every row in the file then kept `as_of_month=None` and
                # `_upsert_constituents` silently drops any row with no
                # as_of_month, producing zero_rows_upserted_scheme_unresolved
                # even though the scheme name resolves fine. `,?\s*` accepts
                # both the spaced (MIRAE) and unspaced (ICICI) comma form.
                if as_of_month is None:
                    month_m = re.search(r"([A-Za-z]+)\s+(\d{1,2}),?\s*(\d{4})", " ".join(row_strs))
                    if month_m:
                        try:
                            from datetime import datetime as _dt

                            as_of_month = (
                                _dt.strptime(
                                    f"01-{month_m.group(1)[:3]}-{month_m.group(3)}", "%d-%b-%Y"
                                )
                                .date()
                                .replace(day=1)
                            )
                        except ValueError:
                            pass

            # Detect scheme name rows (usually bold / standalone text rows).
            # HDFC's per-scheme disclosure files (manual-ingest inbox, ~88 files) title
            # each sheet with a scheme-name banner MERGED across MOST (not necessarily
            # all) columns — e.g. cols 0-9 all literally read "HDFC Liquid Fund (An Open
            # ended Liquid scheme)" while cols 10-11 hold unrelated trailing metadata
            # ("Income", "Hybrid"). openpyxl's read_only=True mode returns the SAME value
            # for every cell in a merge instead of None-padding it, so `non_empty` here
            # has the banner text repeated N times PLUS a couple of unrelated trailing
            # values — not the single entry the original AMCs' banner rows produce.
            # Using the row's MOST COMMON value (appearing >= 2x, or the sole value when
            # there's exactly one) as the candidate accepts this shape without weakening
            # the check for anyone else: a genuine holdings row never has any value
            # repeated across 2+ of its columns (name/ISIN/quantity/value/etc. are all
            # distinct), so this can't misfire on real data.
            non_empty = [s for s in row_strs if s and s.lower() not in ("none", "")]
            if non_empty and not col_map:
                _val_counts = Counter(non_empty)
                _top_val, _top_count = _val_counts.most_common(1)[0]
                if _top_count >= 2 or len(non_empty) == 1:
                    candidate = _top_val
                else:
                    candidate = ""
                    # SBI's per-scheme "Portfolio Details" sheets (manual-ingest
                    # inbox, confirmed 2026-07-07 — 461 files) put the scheme
                    # name in a SEPARATE cell from its "SCHEME NAME :" label
                    # (e.g. row = ['', '', 'SCHEME NAME :', 'SBI-ETF Nifty 50',
                    # ...]) rather than UTI's single-cell "SCHEME:<name>" form
                    # already handled above — a genuine 2-distinct-value row,
                    # so the merged-banner heuristic above correctly rejects it
                    # (candidate stays ""). Recover it explicitly: find the
                    # label cell and take the next non-empty cell in the row.
                    import re as _re_schemename  # local import — `re` is shadowed as a
                    # function-local name elsewhere in this function (other branches
                    # do `import re` inline), so the module-level import can't be
                    # relied on here; alias avoids re-triggering that shadowing.

                    for _ci, _cell in enumerate(row_strs):
                        if _re_schemename.fullmatch(
                            r"scheme\s*name\s*:?", _cell.strip(), _re_schemename.IGNORECASE
                        ):
                            for _nv in row_strs[_ci + 1 :]:
                                if _nv and _nv.lower() not in ("none", ""):
                                    candidate = _nv.strip()
                                    break
                            break
                    # ABSL's Index-sheet-style per-scheme banner (confirmed
                    # 2026-07-08, ~103 schemes): a genuine 2-distinct-value row
                    # pairing a short fund CODE with the real scheme name
                    # (e.g. ['ABBSEIIF', 'ADITYA BIRLA SUN LIFE BSE INDIA
                    # INFRASTRUCTURE INDEX FUND', ...]) — a fund code never
                    # contains a scheme-name keyword, so when exactly ONE of
                    # the two values does, that one is unambiguously the name.
                    if not candidate and len(non_empty) == 2:
                        _scheme_kws = (
                            "fund",
                            "scheme",
                            "plan",
                            "etf",
                            "index",
                            "growth",
                            "idcw",
                            "direct",
                            "regular",
                        )
                        _kw_hits = [
                            v for v in non_empty if any(kw in v.lower() for kw in _scheme_kws)
                        ]
                        # ICICI's per-scheme files pair a "Figures as on <date>"
                        # cell with a "Fund Size Rs. <n> in Lakhs" cell (confirmed
                        # 2026-07-09, 134 files) — "Fund Size" contains the same
                        # "fund" keyword a real scheme name would, so it wins the
                        # single-keyword-hit check below and overwrites the
                        # correctly-detected current_scheme from an earlier row.
                        # It is never itself a scheme name, so exclude it before
                        # the keyword-hit disambiguation runs.
                        _kw_hits = [v for v in _kw_hits if not _FUND_SIZE_BANNER_RE.match(v)]
                        if len(_kw_hits) == 1:
                            candidate = _kw_hits[0]
                        else:
                            # Ambiguous (0 or both matched) — e.g. a fund CODE
                            # can coincidentally contain "ETF"/"FUND" as a bare
                            # substring ("C10YGETF" ends in "ETF"). Fall back
                            # to whitespace: a fund code is always a single
                            # space-free token; a real AMFI/SEBI scheme name
                            # always has spaces.
                            _spaced = [
                                v
                                for v in non_empty
                                if " " in v.strip() and not _FUND_SIZE_BANNER_RE.match(v)
                            ]
                            if len(_spaced) == 1:
                                candidate = _spaced[0]
            else:
                candidate = ""
            # Reject a candidate that is itself a SEBI section-header /
            # subtotal marker ("a) Listed/awaiting listing...", "Sub Total",
            # etc. — same _SECTION_HEADER_RE already used to keep these out of
            # constituent rows below). Without this, a genuine section header
            # inside a multi-asset-class SBI file (e.g. "a) Mutual Fund Units /
            # Exchange Traded Funds") satisfies the single-value + keyword
            # ("fund") checks below and silently OVERWRITES the real
            # current_scheme mid-file (confirmed 2026-07-07, SBI Multi Asset
            # Allocation Fund.xlsx).
            if candidate and _SECTION_HEADER_RE.search(candidate):
                candidate = ""
            # Reject a bare scheme-TYPE description row (see
            # _SCHEME_TYPE_DESCRIPTION_RE docstring) — confirmed 2026-07-08 in
            # ABSL's legacy .xls files, same overwrite-current_scheme failure
            # mode as the section-header case just above.
            if candidate and _SCHEME_TYPE_DESCRIPTION_RE.search(candidate):
                candidate = ""
            if candidate:
                # Strip "SCHEME:" prefix used by UTI and some other AMCs.
                if candidate.upper().startswith("SCHEME:"):
                    candidate = candidate[7:].strip()
                # Strip "Portfolio of ..." / "Portfolio statement of ..." banner
                # prefixes (KOTAK/EDELWEISS sheet titles, 2026-07-07) so the
                # pg_trgm scheme-name resolution matches the bare scheme name.
                low_candidate = candidate.lower()
                for banner in ("portfolio statement of ", "portfolio of "):
                    if low_candidate.startswith(banner):
                        candidate = candidate[len(banner) :].strip()
                        break
                # Strip a trailing "as on <date>" clause from the same banners
                # ("Kotak X Fund as on 31-May-2026" / "EDELWEISS Y AS ON MAY 31,
                # 2026") — the date is period metadata, not part of the scheme
                # name the pg_trgm ISIN resolution should match.
                import re as _re_ason

                candidate = _re_ason.sub(
                    r"\s+as\s+on\s+.+$", "", candidate, flags=_re_ason.IGNORECASE
                ).strip()
                # Reject noise rows like "SCHEME CODE002STARTS" that aren't real names.
                if any(kw in candidate.upper() for kw in ("CODE002", "STARTS", "ENDS")):
                    candidate = ""
                # Reject rows starting with parentheses or other indicators of descriptions (e.g. MIRAE).
                if candidate and not candidate[0].isalnum():
                    candidate = ""
                # HDFC appends a boilerplate SEBI scheme-TYPE disclaimer in parens, e.g.
                # "(An Open ended Liquid scheme)" / "(An open ended Scheme replicating/
                # tracking NIFTY 1D Rate Index TRI)" — never a Direct/Regular/Growth
                # disambiguator (unlike some other AMCs' legitimate "(Direct Plan)"
                # qualifiers, which this deliberately does NOT touch), so stripping it
                # only improves the pg_trgm scheme-name match in _resolve_scheme_isins,
                # never loses information needed to pick the right ISIN.
                if candidate:
                    import re as _re_local  # local import — `re` is shadowed as a

                    # function-local name elsewhere in this function (other branches
                    # do `import re` inline), so the module-level import can't be
                    # relied on here; alias avoids re-triggering that shadowing.
                    # Close-ended schemes are correctly written "(A Close
                    # Ended...)" not "(An Close Ended...)" per English grammar
                    # (a/an depends on the following word's sound) — confirmed
                    # 2026-07-09 in HDFC's FMP and "Charity Fund for Cancer
                    # Cure" close-ended disclosures, where the literal "an"
                    # requirement left the whole tenure/risk disclaimer
                    # attached and broke pg_trgm scheme-name resolution.
                    candidate = _re_local.sub(
                        r"\s*\((?:a|an)\s+(open|close)[\s-]*ended.*$",
                        "",
                        candidate,
                        flags=_re_local.IGNORECASE,
                    ).strip()
                # Scheme rows often start with scheme-type keywords.
                if candidate and any(
                    kw in candidate.lower()
                    for kw in (
                        "fund",
                        "scheme",
                        "plan",
                        "etf",
                        "index",
                        "growth",
                        "idcw",
                        "direct",
                        "regular",
                    )
                ):
                    current_scheme = candidate

            # Fallback for per-scheme files (e.g. MIRAE): use sheet name as scheme if no scheme detected yet.
            # For MIRAE, also use sheet name if header was detected but no explicit scheme row was found.
            if not current_scheme and sheet_scheme and col_map:
                current_scheme = sheet_scheme
                if amc_name == "MIRAE":
                    logger.info(
                        "mf_constituents_fetch MIRAE using sheet_scheme: '%s'", current_scheme
                    )
            # For MIRAE specifically, if we have a header but still no scheme name,
            # and the sheet looks like a scheme name, use the sheet name directly.
            if not current_scheme and amc_name == "MIRAE" and sheet and col_map:
                current_scheme = sheet
                logger.info(
                    "mf_constituents_fetch MIRAE fallback using sheet name directly: '%s'", sheet
                )

            # Detect header row (contains "Name of Instrument" or similar).
            if not col_map and any(
                "name" in s.lower()
                and ("instrument" in s.lower() or "security" in s.lower() or "stock" in s.lower())
                for s in row_strs
            ):
                for ci, cell in enumerate(row_strs):
                    col_map[_normalize_col(cell)] = ci
                rows_since_header = 0
                continue

            # Data rows — only after header detected.
            if col_map and current_scheme:
                row_dict = _extract_sebi_row(
                    row_strs, col_map, current_scheme, amc_name, as_of_month
                )
                if row_dict:
                    result.append(row_dict)
                    rows_since_header += 1

                # Reset on blank rows (new scheme section upcoming) — but only
                # AFTER at least one holding was extracted under this header.
                # EDELWEISS (2026-07-07) pads 1-2 blank rows between the header
                # and the first holding; resetting on those wiped col_map before
                # any data row was ever read (whole workbook parsed to 0 rows).
                if not any(s for s in row_strs if s and s.lower() not in ("none", "")):
                    if rows_since_header:
                        col_map = {}
                        rows_since_header = 0

    return result


def _parse_sebi_csv(csv_text: str, amc_name: str) -> list[dict]:
    """Parse a SEBI-format monthly portfolio disclosure CSV.

    Same column-matching logic as the XLSX parser.
    """
    result: list[dict] = []
    current_scheme: str | None = None
    as_of_month: date | None = None
    col_map: dict[str, int] = {}

    reader = csv.reader(io.StringIO(csv_text))
    for row in reader:
        row_strs = [c.strip() for c in row]

        joined = " ".join(row_strs).lower()
        if as_of_month is None and (
            "portfolio as on" in joined
            or "as at" in joined
            or "as of" in joined
            or "as on" in joined
            or "month end" in joined
        ):
            import re

            date_m = re.search(r"(\d{1,2})[- ](\w+)[- ](\d{4})", " ".join(row_strs), re.IGNORECASE)
            if date_m:
                try:
                    from datetime import datetime as _dt

                    as_of_month = (
                        _dt.strptime(f"01-{date_m.group(2)[:3]}-{date_m.group(3)}", "%d-%b-%Y")
                        .date()
                        .replace(day=1)
                    )
                except ValueError:
                    pass
            # Fallback: "Month DD, YYYY" format (MIRAE style: "May 31, 2026").
            # `,?\s*` (not `,?\s+`) so an unspaced comma form (ICICI_PRU's own
            # banners: "Portfolio as on May 31,2026") still matches — see the
            # matching comment in _parse_sebi_xlsx above, same fix mirrored here
            # for CSV-format disclosures.
            if as_of_month is None:
                month_m = re.search(r"([A-Za-z]+)\s+(\d{1,2}),?\s*(\d{4})", " ".join(row_strs))
                if month_m:
                    try:
                        from datetime import datetime as _dt

                        as_of_month = (
                            _dt.strptime(
                                f"01-{month_m.group(1)[:3]}-{month_m.group(3)}", "%d-%b-%Y"
                            )
                            .date()
                            .replace(day=1)
                        )
                    except ValueError:
                        pass

        non_empty = [s for s in row_strs if s]
        if len(non_empty) == 1 and not col_map:
            candidate = non_empty[0]
            if any(
                kw in candidate.lower()
                for kw in (
                    "fund",
                    "scheme",
                    "plan",
                    "etf",
                    "index",
                    "growth",
                    "idcw",
                    "direct",
                    "regular",
                )
            ):
                current_scheme = candidate

        if not col_map and any(
            "name" in s.lower() and ("instrument" in s.lower() or "security" in s.lower())
            for s in row_strs
        ):
            for ci, cell in enumerate(row_strs):
                col_map[_normalize_col(cell)] = ci
            continue

        if col_map and current_scheme:
            row_dict = _extract_sebi_row(row_strs, col_map, current_scheme, amc_name, as_of_month)
            if row_dict:
                result.append(row_dict)

            if not any(s for s in row_strs if s):
                col_map = {}

    return result


# Section-header / subtotal / grand-total rows that some AMC disclosure sheets
# interleave among actual holding rows — e.g. "(a)  Listed/awaiting listing on
# Stock Exchanges", "(b) Unlisted", "Sub Total". Name-pattern backstop for the
# 2026-07 INF789F01WY2 incident (docs/rca/README.md) — structural detection in
# _extract_sebi_row (missing ISIN + no numbers) is the primary guard; this
# catches header rows that DO carry a subtotal weight/value of their own.
_SECTION_HEADER_RE = re.compile(
    r"^\s*\(?[a-z]\)|^\s*(sub\s*)?total|listed/awaiting|^unlisted$",
    re.IGNORECASE,
)

# A bare SEBI scheme-TYPE description row — e.g. ABSL's "An open ended Index
# Fund replicating the BSE India Infrastructure Total Return Index" (its own
# row, no scheme name attached) or SBI's "An Open Ended Exchange Traded
# Scheme" — every SEBI-format scheme-type description starts with this exact
# "An open/close ended ..." phrasing per the mandated circular wording; a real
# scheme name never starts this way (it starts with the AMC's own brand name).
# Confirmed 2026-07-08: this row satisfies the single-value candidate check
# below AND contains "fund"/"scheme", so without this rejection it silently
# overwrites the real (correctly-detected, earlier-row) current_scheme with
# the scheme's TYPE description instead of its NAME.
_SCHEME_TYPE_DESCRIPTION_RE = re.compile(r"^\s*an?\s+(open|close)[\s-]*ended\b", re.IGNORECASE)

# ICICI's per-scheme portfolio sheets always pair an "as on <date>" cell with a
# "Fund Size Rs. <amount> in Lakhs" / "AUM of the Scheme as on <date> ..." cell
# in the same 2-value row (confirmed 2026-07-09, 134 files) — this is AUM
# metadata, never a scheme name, but contains the same "fund" keyword the
# 2-value scheme-name-vs-fund-code disambiguation (see ABSL comment below)
# looks for, so it must be excluded before that check runs.
_FUND_SIZE_BANNER_RE = re.compile(r"^(fund\s+size|aum\s+of\s+the\s+scheme)\b", re.IGNORECASE)

# SEBI sheets sometimes prefix "Name of Instrument" with a short instrument-type
# code, e.g. UTI writes "EQ - ABB INDIA LTD.". 2-4 caps + " - " is never how a
# real company name starts, so this can't over-strip a genuine holding name.
_INSTRUMENT_PREFIX_RE = re.compile(r"^[A-Z]{2,4}\s*-\s*")


def _extract_sebi_row(
    row_strs: list[str],
    col_map: dict[str, int],
    scheme_name: str,
    amc_name: str,
    as_of_month: date | None,
) -> dict | None:
    """Extract one constituent row from a parsed SEBI row using loose column matching.

    Returns None if the row has no constituent name, or is a section-header /
    subtotal / grand-total row rather than an actual holding — see
    docs/rca/README.md (INF789F01WY2, 2026-05, weight_pct sum ~200%).
    §8.4: market_value_cr and weight_pct are taken directly from the file — never computed
    from AMC-level totals.
    """

    # amc_name is carried through for source provenance on the returned dict.
    def _get(keys: list[str]) -> str:
        for key in keys:
            for col_name, ci in col_map.items():
                if key in col_name and ci < len(row_strs):
                    val = row_strs[ci]
                    if val and val.lower() not in ("none", "n/a", "-", ""):
                        return val
        return ""

    constituent_name = _get(
        [
            "name of instrument",
            "name of security",
            "name of stock",
            "name of the instrument",
        ]
    )
    if not constituent_name:
        # KOTAK (2026-07-07): the name header sits in a merged block whose LABEL
        # lands in col 0 while every VALUE lands 2 cols right (openpyxl read-only
        # gives the label to the first merged cell only). Generic recovery: when
        # the mapped name column is empty but the row carries a plausible ISIN,
        # the holding's name is the nearest non-empty cell LEFT of the ISIN cell
        # (the SEBI format always places the name immediately before the ISIN).
        isin_ci = next((ci for col_name, ci in col_map.items() if "isin" in col_name), None)
        if isin_ci is not None and isin_ci < len(row_strs):
            isin_val = row_strs[isin_ci].strip().upper()
            if len(isin_val) == 12 and isin_val[:2].isalpha():
                for ci in range(isin_ci - 1, -1, -1):
                    cell = row_strs[ci].strip()
                    if cell and cell.lower() not in ("none", "n/a", "-"):
                        constituent_name = cell
                        break
    if not constituent_name:
        return None

    # Strip display-only instrument-type prefix (confirmed on UTI's "EQ - " rows).
    constituent_name = _INSTRUMENT_PREFIX_RE.sub("", constituent_name).strip() or constituent_name

    # Skip total/sub-total rows AS HOLDINGS -- but keep them in the returned dict
    # (flagged is_total_row) when they carry a market_value_cr, so the caller can
    # find the scheme's overall AUM. Root-cause fix 2026-07-05: total-labeled rows
    # were previously dropped here unconditionally, silently discarding the very
    # rows that carry AUM -- the real reason `mf_funds.aum_crore` stayed 100% NULL
    # in prod despite `_upsert_constituents` already having AUM-detection code.
    # Deliberately NARROW: only "grand total" / "net assets" -- a bare "total"
    # match is unsafe, since several AMCs (UTI) label EVERY asset-class subtotal
    # "TOTAL:(a) Listed..." / "TOTAL: EQUITY AND EQUITY RELATED" / "TOTAL: MONEY
    # MARKET INSTRUMENTS" etc. within the SAME table shape as real holdings --
    # broadening this to bare "total" was tried and reverted 2026-07-05 after it
    # caused a sub-category subtotal (not the scheme's true AUM) to be written as
    # if it were the grand total for UTI schemes (verified against a real UTI
    # file: the scheme's actual "Total Net Assets" row lives in a differently
    # shaped summary block that isn't captured by this row-level parse at all,
    # so UTI schemes correctly get no aum_map entry from this source rather than
    # a wrong one). NIPPON's literal "GRAND TOTAL" row IS in the same table shape
    # as real holdings and is captured correctly by this narrower match.
    is_total_row = any(kw in constituent_name.lower() for kw in ("grand total", "net assets"))

    isin_col = _get(["isin", "isin code"])

    weight_pct_raw = _get(["% to nav", "% of net assets", "% to net assets", "weight", "% of nav"])
    weight_pct: float | None = None
    if weight_pct_raw:
        try:
            weight_pct = float(weight_pct_raw.replace(",", "").replace("%", "").strip())
        except ValueError:
            pass

    market_value_raw = _get(
        [
            "market value",
            "market val",
            "value (rs. in lakhs)",
            "value (lakhs)",
            "mkt value",
            # NIPPON debt-scheme disclosures header this column "Market/Fair Value
            # ( Rs. in Lacs)" (embedded newline, "/fair" breaks the "market value"
            # substring match above) -- found 2026-07-05 fixing the grand-total AUM
            # extraction; without this the column never resolves for ANY row in
            # this file, not only the grand-total one.
            "market/fair value",
            # HDFC's per-scheme manual-ingest files header this column "Market/
            # Fair Value (Rs. in Lacs.)" — WITH a space after the slash (unlike
            # NIPPON's no-space "Market/Fair Value" above), which breaks that
            # exact substring match too. Found 2026-07-06 fixing HDFC per-scheme
            # AUM extraction (manual-ingest inbox).
            "market/ fair value",
            # UTI hyphenates: "MARKET-VALUE" (found 2026-07-05, same fix).
            "market-value",
        ]
    )
    market_value_cr: float | None = None
    if market_value_raw:
        try:
            # SEBI files report in Lakhs; convert to Crores (÷100).
            market_value_cr = float(market_value_raw.replace(",", "").strip()) / 100.0
        except ValueError:
            pass

    # Skip section-header rows. Structural signal first: a genuine holding
    # always carries an ISIN; a label-only row with no ISIN and no number is
    # pure section text, not a holding. Name-pattern backstop second, for
    # headers that DO carry the section's own subtotal weight/value (the
    # INF789F01WY2 case — see _SECTION_HEADER_RE docstring above). A total-labeled
    # row that DOES carry a market_value_cr is deliberately let through here (not
    # a real holding, but needed downstream for the AUM max-value heuristic above).
    if not isin_col and weight_pct is None and market_value_cr is None:
        return None
    if _SECTION_HEADER_RE.search(constituent_name) and not (
        is_total_row and market_value_cr is not None
    ):
        return None

    sector = _get(["sector", "industry", "industry/sector"])
    rating = _get(["rating", "credit rating", "instrument rating"])

    # Use first-of-month date if as_of_month was parsed, else None (never fabricated).
    effective_month = as_of_month

    return {
        "scheme_name": scheme_name,
        "constituent_name": constituent_name,
        "constituent_isin": isin_col or None,
        "sector": sector or None,
        "rating": rating or None,
        "weight_pct": weight_pct,
        "market_value_cr": market_value_cr,
        "as_of_month": effective_month,
        "source_amc": amc_name,
        "is_total_row": is_total_row,
    }


def _drop_over_covered_funds(constituent_batch: list[dict], amc_name: str) -> list[dict]:
    """Fail-closed fund-level sanity check (ADR-0039).

    Sums weight_pct per isin; a fund whose holdings sum past 105% almost
    certainly still has a section-header/subtotal row leaking through (see
    docs/rca/README.md, INF789F01WY2 2026-05 — 107 rows summed to ~199.66%).
    Rather than write a fund we know is wrong, drop its rows entirely and log
    a structured warning — null-over-wrong-number.
    """
    weight_by_isin: dict[str, float] = {}
    for r in constituent_batch:
        if r["weight_pct"] is not None:
            weight_by_isin[r["isin"]] = weight_by_isin.get(r["isin"], 0.0) + r["weight_pct"]

    bad_isins = {isin for isin, total in weight_by_isin.items() if total > 105}
    for isin in bad_isins:
        logger.warning(
            "mf_constituents_fetch amc=%s isin=%s weight_pct_sum=%.2f exceeds 105%% "
            "— skipping fund, suspected header/subtotal row leak",
            amc_name,
            isin,
            weight_by_isin[isin],
        )
    if not bad_isins:
        return constituent_batch
    return [r for r in constituent_batch if r["isin"] not in bad_isins]


async def _resolve_scheme_isins(scheme_names: set[str], amc_name: str) -> dict[str, str]:
    """Resolve a set of scheme NAMES → ISINs via pg_trgm fuzzy match, restricted
    to the same AMC's funds (by name prefix) to avoid cross-AMC false positives
    (e.g. "UTI - Liquid Fund" vs "HSBC Liquid Fund").

    Extracted out of `_upsert_constituents` (Phase 6 fund-manager rebuild,
    2026-07) so the SAME fuzzy matcher is shared by the constituents pipeline
    and any other scheme-name-keyed source (currently: the UTI JSON-API and
    NIPPON factsheet-PDF fund-manager fetchers in tasks/mf_fund_manager.py) —
    never a second matcher. Behavior is unchanged from the original inline loop.

    Returns {scheme_name: isin} for names that matched with similarity > 0.35
    (or > 0.25 for MIRAE's shorter extracted names); unmatched names are simply
    absent from the returned dict (never guessed).
    """
    from sqlalchemy import text as sa_text

    from dhanradar.db import TaskSessionLocal

    if not scheme_names:
        return {}

    scheme_isin_map: dict[str, str] = {}
    amc_prefix_map = {
        "MIRAE": "Mirae Asset%",
        "PPFAS": "Parag Parikh%",
        "ABSL": "Aditya Birla%",
    }
    amc_prefix = amc_prefix_map.get(amc_name) or (
        amc_name.split("_")[0] + "%"
    )  # "ICICI_PRU" → "ICICI%"
    async with TaskSessionLocal() as db:
        for sname in scheme_names:
            # Use pg_trgm similarity to fuzzy-match scheme names, restricted to
            # same-AMC funds to prevent false positives across AMC name overlap.
            result = await db.execute(
                sa_text(
                    "SELECT isin, scheme_name, similarity(scheme_name, :sname) as sim FROM mf.mf_funds "
                    "WHERE scheme_name ILIKE :prefix "
                    "ORDER BY similarity(scheme_name, :sname) DESC "
                    "LIMIT 3"
                ),
                {"sname": sname, "prefix": amc_prefix},
            )
            rows = result.fetchall()
            if rows:
                # Log top matches for debugging
                if amc_name == "MIRAE":
                    logger.info(
                        "mf_constituents_fetch amc=MIRAE scheme='%s' matches: %s",
                        sname,
                        [(r[1], f"{r[2]:.2f}") for r in rows],
                    )
                # Use first match if similarity > 0.35
                if rows[0][2] > 0.35:
                    scheme_isin_map[sname] = rows[0][0]
                elif amc_name == "MIRAE" and rows[0][2] > 0.25:
                    # MIRAE per-scheme files produce shorter extracted names; relax threshold.
                    logger.info(
                        "mf_constituents_fetch MIRAE scheme '%s' matched via relaxed threshold: %s",
                        sname,
                        [(r[1], f"{r[2]:.2f}") for r in rows[:3]],
                    )
                    scheme_isin_map[sname] = rows[0][0]
                else:
                    logger.debug(
                        "mf_constituents_fetch amc=%s scheme '%s' top match similarity=%.2f (too low)",
                        amc_name,
                        sname,
                        rows[0][2],
                    )
            else:
                logger.debug(
                    "mf_constituents_fetch amc=%s no matches for scheme '%s' with prefix '%s'",
                    amc_name,
                    sname,
                    amc_prefix,
                )
    return scheme_isin_map


async def _upsert_constituents(
    parsed_rows: list[dict], amc_name: str, run_id: int | None = None
) -> tuple[int, int]:
    """Resolve scheme names → ISINs via pg_trgm, upsert constituent rows.

    Returns (rows_upserted, aum_updates).
    §8.4: aum_crore is written from the file's per-scheme net-assets row only;
    never derived from AMC-level totals.
    `run_id`: optional `mf.ingestion_runs.run_id` to stamp onto `mf.aum_history` rows.
    None for all current callers (none of them open an `ingestion_run()` context yet) —
    the column is nullable specifically to allow this.
    """
    from sqlalchemy import func
    from sqlalchemy import text as sa_text
    from sqlalchemy.dialects.postgresql import insert as pg_insert

    from dhanradar.db import TaskSessionLocal
    from dhanradar.models.mf import MfAumHistory, MfFundConstituent

    # Group rows by scheme_name to resolve ISINs in bulk.
    scheme_names: set[str] = {r["scheme_name"] for r in parsed_rows if r.get("scheme_name")}
    if not scheme_names:
        return 0, 0

    # amc_name may be "UTI", "NIPPON", etc.; fund names in mf_funds start with
    # the AMC's short prefix. Shared matcher — see `_resolve_scheme_isins`.
    scheme_isin_map = await _resolve_scheme_isins(scheme_names, amc_name)

    # Resolve ISINs and split into constituent rows vs aum updates.
    constituent_batch: list[dict] = []
    # Map isin -> (net_assets_cr, as_of_month) for aum updates. Takes the MAX
    # market_value_cr among all is_total_row rows per scheme (narrowed to
    # "grand total" / "net assets" only -- see _extract_sebi_row's is_total_row
    # docstring); as_of_month is carried through so aum_as_of reflects the
    # disclosure file's own month, never the ingestion run time (§8.4).
    aum_map: dict[str, tuple[float, date | None]] = {}

    for row in parsed_rows:
        sname = row.get("scheme_name")
        isin = scheme_isin_map.get(sname or "")
        if not isin:
            continue

        if row.get("is_total_row"):
            mv = row.get("market_value_cr")
            if mv is not None and mv > aum_map.get(isin, (-1.0, None))[0]:
                aum_map[isin] = (mv, row.get("as_of_month"))
            continue

        if row.get("as_of_month") is None:
            continue

        constituent_batch.append(
            {
                "isin": isin,
                "constituent_name": row.get("constituent_name", ""),
                "as_of_month": row["as_of_month"],
                "constituent_isin": row.get("constituent_isin"),
                "sector": row.get("sector"),
                "rating": row.get("rating"),
                "weight_pct": row.get("weight_pct"),
                "market_value_cr": row.get("market_value_cr"),
                "source_amc": amc_name,
            }
        )

    # Deduplicate by ON CONFLICT key — some AMC files (e.g. NIPPON) have duplicate
    # rows for the same (isin, constituent_name, as_of_month); a single upsert
    # statement cannot update the same row twice (CardinalityViolationError).
    seen_keys: set[tuple] = set()
    deduped: list[dict] = []
    for r in constituent_batch:
        key = (r["isin"], r["constituent_name"], r["as_of_month"])
        if key not in seen_keys:
            seen_keys.add(key)
            deduped.append(r)
    constituent_batch = deduped

    # Fail-closed fund-level sanity check (ADR-0039, docs/rca/README.md
    # INF789F01WY2 incident): if a fund's weight_pct rows still sum past 105%
    # — an AMC quirk the row filter above doesn't catch — drop that fund's
    # rows entirely rather than write a partially-garbage portfolio. A missing
    # constituents row beats a wrong one.
    constituent_batch = _drop_over_covered_funds(constituent_batch, amc_name)

    rows_upserted = 0
    aum_updates = 0

    async with TaskSessionLocal() as db:
        # Upsert constituent rows in chunks.
        for i in range(0, len(constituent_batch), _UPSERT_CHUNK):
            chunk = constituent_batch[i : i + _UPSERT_CHUNK]
            if not chunk:
                continue
            stmt = (
                pg_insert(MfFundConstituent)
                .values(chunk)
                .on_conflict_do_update(
                    index_elements=["isin", "constituent_name", "as_of_month"],
                    set_={
                        "constituent_isin": pg_insert(MfFundConstituent).excluded.constituent_isin,
                        "sector": pg_insert(MfFundConstituent).excluded.sector,
                        "rating": pg_insert(MfFundConstituent).excluded.rating,
                        "weight_pct": pg_insert(MfFundConstituent).excluded.weight_pct,
                        "market_value_cr": pg_insert(MfFundConstituent).excluded.market_value_cr,
                        "source_amc": pg_insert(MfFundConstituent).excluded.source_amc,
                        "ingested_at": func.now(),
                    },
                )
            )
            await db.execute(stmt)
            rows_upserted += len(chunk)
        await db.commit()

        # Update aum_crore from per-scheme net-assets (§8.4 — genuine scheme-level data only).
        for isin, (net_assets_cr, as_of_month) in aum_map.items():
            await db.execute(
                sa_text(
                    "UPDATE mf.mf_funds SET aum_crore = :v, aum_as_of = :as_of WHERE isin = :isin"
                ),
                {"v": net_assets_cr, "as_of": as_of_month, "isin": isin},
            )
            aum_updates += 1
            # Also preserve the month-over-month history (mf.aum_history) so the
            # aum_change What-Changed event can diff the latest two rows — same
            # transaction, no second query pass. as_of_month is required (NOT NULL)
            # on this table; a row with no disclosure month can't be dated, so it's
            # skipped here (mf_funds.aum_crore above is still updated regardless).
            if as_of_month is not None:
                stmt = (
                    pg_insert(MfAumHistory)
                    .values(
                        isin=isin,
                        aum_crore=net_assets_cr,
                        as_of_month=as_of_month,
                        source=amc_name,
                        run_id=run_id,
                    )
                    .on_conflict_do_update(
                        index_elements=["isin", "as_of_month"],
                        set_={
                            "aum_crore": pg_insert(MfAumHistory).excluded.aum_crore,
                            "source": pg_insert(MfAumHistory).excluded.source,
                            "run_id": pg_insert(MfAumHistory).excluded.run_id,
                            "ingested_at": func.now(),
                        },
                    )
                )
                await db.execute(stmt)
        if aum_map:
            await db.commit()

    return rows_upserted, aum_updates


# ---------------------------------------------------------------------------
# Kite Connect MF instrument enrichment (ADR-0033 extension)
# ---------------------------------------------------------------------------
# Redis keys for the daily Kite access_token cache + refresh lock.
# Token lifetime: Kite invalidates all tokens at 06:00 IST each morning.
_KITE_TOKEN_KEY = "kite:mf:access_token"
_KITE_LOCK_KEY = "kite:mf:token_lock"


async def _kite_login_and_get_token() -> str:
    """TOTP-automated Zerodha login → fresh Kite Connect access_token.

    Flow: POST /api/login → POST /api/twofa (TOTP) → follow redirects to
    extract request_token from callback URL → generate_session.
    """
    from urllib.parse import parse_qs, urlparse

    import pyotp
    from kiteconnect import KiteConnect

    from dhanradar.config import settings as _s

    kite = KiteConnect(api_key=_s.KITE_API_KEY)
    base = "https://kite.zerodha.com"

    async with httpx.AsyncClient(follow_redirects=False, timeout=30.0) as client:
        # 1. POST credentials
        r1 = await client.post(
            f"{base}/api/login",
            data={"user_id": _s.KITE_USER_ID, "password": _s.KITE_USER_PASSWORD},
        )
        r1.raise_for_status()
        d1 = r1.json()
        if d1.get("status") != "success":
            raise RuntimeError(f"Kite login failed: {d1.get('message')}")
        request_id = d1["data"]["request_id"]

        # 2. POST TOTP 2FA
        totp_code = pyotp.TOTP(_s.KITE_TOTP_SECRET).now()
        r2 = await client.post(
            f"{base}/api/twofa",
            data={
                "request_id": request_id,
                "twofa_value": totp_code,
                "twofa_type": "app_code",
                "user_id": _s.KITE_USER_ID,
            },
        )
        r2.raise_for_status()
        d2 = r2.json()
        if d2.get("status") != "success":
            raise RuntimeError(f"Kite 2FA failed: {d2.get('message')}")

        # 3. Follow redirect chain from connect/login (with active session cookies)
        #    until the callback URL carries ?request_token=...
        location: str | None = f"{base}/connect/login?api_key={_s.KITE_API_KEY}&v=3"
        request_token: str | None = None
        for _ in range(8):
            if not location:
                break
            qs = parse_qs(urlparse(location).query)
            if "request_token" in qs:
                request_token = qs["request_token"][0]
                break
            rn = await client.get(location)
            location = rn.headers.get("location")

        if request_token is None:
            raise RuntimeError("Kite TOTP login: request_token not found in redirect chain")

        # 4. Exchange for access_token (synchronous SDK call is fine here)
        session = kite.generate_session(request_token, api_secret=_s.KITE_API_SECRET)
        return str(session["access_token"])


async def _kite_get_access_token() -> str:
    """Return a valid Kite access_token from Redis cache; refresh via TOTP if absent.

    TTL is computed to 06:00 IST next morning (Kite's daily invalidation boundary).
    TOCTOU guard: one worker holds a 60-second NX lock; others wait then re-read.
    """
    from dhanradar.redis_client import get_redis

    redis = get_redis()

    cached = await redis.get(_KITE_TOKEN_KEY)
    if cached:
        return cached.decode() if isinstance(cached, bytes) else cached

    # Acquire refresh lock so only one worker calls Zerodha at a time.
    lock_id = os.urandom(8).hex()
    acquired = await redis.set(_KITE_LOCK_KEY, lock_id, nx=True, ex=60)
    if not acquired:
        await asyncio.sleep(6)
        val = await redis.get(_KITE_TOKEN_KEY)
        if val:
            return val.decode() if isinstance(val, bytes) else val
        raise RuntimeError("Kite token refresh lock held too long; retry later")

    try:
        token = await _kite_login_and_get_token()
        # TTL: seconds until next 06:00 IST (minimum 5 min for safety).
        now_ist = datetime.now(tz=timezone(timedelta(hours=5, minutes=30)))
        next_reset = now_ist.replace(hour=6, minute=0, second=0, microsecond=0)
        if now_ist.hour >= 6:
            next_reset += timedelta(days=1)
        ttl = max(int((next_reset - now_ist).total_seconds()), 300)
        await redis.set(_KITE_TOKEN_KEY, token, ex=ttl)
        return token
    finally:
        val = await redis.get(_KITE_LOCK_KEY)
        if val and (val.decode() if isinstance(val, bytes) else val) == lock_id:
            await redis.delete(_KITE_LOCK_KEY)


def _kite_norm_name(name: str) -> str:
    """Lowercase + collapse whitespace for scheme-name matching."""
    import re

    return re.sub(r"\s+", " ", name.lower().strip())


@celery_app.task(name="dhanradar.tasks.mf.mf_kite_enrich")
def mf_kite_enrich() -> str:
    """Enrich mf_funds.plan_type / option_type via Kite MF instruments for NULL rows.

    Access token is refreshed automatically via TOTP when expired (cached in Redis).
    No-op when KITE_* credentials are not configured.
    """
    try:
        return asyncio.run(_mf_kite_enrich_pipeline())
    except Exception:  # noqa: BLE001
        logger.exception("mf_kite_enrich pipeline error")
        return "mf_kite_enrich: failed — see worker logs"


async def _mf_kite_enrich_pipeline() -> str:
    from sqlalchemy import text as sa_text

    from dhanradar.config import settings as _s
    from dhanradar.db import TaskSessionLocal

    if not all(
        [
            _s.KITE_API_KEY,
            _s.KITE_API_SECRET,
            _s.KITE_TOTP_SECRET,
            _s.KITE_USER_ID,
            _s.KITE_USER_PASSWORD,
        ]
    ):
        logger.info("mf_kite_enrich: Kite credentials not configured, skipping")
        return "mf_kite_enrich: skipped (credentials not configured)"

    # Fetch MF instruments from Kite (blocking SDK call; only one coroutine running).
    from kiteconnect import KiteConnect

    access_token = await _kite_get_access_token()
    kite = KiteConnect(api_key=_s.KITE_API_KEY)
    kite.set_access_token(access_token)
    instruments: list[dict] = kite.mf_instruments()  # type: ignore[assignment]

    # Build normalised-name → {plan_type, option_type} lookup from Kite data.
    kite_lookup: dict[str, dict] = {}
    for inst in instruments:
        name = (inst.get("name") or "").strip()
        if not name:
            continue
        scheme_type = str(inst.get("scheme_type") or "").lower()
        div_type = str(inst.get("dividend_type") or "").lower()
        plan_type = "direct" if "direct" in scheme_type else "regular"
        option_type = "growth" if "growth" in div_type else "idcw"
        kite_lookup[_kite_norm_name(name)] = {
            "plan_type": plan_type,
            "option_type": option_type,
        }

    if not kite_lookup:
        return "mf_kite_enrich: no instruments returned from Kite"

    # Load mf_funds rows with any NULL plan/option field.
    async with TaskSessionLocal() as db:
        result = await db.execute(
            sa_text(
                "SELECT isin, scheme_name FROM mf.mf_funds"
                " WHERE plan_type IS NULL OR option_type IS NULL"
            )
        )
        null_rows: list[tuple[str, str]] = result.fetchall()

    if not null_rows:
        return "mf_kite_enrich: no NULL rows to enrich"

    # Match by normalised scheme name; only enrich where we have a confident hit.
    updates: list[dict] = []
    for isin, scheme_name in null_rows:
        match = kite_lookup.get(_kite_norm_name(scheme_name))
        if match:
            updates.append({"isin": isin, **match})

    if updates:
        async with TaskSessionLocal() as db:
            for chunk in [updates[i : i + 500] for i in range(0, len(updates), 500)]:
                for row in chunk:
                    await db.execute(
                        sa_text(
                            "UPDATE mf.mf_funds"
                            " SET plan_type = :plan_type, option_type = :option_type"
                            " WHERE isin = :isin"
                            "   AND (plan_type IS NULL OR option_type IS NULL)"
                        ),
                        row,
                    )
            await db.commit()

    logger.info(
        "mf_kite_enrich complete",
        null_rows=len(null_rows),
        kite_instruments=len(kite_lookup),
        enriched=len(updates),
    )
    return (
        f"mf_kite_enrich: enriched {len(updates)}/{len(null_rows)} NULL rows"
        f" ({len(kite_lookup)} Kite instruments)"
    )


# ---------------------------------------------------------------------------
# M2.2 — daily portfolio valuation series
# ---------------------------------------------------------------------------


@celery_app.task(
    name="dhanradar.tasks.mf.compute_portfolio_daily_valuations",
    bind=True,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_backoff_max=300,
    retry_jitter=True,
    max_retries=2,
)
def compute_portfolio_daily_valuations(self: Any) -> str:  # type: ignore[override]
    """Nightly task (04:00 IST) — write one `mf_portfolio_daily_values` row per portfolio.

    For each portfolio that has holdings, fetch the latest available NAV for each
    ISIN from `mf_nav_history` (the most recent row, which after the 23:30 NAV
    fetch is today's date), compute total market value, and upsert.

    Idempotent: ON CONFLICT (portfolio_id, valuation_date) DO UPDATE refreshes
    total_value and total_invested, so a re-run on the same day is safe.

    Runs AFTER:
      - mf-nav-daily-fetch    (23:30 IST) — fresh NAV available
      - mf-daily-portfolio-refresh (01:30 IST) — cached reports rebuilt

    DPDP: only portfolio_id + aggregate ₹ totals are written.  No ISIN-level
    values leave this task.  Log messages use portfolio_id only (not user_id,
    which is a personal identifier).
    """
    return asyncio.run(_compute_portfolio_daily_valuations_async())


async def _compute_portfolio_daily_valuations_async() -> str:
    from sqlalchemy import select, text
    from sqlalchemy.dialects.postgresql import insert as pg_insert

    from dhanradar.db import admin_task_session
    from dhanradar.mf.valuation import compute_daily_value
    from dhanradar.models.mf import MfPortfolio, MfPortfolioDailyValue, MfUserHolding

    today = date.today()
    upserted = 0
    skipped = 0
    errors = 0

    async with admin_task_session() as db:
        # Fetch all portfolio ids (BYPASSRLS — spans all users).
        portfolio_rows = (await db.execute(select(MfPortfolio.id, MfPortfolio.user_id))).all()

    for portfolio_id, user_id in portfolio_rows:
        try:
            async with admin_task_session() as db:
                # S20 (§39.4): same per-portfolio lock the CAS pipeline takes before its write
                # transaction — serializes this nightly upsert against a concurrent upload.
                await db.execute(
                    text("SELECT pg_advisory_xact_lock(hashtext(:pid))"), {"pid": str(portfolio_id)}
                )

                # Holdings for this portfolio.
                holdings = (
                    await db.execute(
                        select(
                            MfUserHolding.isin, MfUserHolding.units, MfUserHolding.invested_amount
                        ).where(MfUserHolding.portfolio_id == portfolio_id)
                    )
                ).all()

                if not holdings:
                    skipped += 1
                    continue

                isins = [h.isin for h in holdings]

                # Latest NAV for each ISIN.
                nav_rows = await db.execute(
                    text(
                        "SELECT DISTINCT ON (isin) isin, nav"
                        " FROM mf.mf_nav_history"
                        " WHERE isin = ANY(:isins)"
                        " ORDER BY isin, nav_date DESC"
                    ),
                    {"isins": isins},
                )
                nav_map: dict[str, float] = {r.isin: float(r.nav) for r in nav_rows}

                total_invested = sum(float(h.invested_amount or 0) for h in holdings)
                units_nav_pairs = [
                    (float(h.units or 0), nav_map.get(h.isin, 0.0)) for h in holdings
                ]

                point = compute_daily_value(units_nav_pairs, total_invested, today)

                # Upsert: update value + invested if a row for today already exists.
                stmt = (
                    pg_insert(MfPortfolioDailyValue)
                    .values(
                        portfolio_id=portfolio_id,
                        user_id=user_id,
                        valuation_date=point.valuation_date,
                        total_value=point.total_value,
                        total_invested=point.total_invested,
                    )
                    .on_conflict_do_update(
                        constraint="uq_mf_portfolio_daily_value",
                        set_={
                            "total_value": point.total_value,
                            "total_invested": point.total_invested,
                            "computed_at": text("now()"),
                        },
                    )
                )
                await db.execute(stmt)
                await db.commit()
                upserted += 1

        except Exception:
            errors += 1
            logger.exception(
                "compute_portfolio_daily_valuations: error for portfolio %s", portfolio_id
            )

    logger.info(
        "compute_portfolio_daily_valuations complete",
        date=today.isoformat(),
        upserted=upserted,
        skipped=skipped,
        errors=errors,
    )
    return (
        f"compute_portfolio_daily_valuations {today}: "
        f"upserted={upserted} skipped={skipped} errors={errors}"
    )


# ---------------------------------------------------------------------------
# Benchmark price-index daily-close series (ADR-0037 — benchmark for chart;
# generalized to a category-benchmark registry, item 3 2026-07)
# ---------------------------------------------------------------------------

#: Benchmark key stored in mf_benchmark_daily.benchmark (ADR-0037 part b).
BENCHMARK_KEY_NIFTY50 = "nifty50_price"

#: Yahoo Finance ticker for the Nifty 50 price index.
_NIFTY50_TICKER = "^NSEI"


@dataclass(frozen=True)
class BenchmarkSpec:
    """One benchmark index's Yahoo source + public metadata.

    ``storage_key`` is the value written to ``mf_benchmark_daily.benchmark`` —
    nifty50 keeps its legacy ``BENCHMARK_KEY_NIFTY50`` value ("nifty50_price",
    ADR-0037) so existing rows stay valid; every other benchmark uses its own
    registry key as the storage key (no legacy rows to reconcile).
    """

    yahoo_symbol: str
    display_name: str
    storage_key: str


#: Category-benchmark registry (item 3, 2026-07). Each Yahoo symbol was
#: live-verified against query1.finance.yahoo.com/v8/finance/chart/<symbol>
#: (real historical daily closes, not just a same-day snapshot) before being
#: added here. Nifty Smallcap 250 has NO working historical-series symbol on
#: Yahoo — every candidate (NIFTYSMLCAP250.NS, NIFTYSMCAP250.NS, ^CNXSC
#: [= Smallcap 100, wrong index], NIFTY_SMLCAP_250.NS) returns either 404 or a
#: single-point stub (validRanges=['1d','5d'], firstTradeDate=null) — so it was
#: DROPPED; the frontend category map falls back to nifty50 for Small Cap funds.
BENCHMARK_REGISTRY: dict[str, BenchmarkSpec] = {
    "nifty50": BenchmarkSpec("^NSEI", "Nifty 50", BENCHMARK_KEY_NIFTY50),
    "nifty100": BenchmarkSpec("^CNX100", "Nifty 100", "nifty100"),
    "nifty500": BenchmarkSpec("^CRSLDX", "Nifty 500", "nifty500"),
    "nifty_midcap_150": BenchmarkSpec("NIFTYMIDCAP150.NS", "Nifty Midcap 150", "nifty_midcap_150"),
}


def _fetch_index_closes(
    symbol: str,
    start: date,
    end: date,
) -> list[tuple[date, float]]:
    """Synchronous yfinance fetch of `symbol`'s daily closes in [start, end].

    Generalized core shared by every BENCHMARK_REGISTRY entry (and the
    back-compat ^NSEI-only `_fetch_nifty_closes` wrapper below).
    Returns a list of (close_date, close_value) pairs with NaN rows dropped.
    Raises on import failure; returns [] on empty data (weekend / holiday range).
    Kept as a module-level helper so tests can monkeypatch it without touching
    the Celery task boundary.
    """
    import yfinance as yf  # lazy — not loaded at worker startup

    raw = yf.download(
        symbol,
        start=start.isoformat(),
        end=(end + timedelta(days=1)).isoformat(),  # yfinance end is exclusive
        progress=False,
        auto_adjust=False,  # price index; no dividend adjustment needed or wanted
    )
    if raw.empty:
        return []
    close_col = raw["Close"]
    # Recent yfinance returns MultiIndex columns even for a single ticker, so raw["Close"]
    # is a 1-column DataFrame (not a Series) and float(val) below would get a Series.
    # Collapse to the ticker column so each val is a scalar.
    if hasattr(close_col, "columns"):
        close_col = close_col.iloc[:, 0]
    rows: list[tuple[date, float]] = []
    for ts, val in close_col.items():
        fval = float(val)
        if math.isnan(fval):
            continue
        # yfinance index is a pandas Timestamp; .date() gives a stdlib date.
        rows.append((ts.date(), fval))  # type: ignore[union-attr]
    return rows


def _fetch_nifty_closes(
    start: date,
    end: date,
) -> list[tuple[date, float]]:
    """Back-compat wrapper — ^NSEI only. Existing tests import this name directly
    and monkeypatch the `yfinance` module, which still works unchanged since this
    just delegates to `_fetch_index_closes`."""
    return _fetch_index_closes(_NIFTY50_TICKER, start, end)


async def _upsert_benchmark_rows(
    rows: list[tuple[date, float]],
    benchmark: str = BENCHMARK_KEY_NIFTY50,
) -> int:
    """Bulk-upsert (close_date, close_value) rows into mf_benchmark_daily.

    Returns the number of rows upserted.  Idempotent: ON CONFLICT DO UPDATE
    refreshes close_value so a re-run is safe.
    """
    from sqlalchemy.dialects.postgresql import insert as pg_insert

    from dhanradar.db import task_session
    from dhanradar.models.mf import MfBenchmarkDaily

    if not rows:
        return 0

    async with task_session() as db:
        stmt = (
            pg_insert(MfBenchmarkDaily)
            .values([{"benchmark": benchmark, "close_date": d, "close_value": v} for d, v in rows])
            .on_conflict_do_update(
                constraint="uq_mf_benchmark_daily",
                set_={"close_value": pg_insert(MfBenchmarkDaily).excluded.close_value},
            )
        )
        await db.execute(stmt)
        await db.commit()
    return len(rows)


async def _daily_close_upsert_async(key: str) -> str:
    """Fetch + upsert one BENCHMARK_REGISTRY benchmark's latest close (5-day
    catch-up lookback across weekends/holidays, keep most-recent row only —
    idempotent). Shared core for the nifty50 back-compat entry point and the
    registry-wide daily task; network/parse failures are caught so one bad
    symbol doesn't block the rest of the registry.

    The nifty50 label is kept literally "nifty_close_daily" (not the registry
    key) so its return string — and every existing test asserting on it — is
    byte-identical to the pre-registry single-benchmark implementation.
    """
    spec = BENCHMARK_REGISTRY[key]
    label = "nifty_close_daily" if key == "nifty50" else key
    today = date.today()
    start = today - timedelta(days=5)  # catch up across weekends / holidays
    try:
        rows = _fetch_index_closes(spec.yahoo_symbol, start, today)
    except Exception:
        logger.exception("%s: yfinance fetch failed", label)
        return f"{label}: fetch_failed"

    if not rows:
        logger.warning("%s: no closes returned from yfinance for %s", label, today)
        return f"{label}: no_data as_of={today}"

    # Keep only the most-recent close (we re-run daily so yesterday's row
    # is already stored from the previous run).
    latest_date, latest_value = max(rows, key=lambda r: r[0])
    upserted = await _upsert_benchmark_rows(
        [(latest_date, latest_value)], benchmark=spec.storage_key
    )
    logger.info(
        "%s: close_date=%s close_value=%.2f upserted=%d",
        label,
        latest_date,
        latest_value,
        upserted,
    )
    return f"{label}: close_date={latest_date} close_value={latest_value}"


async def _nifty_close_daily_async() -> str:
    """Back-compat nifty50-only entry point — kept for existing direct-import
    tests / external callers. Production now runs through
    `_all_benchmarks_daily_close_async` (below), which processes nifty50 via
    this exact same `_daily_close_upsert_async` helper.

    Runs on the 640 MB batch queue — do NOT call from celery-mood (192 MB).
    """
    return await _daily_close_upsert_async("nifty50")


async def _all_benchmarks_daily_close_async() -> str:
    """Async core for the nifty_close_daily Celery task — iterates every
    BENCHMARK_REGISTRY key (one Yahoo call each, same upsert). Per-benchmark
    failures don't block the rest (mirrors the ticker/indices per-symbol
    tolerance pattern).
    """
    results = [await _daily_close_upsert_async(key) for key in BENCHMARK_REGISTRY]
    return "; ".join(results)


@celery_app.task(
    name="dhanradar.tasks.mf.nifty_close_daily",
    bind=True,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_backoff_max=300,
    retry_jitter=True,
    max_retries=2,
)
def nifty_close_daily(self: Any) -> str:  # type: ignore[override]
    """EOD task (23:45 IST) — fetch each registered benchmark's latest
    price-index close and upsert (ADR-0037; generalized to the registry,
    item 3 2026-07). Task name/beat schedule unchanged.

    Sync Celery wrapper; the async core lives in
    ``_all_benchmarks_daily_close_async`` (mirrors the M2.2
    compute_portfolio_daily_valuations pattern) so tests can await the async
    core without hitting the asyncio.run-inside-running-loop error.
    """
    return asyncio.run(_all_benchmarks_daily_close_async())


async def _backfill_benchmark_async(key: str, years: int = 10) -> str:
    """Bulk-backfill one BENCHMARK_REGISTRY benchmark's daily closes for the
    last `years` years. Idempotent; safe to re-run. Shared core for the
    nifty50 back-compat entry point and the registry-wide 'all' backfill.

    The nifty50 label is kept literally "backfill_nifty_close_series" (not the
    registry key) for byte-identical return strings vs the pre-registry
    implementation.
    """
    spec = BENCHMARK_REGISTRY[key]
    label = "backfill_nifty_close_series" if key == "nifty50" else key
    today = date.today()
    start = today - timedelta(days=years * 365)
    logger.info("%s: start=%s end=%s", label, start, today)
    try:
        rows = _fetch_index_closes(spec.yahoo_symbol, start, today)
    except Exception:
        logger.exception("%s: yfinance fetch failed", label)
        return f"{label}: fetch_failed"

    if not rows:
        return f"{label}: no_data"

    upserted = await _upsert_benchmark_rows(rows, benchmark=spec.storage_key)
    logger.info(
        "%s: upserted=%d range=%s..%s",
        label,
        upserted,
        start,
        today,
    )
    return f"{label}: upserted={upserted} range={start}..{today}"


async def _backfill_nifty_close_series_async(years: int = 10) -> str:
    """Back-compat nifty50-only entry point — kept for existing direct-import
    tests / external callers. Delegates to the generalized per-benchmark
    backfill core."""
    return await _backfill_benchmark_async("nifty50", years)


async def _backfill_all_benchmarks_async(years: int = 10) -> str:
    """Backfill every BENCHMARK_REGISTRY key. Per-benchmark failures don't
    block the rest (mirrors the daily task's tolerance)."""
    results = [await _backfill_benchmark_async(key, years) for key in BENCHMARK_REGISTRY]
    return "; ".join(results)


def backfill_nifty_close_series(benchmark: str = "all", years: int = 10) -> str:
    """One-time backfill of benchmark daily closes.  Sync wrapper — calls the
    async core.

    `benchmark`: a BENCHMARK_REGISTRY key (e.g. "nifty50", "nifty100"), or the
    default "all" to backfill every registered benchmark.

    Example (KVM4 shell):
        python -c "from dhanradar.tasks.mf import backfill_nifty_close_series; print(backfill_nifty_close_series('nifty50'))"
    Or via Celery:
        celery -A dhanradar.celery_app call dhanradar.tasks.mf.backfill_nifty_close_series_task
    """
    if benchmark == "all":
        return asyncio.run(_backfill_all_benchmarks_async(years))
    if benchmark not in BENCHMARK_REGISTRY:
        return f"backfill_nifty_close_series: unknown_benchmark={benchmark}"
    return asyncio.run(_backfill_benchmark_async(benchmark, years))


@celery_app.task(
    name="dhanradar.tasks.mf.backfill_nifty_close_series_task",
    bind=False,
)
def backfill_nifty_close_series_task(benchmark: str = "all", years: int = 10) -> str:
    """Celery-triggerable wrapper around backfill_nifty_close_series.

    Invoke via the admin ops console or directly:
        celery -A dhanradar.celery_app call dhanradar.tasks.mf.backfill_nifty_close_series_task
    """
    return backfill_nifty_close_series(benchmark=benchmark, years=years)
